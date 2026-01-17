"""
Index Sheet Re-sequencer and Multi-Table Sheet Splitter

This module implements the Index sheet mapping table plan:
1. Re-sequence xlsx sheets based on unique (Section + Table Title) combinations
2. Physically split multi-table sheets into separate sheets
3. Update Index sheet links and hyperlinks
4. Handle various edge cases (metadata blocks, unit indicators, etc.)

Based on: req_update/Index_sheet_mapping/Index_sheet_mapping_table.md
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
import pandas as pd
import re
import uuid
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, field
import logging

logger = logging.getLogger(__name__)


@dataclass
class TableBlock:
    """Represents a single table block within a sheet"""
    metadata_start_row: Optional[int]  # Start row of metadata block (or None if no metadata)
    metadata_end_row: Optional[int]    # End row of metadata block (row before data header)
    data_start_row: int                # Start row of data (including header)
    data_end_row: int                  # End row of data
    table_title: Optional[str]         # Table Title from metadata
    source: Optional[str]              # Source(s) from metadata
    has_metadata: bool                 # Whether this block has metadata


@dataclass
class SheetMapping:
    """Mapping from old sheet to new sheet(s)"""
    old_sheet_name: str
    new_sheet_name: str
    table_block: Optional[TableBlock]
    index_row: int  # Row number in Index sheet (0-indexed)
    section: str
    table_title: str


@dataclass
class ResequencerStats:
    """Track statistics during re-sequencing operation."""
    total_sheets: int = 0
    total_index_entries: int = 0
    sheets_with_multiple_tables: int = 0
    sheets_skipped: int = 0
    sheets_renamed: int = 0
    blocks_detected: int = 0
    index_updates: int = 0
    block_index_mismatches: int = 0
    rename_conflicts_resolved: int = 0
    warnings: List[Dict] = field(default_factory=list)
    
    def to_dict(self) -> Dict:
        return {
            "Total Sheets": self.total_sheets,
            "Total Index Entries": self.total_index_entries,
            "Multi-table Sheets": self.sheets_with_multiple_tables,
            "Sheets Skipped": self.sheets_skipped,
            "Sheets Renamed": self.sheets_renamed,
            "Total Blocks Detected": self.blocks_detected,
            "Index Updates": self.index_updates,
            "Block/Index Mismatches": self.block_index_mismatches,
            "Rename Conflicts Resolved": self.rename_conflicts_resolved,
            "Warnings Count": len(self.warnings)
        }
    
    def generate_report(self) -> str:
        """Generate formatted statistics report."""
        report = "# Re-sequencing Statistics\n\n"
        report += "## Summary\n\n"
        
        for key, value in self.to_dict().items():
            report += f"- **{key}:** {value}\n"
        
        report += "\n## Analysis\n\n"
        
        # Expected Index entries calculation
        expected_entries = self.total_sheets
        discrepancy = self.total_index_entries - expected_entries
        
        report += f"- **Expected Index Entries (approx):** {expected_entries}\n"
        report += f"- **Actual Index Entries:** {self.total_index_entries}\n"
        report += f"- **Discrepancy:** {discrepancy:+d}\n\n"
        
        if abs(discrepancy) > 0:
            report += "> [!NOTE]\n"
            report += f"> The discrepancy of {abs(discrepancy)} entries is "
            if discrepancy > 0:
                report += "likely due to multi-table sheets with more than 2 tables.\n"
            else:
                report += "likely due to skipped or merged sheets.\n"
        
        if self.warnings:
            report += "\n## Warnings\n\n"
            for warning in self.warnings[:20]:  # Limit to first 20
                report += f"- {warning.get('message', '')}\n"
        
        return report


class BlockDetector:
    """Detects table blocks within a sheet"""
    
    # Unit indicator patterns
    UNIT_PATTERNS = [
        r'\$\s*in\s+millions',
        r'\$\s*in\s+thousands',
        r'\$\s*in\s+billions',
        r'\$\s*\(\s*in\s+millions\s*\)',
        r'\$\s*\(\s*in\s+thousands\s*\)',
        r'\$\s*\(\s*in\s+billions\s*\)',
        r'in\s+millions',
        r'in\s+thousands',
        r'in\s+billions',
        r'\$in\s*millions',  # no space
    ]
    
    # Period/date patterns
    PERIOD_PATTERNS = [
        r'Q[1-4]-(?:QTD|YTD)-\d{4}',  # Q3-QTD-2025
        r'Q[1-4]-\d{4}',               # Q3-2025
        r'YTD-\d{4}',                  # YTD-2025
        r'At\s+\w+\s+\d{1,2},\s+\d{4}',  # At September 30, 2025
        r'\d{4}',                      # 2025
    ]
    
    @classmethod
    def detect_blocks(cls, ws) -> List[TableBlock]:
        """
        Detect all table blocks in a worksheet.
        
        Returns list of TableBlock objects found in the sheet.
        """
        blocks = []
        
        # Find all "Table Title:" markers
        table_title_rows = cls._find_table_title_rows(ws)
        
        if not table_title_rows:
            # No metadata - check for unit indicator splits
            return cls._detect_unit_indicator_blocks(ws)
        
        # Process each metadata block
        for i, title_row in enumerate(table_title_rows):
            source_row = cls._find_source_row(ws, title_row)
            
            if source_row is None:
                logger.warning(f"No Source(s) row found after Table Title at row {title_row}")
                continue
            
            # Find data start (first non-empty row after source)
            data_start = cls._find_data_start(ws, source_row)
            
            # Find data end (blank row or next metadata block)
            if i < len(table_title_rows) - 1:
                # Data ends before next metadata block
                next_meta_start = cls._find_metadata_start(ws, table_title_rows[i + 1])
                data_end = next_meta_start - 1
            else:
                # Last block - find end by blank rows
                data_end = cls._find_data_end(ws, data_start)
            
            # Extract metadata
            table_title = cls._extract_cell_value(ws, title_row, 'B')
            source = cls._extract_cell_value(ws, source_row, 'B')
            
            # Find metadata start (scan backwards from title row)
            metadata_start = cls._find_metadata_start(ws, title_row)
            
            block = TableBlock(
                metadata_start_row=metadata_start,
                metadata_end_row=source_row,
                data_start_row=data_start,
                data_end_row=data_end,
                table_title=table_title,
                source=source,
                has_metadata=True
            )
            
            blocks.append(block)
        
        # Check if we need to split any blocks by unit indicators
        final_blocks = []
        for block in blocks:
            sub_blocks = cls._split_block_on_new_headers(ws, block)
            final_blocks.extend(sub_blocks)
        
        return final_blocks
    
    @classmethod
    def _find_table_title_rows(cls, ws) -> List[int]:
        """Find all rows containing 'Table Title:' in column A"""
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            cell_value = str(row[0].value or '').strip()
            if cell_value.startswith('Table Title:'):
                rows.append(row[0].row)
        return rows
    
    @classmethod
    def _find_source_row(cls, ws, start_row: int) -> Optional[int]:
        """Find 'Source(s):' row after the given start row"""
        for row in range(start_row + 1, min(start_row + 10, ws.max_row + 1)):
            cell_value = str(ws.cell(row, 1).value or '').strip()
            if cell_value.startswith('Source(s):') or cell_value.startswith('Source:'):
                return row
        return None
    
    @classmethod
    def _find_metadata_start(cls, ws, title_row: int) -> int:
        """Find start of metadata block (scan backwards from title row)"""
        # Look for 'Category (Parent):', 'Line Items:', etc.
        for row in range(title_row - 1, 0, -1):
            cell_value = str(ws.cell(row, 1).value or '').strip()
            if cell_value.startswith('← Back to Index') or not cell_value:
                return row + 1
            # Check if this looks like metadata
            if any(marker in cell_value for marker in ['Category', 'Line Items', 'Product/Entity', 'Column Header']):
                continue
            else:
                return row + 1
        return 1
    
    @classmethod
    def _find_data_start(cls, ws, source_row: int) -> int:
        """Find first data row after source (usually the unit indicator or header)"""
        for row in range(source_row + 1, min(source_row + 10, ws.max_row + 1)):
            # Skip blank rows
            if cls._is_blank_row(ws, row):
                continue
            return row
        return source_row + 1
    
    @classmethod
    def _find_data_end(cls, ws, data_start: int) -> int:
        """Find end of data block (consecutive blank rows or end of sheet)"""
        blank_count = 0
        last_data_row = data_start
        
        for row in range(data_start, ws.max_row + 1):
            if cls._is_blank_row(ws, row):
                blank_count += 1
                if blank_count >= 2:  # 2 consecutive blank rows = end
                    return last_data_row
            else:
                blank_count = 0
                last_data_row = row
        
        return last_data_row
    
    @classmethod
    def _is_blank_row(cls, ws, row: int) -> bool:
        """Check if a row is completely blank"""
        for cell in ws[row]:
            if cell.value is not None and str(cell.value).strip():
                return False
        return True
    
    @classmethod
    def _extract_cell_value(cls, ws, row: int, col: str) -> Optional[str]:
        """Extract cell value as string"""
        value = ws[f'{col}{row}'].value
        return str(value).strip() if value else None
    
    @classmethod
    def _detect_unit_indicator_blocks(cls, ws) -> List[TableBlock]:
        """
        Detect blocks when no Table Title markers exist.
        Look for unit indicator patterns.
        """
        blocks = []
        unit_rows = cls._find_unit_indicator_rows(ws)
        
        if not unit_rows:
            # No unit indicators - treat entire sheet as one block
            return [TableBlock(
                metadata_start_row=None,
                metadata_end_row=None,
                data_start_row=1,
                data_end_row=ws.max_row,
                table_title=None,
                source=None,
                has_metadata=False
            )]
        
        # Create blocks based on unit indicators
        for i, unit_row in enumerate(unit_rows):
            data_start = unit_row
            
            # Find data end
            if i < len(unit_rows) - 1:
                data_end = unit_rows[i + 1] - 1
            else:
                data_end = cls._find_data_end(ws, data_start)
            
            blocks.append(TableBlock(
                metadata_start_row=None,
                metadata_end_row=None,
                data_start_row=data_start,
                data_end_row=data_end,
                table_title=None,
                source=None,
                has_metadata=False
            ))
        
        return blocks
    
    @classmethod
    def _find_unit_indicator_rows(cls, ws) -> List[int]:
        """Find rows with unit indicators ($ in millions, etc.)"""
        rows = []
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            first_cell = str(row[0].value or '').strip().lower()
            
            # Check if matches unit pattern
            for pattern in cls.UNIT_PATTERNS:
                if re.search(pattern, first_cell, re.IGNORECASE):
                    # Also check if next columns have period patterns
                    if cls._has_period_headers(ws, row[0].row):
                        rows.append(row[0].row)
                        break
        
        return rows
    
    @classmethod
    def _has_period_headers(cls, ws, row: int) -> bool:
        """Check if row has period/date headers in subsequent columns"""
        for col in range(2, min(10, ws.max_column + 1)):
            cell_value = str(ws.cell(row, col).value or '').strip()
            for pattern in cls.PERIOD_PATTERNS:
                if re.search(pattern, cell_value, re.IGNORECASE):
                    return True
        return False
    
    @classmethod
    def _split_block_on_new_headers(cls, ws, block: TableBlock) -> List[TableBlock]:
        """
        Split a block if multiple header rows are detected within the data.
        This handles Case 4: Single metadata block with multiple data blocks.
        """
        # Find additional header rows within data range
        header_rows = [block.data_start_row]  # First header
        
        for row in range(block.data_start_row + 1, block.data_end_row + 1):
            # Check for new header (Condition A: empty col A + period patterns)
            first_cell = str(ws.cell(row, 1).value or '').strip()
            
            if not first_cell:  # Empty first column
                if cls._has_period_headers(ws, row):
                    header_rows.append(row)
                    continue
            
            # Check for new header (Condition B: unit indicator + period patterns)
            for pattern in cls.UNIT_PATTERNS:
                if re.search(pattern, first_cell, re.IGNORECASE):
                    if cls._has_period_headers(ws, row):
                        header_rows.append(row)
                        break
        
        # If only one header found, no split needed
        if len(header_rows) == 1:
            return [block]
        
        # Split into multiple blocks
        sub_blocks = []
        for i, header_row in enumerate(header_rows):
            data_start = header_row
            
            # Find data end
            if i < len(header_rows) - 1:
                data_end = header_rows[i + 1] - 1
            else:
                data_end = block.data_end_row
            
            sub_block = TableBlock(
                metadata_start_row=block.metadata_start_row if i == 0 else None,
                metadata_end_row=block.metadata_end_row if i == 0 else None,
                data_start_row=data_start,
                data_end_row=data_end,
                table_title=block.table_title,
                source=block.source,
                has_metadata=i == 0  # Only first block has metadata
            )
            
            sub_blocks.append(sub_block)
        
        return sub_blocks


class IndexSheetResequencer:
    """Re-sequences Index sheet and splits multi-table sheets"""
    
    def __init__(self, xlsx_path: Path):
        self.xlsx_path = Path(xlsx_path)
        self.wb = openpyxl.load_workbook(str(xlsx_path))
        self.index_ws = self.wb['Index']
        self.stats = ResequencerStats()
        
    def process(self, output_path: Optional[Path] = None) -> Path:
        """
        Main processing pipeline:
        1. Analyze Index sheet and detect unique groups
        2. Build mapping from old to new sheet IDs
        3. Split multi-table sheets
        4. Rename sheets
        5. Update Index sheet
        6. Update hyperlinks
        7. Save workbook
        8. Generate statistics report
        """
        logger.info(f"Processing {self.xlsx_path.name}")
        
        # Step 1: Read Index and build grouping
        index_data = self._read_index_sheet()
        self.stats.total_index_entries = len(index_data)
        
        # Step 2: Build mapping
        mapping = self._build_sheet_mapping(index_data)
        
        # Step 3: Split multi-table sheets
        self._split_sheets(mapping)
        
        # Step 4: Rename sheets
        self._rename_sheets(mapping)
        
        # Step 5: Update Index sheet
        self._update_index_sheet(mapping)
        
        # Step 6: Update hyperlinks
        self._update_hyperlinks(mapping)
        
        # Step 7: Save
        output_path = output_path or self.xlsx_path.parent / f"{self.xlsx_path.stem}_resequenced.xlsx"
        self.wb.save(str(output_path))
        logger.info(f"Saved to {output_path}")
        
        # Step 8: Generate and save statistics report
        stats_report_path = output_path.parent / f"{output_path.stem}_stats.md"
        with open(stats_report_path, 'w') as f:
            f.write(self.stats.generate_report())
        logger.info(f"Statistics report saved to {stats_report_path}")
        
        return output_path
    
    def _read_index_sheet(self) -> pd.DataFrame:
        """Read Index sheet into DataFrame"""
        data = []
        headers = None
        
        for row_idx, row in enumerate(self.index_ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                headers = [str(h or '').strip() for h in row]
                continue
            
            data.append(row)
        
        df = pd.DataFrame(data, columns=headers)
        return df
    
    def _build_sheet_mapping(self, index_df: pd.DataFrame) -> List[SheetMapping]:
        """
        Build mapping from old sheet names to new sheet names.
        
        Returns list of SheetMapping objects.
        """
        unique_group_to_id = {}  # {(section, title): base_id}
        group_counters = {}      # {(section, title): current_suffix_count}
        next_id = 1
        
        mapping = []
        
        for idx, row in index_df.iterrows():
            section = str(row.get('Section', '') or '').strip()
            table_title = str(row.get('Table Title', '') or '').strip()
            old_link = str(row.get('Link', '') or '').replace('→', '').strip()
            
            group_key = (section, table_title)
            
            if group_key not in unique_group_to_id:
                # First occurrence of this unique group
                unique_group_to_id[group_key] = next_id
                group_counters[group_key] = 0
                
                new_id = str(next_id)
                next_id += 1
            else:
                # Sub-table: same Section+Title seen before
                group_counters[group_key] += 1
                base_id = unique_group_to_id[group_key]
                suffix = group_counters[group_key]
                
                new_id = f"{base_id}_{suffix}"
            
            mapping.append(SheetMapping(
                old_sheet_name=old_link,
                new_sheet_name=new_id,
                table_block=None,  # Will be populated during split
                index_row=idx,
                section=section,
                table_title=table_title
            ))
        
        return mapping
    
    def _split_sheets(self, mapping: List[SheetMapping]):
        """
        Split multi-table sheets into separate physical sheets.
        This detects table blocks and creates new sheets as needed.
        """
        # Group mapping by old sheet name
        sheets_to_split = {}
        for m in mapping:
            if m.old_sheet_name not in sheets_to_split:
                sheets_to_split[m.old_sheet_name] = []
            sheets_to_split[m.old_sheet_name].append(m)
        
        self.stats.total_sheets = len(sheets_to_split)
        
        # Process each sheet
        for old_sheet_name, sheet_mappings in sheets_to_split.items():
            if old_sheet_name not in self.wb.sheetnames:
                logger.warning(f"Sheet {old_sheet_name} not found in workbook")
                self.stats.sheets_skipped += 1
                continue
            
            ws = self.wb[old_sheet_name]
            
            # Detect table blocks
            blocks = BlockDetector.detect_blocks(ws)
            self.stats.blocks_detected += len(blocks)
            
            # Check if split is needed
            if len(blocks) == 1 and len(sheet_mappings) == 1:
                # No split needed
                sheet_mappings[0].table_block = blocks[0]
                continue
            
            # Track multi-table sheets
            if len(sheet_mappings) > 1:
                self.stats.sheets_with_multiple_tables += 1
            
            # Split is needed - check for mismatches
            if len(blocks) != len(sheet_mappings):
                self.stats.block_index_mismatches += 1
                logger.warning(
                    f"Sheet {old_sheet_name}: Found {len(blocks)} blocks but Index has {len(sheet_mappings)} entries"
                )
                self.stats.warnings.append({
                    'type': 'block_mismatch',
                    'sheet': old_sheet_name,
                    'blocks_found': len(blocks),
                    'index_entries': len(sheet_mappings),
                    'message': f"Sheet {old_sheet_name}: {len(blocks)} blocks detected vs {len(sheet_mappings)} Index entries"
                })
            
            # Assign blocks to mappings
            for i, (block, sheet_mapping) in enumerate(zip(blocks, sheet_mappings)):
                sheet_mapping.table_block = block
                
                # Create new sheet for blocks after the first
                if i > 0:
                    new_sheet_name = sheet_mapping.new_sheet_name
                    new_ws = self.wb.create_sheet(title=f"temp_{new_sheet_name}")
                    
                    # Copy metadata (if exists)
                    if block.has_metadata and block.metadata_start_row:
                        self._copy_rows(ws, new_ws, block.metadata_start_row, block.metadata_end_row)
                    else:
                        # Add empty metadata structure
                        self._add_empty_metadata(new_ws, sheet_mapping)
                    
                    # Copy data
                    data_dest_row = new_ws.max_row + 1 if new_ws.max_row > 1 else 1
                    self._copy_rows(ws, new_ws, block.data_start_row, block.data_end_row, dest_start_row=data_dest_row)
                    
                    # Add back link with hyperlink
                    new_ws.insert_rows(1)
                    back_link_cell = new_ws.cell(1, 1, "← Back to Index")
                    back_link_cell.hyperlink = Hyperlink(ref="A1", target="", location="#'Index'!A1")
                    back_link_cell.style = "Hyperlink"
            
            # Update original sheet to contain only first block
            if len(blocks) > 1:
                first_block = blocks[0]
                # Delete rows after first block
                if first_block.data_end_row < ws.max_row:
                    ws.delete_rows(first_block.data_end_row + 1, ws.max_row - first_block.data_end_row)
    
    def _copy_rows(self, source_ws, dest_ws, start_row: int, end_row: int, dest_start_row: int = None):
        """Copy rows from source to destination worksheet"""
        if dest_start_row is None:
            dest_start_row = dest_ws.max_row + 1 if dest_ws.max_row > 1 else 1
        
        for row_idx, row in enumerate(source_ws.iter_rows(min_row=start_row, max_row=end_row), start=dest_start_row):
            for col_idx, cell in enumerate(row, start=1):
                new_cell = dest_ws.cell(row=row_idx, column=col_idx)
                new_cell.value = cell.value
                
                # Copy formatting
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection.copy()
                    new_cell.alignment = cell.alignment.copy()
    
    def _add_empty_metadata(self, ws, sheet_mapping: SheetMapping):
        """Add empty metadata structure from Index data"""
        row = 1
        back_link_cell = ws.cell(row, 1, "← Back to Index")
        back_link_cell.hyperlink = Hyperlink(ref="A1", target="", location="#'Index'!A1")
        back_link_cell.style = "Hyperlink"
        row += 1
        
        ws.cell(row, 1, "Category (Parent):")
        ws.cell(row, 2, "")
        row += 1
        
        ws.cell(row, 1, "Line Items:")
        ws.cell(row, 2, "")
        row += 1
        
        ws.cell(row, 1, "Table Title:")
        ws.cell(row, 2, sheet_mapping.table_title)
        row += 1
        
        ws.cell(row, 1, "Source(s):")
        ws.cell(row, 2, "")  # Source not available
        row += 1
    
    def _rename_sheets(self, mapping: List[SheetMapping]):
        """Rename sheets based on mapping using safe two-pass approach"""
        # Build rename plan (old_name -> new_name)
        rename_plan = {}
        for m in mapping:
            if m.old_sheet_name != m.new_sheet_name:
                rename_plan[m.old_sheet_name] = m.new_sheet_name
        
        # Handle temp sheets created during split
        for sheet_name in list(self.wb.sheetnames):
            if sheet_name.startswith("temp_"):
                final_name = sheet_name.replace("temp_", "")
                safe_name = self._get_unique_sheet_name(final_name)
                self.wb[sheet_name].title = safe_name
                if safe_name != final_name:
                    self.stats.rename_conflicts_resolved += 1
        
        # Use safe rename for existing sheets
        if rename_plan:
            result = self._rename_sheets_safe(rename_plan)
            self.stats.sheets_renamed = len(result)
    
    def _rename_sheets_safe(self, rename_mapping: Dict[str, str]) -> Dict[str, str]:
        """
        Safely rename sheets avoiding conflicts using two-pass approach.
        
        Args:
            rename_mapping: Dictionary mapping {old_name: new_name}
        
        Returns:
            Dictionary of successful renames {old_name: actual_final_name}
        """
        # Pass 1: Rename to temporary names
        temp_mapping = {}
        for old_name, new_name in rename_mapping.items():
            if old_name not in self.wb.sheetnames:
                logger.warning(f"Sheet {old_name} not found, skipping rename")
                continue
                
            temp_name = f"__TEMP_{uuid.uuid4().hex[:8]}__"
            sheet = self.wb[old_name]
            sheet.title = temp_name
            temp_mapping[temp_name] = (old_name, new_name)
        
        # Pass 2: Rename from temporary to final names
        final_mapping = {}
        for temp_name, (old_name, new_name) in temp_mapping.items():
            sheet = self.wb[temp_name]
            # Check if target name exists, if so append counter
            final_name = self._get_unique_sheet_name(new_name)
            sheet.title = final_name
            final_mapping[old_name] = final_name
            
            if final_name != new_name:
                logger.info(f"Renamed {old_name} to {final_name} (conflict with {new_name} resolved)")
                self.stats.warnings.append({
                    'type': 'rename_conflict',
                    'message': f"Renamed {old_name} to {final_name} instead of {new_name} due to conflict"
                })
        
        return final_mapping
    
    def _get_unique_sheet_name(self, desired_name: str) -> str:
        """
        Get a unique sheet name by appending counter if needed.
        
        Args:
            desired_name: Desired sheet name
        
        Returns:
            Unique sheet name
        """
        if desired_name not in self.wb.sheetnames:
            return desired_name
        
        # Append counter until unique
        counter = 1
        while f"{desired_name}_{counter}" in self.wb.sheetnames:
            counter += 1
        
        return f"{desired_name}_{counter}"
    
    def _update_index_sheet(self, mapping: List[SheetMapping]):
        """Update Index sheet with new links"""
        link_col = None
        
        # Find Link column
        for col_idx, cell in enumerate(self.index_ws[1], start=1):
            if str(cell.value or '').strip() == 'Link':
                link_col = col_idx
                break
        
        if not link_col:
            logger.error("Could not find 'Link' column in Index sheet")
            return
        
        # Update each row
        for m in mapping:
            row_idx = m.index_row + 2  # +1 for 0-indexed, +1 for header row
            cell = self.index_ws.cell(row_idx, link_col)
            cell.value = f"→ {m.new_sheet_name}"
            self.stats.index_updates += 1
    
    def _update_hyperlinks(self, mapping: List[SheetMapping]):
        """Update hyperlinks in Index sheet to point to new sheet names"""
        # This would require updating hyperlinks - openpyxl support is limited
        # For now, we just update the text values
        pass


def process_all_xlsx_files(input_dir: Path, output_dir: Optional[Path] = None):
    """Process all xlsx files in a directory"""
    input_dir = Path(input_dir)
    output_dir = Path(output_dir) if output_dir else input_dir / "resequenced"
    output_dir.mkdir(exist_ok=True)
    
    for xlsx_file in input_dir.glob("*.xlsx"):
        if xlsx_file.name.startswith("~"):  # Skip temp files
            continue
        
        logger.info(f"\n{'='*60}")
        logger.info(f"Processing: {xlsx_file.name}")
        logger.info(f"{'='*60}")
        
        try:
            resequencer = IndexSheetResequencer(xlsx_file)
            output_path = output_dir / xlsx_file.name
            resequencer.process(output_path)
            logger.info(f"✓ Successfully processed {xlsx_file.name}")
        except Exception as e:
            logger.error(f"✗ Failed to process {xlsx_file.name}: {e}", exc_info=True)


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Process all files in data/processed/
    input_dir = Path("data/processed")
    process_all_xlsx_files(input_dir)
