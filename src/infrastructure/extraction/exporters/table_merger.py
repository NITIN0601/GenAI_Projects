"""
Advanced Table Merger for processed Excel files.

Merges tables within the same sheet that share identical row labels (Column A)
into a single horizontally-combined table.

Source: ./data/processed/
Destination: ./data/processed_advanced/
"""

import re
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional, Set
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from copy import copy
from functools import lru_cache

from src.utils import get_logger
from src.core import get_paths
from src.infrastructure.extraction.consolidation.table_grouping import TableGrouper
from src.utils.metadata_labels import MetadataLabels
from src.utils.quarter_mapper import QuarterDateMapper
from src.utils.header_parser import MultiLevelHeaderParser
from src.utils.metadata_builder import MetadataBuilder
from src.utils.financial_domain import (
    TABLE_HEADER_PATTERNS, 
    DATA_LABEL_PATTERNS,
    METADATA_BOUNDARY_MARKERS,
    is_new_table_header_row,
)
# Import from new focused modules
from src.infrastructure.extraction.exporters.block_detection import BlockDetector
from src.infrastructure.extraction.exporters.index_manager import IndexManager
from src.utils.constants import (
    TABLE_FILE_PATTERN,
    TABLE_MERGER_MAX_HEADER_SCAN_ROWS as MAX_HEADER_SCAN_ROWS,
    TABLE_MERGER_MAX_COL_SCAN as MAX_COL_SCAN,
    TABLE_MERGER_MIN_YEAR_COUNT_FOR_HEADER as MIN_YEAR_COUNT_FOR_HEADER,
    YEAR_STRING_LENGTH,
)

logger = get_logger(__name__)


class TableMerger:
    """
    Merge tables with identical row labels within the same Excel sheet.
    
    Features:
    - Detects tables within a sheet that share the same Column A values
    - Merges horizontally (appends columns from matching tables)
    - Only merges if 100% row labels match (entire table must match)
    - Preserves column headers from each table
    """
    
    def __init__(self):
        """Initialize merger with paths."""
        self.paths = get_paths()
        self.source_dir = self.paths.data_dir / "processed"
        self.dest_dir = self.paths.data_dir / "processed_advanced"
        
        # Ensure destination exists
        self.dest_dir.mkdir(parents=True, exist_ok=True)
    
    def _is_sheet_near_empty(self, ws: Worksheet, min_data_rows: int = 3) -> bool:
        """
        Check if a sheet has minimal content (near-empty).
        
        A sheet is considered near-empty if it has fewer than min_data_rows
        of actual data (excluding metadata rows and blank rows).
        
        Args:
            ws: Worksheet to check
            min_data_rows: Minimum data rows required (default: 3)
            
        Returns:
            True if sheet is near-empty
        """
        data_row_count = 0
        
        for row_num in range(1, ws.max_row + 1):
            row_has_data = False
            first_col_val = None
            
            for col_num in range(1, min(ws.max_column + 1, 20)):
                cell_val = ws.cell(row=row_num, column=col_num).value
                if col_num == 1:
                    first_col_val = cell_val
                if cell_val is not None and str(cell_val).strip():
                    row_has_data = True
                    break
            
            if not row_has_data:
                continue
            
            # Skip metadata rows (check first column for metadata labels)
            if first_col_val:
                first_str = str(first_col_val).strip()
                if any(label in first_str for label in [
                    'Category', 'Line Items', 'Product/Entity', 
                    'Column Header', 'Year/Quarter', 'Table Title', 'Source',
                    '← Back to Index'
                ]):
                    continue
            
            data_row_count += 1
            if data_row_count >= min_data_rows:
                return False  # Has enough data
        
        return data_row_count < min_data_rows
    
    def _get_first_block_data(self, ws: Worksheet, first_block: Dict) -> List[List[Any]]:
        """
        Extract all data from the first table block (metadata + data).
        
        Args:
            ws: Source worksheet
            first_block: First table block dict
            
        Returns:
            List of rows (each row is list of cell values)
        """
        rows_data = []
        meta_start = first_block.get('metadata_start_row', first_block.get('source_row', 2))
        data_end = first_block.get('end_row', ws.max_row)
        
        # Include row 1 (Back to Index link)
        for row_num in range(1, data_end + 1):
            row_values = []
            for col_num in range(1, ws.max_column + 1):
                row_values.append(ws.cell(row=row_num, column=col_num).value)
            rows_data.append(row_values)
        
        return rows_data
    
    def process_all_files(self, max_workers: int = 4) -> Dict[str, Any]:
        """
        Process all xlsx files in source directory with parallel processing.
        
        Args:
            max_workers: Maximum number of parallel file processors (default: 4)
        
        Returns:
            Dict with processing results
        """
        from concurrent.futures import ThreadPoolExecutor, as_completed
        
        results = {
            'files_processed': 0,
            'files_with_merges': 0,
            'total_tables_merged': 0,
            'output_files': [],
            'errors': []
        }
        
        xlsx_files = [f for f in self.source_dir.glob(TABLE_FILE_PATTERN) 
                      if not f.name.startswith('~$')]
        
        if not xlsx_files:
            logger.warning(f"No xlsx files found in {self.source_dir}")
            return results
        
        logger.info(f"Processing {len(xlsx_files)} files with {max_workers} workers...")
        
        def process_single_file(xlsx_path: Path) -> Dict[str, Any]:
            """Process a single file and return result dict."""
            try:
                merge_result = self.process_file(xlsx_path)
                return {
                    'success': True,
                    'path': xlsx_path,
                    'result': merge_result
                }
            except Exception as e:
                logger.error(f"Error processing {xlsx_path}: {e}")
                return {
                    'success': False,
                    'path': xlsx_path,
                    'error': str(e)
                }
        
        # Process files in parallel
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_path = {
                executor.submit(process_single_file, path): path 
                for path in xlsx_files
            }
            
            for future in as_completed(future_to_path):
                file_result = future.result()
                
                if file_result['success']:
                    results['files_processed'] += 1
                    merge_result = file_result['result']
                    
                    if merge_result['tables_merged'] > 0:
                        results['files_with_merges'] += 1
                        results['total_tables_merged'] += merge_result['tables_merged']
                    
                    if merge_result['output_path']:
                        results['output_files'].append(merge_result['output_path'])
                else:
                    results['errors'].append(
                        f"{file_result['path'].name}: {file_result['error']}"
                    )
        
        logger.info(f"Parallel processing complete: {results['files_processed']} files, "
                    f"{results['total_tables_merged']} tables merged")
        return results
    
    def process_file(self, source_path: Path) -> Dict[str, Any]:
        """
        Process a single xlsx file.
        
        Args:
            source_path: Path to source xlsx file
            
        Returns:
            Dict with processing result
        """
        result = {
            'source': str(source_path),
            'output_path': None,
            'tables_merged': 0,
            'tables_split': 0,
            'sheets_processed': 0
        }
        
        try:
            # Load workbook
            wb = load_workbook(source_path)
            
            # Track new sheets created during splits
            new_sheets_created = []
            
            # Process each sheet (skip Index)
            # Use list() to avoid modification during iteration
            for sheet_name in list(wb.sheetnames):
                if sheet_name.lower() == 'index':
                    continue
                
                ws = wb[sheet_name]
                merge_count, split_sheets = self._process_sheet(wb, ws, sheet_name)
                result['tables_merged'] += merge_count
                result['tables_split'] += len(split_sheets)
                result['sheets_processed'] += 1
                new_sheets_created.extend(split_sheets)
            
            # NOTE: Index updates are already handled per-sheet by _update_index_for_splits()
            # in _split_non_mergeable_to_new_sheets(). No additional call needed here.
            
            # Save to destination
            output_path = self.dest_dir / source_path.name
            wb.save(output_path)
            result['output_path'] = str(output_path)
            
            # Apply currency/percentage formatting to output file
            try:
                self._apply_number_formatting(output_path)
            except Exception as e:
                logger.debug(f"Could not apply number formatting to {output_path}: {e}")
            
            logger.info(f"Processed {source_path.name}: {result['tables_merged']} merges, {result['tables_split']} splits across {result['sheets_processed']} sheets")
            
        except Exception as e:
            logger.error(f"Failed to process {source_path}: {e}")
            raise
        
        return result
    
    def _process_sheet(self, wb, ws: Worksheet, sheet_name: str) -> tuple:
        """
        Process a single worksheet, merging tables with matching row labels.
        Splits non-mergeable tables to new sheets with their metadata.
        
        Args:
            wb: openpyxl Workbook
            ws: openpyxl Worksheet
            sheet_name: Name of sheet for logging
            
        Returns:
            Tuple of (merge_count, list of new sheet names from splits)
        """
        # Find all table blocks in the sheet
        table_blocks = self._find_table_blocks(ws)
        
        if len(table_blocks) < 2:
            return (0, [])  # Need at least 2 tables to merge/split
        
        logger.info(f"Sheet '{sheet_name}': Found {len(table_blocks)} table blocks")
        
        # --- PHASE 1: VERTICAL MERGING (Stacking split tables) ---
        # Find table parts that should be stacked vertically (e.g. Part 1, Part 2)
        vertical_groups = self._find_vertical_groups(ws, table_blocks)
        
        if vertical_groups:
            logger.info(f"Sheet '{sheet_name}': Found {len(vertical_groups)} vertical merge groups")
            # Perform vertical merge (stacking)
            self._merge_tables_vertically(ws, vertical_groups)
            
            # CRITICAL: Re-scan for blocks because block positions and counts have changed
            table_blocks = self._find_table_blocks(ws)
            if len(table_blocks) < 2:
                 # If we merged everything into one table, we might be done
                 # But we still check if it needs splitting (unlikely for single table)
                 pass
        
        # --- PHASE 2: HORIZONTAL MERGING (Joining periods) ---
        # Find groups of tables that can be merged (same row labels)
        mergeable_groups = self._find_mergeable_tables(table_blocks)
        
        merge_count = 0
        merged_blocks = set()  # Track which blocks were merged
        
        for group in mergeable_groups:
            if len(group) >= 2:
                logger.info(f"Sheet '{sheet_name}': Merging {len(group)} tables with {len(group[0]['row_labels'])} matching rows")
                self._merge_tables_horizontally(ws, group)
                merge_count += len(group) - 1
                
                # Track merged blocks (by source_row which is unique)
                for block in group[1:]:  # Skip first block (it's kept)
                    merged_blocks.add(block['source_row'])
        
        # Split non-mergeable tables to new sheets (instead of just clearing metadata)
        new_sheets = self._split_non_mergeable_to_new_sheets(wb, ws, sheet_name, table_blocks, merged_blocks)
        
        return (merge_count, new_sheets)

    def _extract_block_definition(self, ws: Worksheet, block: Dict) -> Dict[str, Any]:
        """
        Extract definition details from a block for matching.
        
        Extracts:
        - Table Title (Exact)
        - Section (Exact) 
        - Column Headers (List)
        - Metadata Fingerprint (Category + Line Items)
        """
        definition = {
            'title': '',
            'section': '',
            'headers': [],
            'metadata_fingerprint': ''
        }
        
        # Parse metadata rows
        meta_start = block.get('metadata_start_row', block['source_row'])
        meta_end = block['start_row'] - 1
        
        category = ''
        line_items = ''
        
        for row_num in range(meta_start, meta_end + 1):
            cell_val = ws.cell(row=row_num, column=1).value
            if not cell_val:
                continue
            
            val_str = str(cell_val).strip()
            
            if MetadataLabels.TABLE_TITLE in val_str:
                raw_title = val_str.replace(MetadataLabels.TABLE_TITLE, '').strip()
                # Strip (Part N) suffix for matching purposes
                # Example: "Table Title (Part 1)" -> "Table Title"
                definition['title'] = re.sub(r'\s*\(Part\s+\d+\)$', '', raw_title, flags=re.IGNORECASE).strip()
            elif 'Section:' in val_str: # or matching your section label constant
                definition['section'] = val_str.split(':', 1)[1].strip() if ':' in val_str else ''
            elif MetadataLabels.CATEGORY_PARENT in val_str:
                category = val_str.split(':', 1)[1].strip() if ':' in val_str else ''
            elif MetadataLabels.LINE_ITEMS in val_str:
                line_items = val_str.split(':', 1)[1].strip() if ':' in val_str else ''
        
        definition['metadata_fingerprint'] = f"{category}|{line_items}"
        
        # Extract Column Headers (Structure)
        # Use existing helper (we'll ensure it exists or impl it) similar to existing logic
        # Here we look at the specific header row determined by _identify_header_and_data_rows
        # which sets 'data_start_row'. The row before typically has headers.
        header_rows = block.get('header_rows', [])
        if header_rows:
            # Use the last header row for comparison
            # Or perhaps better: collect all header rows to be safe
            all_headers = []
            for h_row in header_rows:
                row_vals = []
                has_dollar = False
                has_text = False
                for col in range(2, min(MAX_COL_SCAN, ws.max_column + 1)):
                    v = ws.cell(row=h_row, column=col).value
                    v_str = str(v).strip() if v else ''
                    row_vals.append(v_str)
                    if '$' in v_str:
                        has_dollar = True
                    # Check for non-numeric text (ignoring empty, years, pipes)
                    # If it has specific words like 'Trading', 'Revenues' they are text.
                    # '2025' is not text in this context (it's year).
                    if v_str and not v_str.isdigit() and not re.match(r'^(19|20)\d{2}$', v_str) and v_str not in ['|', '-']:
                        has_text = True
                
                # Filter out if:
                # 1. Has '$' (likely data)
                # 2. Has NO text (only years/numbers/empty) -> Likely year row or garbage
                if not has_dollar and has_text:
                    all_headers.append('|'.join(row_vals))
                    
            definition['headers'] = all_headers
        
        return definition

    def _find_vertical_groups(self, ws: Worksheet, blocks: List[Dict]) -> List[List[Dict]]:
        """
        Find separate table blocks that should be vertically merged (stacked).
        
        Criteria:
        1. Identical Column Headers.
        2. Identical Section & Table Title (100% Exact Match).
        3. High Metadata Similarity (>= 80%).
        """
        if len(blocks) < 2:
            return []
            
        groups = []
        current_group = [blocks[0]]
        
        # Pre-extract definitions to avoid re-parsing
        block_defs = [self._extract_block_definition(ws, b) for b in blocks]
        
        for i in range(1, len(blocks)):
            prev_block = blocks[i-1]
            curr_block = blocks[i]
            
            prev_def = block_defs[i-1]
            curr_def = block_defs[i]
            
            # Check 1: Headers must match exactly OR current block has no headers (continuation)
            headers_match = (prev_def['headers'] == curr_def['headers']) or (not curr_def['headers'])
            
            # Check 2: Section & Title must match EXACTLY (100%)
            # We strip to be safe against minor whitespace, but string must match
            title_match = (prev_def['title'] == curr_def['title'])
            section_match = (prev_def['section'] == curr_def['section'])
            
            # Check 3: Metadata fuzzy match
            meta_sim = TableGrouper.calculate_similarity(prev_def['metadata_fingerprint'], curr_def['metadata_fingerprint'])
            # Lower threshold to 0.75 to accommodate minor footer/note differences in split parts
            meta_match = (meta_sim >= 0.75)
            
            if headers_match and title_match and section_match and meta_match:
                current_group.append(curr_block)
                logger.debug(f"Vertical match found: {prev_def['title']} -> {curr_def['title']}")
            else:
                if not headers_match:
                    logger.debug(f"Vertical mismatch headers: {prev_def['title']} vs {curr_def['title']}")
                if not title_match:
                    logger.debug(f"Vertical mismatch title: {prev_def['title']} vs {curr_def['title']}")
                if not section_match:
                    logger.debug(f"Vertical mismatch section: {prev_def['title']} vs {curr_def['title']}")
                if not meta_match:
                    logger.debug(f"Vertical mismatch metadata: {prev_def['title']} ({meta_sim:.2f})")
                
                if len(current_group) > 1:
                    groups.append(current_group)
                current_group = [curr_block]
        
        if len(current_group) > 1:
            groups.append(current_group)
            
        return groups

    def _merge_tables_vertically(self, ws: Worksheet, groups: List[List[Dict]]) -> None:
        """
        Execute vertical merge: Move data rows from tail tables to head table.
        """
        # Process groups in reverse order to avoid index shifts affecting unprocessed groups?
        # Actually within a sheet, moving cells might affect rows below.
        # But we are collapsing blocks UPWARDS. 
        # Safest is to handle matching groups.
        
        # We perform moves. NOTE: Moving cells in openpyxl can be tricky if we don't manage references.
        # Ideally:
        # 1. Copy data from Block B to immediately after Block A.
        # 2. Clear Block B.
        
        for group in groups:
            head_block = group[0]
            
            # Current insertion point is end of head block
            insert_row = head_block['end_row'] + 1
            
            for block in group[1:]:
                # Rows to copy: data_start_row to end_row
                # We skip the headers/metadata of the tail block!
                rows_to_copy = block['end_row'] - block['data_start_row'] + 1
                
                # Copy rows
                for r_offset in range(rows_to_copy):
                    src_row = block['data_start_row'] + r_offset
                    dest_row = insert_row + r_offset
                    
                    # Ensure we aren't overwriting? (We are conceptually "moving" up)
                    # But actually the blocks are sequential in the file.
                    # Block B is further down.
                    # So we are copying from lower row to higher row (closer to top).
                    
                    for col in range(1, ws.max_column + 1):
                        src_cell = ws.cell(row=src_row, column=col)
                        dest_cell = ws.cell(row=dest_row, column=col)
                        
                        dest_cell.value = src_cell.value
                        self._copy_cell_style(src_cell, dest_cell)
                
                # Update insertion point for next block
                insert_row += rows_to_copy
                
                # Update head block definition?
                head_block['end_row'] += rows_to_copy
                
                # Clear the source block (metadata + data)
                self._clear_block(ws, block, include_metadata=True)
                
                # Log
                logger.debug(f"Vertically merged block at {block['start_row']} into block at {head_block['start_row']}")
    
    def _clear_block_metadata_only(self, ws: Worksheet, block: Dict) -> None:
        """
        Clear only the metadata rows for a table block (keep the data).
        
        This removes Category, Line Items, Product/Entity, Period Type, Year, 
        Table Title, and Source rows, but keeps the actual table data.
        
        Args:
            ws: Worksheet
            block: Block dict with metadata_start_row, source_row, start_row
        """
        # Clear from metadata_start_row to just before start_row (data area)
        clear_start = block.get('metadata_start_row', block['source_row'])
        clear_end = block['start_row'] - 1  # Don't clear the data
        
        if clear_end < clear_start:
            clear_end = block['source_row']  # At minimum, clear the source row
        
        for row_num in range(clear_start, clear_end + 1):
            for col_num in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=row_num, column=col_num)
                    if hasattr(cell, 'value') and not isinstance(cell, type(None)):
                        cell.value = None
                except AttributeError:
                    pass
    
    def _split_non_mergeable_to_new_sheets(self, wb, ws: Worksheet, sheet_name: str,
                                            table_blocks: List[Dict], merged_blocks: set) -> List[str]:
        """
        Split non-mergeable tables to new sheets with their metadata.
        
        When splitting occurs:
        - Rename original sheet to {sheet_name}_1 (first table stays there)
        - Create new sheets {sheet_name}_2, _3, etc. for subsequent tables
        - Update Index: remove original entry, add entries for all _N sheets
        - Uses first row label as subtable name for descriptive titles
        
        Args:
            wb: openpyxl Workbook
            ws: Source worksheet
            sheet_name: Source sheet name
            table_blocks: All table blocks found in sheet
            merged_blocks: Set of source_rows that were merged (skip these)
            
        Returns:
            List of ALL split sheet names (including renamed original)
        """
        # Find blocks that need to be split (not merged with first)
        blocks_to_split = []
        for block in table_blocks[1:]:  # Skip first block
            if block['source_row'] not in merged_blocks:
                blocks_to_split.append(block)
        
        if not blocks_to_split:
            return []  # Nothing to split
        
        all_split_sheets = []
        split_subtable_info = {}  # Track subtable names for each split sheet
        
        # Get first row label from first block as its subtable name
        first_block = table_blocks[0] if table_blocks else None
        first_block_subtitle = ''
        if first_block and first_block.get('row_labels'):
            # Use first row label (title case) as subtitle
            first_block_subtitle = first_block['row_labels'][0].strip().title()
        
        # Step 1: Rename original sheet to _1 (first table stays there)
        original_new_name = f"{sheet_name}_1"
        if len(original_new_name) > 31:
            original_new_name = f"{sheet_name[:25]}_1"
        
        ws.title = original_new_name
        all_split_sheets.append(original_new_name)
        split_subtable_info[original_new_name] = first_block_subtitle
        logger.info(f"Renamed sheet '{sheet_name}' to '{original_new_name}'")
        
        # Step 2: Create new sheets for subsequent blocks
        split_index = 2  # Start from _2
        
        for block in blocks_to_split:
            # Get first row label from this block as its subtable name
            block_subtitle = ''
            if block.get('row_labels'):
                block_subtitle = block['row_labels'][0].strip().title()
            
            # Generate new sheet name
            new_sheet_name = f"{sheet_name}_{split_index}"
            if len(new_sheet_name) > 31:
                new_sheet_name = f"{sheet_name[:25]}_{split_index}"
            
            # Track subtable info for Index update
            split_subtable_info[new_sheet_name] = block_subtitle
            
            # Get the position of the original (now renamed) sheet to insert after it
            original_pos = wb.sheetnames.index(original_new_name)
            
            # Create new sheet at end first
            new_ws = wb.create_sheet(title=new_sheet_name)
            
            # Move the new sheet to right after original position  
            # Formula: target_position - current_position gives offset
            current_pos = len(wb.sheetnames) - 1  # It's at the end
            target_pos = original_pos + split_index - 1  # Right after previous splits
            offset = target_pos - current_pos
            wb.move_sheet(new_ws, offset=offset)
            
            # Copy "← Back to Index" link (row 1 of original) WITH hyperlink
            back_link_cell = ws.cell(row=1, column=1)
            new_back_cell = new_ws.cell(row=1, column=1)
            new_back_cell.value = back_link_cell.value
            # Create hyperlink to Index
            from openpyxl.worksheet.hyperlink import Hyperlink
            new_back_cell.hyperlink = Hyperlink(ref=new_back_cell.coordinate, target="#'Index'!A1")
            
            # Copy metadata rows
            # For sub-blocks (created by split_block_on_new_headers), they share metadata
            # with the first block, so copy from the first block's metadata range
            if block.get('_is_sub_block') and table_blocks:
                first_block = table_blocks[0]
                meta_start = first_block.get('metadata_start_row', first_block['source_row'])
                meta_end = first_block['source_row']
            else:
                meta_start = block.get('metadata_start_row', block['source_row'])
                meta_end = block['source_row']
            
            dest_row = 2  # Start after Back to Index
            
            for src_row in range(meta_start, meta_end + 1):
                if src_row == 1:
                    continue  # Already copied Back to Index
                
                for col in range(1, min(ws.max_column + 1, 50)):
                    src_val = ws.cell(row=src_row, column=col).value
                    if src_val is not None:
                        new_ws.cell(row=dest_row, column=col).value = src_val
                dest_row += 1
            
            # Add a blank row between metadata and data
            dest_row += 1
            table_data_start_row = dest_row
            
            # Copy table data (from block start to end)
            for src_row in range(block['start_row'], block['end_row'] + 1):
                for col in range(1, min(ws.max_column + 1, 50)):
                    src_val = ws.cell(row=src_row, column=col).value
                    if src_val is not None:
                        new_ws.cell(row=dest_row, column=col).value = src_val
                dest_row += 1
            
            # Extract column headers from the copied table data and update metadata
            self._update_split_sheet_column_headers(new_ws, table_data_start_row)
            
            # Add unit indicator to first column of header row if missing
            # This copies '$ in billions/millions' from the first block's header
            self._add_unit_indicator_to_split_sheet(ws, new_ws, first_block, table_data_start_row)
            
            # Update the Table Title metadata row inside split sheet to include subtable name
            if block_subtitle:
                self._update_split_sheet_table_title(new_ws, block_subtitle)
            
            # Clear original rows from source sheet (metadata + data)
            self._clear_block_rows(ws, block)
            
            all_split_sheets.append(new_sheet_name)
            split_index += 1
            logger.info(f"Split table to new sheet '{new_sheet_name}' (subtitle: {block_subtitle})")
        
        # Step 2.5: Validate _1 sheet still has content after clearing split blocks
        # The first block should remain in the _1 sheet
        # Only warn if the sheet has fewer data rows than expected from first block
        first_block = table_blocks[0] if table_blocks else None
        expected_data_rows = len(first_block.get('row_labels', [])) if first_block else 0
        
        # Use a lower threshold - only warn if truly empty (0 or 1 data rows)
        if self._is_sheet_near_empty(ws, min_data_rows=1):
            if first_block:
                logger.debug(
                    f"Sheet '{original_new_name}' has minimal data after split. "
                    f"First block: rows {first_block.get('start_row')}-{first_block.get('end_row')} "
                    f"with {expected_data_rows} row labels."
                )
        
        # Step 3: Update Index - remove original entry and add all split entries with subtable names
        self._update_index_for_splits(wb, sheet_name, all_split_sheets, split_subtable_info)
        
        return all_split_sheets
    
    def _ensure_metadata_labels(self, ws: Worksheet, data_start_row: int) -> None:
        """
        Ensure all required metadata labels exist in the split sheet.
        
        If metadata labels are missing, inserts them at appropriate positions
        before the data starts. This handles cases where block detection
        didn't properly identify metadata rows.
        
        Args:
            ws: The worksheet to check/update
            data_start_row: Row where table data starts
        """
        # Required metadata labels in order
        REQUIRED_LABELS = [
            MetadataLabels.CATEGORY_PARENT,
            MetadataLabels.LINE_ITEMS,
            MetadataLabels.PRODUCT_ENTITY,
            MetadataLabels.COLUMN_HEADER_L1,
            MetadataLabels.COLUMN_HEADER_L2,
            MetadataLabels.COLUMN_HEADER_L3,
            MetadataLabels.YEAR_QUARTER,
            MetadataLabels.TABLE_TITLE,
            MetadataLabels.SOURCES,
        ]
        
        # Find existing labels
        existing_labels = {}
        for row_num in range(1, data_start_row):
            cell_val = ws.cell(row=row_num, column=1).value
            if cell_val:
                cell_str = str(cell_val).strip()
                for label in REQUIRED_LABELS:
                    if label in cell_str:
                        existing_labels[label] = row_num
                        break
        
        # Check what's missing
        missing_labels = [l for l in REQUIRED_LABELS if l not in existing_labels]
        
        if not missing_labels:
            return  # All labels present
        
        # Find first blank row after row 1 (Back to Index) where we can add labels
        # If no blank rows, we'll need to use existing structure
        insert_row = 2
        for row_num in range(2, data_start_row):
            cell_val = ws.cell(row=row_num, column=1).value
            if not cell_val or not str(cell_val).strip():
                insert_row = row_num
                break
        
        # Insert missing labels at blank rows (don't shift data)
        # Find all blank rows we can use
        blank_rows = []
        for row_num in range(2, data_start_row):
            cell_val = ws.cell(row=row_num, column=1).value
            if not cell_val or not str(cell_val).strip():
                blank_rows.append(row_num)
        
        # Assign missing labels to blank rows
        for i, label in enumerate(missing_labels):
            if i < len(blank_rows):
                ws.cell(row=blank_rows[i], column=1).value = label
                logger.debug(f"Added missing metadata label '{label}' at row {blank_rows[i]}")
            else:
                # No more blank rows - log warning
                logger.warning(f"No blank row available for metadata label '{label}'")

    def _update_split_sheet_table_title(self, ws: Worksheet, subtitle: str) -> None:
        """
        NOTE: This method is kept for backwards compatibility but no longer appends subtitles.
        
        The table title should already include qualifiers like "by Property Type" or "by Region"
        from the original title extraction. We should NOT append first row labels as suffixes.
        
        Args:
            ws: The worksheet to update
            subtitle: Ignored - kept for API compatibility
        """
        # Do nothing - titles should not have suffixes appended
        # The full title including "by Property Type" etc. should come from title extraction
        pass

    def _add_unit_indicator_to_split_sheet(self, source_ws: Worksheet, new_ws: Worksheet, 
                                            first_block: Dict, data_start_row: int) -> None:
        """
        Add unit indicator ('$ in billions/millions') to split sheet header row.
        
        When a table is split, the new sub-table may have an empty first column in its
        header row. This copies the unit indicator from the first block's header.
        
        Args:
            source_ws: Original worksheet with unit indicator
            new_ws: New split worksheet
            first_block: First table block dict (contains unit indicator)
            data_start_row: Row where table data starts in new_ws
        """
        if not first_block:
            return
        
        # Find unit indicator in first block (usually in the header row)
        unit_indicator = None
        
        # Search in the first block's data area for unit indicator
        block_start = first_block.get('start_row', first_block.get('data_start_row', 13))
        for row_num in range(block_start, min(block_start + 4, source_ws.max_row + 1)):
            cell_val = source_ws.cell(row=row_num, column=1).value
            if cell_val:
                cell_str = str(cell_val).strip().lower()
                if '$ in' in cell_str or 'in millions' in cell_str or 'in billions' in cell_str:
                    unit_indicator = source_ws.cell(row=row_num, column=1).value
                    break
        
        if not unit_indicator:
            return
        
        # Find header row in new sheet (row with empty first col but has Q-codes in other cols)
        for row_num in range(data_start_row, min(data_start_row + 4, new_ws.max_row + 1)):
            first_col = new_ws.cell(row=row_num, column=1).value
            first_val = str(first_col).strip() if first_col else ''
            
            # If first column is empty, check if other columns have Q-codes
            if not first_val or first_val.lower() in ['', 'nan', 'none']:
                # Check column 2 for Q-code pattern
                col2_val = new_ws.cell(row=row_num, column=2).value
                if col2_val:
                    col2_str = str(col2_val).lower()
                    if any(p in col2_str for p in ['q1-', 'q2-', 'q3-', 'q4-', 'ytd-']):
                        # Found the header row - add unit indicator
                        new_ws.cell(row=row_num, column=1).value = unit_indicator
                        logger.debug(f"Added unit indicator '{unit_indicator}' to split sheet row {row_num}")
                        return

    def _update_split_sheet_column_headers(self, ws: Worksheet, data_start_row: int) -> None:
        """
        Extract column headers from the first rows of split table and update metadata.
        
        Uses MultiLevelHeaderParser to dynamically detect L1/L2/L3/L4 levels from
        multi-row headers. Updates all column header metadata rows.
        
        Args:
            ws: The new worksheet with split table
            data_start_row: Row where table data starts
        """
        # First, ensure all required metadata labels exist
        self._ensure_metadata_labels(ws, data_start_row)
        
        # Find metadata rows to update
        l1_row = None
        l2_row = None
        l3_row = None
        year_quarter_row = None
        category_row = None
        line_items_row = None
        product_entity_row = None
        table_title_row = None
        source_row = None
        
        for row_num in range(1, data_start_row):
            cell_val = ws.cell(row=row_num, column=1).value
            if not cell_val:
                continue
            cell_str = str(cell_val)
            
            if MetadataLabels.COLUMN_HEADER_L1 in cell_str:
                l1_row = row_num
            elif MetadataLabels.COLUMN_HEADER_L2 in cell_str:
                l2_row = row_num
            elif MetadataLabels.COLUMN_HEADER_L3 in cell_str:
                l3_row = row_num
            elif MetadataLabels.YEAR_QUARTER in cell_str:
                year_quarter_row = row_num
            elif MetadataLabels.CATEGORY_PARENT in cell_str:
                category_row = row_num
            elif MetadataLabels.LINE_ITEMS in cell_str:
                line_items_row = row_num
            elif MetadataLabels.PRODUCT_ENTITY in cell_str:
                product_entity_row = row_num
            elif MetadataLabels.TABLE_TITLE in cell_str:
                table_title_row = row_num
            elif MetadataLabels.SOURCES in cell_str:
                source_row = row_num
        
        # Log if critical metadata rows are missing
        missing = []
        if not l3_row:
            missing.append('Column Header L3')
        if not category_row:
            missing.append('Category')
        if not source_row:
            missing.append('Source')
        
        if missing:
            logger.warning(f"Split sheet missing metadata rows: {missing}")
        
        # Extract multi-row headers from table data area
        header_rows = []
        for scan_row in range(data_start_row, min(data_start_row + 4, ws.max_row + 1)):
            first_col = ws.cell(row=scan_row, column=1).value
            first_val = str(first_col).strip() if first_col else ''
            
            # If first column is empty or "$", this is likely a header row
            if not first_val or first_val.lower() in ['', 'nan', 'none'] or first_val.startswith('$'):
                row_values = []
                for col in range(1, min(ws.max_column + 1, 20)):
                    val = ws.cell(row=scan_row, column=col).value
                    row_values.append(str(val).strip() if val else '')
                header_rows.append(row_values)
            else:
                break  # Hit data row, stop
        
        if not header_rows:
            return
        
        # Get source filename for 10-K detection
        source_filename = ''
        for row_num in range(1, 15):
            cell_val = ws.cell(row=row_num, column=1).value
            if cell_val and MetadataLabels.SOURCES in str(cell_val):
                source_filename = str(cell_val)
                break
        
        # Parse headers using MultiLevelHeaderParser
        parsed = MultiLevelHeaderParser.parse_multi_row_headers(header_rows, source_filename)
        
        # Collect unique L1, L2, L3 values and codes
        l1_values = set()
        l2_values = set()
        l3_values = set()
        codes = set()
        
        for col_info in parsed.get('columns', []):
            if col_info.get('l1'):
                l1_values.add(col_info['l1'])
            if col_info.get('l2'):
                l2_values.add(col_info['l2'])
            if col_info.get('l3'):
                l3_values.add(col_info['l3'])
            if col_info.get('code') and col_info['code'] != 'STATIC':
                codes.add(col_info['code'])
        
        # Update metadata rows
        if l1_row and l1_values:
            l1_text = ', '.join(sorted(l1_values)[:5])
            ws.cell(row=l1_row, column=1).value = f"{MetadataLabels.COLUMN_HEADER_L1} {l1_text}"
        
        if l2_row and l2_values:
            l2_text = ', '.join(sorted(l2_values)[:5])
            ws.cell(row=l2_row, column=1).value = f"{MetadataLabels.COLUMN_HEADER_L2} {l2_text}"
        
        if l3_row:
            if l3_values:
                l3_text = ', '.join(sorted(l3_values)[:5])
                ws.cell(row=l3_row, column=1).value = f"{MetadataLabels.COLUMN_HEADER_L3} {l3_text}"
            elif parsed.get('spanning_l3'):
                ws.cell(row=l3_row, column=1).value = f"{MetadataLabels.COLUMN_HEADER_L3} {parsed['spanning_l3']}"
        
        # Update Year/Quarter with standardized codes
        if year_quarter_row and codes:
            codes_text = ', '.join(sorted(codes)[:5])
            ws.cell(row=year_quarter_row, column=1).value = f"{MetadataLabels.YEAR_QUARTER} {codes_text}"
    
    def _clear_block_rows(self, ws: Worksheet, block: Dict) -> None:
        """
        Clear all rows for a table block (metadata + data).
        
        Args:
            ws: Worksheet
            block: Block dict with metadata_start_row, source_row, start_row, end_row
        """
        clear_start = block.get('metadata_start_row', block['source_row'])
        clear_end = block['end_row']
        
        for row_num in range(clear_start, clear_end + 1):
            for col_num in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=row_num, column=col_num)
                    if hasattr(cell, 'value'):
                        cell.value = None
                except AttributeError:
                    pass
    
    def _update_index_for_split_sheets(self, wb, new_sheets: List[str]) -> None:
        """
        Add entries to Index for newly split sheets.
        
        Delegates to IndexManager for centralized logic.
        """
        IndexManager.update_index_for_split_sheets(wb, new_sheets)
    
    def _update_index_for_splits(self, wb, original_sheet_name: str, split_sheets: List[str], 
                                   subtable_info: Dict[str, str] = None) -> None:
        """
        Update Index when a sheet is split: update original row and insert new rows.
        
        Delegates to IndexManager for centralized logic.
        
        Args:
            wb: Workbook
            original_sheet_name: Original sheet name before split
            split_sheets: List of new sheet names after split
            subtable_info: Dict mapping sheet name to first row label (subtitle)
        """
        IndexManager.update_index_for_splits(wb, original_sheet_name, split_sheets, subtable_info)
    
    def _extract_block_metadata(self, ws: Worksheet, block: Dict) -> Dict[str, Set[str]]:
        """
        Extract metadata values from a table block's metadata rows.
        
        Uses centralized extract_metadata_from_cell from MetadataBuilder.
        
        Returns dict with keys: 'period_type', 'years', 'sources', 'main_header'
        """
        combined = {
            'period_type': set(),
            'years': set(),
            'sources': set(),
            'main_header': set(),
        }
        
        # Scan metadata rows (from metadata_start_row to source_row)
        meta_start = block.get('metadata_start_row', block['source_row'])
        meta_end = block['source_row']
        
        for row_num in range(meta_start, meta_end + 1):
            cell_val = ws.cell(row=row_num, column=1).value
            if cell_val:
                # Use centralized extraction function from MetadataBuilder
                cell_meta = MetadataBuilder.extract_metadata_from_cell(str(cell_val))
                combined = MetadataBuilder.merge_metadata_sets(combined, cell_meta)
        
        return combined
    
    def _normalize_column_header_to_code(self, header: str, source_filename: str = '') -> dict:
        """
        Normalize column header to standard quarter code format.
        
        Uses QuarterDateMapper for conversion. Also detects and splits L1 headers
        like 'Assets', 'Liability', 'Average Balances' that may be merged with L2.
        
        Args:
            header: Column header string (e.g., "Three Months Ended March 31, 2025")
            source_filename: Source filename for 10-K detection
            
        Returns:
            Dict with 'code' (Q1-QTD-2025), 'l1' (optional), 'display' (original)
        """
        result = {
            'code': '',
            'l1': '',       # Optional L1 component (Assets, Liability, etc.)
            'display': header,
            'l2': '',       # Period type (Three Months Ended)
            'l3': '',       # Date (March 31, 2025)
        }
        
        if not header:
            return result
        
        header_str = str(header).strip()
        
        # Known L1 prefixes that may be merged with L2
        L1_PATTERNS = [
            'assets', 'liability', 'liabilities', 'average balance', 'average balances',
            'average monthly', 'ending balance', 'beginning balance',
            'amortized cost', 'fair value', 'net revenue', 'net revenues',
        ]
        
        # Check if header starts with L1 pattern
        header_lower = header_str.lower()
        l1_found = ''
        remaining = header_str
        
        for pattern in L1_PATTERNS:
            if header_lower.startswith(pattern):
                # Find where L1 ends and L2 begins
                # L2 typically starts with 'Three Months', 'Six Months', 'At', 'As of'
                l2_markers = ['three months', 'six months', 'nine months', 'year ended', 'at ', 'as of ']
                
                for marker in l2_markers:
                    idx = header_lower.find(marker)
                    if idx > 0:
                        l1_found = header_str[:idx].strip()
                        remaining = header_str[idx:].strip()
                        break
                
                if l1_found:
                    break
        
        result['l1'] = l1_found
        
        # Convert remaining to code using QuarterDateMapper
        report_type = '10k' if source_filename and '10k' in source_filename.lower() else ''
        code = QuarterDateMapper.display_to_code(remaining, report_type)
        
        result['code'] = code
        result['display'] = remaining
        
        # Determine L2 and L3 from the code
        parts = code.split('-')
        if len(parts) >= 2:
            quarter = parts[0]
            year = parts[-1]
            
            if len(parts) == 2:
                # Point-in-time: Q1-2025
                result['l2'] = ''
                result['l3'] = QuarterDateMapper.code_to_display(code)
            elif len(parts) == 3:
                # Period: Q1-QTD-2025 or Q2-YTD-2025
                period = parts[1]
                if period == 'QTD':
                    result['l2'] = 'Three Months Ended'
                elif period == 'YTD' and quarter == 'Q2':
                    result['l2'] = 'Six Months Ended'
                elif period == 'YTD' and quarter == 'Q3':
                    result['l2'] = 'Nine Months Ended'
                month, day = QuarterDateMapper.QUARTER_END_DATES.get(quarter, ('', ''))
                result['l3'] = f"{month} {day}, {year}" if month else year
        
        return result
    
    def _merge_metadata_rows(self, ws: Worksheet, first_block: Dict, all_blocks: List[Dict]) -> None:
        """
        Merge metadata from all blocks into the first block's metadata rows.
        
        Uses centralized MetadataLabels for consistent labeling and 
        MetadataBuilder.format_merged_metadata for formatting.
        
        Updates the first block's metadata to include values from all merged tables:
        - Period Type: combined unique period types
        - Year(s): combined unique years, sorted descending
        - Sources: combined unique sources
        """
        if len(all_blocks) < 2:
            return
        
        # Collect metadata from all blocks using centralized functions
        all_metadata = [self._extract_block_metadata(ws, block) for block in all_blocks]
        combined = MetadataBuilder.merge_metadata_sets(*all_metadata)
        
        # Format the merged metadata
        formatted = MetadataBuilder.format_merged_metadata(combined)
        
        # Update first block's metadata rows
        meta_start = first_block.get('metadata_start_row', first_block['source_row'])
        meta_end = first_block['source_row']
        
        for row_num in range(meta_start, meta_end + 1):
            cell = ws.cell(row=row_num, column=1)
            cell_val = cell.value
            if not cell_val:
                continue
            cell_str = str(cell_val).strip()
            
            # Update Column Header L2 (Period Type)
            if MetadataLabels.is_column_header_l2(cell_str):
                if formatted['period_type']:
                    cell.value = f"{MetadataLabels.COLUMN_HEADER_L2} {formatted['period_type']}"
            
            # Update Column Header L3 (Year(s))
            elif MetadataLabels.is_column_header_l3(cell_str):
                if formatted['years']:
                    cell.value = f"{MetadataLabels.COLUMN_HEADER_L3} {formatted['years']}"
            
            # Update Sources
            elif MetadataLabels.is_sources(cell_str):
                if formatted['sources']:
                    cell.value = f"{MetadataLabels.SOURCES} {formatted['sources']}"
            
            # Update Column Header L1 (Main Header)
            elif MetadataLabels.is_column_header_l1(cell_str):
                if formatted['main_header']:
                    cell.value = f"{MetadataLabels.COLUMN_HEADER_L1} {formatted['main_header']}"
    
    def _find_table_blocks(self, ws) -> List[Dict[str, Any]]:
        """
        Find all table blocks in a worksheet.
        
        Delegates to BlockDetector for centralized logic.
        """
        return BlockDetector.find_table_blocks(ws, self._extract_row_labels)
    
    def _split_block_on_new_headers(self, ws: Worksheet, block: Dict) -> List[Dict]:
        """
        Check if a table block contains new column headers mid-table.
        
        If a row has empty first column but contains date/period patterns
        in other columns (e.g., "Three Months Ended June 30"), it indicates
        a new sub-table and the block should be split.
        
        Args:
            ws: Worksheet
            block: Table block dict
            
        Returns:
            List of blocks (original if no split needed, or split blocks)
        """
        data_start = block.get('data_start_row', block['start_row'])
        data_end = block['end_row']
        
        # Scan for new header rows (after at least one data row)
        split_points = []
        seen_data_row = False
        
        for row_num in range(data_start, data_end + 1):
            first_col = ws.cell(row=row_num, column=1).value
            
            # Get all values in the row
            row_values = []
            for col in range(1, min(ws.max_column + 1, 15)):
                row_values.append(ws.cell(row=row_num, column=col).value)
            
            # Check if first column has data (this is a data row)
            first_val = str(first_col).strip() if first_col else ''
            if first_val and first_val.lower() not in ['', 'nan', 'none']:
                seen_data_row = True
                continue
            
            # If we've seen data rows and this row is a new header, mark split point
            if seen_data_row and is_new_table_header_row(row_values, first_col):
                split_points.append(row_num)
        
        # If no splits needed, return original block
        if not split_points:
            return [block]
        
        # Create split blocks
        result_blocks = []
        current_start = block['start_row']
        
        for split_row in split_points:
            # Create block ending before the split point
            if split_row > current_start:
                new_block = block.copy()
                new_block['start_row'] = current_start
                new_block['end_row'] = split_row - 1
                new_block['data_start_row'] = current_start
                new_block['row_labels'] = self._extract_row_labels(ws, new_block)
                if new_block['row_labels']:
                    result_blocks.append(new_block)
            
            # Next block starts at the split point
            current_start = split_row
        
        # Create final block from last split to end
        if current_start <= data_end:
            new_block = block.copy()
            new_block['start_row'] = current_start
            new_block['end_row'] = data_end
            new_block['data_start_row'] = current_start
            new_block['row_labels'] = self._extract_row_labels(ws, new_block)
            if new_block['row_labels']:
                result_blocks.append(new_block)
        
        if result_blocks:
            logger.debug(f"Split table block into {len(result_blocks)} sub-tables at rows {split_points}")
        
        return result_blocks if result_blocks else [block]
    
    def _identify_header_and_data_rows(self, ws: Worksheet, block: Dict) -> None:
        """
        Identify which rows are headers vs data in a table block.
        Updates block['data_start_row'] to point to first data row.
        
        Header rows typically:
        - Have empty first column but values in other columns (column headers)
        - First column contains currency units like "$ in millions"
        - First column is a year (4 digits)
        - Non-first columns contain year values
        
        Data rows have:
        - Non-empty first column with descriptive text (row labels)
        """
        # Use centralized patterns from domain_patterns.py
        header_patterns = TABLE_HEADER_PATTERNS
        
        data_start = block['start_row']
        header_rows_found = []
        
        # Only scan first few rows for headers (typically 1-3 header rows max)
        max_header_rows = min(4, block['end_row'] - block['start_row'] + 1)
        
        for row_num in range(block['start_row'], block['start_row'] + max_header_rows):
            if row_num > block['end_row']:
                break
                
            row_values = []
            first_col_value = None
            
            for col in range(1, min(MAX_COL_SCAN, ws.max_column + 1)):
                cell_val = ws.cell(row=row_num, column=col).value
                if cell_val:
                    row_values.append(str(cell_val).strip())
                    if col == 1:
                        first_col_value = str(cell_val).strip().lower()
            
            is_header_row = False
            
            # Empty first column but has values in other columns = header row (column headers)
            if not first_col_value and len(row_values) > 0:
                is_header_row = True
            
            # First column has explicit header pattern ($ in millions, etc.)
            elif first_col_value:
                if any(pattern in first_col_value for pattern in header_patterns):
                    is_header_row = True
                # First column is a year (4 digits)
                elif first_col_value.isdigit() and len(first_col_value) == YEAR_STRING_LENGTH:
                    is_header_row = True
                else:
                    # First column has a value that's NOT a header pattern
                    # This is the start of data rows
                    data_start = row_num
                    break
            
            # Check if non-first columns have multiple year values (date header row)
            if not is_header_row:
                year_count = sum(1 for v in row_values if v.isdigit() and len(v) == YEAR_STRING_LENGTH)
                if year_count >= MIN_YEAR_COUNT_FOR_HEADER:
                    is_header_row = True
            
            if is_header_row:
                header_rows_found.append(row_num)
                data_start = row_num + 1
        
        block['header_rows'] = header_rows_found
        block['data_start_row'] = data_start
    
    def _extract_row_labels(self, ws: Worksheet, block: Dict) -> List[str]:
        """
        Extract row labels (Column A values) from a table block's data rows.
        
        Returns:
            List of row labels (normalized for comparison)
        """
        labels = []
        for row_num in range(block['data_start_row'], block['end_row'] + 1):
            val = ws.cell(row=row_num, column=1).value
            if val is not None:
                labels.append(str(val).strip())
            else:
                labels.append('')
        return labels
    
    def _find_mergeable_tables(self, blocks: List[Dict]) -> List[List[Dict]]:
        """
        Group table blocks that can be merged (identical row labels for entire table).
        
        Returns:
            List of groups, where each group is a list of mergeable blocks
        """
        if not blocks:
            return []
        
        groups = []
        used = set()
        
        for i, block_a in enumerate(blocks):
            if i in used:
                continue
            
            group = [block_a]
            used.add(i)
            
            for j, block_b in enumerate(blocks):
                if j in used:
                    continue
                
                if self._tables_match(block_a, block_b):
                    group.append(block_b)
                    used.add(j)
            
            if len(group) >= 2:
                groups.append(group)
        
        return groups
    
    def _tables_match(self, block_a: Dict, block_b: Dict) -> bool:
        """
        Check if two table blocks can be merged.
        
        Requirements:
        - Same number of row labels
        - 100% match of row labels (same text in same order)
        
        Returns:
            True if tables can be merged
        """
        labels_a = block_a['row_labels']
        labels_b = block_b['row_labels']
        
        if len(labels_a) != len(labels_b):
            return False
        
        if len(labels_a) == 0:
            return False
        
        for a, b in zip(labels_a, labels_b):
            # Normalize for comparison
            a_norm = a.lower().strip()
            b_norm = b.lower().strip()
            
            if a_norm != b_norm:
                return False
        
        return True
    
    def _merge_tables_horizontally(self, ws: Worksheet, blocks: List[Dict]) -> None:
        """
        Merge multiple table blocks horizontally in the worksheet.
        
        Process:
        1. Keep first table in place
        2. For each subsequent table, append its columns (except Column A) to the first table
        3. Clear the merged tables after copying
        
        Args:
            ws: openpyxl Worksheet
            blocks: List of table blocks to merge (must have matching row labels)
        """
        if len(blocks) < 2:
            return
        
        # Sort blocks by row position (first table stays in place)
        blocks = sorted(blocks, key=lambda b: b['start_row'])
        
        first_block = blocks[0]
        
        # Find the current rightmost column of the first block
        max_col = self._get_block_column_count(ws, first_block)
        insert_col = max_col + 1
        
        # Track existing columns in first block (for deduplication)
        # Key: tuple of column values, Value: column index
        existing_columns = self._extract_column_signatures(ws, first_block, 2, max_col)
        
        # Merge each subsequent block into the first
        for block in blocks[1:]:
            block_cols = self._get_block_column_count(ws, block)
            
            if block_cols <= 1:
                continue  # No data columns to copy
            
            # For each column in source block, check if it's a duplicate
            cols_to_copy = []  # List of (source_col_index, source_signature)
            
            for src_col in range(2, block_cols + 1):  # Skip column 1 (row labels)
                col_signature = self._get_column_signature(ws, block, src_col)
                
                # Check if this column already exists
                if col_signature not in existing_columns:
                    cols_to_copy.append((src_col, col_signature))
                    existing_columns[col_signature] = insert_col + len(cols_to_copy) - 1
                else:
                    logger.debug(f"Skipping duplicate column at {block['start_row']}, col {src_col}")
            
            if not cols_to_copy:
                # All columns are duplicates, skip this block
                self._clear_block(ws, block)
                continue
            
            # Copy only non-duplicate columns
            for copy_idx, (src_col, _) in enumerate(cols_to_copy):
                target_col = insert_col + copy_idx
                
                # Copy header rows
                for row_offset in range(block['start_row'], block['data_start_row']):
                    target_row = first_block['start_row'] + (row_offset - block['start_row'])
                    
                    try:
                        source_cell = ws.cell(row=row_offset, column=src_col)
                        target_cell = ws.cell(row=target_row, column=target_col)
                        
                        source_value = source_cell.value if hasattr(source_cell, 'value') else None
                        target_cell.value = source_value
                        self._copy_cell_style(source_cell, target_cell)
                    except AttributeError:
                        pass
                
                # Copy data rows
                for row_offset, src_row in enumerate(range(block['data_start_row'], block['end_row'] + 1)):
                    target_row = first_block['data_start_row'] + row_offset
                    
                    try:
                        source_cell = ws.cell(row=src_row, column=src_col)
                        target_cell = ws.cell(row=target_row, column=target_col)
                        
                        source_value = source_cell.value if hasattr(source_cell, 'value') else None
                        target_cell.value = source_value
                        self._copy_cell_style(source_cell, target_cell)
                    except AttributeError:
                        pass
            
            # Update insert column for next block
            insert_col += len(cols_to_copy)
            
            # Clear the merged block (to avoid duplicate data)
            self._clear_block(ws, block)
        
        # IMPORTANT: Merge metadata from all blocks into the first block
        # This ensures Period Type, Years, and Sources reflect ALL merged tables
        self._merge_metadata_rows(ws, first_block, blocks)
    
    def _extract_column_signatures(self, ws: Worksheet, block: Dict, start_col: int, end_col: int) -> Dict[tuple, int]:
        """
        Extract column signatures for all columns in a block.
        Returns dict mapping column signature (tuple of values) to column index.
        """
        signatures = {}
        for col in range(start_col, end_col + 1):
            sig = self._get_column_signature(ws, block, col)
            signatures[sig] = col
        return signatures
    
    def _get_column_signature(self, ws: Worksheet, block: Dict, col: int) -> tuple:
        """
        Get a signature (tuple of all values) for a column in a block.
        Includes header rows and data rows.
        """
        values = []
        
        # Include header row values
        for row_num in range(block['start_row'], block['data_start_row']):
            cell = ws.cell(row=row_num, column=col)
            val = cell.value if hasattr(cell, 'value') else None
            # Normalize value for comparison
            if val is not None:
                values.append(str(val).strip().lower())
            else:
                values.append('')
        
        # Include data row values
        for row_num in range(block['data_start_row'], block['end_row'] + 1):
            cell = ws.cell(row=row_num, column=col)
            val = cell.value if hasattr(cell, 'value') else None
            if val is not None:
                values.append(str(val).strip().lower())
            else:
                values.append('')
        
        return tuple(values)
    
    def _copy_cell_style(self, source_cell, target_cell) -> None:
        """Copy cell styling from source to target."""
        try:
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            if source_cell.border:
                target_cell.border = copy(source_cell.border)
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
        except Exception as e:
            logger.debug(f"Style copy error (non-critical): {e}")
    
    def _clear_block(self, ws: Worksheet, block: Dict, include_metadata: bool = True) -> None:
        """
        Clear a table block after it has been merged.
        
        Args:
            ws: Worksheet
            block: Block dict with metadata_start_row, source_row, start_row, end_row
            include_metadata: If True, also clear metadata rows above the table
        """
        # Use dynamic metadata_start_row if available and include_metadata is True
        if include_metadata and 'metadata_start_row' in block:
            clear_start = block['metadata_start_row']
        elif include_metadata and 'source_row' in block:
            clear_start = block['source_row']
        else:
            clear_start = block['start_row']
        
        # Clear from clear_start to end_row
        for row_num in range(clear_start, block['end_row'] + 1):
            for col_num in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=row_num, column=col_num)
                    # Skip merged cells (they are read-only)
                    if hasattr(cell, 'value') and not isinstance(cell, type(None)):
                        cell.value = None
                except AttributeError:
                    pass  # MergedCell objects are read-only, skip them
    
    def _get_block_column_count(self, ws: Worksheet, block: Dict) -> int:
        """Get the number of columns with data in a table block."""
        max_col = 0
        for row_num in range(block['start_row'], block['end_row'] + 1):
            for col_num in range(1, ws.max_column + 1):
                if ws.cell(row=row_num, column=col_num).value is not None:
                    max_col = max(max_col, col_num)
        return max_col
    
    def _apply_number_formatting(self, output_path: Path) -> None:
        """
        Apply currency and percentage formatting to merged output file.
        
        Uses row label heuristics and value-based detection similar to
        ExcelFormatter.apply_currency_format().
        """
        from openpyxl import load_workbook
        
        # US currency accounting format
        CURRENCY_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        PERCENTAGE_FORMAT = '0.00%'
        
        # Indicators for determining format type
        CURRENCY_INDICATORS = ['$', 'dollar', 'revenue', 'income', 'expense', 'cost', 'assets', 'liabilities', 'balance']
        PERCENTAGE_INDICATORS = ['%', 'percent', 'ratio', 'margin', 'return', 'rate', 'yield', 'roe', 'roa', 'rotce']
        
        wb = load_workbook(output_path)
        
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() == 'index':
                continue
            
            ws = wb[sheet_name]
            
            # Find where data starts (after Row Label row)
            data_start_row = None
            for r in range(1, min(20, ws.max_row + 1)):
                cell_val = str(ws.cell(row=r, column=1).value or '').strip()
                # Look for data rows (non-metadata, non-empty first column)
                if cell_val and not any(label in cell_val for label in [
                    'Category:', 'Line Items:', 'Product/Entity:', 'Period Type:',
                    'Year:', 'Table Title:', 'Source:', 'Column Header', '← Back to Index'
                ]):
                    # Check if this looks like a data row (has values in other columns)
                    has_data = any(ws.cell(row=r, column=c).value for c in range(2, min(10, ws.max_column + 1)))
                    if has_data:
                        data_start_row = r
                        break
            
            if data_start_row is None:
                continue
            
            # Apply formatting to data rows
            for row in range(data_start_row, ws.max_row + 1):
                row_label = str(ws.cell(row=row, column=1).value or '').lower()
                
                # Collect numeric values in this row
                row_values = []
                for col in range(2, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float)) and cell.value is not None:
                        row_values.append(cell.value)
                
                # Determine if percentage or currency based on row label
                is_pct_label = any(ind in row_label for ind in PERCENTAGE_INDICATORS)
                is_currency_label = any(ind in row_label for ind in CURRENCY_INDICATORS)
                
                if is_pct_label:
                    is_percentage_row = True
                elif is_currency_label:
                    is_percentage_row = False
                else:
                    # Value-based detection for unlabeled rows
                    is_percentage_row = False
                    if row_values:
                        non_zero_values = [v for v in row_values if v != 0]
                        if non_zero_values:
                            all_in_pct_range = all(-1 <= v <= 1 for v in non_zero_values)
                            has_decimal = any(0 < abs(v) < 1 for v in non_zero_values)
                            is_percentage_row = all_in_pct_range and has_decimal
                
                # Apply format to all numeric cells in this row
                format_to_apply = PERCENTAGE_FORMAT if is_percentage_row else CURRENCY_FORMAT
                for col in range(2, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float)) and cell.value is not None:
                        cell.number_format = format_to_apply
        
        wb.save(output_path)
        logger.debug(f"Applied number formatting to {output_path.name}")

# =============================================================================
# FACTORY FUNCTION (not a singleton - allows directory overrides)
# =============================================================================

_merger_instance: Optional[TableMerger] = None


def get_table_merger(fresh: bool = False) -> TableMerger:
    """
    Get TableMerger instance.
    
    Args:
        fresh: If True, create a new instance instead of reusing cached one.
               Use this when you need to modify source_dir/dest_dir.
    
    Returns:
        TableMerger instance
    
    Note:
        When overriding source_dir/dest_dir in ProcessAdvancedStep, 
        set fresh=True to ensure a clean instance.
    """
    global _merger_instance
    
    if fresh or _merger_instance is None:
        _merger_instance = TableMerger()
    
    return _merger_instance


def reset_table_merger() -> None:
    """Reset the merger singleton (for testing)."""
    global _merger_instance
    _merger_instance = None
