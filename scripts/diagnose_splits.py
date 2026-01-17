import sys
import os
import re
from pathlib import Path
from difflib import SequenceMatcher
import pandas as pd
from openpyxl import load_workbook

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.append(str(project_root))

from src.infrastructure.extraction.exporters.table_merger import get_table_merger
from src.infrastructure.extraction.consolidation.table_grouping import TableGrouper
from src.utils.metadata_labels import MetadataLabels

def inspect_splits(file_name="10q0925_tables.xlsx"):
    print(f"--- Inspecting Splits in {file_name} ---")
    
    file_path = Path("data/processed") / file_name
    if not file_path.exists():
        print(f"File not found: {file_path}")
        return

    merger = get_table_merger()
    wb = load_workbook(file_path, read_only=False) # read_only=False to access cells easily
    
    # We want to find tables that WOULD be split (or are split in output).
    # Since we are reading INPUT `processed`, they are just blocks in sheets.
    # We need to find sheets that end up having multiple parts.
    
    # Let's iterate all sheets and run the _find_vertical_groups logic MANUALLY 
    # and print WHY they don't group.
    
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Index': continue
        
        ws = wb[sheet_name]
        blocks = merger._find_table_blocks(ws)
        if len(blocks) < 2: continue
        
        # We only care about sheets that have multiple blocks with SAME title
        # to see why they didn't merge.
        
        # Filter for specific table if needed
        # blocks = [b for b in blocks if "Gains" in b.get('title', '')]
        
        block_defs = [merger._extract_block_definition(ws, b) for b in blocks]
        
        # Check adjacent pairs
        for i in range(1, len(blocks)):
            prev_def = block_defs[i-1]
            curr_def = block_defs[i]
            
            # Filter output for Gain/Losses
            if "Gains" not in prev_def['title'] and "Gains" not in curr_def['title']:
                continue
            
            # Simple check: do they look like they SHOULD merge?
            # e.g. same title or "Part" in title
            
            title1 = prev_def['title']
            title2 = curr_def['title']
            
            # Normalize for check
            t1_norm = re.sub(r'\s*\(Part\s+\d+\)$', '', title1, flags=re.IGNORECASE).strip()
            t2_norm = re.sub(r'\s*\(Part\s+\d+\)$', '', title2, flags=re.IGNORECASE).strip()
            
            if t1_norm == t2_norm and t1_norm != "":
                print(f"\nPotential Split Pair in Sheet '{sheet_name}':")
                print(f"  Block A: Title='{title1}' Rows={blocks[i-1]['start_row']}-{blocks[i-1]['end_row']}")
                print(f"  Block B: Title='{title2}' Rows={blocks[i]['start_row']}-{blocks[i]['end_row']}")
                
                # Check merge criteria
                headers_match = (prev_def['headers'] == curr_def['headers']) or (not curr_def['headers'])
                
                title_match = (prev_def['title'] == curr_def['title'])
                section_match = (prev_def['section'] == curr_def['section'])
                
                meta_sim = TableGrouper.calculate_similarity(prev_def['metadata_fingerprint'], curr_def['metadata_fingerprint'])
                meta_match = (meta_sim >= 0.8)
                
                print(f"  > Headers Match: {headers_match}")
                if not headers_match:
                    print(f"    Spec A: {prev_def['headers']}")
                    print(f"    Spec B: {curr_def['headers']}")
                
                print(f"  > Title Match:   {title_match} ('{prev_def['title']}' vs '{curr_def['title']}')")
                print(f"  > Section Match: {section_match} ('{prev_def['section']}' vs '{curr_def['section']}')")
                print(f"  > Metadata Sim:  {meta_sim:.4f} (Match: {meta_match})")
                print(f"    Meta A: {prev_def['metadata_fingerprint'][:100]}...")
                print(f"    Meta B: {curr_def['metadata_fingerprint'][:100]}...")

if __name__ == "__main__":
    inspect_splits()
