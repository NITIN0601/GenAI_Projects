"""
Test script for Index Sheet Resequencer

This script tests the index_sheet_resequencer.py implementation.
"""

import sys
from pathlib import Path

# Add src to path
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from index_sheet_resequencer import IndexSheetResequencer, BlockDetector, process_all_xlsx_files
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

logger = logging.getLogger(__name__)


def test_single_file(xlsx_path: str):
    """Test processing a single xlsx file"""
    logger.info(f"\n{'='*70}")
    logger.info(f"Testing single file: {xlsx_path}")
    logger.info(f"{'='*70}\n")
    
    xlsx_path = Path(xlsx_path)
    
    if not xlsx_path.exists():
        logger.error(f"File not found: {xlsx_path}")
        return False
    
    try:
        resequencer = IndexSheetResequencer(xlsx_path)
        output_path = xlsx_path.parent / "test_output" / xlsx_path.name
        output_path.parent.mkdir(exist_ok=True)
        
        result_path = resequencer.process(output_path)
        logger.info(f"\n✓ Successfully processed file")
        logger.info(f"  Input:  {xlsx_path}")
        logger.info(f"  Output: {result_path}")
        
        return True
        
    except Exception as e:
        logger.error(f"\n✗ Failed to process {xlsx_path.name}: {e}", exc_info=True)
        return False


def test_all_files(input_dir: str = "data/processed"):
    """Test processing all xlsx files in a directory"""
    logger.info(f"\n{'='*70}")
    logger.info(f"Testing all files in: {input_dir}")
    logger.info(f"{'='*70}\n")
    
    input_dir = Path(input_dir)
    
    if not input_dir.exists():
        logger.error(f"Directory not found: {input_dir}")
        return False
    
    output_dir = input_dir / "test_output"
    
    try:
        process_all_xlsx_files(input_dir, output_dir)
        logger.info(f"\n✓ Successfully processed all files")
        logger.info(f"  Output directory: {output_dir}")
        return True
        
    except Exception as e:
        logger.error(f"\n✗ Failed to process files: {e}", exc_info=True)
        return False


def analyze_file_structure(xlsx_path: str):
    """Analyze and report structure of an xlsx file"""
    import openpyxl
    
    logger.info(f"\n{'='*70}")
    logger.info(f"Analyzing file structure: {xlsx_path}")
    logger.info(f"{'='*70}\n")
    
    xlsx_path = Path(xlsx_path)
    
    if not xlsx_path.exists():
        logger.error(f"File not found: {xlsx_path}")
        return
    
    wb = openpyxl.load_workbook(str(xlsx_path))
    
    # Analyze Index sheet
    if 'Index' in wb.sheetnames:
        index_ws = wb['Index']
        logger.info("Index Sheet:")
        logger.info(f"  Total rows: {index_ws.max_row}")
        
        # Count unique (Section + Table Title) combinations
        section_title_map = {}
        
        for row_idx, row in enumerate(index_ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:  # Header row
                headers = [str(h or '').strip() for h in row]
                section_idx = headers.index('Section') if 'Section' in headers else -1
                title_idx = headers.index('Table Title') if 'Table Title' in headers else -1
                link_idx = headers.index('Link') if 'Link' in headers else -1
                continue
            
            if section_idx >= 0 and title_idx >= 0:
                section = str(row[section_idx] or '').strip()
                title = str(row[title_idx] or '').strip()
                link = str(row[link_idx] or '').replace('→', '').strip() if link_idx >= 0 else ''
                
                key = (section, title)
                if key not in section_title_map:
                    section_title_map[key] = []
                section_title_map[key].append(link)
        
        logger.info(f"  Unique (Section, Title) combinations: {len(section_title_map)}")
        logger.info(f"  Total Index entries: {index_ws.max_row - 1}")
        
        # Show groups with multiple entries
        multi_entry_groups = {k: v for k, v in section_title_map.items() if len(v) > 1}
        if multi_entry_groups:
            logger.info(f"\n  Groups with multiple entries ({len(multi_entry_groups)}):")
            for (section, title), links in sorted(multi_entry_groups.items())[:10]:  # Show first 10
                logger.info(f"    [{section}] {title}: {len(links)} entries → sheets {', '.join(set(links))}")
            if len(multi_entry_groups) > 10:
                logger.info(f"    ... and {len(multi_entry_groups) - 10} more")
    
    # Analyze data sheets
    logger.info(f"\n  Total sheets in workbook: {len(wb.sheetnames)}")
    logger.info(f"  Sheet names: {', '.join(wb.sheetnames[:20])}")
    if len(wb.sheetnames) > 20:
        logger.info(f"    ... and {len(wb.sheetnames) - 20} more")
    
    # Analyze a sample sheet for table blocks
    sample_sheets = [name for name in wb.sheetnames if name != 'Index' and not name.startswith('~')][:3]
    
    if sample_sheets:
        logger.info(f"\n  Analyzing sample sheets for table blocks:")
        for sheet_name in sample_sheets:
            ws = wb[sheet_name]
            blocks = BlockDetector.detect_blocks(ws)
            logger.info(f"    Sheet '{sheet_name}': {len(blocks)} block(s) detected")
            for i, block in enumerate(blocks, 1):
                logger.info(f"      Block {i}: rows {block.data_start_row}-{block.data_end_row}, "
                          f"has_metadata={block.has_metadata}, title={block.table_title}")


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Test Index Sheet Resequencer")
    parser.add_argument('--mode', choices=['single', 'all', 'analyze'], default='analyze',
                       help='Test mode: single file, all files, or analyze structure')
    parser.add_argument('--file', type=str, default='data/processed/10q0925_tables.xlsx',
                       help='Path to xlsx file (for single mode)')
    parser.add_argument('--dir', type=str, default='data/processed',
                       help='Directory containing xlsx files (for all mode)')
    
    args = parser.parse_args()
    
    if args.mode == 'single':
        test_single_file(args.file)
    elif args.mode == 'all':
        test_all_files(args.dir)
    elif args.mode == 'analyze':
        # Analyze structure first
        analyze_file_structure(args.file)
        
        # Ask if user wants to proceed with processing
        print("\n" + "="*70)
        response = input("Do you want to proceed with processing this file? (y/n): ")
        if response.lower() == 'y':
            test_single_file(args.file)
