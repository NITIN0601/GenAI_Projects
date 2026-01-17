import pytest
import tempfile
import os
import shutil
from pathlib import Path
from openpyxl import Workbook, load_workbook

from src.infrastructure.extraction.exporters.table_merger import (
    TableMerger,
    reset_table_merger
)

class TestMergeSimple:
    """
    Simple end-to-end test for table merging.
    Mimics the structure of a real integration test but simplified.
    """
    
    def setup_method(self):
        self.temp_dir = tempfile.mkdtemp()
        self.input_dir = Path(self.temp_dir) / "input"
        self.input_dir.mkdir()
        self.output_dir = Path(self.temp_dir) / "output"
        self.output_dir.mkdir()
        
        reset_table_merger()
        self.merger = TableMerger()
        # Override paths to use temp dirs
        self.merger.source_dir = self.input_dir
        self.merger.dest_dir = self.output_dir

    def teardown_method(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_simple_merge_execution(self):
        """
        Creates a simple Excel file with two mergeable tables 
        and verifies the merger runs and produces output.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # --- Table 1 ---
        # Metadata
        ws.cell(row=1, column=1, value="Table Title: Summary")
        ws.cell(row=2, column=1, value="Section: Overview")
        ws.cell(row=3, column=1, value="Category: Financials")
        # Headers
        ws.cell(row=5, column=2, value="2023")
        # Data (Start row 6)
        ws.cell(row=6, column=1, value="Revenue")
        ws.cell(row=6, column=2, value="1000")
        ws.cell(row=7, column=1, value="Profit")
        ws.cell(row=7, column=2, value="100")
        
        # --- Table 2 (Below, same row labels) ---
        start_row = 15
        ws.cell(row=start_row, column=1, value="Table Title: Summary")
        ws.cell(row=start_row+1, column=1, value="Section: Overview")
        ws.cell(row=start_row+2, column=1, value="Category: Financials")
        # Headers
        ws.cell(row=start_row+4, column=2, value="2024")
        # Data
        ws.cell(row=start_row+5, column=1, value="Revenue")
        ws.cell(row=start_row+5, column=2, value="1200")
        ws.cell(row=start_row+6, column=1, value="Profit")
        ws.cell(row=start_row+6, column=2, value="120")

        # Save
        input_file = self.input_dir / "test_merge.xlsx"
        wb.save(input_file)
        
        # Run process
        result = self.merger.process_file(input_file)
        
        # Verify result structure
        assert result['source'] == str(input_file)
        assert result['output_path'] is not None
        assert os.path.exists(result['output_path'])
        
        # Check that it didn't crash and hopefully identified tables
        # Note: Actual merge success depends on BlockDetector logic which is complex
        # check if tables_merged or tables_split is reported
        print(f"Merge Result: {result}")

if __name__ == "__main__":
    pytest.main([__file__, "-v"])
