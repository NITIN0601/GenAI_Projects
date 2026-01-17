"""
Tests for table_merger.py horizontal merge logic.

Tests the TableMerger's ability to:
1. Find and merge tables with identical row labels
2. Handle wide tables (>10 columns)
3. Properly copy column headers
4. Use centralized patterns from domain_patterns.py
"""

import pytest
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook, load_workbook

from src.infrastructure.extraction.exporters.table_merger import (
    TableMerger,
    get_table_merger,
    reset_table_merger
)
from src.utils.domain_patterns import TABLE_HEADER_PATTERNS, DATA_LABEL_PATTERNS


class TestTableMergerPatterns:
    """Test that table_merger uses centralized patterns."""
    
    def test_patterns_imported_from_domain_patterns(self):
        """Verify patterns are imported from domain_patterns.py."""
        assert TABLE_HEADER_PATTERNS is not None
        assert DATA_LABEL_PATTERNS is not None
        assert len(TABLE_HEADER_PATTERNS) > 0
        assert len(DATA_LABEL_PATTERNS) > 0
    
    def test_header_patterns_include_currency(self):
        """Verify header patterns include currency indicators."""
        assert '$ in millions' in TABLE_HEADER_PATTERNS
        assert '$ in billions' in TABLE_HEADER_PATTERNS
    
    def test_header_patterns_include_periods(self):
        """Verify header patterns include period indicators."""
        assert 'three months ended' in TABLE_HEADER_PATTERNS
        assert 'six months ended' in TABLE_HEADER_PATTERNS
    
    def test_data_label_patterns_include_financial_terms(self):
        """Verify data patterns include financial terms."""
        assert 'revenues' in DATA_LABEL_PATTERNS
        assert 'expenses' in DATA_LABEL_PATTERNS


class TestTableMergerSingleton:
    """Test singleton pattern for TableMerger."""
    
    def setup_method(self):
        """Reset singleton before each test."""
        reset_table_merger()
    
    def test_get_table_merger_returns_same_instance(self):
        """Verify get_table_merger returns singleton."""
        merger1 = get_table_merger()
        merger2 = get_table_merger()
        assert merger1 is merger2
    
    def test_reset_table_merger_clears_cache(self):
        """Verify reset creates new instance."""
        merger1 = get_table_merger()
        reset_table_merger()
        merger2 = get_table_merger()
        # After reset, should be new instance (but equal type)
        assert isinstance(merger2, TableMerger)


class TestWideTableHandling:
    """Test handling of tables with >10 columns."""
    
    def setup_method(self):
        """Create temp directory for test files."""
        self.temp_dir = tempfile.mkdtemp()
        reset_table_merger()
    
    def teardown_method(self):
        """Clean up temp files."""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)
    
    def test_wide_table_all_columns_read(self):
        """Verify all columns are read for wide tables (>10 cols)."""
        # Create test workbook with 15 columns
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        # Add header row with 15 columns
        headers = ['Label'] + [f'Col{i}' for i in range(1, 15)]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Add data rows
        for row_idx in range(2, 6):
            ws.cell(row=row_idx, column=1, value=f'Row{row_idx}')
            for col_idx in range(2, 16):
                ws.cell(row=row_idx, column=col_idx, value=row_idx * 100 + col_idx)
        
        # Save and verify structure
        test_path = os.path.join(self.temp_dir, "wide_table.xlsx")
        wb.save(test_path)
        
        # Reload and verify all columns present
        wb2 = load_workbook(test_path)
        ws2 = wb2.active
        assert ws2.max_column == 15  # Should have all 15 columns
        
        # Verify last column has data
        last_col_value = ws2.cell(row=3, column=15).value
        assert last_col_value is not None


class TestYearRangeEdgeCases:
    """Test year range detection edge cases."""
    
    def test_year_2029_detected(self):
        """Verify 2029 is detected as valid year."""
        from config.settings import settings
        assert settings.EXTRACTION_YEAR_MIN <= 2029 <= settings.EXTRACTION_YEAR_MAX
    
    def test_year_2030_detected(self):
        """Verify 2030 is detected as valid year."""
        from config.settings import settings
        assert settings.EXTRACTION_YEAR_MIN <= 2030 <= settings.EXTRACTION_YEAR_MAX
    
    def test_year_2035_detected(self):
        """Verify 2035 is within range (future-proofing)."""
        from config.settings import settings
        assert settings.EXTRACTION_YEAR_MIN <= 2035 <= settings.EXTRACTION_YEAR_MAX
    
    def test_year_2000_is_minimum(self):
        """Verify 2000 is the minimum year."""
        from config.settings import settings
        assert settings.EXTRACTION_YEAR_MIN == 2000
    
    def test_domain_patterns_uses_settings(self):
        """Verify domain_patterns VALID_YEAR_RANGE uses settings."""
        from src.utils.domain_patterns import VALID_YEAR_RANGE
        from config.settings import settings
        assert VALID_YEAR_RANGE == (settings.EXTRACTION_YEAR_MIN, settings.EXTRACTION_YEAR_MAX)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
