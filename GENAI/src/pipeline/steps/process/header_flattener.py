"""
Header flattener for the Process step.

Functions for flattening multi-level headers into single header rows.
"""

from typing import List, Dict, Any

from openpyxl.cell.cell import MergedCell

from src.utils import get_logger
from src.pipeline.steps.process.constants import DEFAULT_HEADER_START_ROW
from src.pipeline.steps.process.table_finder import find_data_start_row
from src.pipeline.steps.process.header_builders import (
    build_combined_headers_3level,
    build_combined_headers_4level,
    build_flattened_headers,
    count_header_rows,
)

logger = get_logger(__name__)


def flatten_table_headers_dynamic(
    ws, 
    header_row: int, 
    data_header_rows: List[List], 
    normalized_headers: List[str], 
    stats: Dict[str, Any]
) -> None:
    """
    Flatten multi-level headers for a specific table starting at header_row.
    
    This is the dynamic version that works with any starting row, enabling
    processing of multiple tables per sheet.
    
    Args:
        ws: openpyxl Worksheet object
        header_row: The row where this table's headers start
        data_header_rows: List of header row values
        normalized_headers: List of normalized header values
        stats: Stats dictionary to update
    """
    num_cols = ws.max_column
    num_header_rows = len(data_header_rows)
    
    if num_header_rows == 0:
        return
    
    logger.debug(f"Flattening {num_header_rows} header rows starting at row {header_row}")
    
    # Unmerge any merged cells in the header rows for this table
    merged_ranges_to_unmerge = []
    for merge_range in ws.merged_cells.ranges:
        if header_row <= merge_range.min_row <= header_row + num_header_rows:
            merged_ranges_to_unmerge.append(merge_range)
    
    for merge_range in merged_ranges_to_unmerge:
        try:
            ws.unmerge_cells(str(merge_range))
        except Exception as e:
            logger.debug(f"Could not unmerge {merge_range}: {e}")
    
    if num_header_rows == 1:
        # 1-level: Write normalized headers to the header row
        # DO NOT add empty separator - data starts immediately after
        for col_idx in range(1, num_cols + 1):
            if col_idx < len(normalized_headers) + 1 and normalized_headers[col_idx - 1]:
                cell = ws.cell(row=header_row, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = normalized_headers[col_idx - 1]
        
        # For 1-level headers, the next row is DATA, not a header to remove
        # So we DON'T set it to empty or add a separator
    
    elif num_header_rows == 2:
        # 2-level: Combine into single row + empty separator
        for col_idx in range(1, num_cols + 1):
            if col_idx < len(normalized_headers) + 1 and normalized_headers[col_idx - 1]:
                cell = ws.cell(row=header_row, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = normalized_headers[col_idx - 1]
        
        # Row header_row+1 becomes empty separator
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=header_row + 1, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = None
        
        stats['headers_flattened'] = stats.get('headers_flattened', 0) + 1
        stats['empty_rows_added'] = stats.get('empty_rows_added', 0) + 1
    
    else:
        # 3+ levels: Build combined headers
        if num_header_rows == 3:
            combined = build_combined_headers_3level(data_header_rows, normalized_headers, num_cols)
        else:
            combined = build_combined_headers_4level(data_header_rows, normalized_headers, num_cols)
        
        # Write combined headers to header_row
        for col_idx, value in enumerate(combined, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            if not isinstance(cell, MergedCell):
                cell.value = value
        
        # Row header_row+1 becomes empty separator
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=header_row + 1, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = None
        
        # Delete remaining header rows (shift data up)
        # We want: header_row (combined) + header_row+1 (empty) + data
        # So delete rows header_row+2 onwards up to the original data start
        rows_to_delete = num_header_rows - 2
        for _ in range(rows_to_delete):
            ws.delete_rows(header_row + 2)
        
        stats['headers_flattened'] = stats.get('headers_flattened', 0) + 1
        stats['empty_rows_added'] = stats.get('empty_rows_added', 0) + 1
    
    logger.debug(f"Flattened headers at row {header_row}")


def flatten_table_headers(
    ws, 
    data_header_rows: List[List], 
    normalized_headers: List[str], 
    stats: Dict[str, Any]
) -> None:
    """
    Flatten multi-level headers into a single header row with empty separator.
    
    Handles:
    - 1-level: Add empty row separator after header
    - 2-level: L1 (period) + L2 (year) → single row + empty separator
    - 3-level: L1 (period) + L2 (year) + L3 (category) → combined row + empty separator
    - 4-level: L1 (period) + L2 (year) + L3 (year) + L4 (category) → combined row + empty separator
    
    Format Rules:
        At/As of dates      → Q1-2025, Q2-2025, Q3-2025, Q4-2025
        Three Months Ended  → Q1-QTD-2025, Q2-QTD-2025
        Six Months Ended    → Q2-YTD-2025
        Nine Months Ended   → Q3-YTD-2025
        Year Ended          → YTD-2025
    
    This ONLY modifies header rows, NOT data values.
    
    Args:
        ws: openpyxl Worksheet object
        data_header_rows: List of header row data
        normalized_headers: List of normalized header values
        stats: Stats dictionary to update
    """
    # Count actual header rows (between Source: and first data row)
    header_count = count_header_rows(ws)
    
    # Find where data starts (first row with numeric values)
    data_start_row = find_data_start_row(ws)
    if not data_start_row:
        return  # Can't determine data start
    
    num_cols = ws.max_column
    header_start = DEFAULT_HEADER_START_ROW  # Headers typically start at row 13
    
    # Calculate rows to flatten
    rows_to_flatten = data_start_row - header_start if data_start_row > header_start else 0
    
    logger.debug(f"Header analysis: {rows_to_flatten} header rows (row {header_start} to {data_start_row - 1}), data starts at row {data_start_row}")
    
    # Unmerge any merged cells in the header rows
    merged_ranges_to_unmerge = []
    for merge_range in ws.merged_cells.ranges:
        if merge_range.min_row >= header_start and merge_range.min_row < data_start_row:
            merged_ranges_to_unmerge.append(str(merge_range))
    
    for merge_range in merged_ranges_to_unmerge:
        try:
            ws.unmerge_cells(merge_range)
        except Exception as e:
            logger.debug(f"Could not unmerge {merge_range}: {e}")
    
    rows_deleted = 0
    
    if rows_to_flatten <= 1:
        # === 1-LEVEL: Already flat, just add empty row separator ===
        # Insert empty row after header (row 14)
        ws.insert_rows(header_start + 1)
        # Clear the new row
        for col in range(1, num_cols + 1):
            ws.cell(row=header_start + 1, column=col).value = None
        logger.debug("1-level: Added empty row separator after header")
        stats['empty_rows_added'] = stats.get('empty_rows_added', 0) + 1
        return
    
    elif rows_to_flatten == 2:
        # === 2-LEVEL: Period + Year → single row + empty separator ===
        flattened = build_flattened_headers(ws, header_start, data_start_row, num_cols, normalized_headers)
        
        # Write flattened headers to row 13
        for col_idx, value in enumerate(flattened, start=1):
            try:
                cell = ws.cell(row=header_start, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
            except Exception as e:
                logger.debug(f"Could not write to row 13, col {col_idx}: {e}")
        
        # Clear row 14 (make it the empty separator) instead of deleting
        for col in range(1, num_cols + 1):
            try:
                ws.cell(row=header_start + 1, column=col).value = None
            except (AttributeError, ValueError):
                pass
        
        logger.debug("2-level: Flattened headers, row 14 is now empty separator")
        
    elif rows_to_flatten == 3:
        # === 3-LEVEL: Period + Year + Category → combined row + empty separator ===
        # Pattern: Row 13 (date), Row 14 (date dup or category labels), Row 15 (categories)
        
        # Collect all header row data
        header_rows_data = []
        for row in range(header_start, data_start_row):
            row_data = [ws.cell(row, c).value for c in range(1, num_cols + 1)]
            header_rows_data.append(row_data)
        
        # Build combined headers: date_code + category
        combined = build_combined_headers_3level(
            header_rows_data, normalized_headers, num_cols
        )
        
        # Write combined headers to row 13
        for col_idx, value in enumerate(combined, start=1):
            try:
                ws.cell(row=header_start, column=col_idx).value = value
            except Exception as e:
                logger.debug(f"Could not write to row 13, col {col_idx}: {e}")
        
        # Row 14 becomes empty separator
        for col in range(1, num_cols + 1):
            try:
                ws.cell(row=header_start + 1, column=col).value = None
            except (AttributeError, ValueError):
                pass
        
        # Delete remaining header rows (row 15 onwards)
        rows_to_delete = rows_to_flatten - 2
        for _ in range(rows_to_delete):
            ws.delete_rows(header_start + 2)
        rows_deleted = rows_to_delete
        
        logger.debug(f"3-level: Combined headers, deleted {rows_deleted} rows")
        
    else:
        # === 4+ LEVEL: Period + Year + Year + Category → combined row + empty separator ===
        # Collect all header row data
        header_rows_data = []
        for row in range(header_start, data_start_row):
            row_data = [ws.cell(row, c).value for c in range(1, num_cols + 1)]
            header_rows_data.append(row_data)
        
        # Build combined headers for 4+ levels
        combined = build_combined_headers_4level(
            header_rows_data, normalized_headers, num_cols
        )
        
        # Write combined headers to row 13
        for col_idx, value in enumerate(combined, start=1):
            try:
                ws.cell(row=header_start, column=col_idx).value = value
            except Exception as e:
                logger.debug(f"Could not write to row 13, col {col_idx}: {e}")
        
        # Row 14 becomes empty separator
        for col in range(1, num_cols + 1):
            try:
                ws.cell(row=header_start + 1, column=col).value = None
            except (AttributeError, ValueError):
                pass
        
        # Delete remaining header rows
        rows_to_delete = rows_to_flatten - 2
        for _ in range(rows_to_delete):
            ws.delete_rows(header_start + 2)
        rows_deleted = rows_to_delete
        
        logger.debug(f"4+-level: Combined headers, deleted {rows_deleted} rows")
    
    stats['headers_flattened'] = stats.get('headers_flattened', 0) + 1
    stats['empty_rows_added'] = stats.get('empty_rows_added', 0) + 1
    logger.debug(f"Flattened {rows_to_flatten} header rows, deleted {rows_deleted} rows, added empty separator")
