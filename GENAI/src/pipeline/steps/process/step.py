"""
Process Step - Data processing and normalization.

Step 2 in pipeline: Extract → Process → Process-Advanced → Consolidate → Transpose

Implements ALL operations per docs/pipeline_operations.md (2.1-2.15):
- 2.1  OCR Broken Words Fix
- 2.2  Footnote Reference Cleanup
- 2.3  Row Label Normalization
- 2.4  Title Normalization
- 2.5  Currency → Float
- 2.6  Negative (Parens) → Float
- 2.7  Percentage → Float
- 2.8  Excel Currency Format ($#,##0.00)
- 2.9  Excel Percent Format (0.00%)
- 2.10 Year String Cleanup
- 2.11 L1/L2/L3/L4 Detection
- 2.12 Y/Q Code Formatting
- 2.13 Y/Q Period Detection
- 2.14 Fill Y/Q Placeholder (Row 8)
- 2.15 Index Update
"""

import re
from pathlib import Path
from typing import Dict, Any, Optional, List

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from src.pipeline.base import StepInterface, StepResult, StepStatus, PipelineContext
from src.utils import get_logger
from src.utils.multi_row_header_normalizer import normalize_headers as normalize_multi_row_headers

# Import from modular sub-modules
from src.pipeline.steps.process.constants import TABLE_FILE_PATTERN
from src.pipeline.steps.process.table_finder import (
    find_all_tables,
    find_first_data_row_after_source,
)
from src.pipeline.steps.process.cell_processor import (
    process_cell_value,
    is_numeric_value,
)
from src.pipeline.steps.process.key_value_handler import (
    is_key_value_table,
    process_data_cells_only,
)
from src.pipeline.steps.process.header_flattener import flatten_table_headers_dynamic

logger = get_logger(__name__)

# Import from consolidated header normalizer module - single source of truth
from src.utils.header_normalizer import (
    MONTH_TO_QUARTER_MAP,
    normalize_point_in_time_header,
    is_valid_date_code,
    convert_year_to_period,
    combine_category_with_period,
    combine_period_with_dates,
    extract_quarter_from_header,
    extract_year_from_header,
)

# Re-export for backward compatibility (other modules import from here)



def is_valid_date_code(code: str) -> bool:
    """
    Check if a string is a valid normalized date code.
    
    Valid codes:
    - Q1-2024, Q2-2024, Q3-2024, Q4-2024 (point-in-time)
    - Q1-QTD-2024, Q2-YTD-2024 (period-based)
    - YTD-2024 (annual)
    
    Args:
        code: String to validate
        
    Returns:
        True if the code is a valid date code
    """
    if not code:
        return False
    return bool(
        re.match(r'^Q[1-4](-QTD|-YTD)?-20\d{2}', code) or
        re.match(r'^YTD-20\d{2}', code)
    )

class ProcessStep(StepInterface):
    """
    Process step - normalize and clean extracted data.
    
    Implements StepInterface following pipeline pattern.
    
    Reads: data/extracted_raw/*.xlsx (output from extract step)
    Writes: data/processed/*.xlsx
    
    Operations performed (per docs/pipeline_operations.md):
    - OCR broken words fix (2.1)
    - Footnote cleanup (2.2)
    - Row label normalization (2.3)
    - Title normalization (2.4)
    - Currency → float conversion with Excel format (2.5, 2.6, 2.8)
    - Percentage → float conversion with Excel format (2.7, 2.9)
    - Year string cleanup (2.10)
    - Multi-level header parsing (2.11)
    - Y/Q code formatting (2.12, 2.13)
    - Y/Q placeholder fill in Row 8 (2.14)
    """
    
    name = "process"
    
    def __init__(self, source_dir: Optional[str] = None, dest_dir: Optional[str] = None):
        """
        Initialize step with optional directory overrides.
        
        Args:
            source_dir: Override source directory (default: data/extracted_raw)
            dest_dir: Override destination directory (default: data/processed)
        """
        self.source_dir = source_dir
        self.dest_dir = dest_dir
    
    def validate(self, context: PipelineContext) -> bool:
        """Validate that source directory exists and has xlsx files."""
        from src.core import get_paths
        
        paths = get_paths()
        # Step 2: reads from extracted_raw (must match execute() source)
        source_path = Path(self.source_dir) if self.source_dir else Path(paths.data_dir) / "extracted_raw"
        
        if not source_path.exists():
            logger.warning(f"Source directory does not exist: {source_path}")
            return False
        
        xlsx_files = list(source_path.glob(TABLE_FILE_PATTERN))
        if not xlsx_files:
            logger.warning(f"No xlsx files in {source_path}")
            return False
        
        return True
    
    def get_step_info(self) -> Dict[str, Any]:
        """Get step metadata."""
        return {
            "name": self.name,
            "description": "Normalize and clean extracted data with Excel formatting",
            "reads": ["data/extracted_raw/*.xlsx"],
            "writes": ["data/processed/*.xlsx"],
            "operations": [
                "2.1 OCR broken words fix",
                "2.2 Footnote cleanup",
                "2.5-2.6 Currency → Float",
                "2.7 Percentage → Float",
                "2.8-2.9 Excel cell formatting",
                "2.12-2.14 Y/Q code formatting and placeholder fill"
            ]
        }
    
    def execute(self, context: PipelineContext) -> StepResult:
        """Execute processing step."""
        from src.core import get_paths
        
        paths = get_paths()
        # Step 2: reads from extracted_raw, writes to processed
        source_path = Path(self.source_dir) if self.source_dir else Path(paths.data_dir) / "extracted_raw"
        dest_path = Path(self.dest_dir) if self.dest_dir else Path(paths.data_dir) / "processed"
        
        dest_path.mkdir(parents=True, exist_ok=True)
        
        xlsx_files = list(source_path.glob(TABLE_FILE_PATTERN))
        
        stats = {
            'files_processed': 0,
            'cells_formatted': 0,
            'yq_filled': 0,
            'errors': []
        }
        
        for xlsx_path in xlsx_files:
            try:
                self._process_file(xlsx_path, dest_path, stats)
                stats['files_processed'] += 1
            except Exception as e:
                logger.error(f"Error processing {xlsx_path.name}: {e}")
                stats['errors'].append(f"{xlsx_path.name}: {str(e)}")
        
        status = StepStatus.SUCCESS if not stats['errors'] else StepStatus.PARTIAL_SUCCESS
        
        return StepResult(
            step_name=self.name,
            status=status,
            data=stats,
            message=f"Processed {stats['files_processed']} files, formatted {stats['cells_formatted']} cells",
            metadata={
                'source_dir': str(source_path),
                'dest_dir': str(dest_path)
            }
        )
    
    def _process_file(self, source_path: Path, dest_path: Path, stats: Dict) -> None:
        """Process a single xlsx file."""
        # Detect if this is a 10K (annual) report
        is_10k = '10k' in source_path.name.lower()
        
        wb = load_workbook(source_path)
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ['Index', 'TOC', 'TOC_Sheet']:
                continue
            
            ws = wb[sheet_name]
            self._process_sheet(ws, stats, is_10k=is_10k)
        
        # Save to destination
        output_path = dest_path / source_path.name
        wb.save(output_path)
        wb.close()
        
        logger.debug(f"Processed {source_path.name} (is_10k={is_10k}) -> {output_path}")
    
    def _process_sheet(self, ws, stats: Dict, is_10k: bool = False) -> None:
        """
        Process a single worksheet - finds ALL tables dynamically and processes each.
        
        Tables are identified by "Table Title:" or "Source:" markers.
        Each table is processed independently for header flattening.
        
        Args:
            ws: Worksheet to process
            stats: Stats dictionary
            is_10k: True if this is from a 10K (annual) report
        """
        # === Check if this is a KEY-VALUE TABLE (should skip header processing) ===
        if is_key_value_table(ws):
            logger.debug("Skipping header processing for key-value table")
            process_data_cells_only(ws, stats, process_cell_value)
            return
        
        # === DYNAMIC TABLE DETECTION ===
        # Find all tables in the sheet by looking for "Source:" markers
        tables = find_all_tables(ws)
        
        if not tables:
            # Fall back to processing the whole sheet as one table
            # Dynamically find where data starts after Source(s): marker
            fallback_start = find_first_data_row_after_source(ws)
            tables = [{'start_row': fallback_start, 'end_row': ws.max_row, 'header_row': fallback_start}]
        
        logger.debug(f"Found {len(tables)} table(s) in sheet (is_10k={is_10k})")
        
        # Process each table independently
        for table_idx, table_info in enumerate(tables):
            self._process_single_table(ws, table_info, stats, is_10k=is_10k)
    
    def _process_single_table(
        self, 
        ws, 
        table_info: Dict, 
        stats: Dict, 
        is_10k: bool = False
    ) -> None:
        """
        Process a single table within the worksheet.
        
        Args:
            ws: The worksheet
            table_info: Dict with 'start_row', 'end_row', 'header_row', 'source_row'
            stats: Stats dict to update
            is_10k: True if this is from a 10K (annual) report
        """
        header_row = table_info.get('header_row', 13)
        end_row = table_info.get('end_row', ws.max_row)
        start_row = table_info.get('start_row', 1)
        
        # === UPDATE METADATA ROWS WITH NORMALIZED HEADERS ===
        # Find and update Column Header L2 row (typically row 6)
        # Also read L1 for date context to combine with L2 category headers
        # Metadata rows are at the TOP of the sheet (rows 1-10), before table data starts
        
        # First, extract L1 content for date context
        l1_content = ''
        l1_normalized = None
        for row_idx in range(1, 12):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val and 'Column Header L1' in str(cell_val):
                l1_content = str(cell_val).split(':', 1)[1].strip() if ':' in str(cell_val) else ''
                # Try to normalize L1 if it has date patterns
                if l1_content:
                    l1_normalized = normalize_point_in_time_header(l1_content)
                    if not l1_normalized and re.match(r'^20\d{2}$', l1_content.strip()):
                        l1_normalized = f"YTD-{l1_content.strip()}" if is_10k else l1_content.strip()
                break
        
        # Now process L2
        for row_idx in range(1, 12):  # Check rows 1-11 for metadata
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val and 'Column Header L2' in str(cell_val):
                # Found the Column Header L2 row - normalize its content
                old_content = str(cell_val)
                if ':' in old_content:
                    prefix, raw_headers = old_content.split(':', 1)
                    raw_list = [h.strip() for h in raw_headers.split(',') if h.strip()]
                    
                    # Recombine split date parts: ['At June 30', '2024'] → ['At June 30, 2024']
                    # This happens when dates like "June 30, 2024" get split by comma
                    recombined = []
                    i = 0
                    while i < len(raw_list):
                        item = raw_list[i]
                        # Check if this is a date prefix without year (At/As/Months pattern + month/day)
                        has_date_prefix = re.search(r'(at|as of|months ended)\s+\w+\s+\d+', item.lower())
                        has_year = bool(re.search(r'20\d{2}', item))
                        
                        if has_date_prefix and not has_year and i + 1 < len(raw_list):
                            # Check if next item is a year
                            next_item = raw_list[i + 1]
                            if re.match(r'^20\d{2}$', next_item.strip()):
                                # Combine: "At June 30" + "2024" → "At June 30, 2024"
                                recombined.append(f"{item}, {next_item}")
                                i += 2
                                continue
                        
                        recombined.append(item)
                        i += 1
                    
                    raw_list = recombined
                    
                    # Determine source type (10-K vs 10-Q) from worksheet name or parent
                    source_doc = ''
                    if hasattr(ws, 'parent') and hasattr(ws.parent, 'path') and ws.parent.path:
                        source_doc = str(ws.parent.path)
                    
                    # Normalize each header using all available methods
                    normalized_list = []
                    for h in raw_list:
                        # Try normalize_point_in_time_header first (handles At, Months Ended, etc.)
                        norm = normalize_point_in_time_header(h)
                        if norm:
                            normalized_list.append(norm)
                        # Handle year-only values - use is_10k parameter for correct period
                        elif re.match(r'^20\d{2}$', h.strip()):
                            year = h.strip()
                            converted = f"YTD-{year}" if is_10k else year
                            normalized_list.append(converted)
                        else:
                            normalized_list.append(h)
                    
                    # Check RAW patterns to see if L3 combination is needed
                    # (Check raw_list, not normalized_list, since At patterns get normalized)
                    needs_l3_combo = any('Months Ended' in h and not re.search(r'20\d{2}', h) for h in raw_list)
                    if needs_l3_combo:
                        # Look for Column Header L3 with years/dates (should be right after L2)
                        for l3_row in range(row_idx + 1, row_idx + 5):
                            l3_val = ws.cell(row=l3_row, column=1).value
                            if l3_val and 'Column Header L3' in str(l3_val):
                                l3_content = str(l3_val).split(':', 1)[1] if ':' in str(l3_val) else ''
                                l3_parts = [y.strip() for y in l3_content.split(',') if y.strip()]
                                
                                # Recombine split date parts in L3: ['December 31', '2024'] → ['December 31, 2024']
                                l3_recombined = []
                                i = 0
                                while i < len(l3_parts):
                                    part = l3_parts[i]
                                    # Check if this is a month+day without year
                                    has_month = bool(re.search(r'(january|february|march|april|may|june|july|august|september|october|november|december)', part.lower()))
                                    has_year = bool(re.search(r'20\d{2}', part))
                                    
                                    if has_month and not has_year and i + 1 < len(l3_parts):
                                        # Next part might be the year
                                        next_part = l3_parts[i + 1]
                                        if re.match(r'^20\d{2}$', next_part.strip()):
                                            l3_recombined.append(f"{part}, {next_part}")
                                            i += 2
                                            continue
                                    
                                    l3_recombined.append(part)
                                    i += 1
                                
                                # Extract both full dates and year-only values
                                full_dates = [d for d in l3_recombined if re.search(r'(january|february|march|april|may|june|july|august|september|october|november|december)', d.lower()) and re.search(r'20\d{2}', d)]
                                years = [y for y in l3_recombined if re.match(r'^20\d{2}$', y.strip())]
                                
                                if full_dates or years:
                                    # Combine period headers with dates/years and normalize
                                    combined_normalized = []
                                    # Prefer full dates, fall back to years
                                    date_values = full_dates if full_dates else years
                                    
                                    for norm_h in normalized_list:
                                        if 'Months Ended' in norm_h or 'At ' in norm_h:
                                            for date_val in date_values:
                                                combined = f"{norm_h.rstrip(',')} {date_val}"
                                                norm_combined = normalize_point_in_time_header(combined)
                                                if norm_combined and norm_combined not in combined_normalized:
                                                    combined_normalized.append(norm_combined)
                                        else:
                                            if norm_h and norm_h not in combined_normalized:
                                                combined_normalized.append(norm_h)
                                    
                                    if combined_normalized:
                                        normalized_list = combined_normalized
                                break
                    
                    # === HANDLE DESCRIPTIVE HEADERS WITH YEARS IN L3 ===
                    # If L2 has descriptive text (not date patterns) and L3 has years,
                    # append date context to descriptive headers
                    has_descriptive = any(
                        not is_valid_date_code(h) and 
                        not re.match(r'^20\d{2}$', h.strip()) and
                        not any(kw in h.lower() for kw in ['months ended', 'at ', 'as of'])
                        for h in normalized_list
                    )
                    
                    if has_descriptive:
                        # Look for years in Column Header L3
                        for l3_row in range(row_idx + 1, row_idx + 5):
                            l3_val = ws.cell(row=l3_row, column=1).value
                            if l3_val and 'Column Header L3' in str(l3_val):
                                l3_content = str(l3_val).split(':', 1)[1] if ':' in str(l3_val) else ''
                                l3_years = [y.strip() for y in l3_content.split(',') if re.match(r'^20\d{2}$', y.strip())]
                                
                                if l3_years:
                                    # Use is_10k from function parameter (already correctly detected)
                                    
                                    # Build new normalized list with date prefixes
                                    appended_list = []
                                    for h in normalized_list:
                                        if is_valid_date_code(h):
                                            # Already normalized - keep as is
                                            if h not in appended_list:
                                                appended_list.append(h)
                                        elif re.match(r'^20\d{2}$', h.strip()):
                                            # Year-only - convert to period
                                            period = f"YTD-{h.strip()}" if is_10k else h.strip()
                                            if period not in appended_list:
                                                appended_list.append(period)
                                        else:
                                            # Descriptive header - append each year
                                            for year in l3_years:
                                                period = f"YTD-{year}" if is_10k else year
                                                combined = f"{period} {h}"
                                                if combined not in appended_list:
                                                    appended_list.append(combined)
                                    
                                    if appended_list:
                                        normalized_list = appended_list
                                break
                    
                    # === COMBINE L1 + L2 WHEN L1 HAS DATE AND L2 HAS CATEGORIES ===
                    # If L1 has date context (l1_normalized) and L2 still has un-normalized
                    # descriptive headers, combine them: "Q4-2024" + "IS, WM" → "Q4-2024 IS, Q4-2024 WM"
                    if l1_normalized:
                        still_descriptive = any(
                            not is_valid_date_code(h) and 
                            not re.match(r'^20\d{2}$', h.strip())
                            for h in normalized_list
                        )
                        
                        if still_descriptive:
                            l1_combined = []
                            for h in normalized_list:
                                if is_valid_date_code(h):
                                    # Already normalized - keep as is
                                    if h not in l1_combined:
                                        l1_combined.append(h)
                                elif re.match(r'^20\d{2}$', h.strip()):
                                    # Year-only - convert using L1 context
                                    period = f"YTD-{h.strip()}" if is_10k else h.strip()
                                    if period not in l1_combined:
                                        l1_combined.append(period)
                                else:
                                    # Descriptive header - prepend L1 normalized date
                                    combined = f"{l1_normalized} {h}"
                                    if combined not in l1_combined:
                                        l1_combined.append(combined)
                            
                            if l1_combined:
                                normalized_list = l1_combined
                    
                    # Update the cell with normalized values
                    new_content = f"{prefix}: {', '.join(normalized_list)}"
                    if new_content != old_content:
                        ws.cell(row=row_idx, column=1).value = new_content
                        stats['cells_formatted'] = stats.get('cells_formatted', 0) + 1
                        
                        # Also update Year/Quarter row if we have normalized values
                        if any(is_valid_date_code(n) for n in normalized_list):
                            for yq_row in range(row_idx + 1, row_idx + 5):
                                yq_val = ws.cell(row=yq_row, column=1).value
                                if yq_val and 'Year/Quarter' in str(yq_val):
                                    # Build new Year/Quarter entries from normalized headers
                                    yq_entries = []
                                    for norm_h in normalized_list:
                                        if is_valid_date_code(norm_h):
                                            # Extract period and year from normalized code
                                            parts = norm_h.split('-')
                                            if len(parts) >= 2:
                                                if 'QTD' in norm_h or 'YTD' in norm_h:
                                                    # Q2-QTD-2024 → Q2-QTD,2024
                                                    period = f"{parts[0]}-{parts[1]}"
                                                    year = parts[2] if len(parts) > 2 else ''
                                                else:
                                                    # Q2-2024 → Q2,2024
                                                    period = parts[0]
                                                    year = parts[1]
                                                yq_entries.append(f"{period},{year}")
                                    
                                    if yq_entries:
                                        new_yq = f"Year/Quarter: {', '.join(yq_entries)}"
                                        ws.cell(row=yq_row, column=1).value = new_yq
                                        stats['cells_formatted'] = stats.get('cells_formatted', 0) + 1
                                    break
                break
        
        # PRE-SCAN: Normalize point-in-time headers ("At June 30, 2024" -> "Q2-2024")
        # This handles headers that appear BEFORE the detected header_row
        # (e.g., when a table has two sections with different header types)
        source_row = table_info.get('source_row', 11)
        for row_idx in range(source_row + 1, end_row + 1):
            for col_idx in range(1, min(ws.max_column + 1, 20)):  # Limit to first 20 columns
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    val = str(cell.value).strip()
                    val_lower = val.lower()
                    
                    normalized_val = normalize_point_in_time_header(val)
                    if normalized_val:
                        cell.value = normalized_val
                        stats['cells_formatted'] = stats.get('cells_formatted', 0) + 1
        
        # Extract header rows (typically 1-4 rows starting from header_row)
        data_header_rows = []
        for offset in range(4):  # Check up to 4 header rows
            row_idx = header_row + offset
            if row_idx > end_row:
                break
            
            row_values = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = str(cell.value) if cell.value else ''
                
                # For 10K reports: Convert year-only headers to YTD-YYYY
                if is_10k and val.strip() and re.match(r'^20\d{2}$', val.strip()):
                    val = f"YTD-{val.strip()}"
                    # Also update the cell directly
                    cell.value = val
                    stats['cells_formatted'] = stats.get('cells_formatted', 0) + 1
                
                # Normalize point-in-time headers: "At June 30, 2024" -> "Q2-2024"
                if val.strip():
                    normalized_val = normalize_point_in_time_header(val)
                    if normalized_val:
                        val = normalized_val
                        cell.value = normalized_val
                        stats['cells_formatted'] = stats.get('cells_formatted', 0) + 1
                
                row_values.append(val)
            
            # Check if this row has numeric data (means we've reached data rows)
            has_numeric = any(
                is_numeric_value(v) for v in row_values[1:5] if v
            )
            if has_numeric:
                break
            
            # Check if row has any content
            if any(row_values[1:]):
                data_header_rows.append(row_values)
        
        if not data_header_rows:
            return
        
        # Normalize headers
        source_filename = ''
        if hasattr(ws, 'parent') and hasattr(ws.parent, 'path') and ws.parent.path:
            source_filename = str(ws.parent.path)
        
        normalized = normalize_multi_row_headers(data_header_rows, source_filename)
        
        # Extract normalized headers
        l1_per_column = normalized.get('l1_headers', [])
        normalized_headers = normalized.get('normalized_headers', [])
        
        # For 10K reports: Update normalized_headers with YTD- prefix for year-only values
        # Also handle year + category (e.g., "2024 Average Monthly Balance" -> "YTD-2024 Average Monthly Balance")
        if is_10k:
            for i, val in enumerate(normalized_headers):
                if val:
                    # Pattern 1: year-only (e.g., "2024")
                    if re.match(r'^20\d{2}$', val.strip()):
                        normalized_headers[i] = f"YTD-{val.strip()}"
                    # Pattern 2: year + category (e.g., "2024 Average Monthly Balance")
                    elif re.match(r'^20\d{2}\s+\w', val.strip()):
                        year_match = re.match(r'^(20\d{2})\s+(.+)$', val.strip())
                        if year_match:
                            year = year_match.group(1)
                            suffix = year_match.group(2)
                            normalized_headers[i] = f"YTD-{year} {suffix}"
        
        num_cols = ws.max_column
        
        def safe_set_cell_value(ws, row: int, col: int, value: Any) -> bool:
            """Safely set cell value, skipping merged cells."""
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = value
                return True
            return False
        
        # === REPLACE actual data header rows with normalized Y/Q codes ===
        # Use dynamic row numbers based on header_row
        # For multi-row headers: put normalized code in first row, clear second row
        if len(data_header_rows) >= 1:
            for col_idx in range(2, num_cols + 1):
                norm_val = normalized_headers[col_idx - 1] if col_idx - 1 < len(normalized_headers) else ''
                
                # Apply normalized code if it's a valid date code OR starts with one
                # Valid codes: Q1-2024, Q2-QTD-2024, Q3-YTD-2024, YTD-2024
                # Also allow: YTD-2024 Average Monthly Balance (code + category suffix)
                is_valid = is_valid_date_code(norm_val) or (
                    norm_val and (
                        re.match(r'^Q[1-4](-QTD|-YTD)?-20\d{2}\s', norm_val) or  # Q-code with suffix
                        re.match(r'^YTD-20\d{2}\s', norm_val)  # YTD-year with suffix
                    )
                )
                if norm_val and is_valid:
                    if safe_set_cell_value(ws, header_row, col_idx, norm_val):
                        stats['cells_formatted'] = stats.get('cells_formatted', 0) + 1
        
        # Clear second header row for multi-row headers to prevent duplicates
        # The normalized code is now in row 1, row 2 should be empty or cleared
        if len(data_header_rows) > 1:
            for col_idx in range(2, num_cols + 1):
                orig_row2_val = data_header_rows[1][col_idx - 1] if col_idx - 1 < len(data_header_rows[1]) else ''
                norm_val = normalized_headers[col_idx - 1] if col_idx - 1 < len(normalized_headers) else ''
                
                # If row 2 has a year (e.g., '2024') and we have a normalized code,
                # clear row 2 since the full code is now in row 1
                if orig_row2_val and norm_val:
                    orig_str = str(orig_row2_val).strip()
                    is_year_only = re.match(r'^20\d{2}$', orig_str)
                    if is_year_only and is_valid_date_code(norm_val):
                        # Clear the year from row 2 since it's merged into row 1
                        safe_set_cell_value(ws, header_row + 1, col_idx, None)
                        stats['cells_formatted'] = stats.get('cells_formatted', 0) + 1
        
        # === FLATTEN HEADERS for this table ===
        flatten_table_headers_dynamic(ws, header_row, data_header_rows, normalized_headers, stats)
        
        # Process data cells for this table
        data_start = header_row + len(data_header_rows) + 1  # After headers + empty separator
        for row_idx in range(data_start, end_row + 1):
            # === DETECT AND NORMALIZE MID-TABLE HEADERS ===
            # Mid-table headers have empty/unit first col but date patterns in other cols
            first_cell = ws.cell(row_idx, 1).value
            first_val = str(first_cell).strip().lower() if first_cell else ''
            is_mid_table_header = first_val in ['', 'nan', 'none'] or first_val.startswith('$')
            
            if is_mid_table_header:
                # Check if this row has date patterns in columns 2+
                mid_header_row = []
                for col_idx in range(1, min(ws.max_column + 1, 15)):
                    cell_val = ws.cell(row_idx, col_idx).value
                    mid_header_row.append(str(cell_val) if cell_val else '')
                
                # If we find date patterns, try to normalize
                has_date_pattern = any(
                    any(kw in str(v).lower() for kw in ['months ended', 'at ', 'as of', 'june', 'march', 'september', 'december'])
                    for v in mid_header_row[1:5] if v
                )
                
                if has_date_pattern:
                    # Check if next row contains year values (e.g., '2024', '2023')
                    # Multi-row headers have period in current row, year in next row
                    header_rows_to_normalize = [mid_header_row]
                    
                    if row_idx + 1 <= end_row:
                        next_row = []
                        for col_idx in range(1, min(ws.max_column + 1, 15)):
                            cell_val = ws.cell(row_idx + 1, col_idx).value
                            next_row.append(str(cell_val) if cell_val else '')
                        
                        # Check if next row has year values (4-digit years like 2024, 2023)
                        has_year = any(
                            re.match(r'^20\d{2}$', str(v).strip()) 
                            for v in next_row[1:5] if v
                        )
                        if has_year:
                            header_rows_to_normalize.append(next_row)
                    
                    # Normalize using multi_row_header_normalizer
                    mid_normalized = normalize_multi_row_headers(header_rows_to_normalize, source_filename)
                    mid_norm_headers = mid_normalized.get('normalized_headers', [])
                    
                    # Write normalized values back to cells (first row gets the codes)
                    for col_idx in range(2, min(len(mid_norm_headers) + 1, ws.max_column + 1)):
                        norm_val = mid_norm_headers[col_idx - 1] if col_idx - 1 < len(mid_norm_headers) else ''
                        if norm_val and is_valid_date_code(norm_val):
                            if safe_set_cell_value(ws, row_idx, col_idx, norm_val):
                                stats['cells_formatted'] = stats.get('cells_formatted', 0) + 1
                            # Clear year from next row if we used it
                            if len(header_rows_to_normalize) > 1:
                                safe_set_cell_value(ws, row_idx + 1, col_idx, None)
                    continue  # Move to next row after normalizing header
            
            # Process regular data cells
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row_idx, col_idx)
                if cell.value is None:
                    continue
                
                original_value = cell.value
                new_value, cell_format = process_cell_value(original_value, row_idx, col_idx)
                
                if new_value != original_value:
                    cell.value = new_value
                    stats['cells_formatted'] = stats.get('cells_formatted', 0) + 1
                
                if cell_format:
                    cell.number_format = cell_format
