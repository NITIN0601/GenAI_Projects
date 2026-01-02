"""
Process Step - Data processing and normalization.

Step 2 in pipeline: Extract → Process → Process-Advanced → Consolidate → Transpose

Implements ALL operations per docs/pipeline_operations.md (2.1-2.15):
- 2.1  OCR Broken Words Fix
- 2.2  Footnote Reference Cleanup
- 2.3  Row Label Normalization
- 2.4  Title Normalization
- 2.5  Currency → Float
- 2.6  Negative (Parens) → Float
- 2.7  Percentage → Float
- 2.8  Excel Currency Format ($#,##0.00)
- 2.9  Excel Percent Format (0.00%)
- 2.10 Year String Cleanup
- 2.11 L1/L2/L3/L4 Detection
- 2.12 Y/Q Code Formatting
- 2.13 Y/Q Period Detection
- 2.14 Fill Y/Q Placeholder (Row 8)
- 2.15 Index Update
"""

from pathlib import Path
from typing import Dict, Any, Optional

from src.pipeline.base import StepInterface, StepResult, StepStatus, PipelineContext
from src.pipeline import PipelineResult, PipelineStep
from src.utils import get_logger
from src.utils.excel_utils import ExcelUtils
from src.utils.metadata_labels import MetadataLabels
from src.utils.quarter_mapper import QuarterDateMapper
from src.utils.header_parser import MultiLevelHeaderParser
from src.utils.multi_row_header_normalizer import normalize_headers as normalize_multi_row_headers
from src.utils.text_normalizer import normalize_text

logger = get_logger(__name__)

# File pattern for processed xlsx files
TABLE_FILE_PATTERN = "*_tables.xlsx"

# Excel number formats
CURRENCY_FORMAT = '$#,##0.00'
PERCENT_FORMAT = '0.00%'
NEGATIVE_CURRENCY_FORMAT = '$#,##0.00;[Red]($#,##0.00)'


class ProcessStep(StepInterface):
    """
    Process step - normalize and clean extracted data.
    
    Implements StepInterface following pipeline pattern.
    
    Reads: data/processed/*.xlsx (output from extract step)
    Writes: data/processed/*.xlsx
    
    Operations performed (per docs/pipeline_operations.md):
    - OCR broken words fix (2.1)
    - Footnote cleanup (2.2)
    - Row label normalization (2.3)
    - Title normalization (2.4)
    - Currency → float conversion with Excel format (2.5, 2.6, 2.8)
    - Percentage → float conversion with Excel format (2.7, 2.9)
    - Year string cleanup (2.10)
    - Multi-level header parsing (2.11)
    - Y/Q code formatting (2.12, 2.13)
    - Y/Q placeholder fill in Row 8 (2.14)
    """
    
    name = "process"
    
    def __init__(self, source_dir: Optional[str] = None, dest_dir: Optional[str] = None):
        """
        Initialize step with optional directory overrides.
        
        Args:
            source_dir: Override source directory (default: data/processed)
            dest_dir: Override destination directory (default: data/processed)
        """
        self.source_dir = source_dir
        self.dest_dir = dest_dir
    
    def validate(self, context: PipelineContext) -> bool:
        """Validate that source directory exists and has xlsx files."""
        from src.core import get_paths
        
        paths = get_paths()
        # Step 2: reads from extracted_raw (must match execute() source)
        source_path = Path(self.source_dir) if self.source_dir else Path(paths.data_dir) / "extracted_raw"
        
        if not source_path.exists():
            logger.warning(f"Source directory does not exist: {source_path}")
            return False
        
        xlsx_files = list(source_path.glob(TABLE_FILE_PATTERN))
        if not xlsx_files:
            logger.warning(f"No xlsx files in {source_path}")
            return False
        
        return True
    
    def get_step_info(self) -> Dict[str, Any]:
        """Get step metadata."""
        return {
            "name": self.name,
            "description": "Normalize and clean extracted data with Excel formatting",
            "reads": ["data/processed/*.xlsx"],
            "writes": ["data/processed/*.xlsx"],
            "operations": [
                "2.1 OCR broken words fix",
                "2.2 Footnote cleanup",
                "2.5-2.6 Currency → Float",
                "2.7 Percentage → Float",
                "2.8-2.9 Excel cell formatting",
                "2.12-2.14 Y/Q code formatting and placeholder fill"
            ]
        }
    
    def execute(self, context: PipelineContext) -> StepResult:
        """Execute processing step."""
        from openpyxl import load_workbook
        from src.core import get_paths
        
        paths = get_paths()
        # Step 2: reads from extracted_raw, writes to processed
        source_path = Path(self.source_dir) if self.source_dir else Path(paths.data_dir) / "extracted_raw"
        dest_path = Path(self.dest_dir) if self.dest_dir else Path(paths.data_dir) / "processed"
        
        dest_path.mkdir(parents=True, exist_ok=True)
        
        xlsx_files = list(source_path.glob(TABLE_FILE_PATTERN))
        
        stats = {
            'files_processed': 0,
            'cells_formatted': 0,
            'yq_filled': 0,
            'errors': []
        }
        
        for xlsx_path in xlsx_files:
            try:
                self._process_file(xlsx_path, dest_path, stats)
                stats['files_processed'] += 1
            except Exception as e:
                logger.error(f"Error processing {xlsx_path.name}: {e}")
                stats['errors'].append(f"{xlsx_path.name}: {str(e)}")
        
        status = StepStatus.SUCCESS if not stats['errors'] else StepStatus.PARTIAL_SUCCESS
        
        return StepResult(
            step_name=self.name,
            status=status,
            data=stats,
            message=f"Processed {stats['files_processed']} files, formatted {stats['cells_formatted']} cells",
            metadata={
                'source_dir': str(source_path),
                'dest_dir': str(dest_path)
            }
        )
    
    def _process_file(self, source_path: Path, dest_path: Path, stats: Dict):
        """Process a single xlsx file."""
        from openpyxl import load_workbook
        
        # Detect if this is a 10K (annual) report
        is_10k = '10k' in source_path.name.lower()
        
        wb = load_workbook(source_path)
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ['Index', 'TOC', 'TOC_Sheet']:
                continue
            
            ws = wb[sheet_name]
            self._process_sheet(ws, stats, is_10k=is_10k)
        
        # Save to destination
        output_path = dest_path / source_path.name
        wb.save(output_path)
        wb.close()
        
        logger.debug(f"Processed {source_path.name} (is_10k={is_10k}) -> {output_path}")
    
    def _process_sheet(self, ws, stats: Dict, is_10k: bool = False):
        """
        Process a single worksheet - finds ALL tables dynamically and processes each.
        
        Tables are identified by "Table Title:" or "Source:" markers.
        Each table is processed independently for header flattening.
        
        Args:
            ws: Worksheet to process
            stats: Stats dictionary
            is_10k: True if this is from a 10K (annual) report
        """
        from openpyxl.styles import numbers
        
        # === Check if this is a KEY-VALUE TABLE (should skip header processing) ===
        if self._is_key_value_table(ws):
            logger.debug("Skipping header processing for key-value table")
            self._process_data_cells_only(ws, stats)
            return
        
        # === DYNAMIC TABLE DETECTION ===
        # Find all tables in the sheet by looking for "Source:" markers
        tables = self._find_all_tables(ws)
        
        if not tables:
            # Fall back to processing the whole sheet as one table
            # Dynamically find where data starts after Source(s): marker
            fallback_start = self._find_first_data_row_after_source(ws)
            tables = [{'start_row': fallback_start, 'end_row': ws.max_row, 'header_row': fallback_start}]
        
        logger.debug(f"Found {len(tables)} table(s) in sheet (is_10k={is_10k})")
        
        # Process each table independently
        for table_idx, table_info in enumerate(tables):
            self._process_single_table(ws, table_info, stats, is_10k=is_10k)
    
    def _find_all_tables(self, ws) -> list:
        """
        Find all tables in a worksheet by identifying table boundaries.
        
        Tables are marked by:
        - "Table Title:" rows
        - "Source:" or "Source(s):" rows
        - Empty rows between tables
        
        Returns list of dicts with 'start_row', 'end_row', 'header_row'
        """
        tables = []
        current_table_start = None
        
        for row in range(1, ws.max_row + 1):
            col_a = ws.cell(row, 1).value
            
            if col_a:
                col_a_str = str(col_a).strip()
                col_a_lower = col_a_str.lower()
                
                # Found a new table marker - must START with "Source" (not contain it elsewhere)
                is_source_row = (
                    col_a_lower.startswith('source(s):') or
                    col_a_lower.startswith('source:')
                )
                
                if is_source_row:
                    # If we had a previous table, close it
                    if current_table_start is not None:
                        tables.append({
                            'start_row': current_table_start,
                            'end_row': row - 1,
                            'source_row': current_table_start  # Use the table's OWN source row
                        })
                    current_table_start = row
        
        # Don't forget the last table
        if current_table_start is not None:
            tables.append({
                'start_row': current_table_start, 
                'end_row': ws.max_row,
                'source_row': current_table_start
            })
        
        # For each table, find the actual data header row (first row after source with content in col B)
        for table in tables:
            source_row = table['source_row']
            header_row = None
            
            for row in range(source_row + 1, min(table['end_row'] + 1, source_row + 10)):
                col_b = ws.cell(row, 2).value
                if col_b and str(col_b).strip():
                    header_row = row
                    break
            
            table['header_row'] = header_row if header_row else source_row + 2
        
        return tables
    
    def _process_single_table(self, ws, table_info: dict, stats: Dict, is_10k: bool = False):
        """
        Process a single table within the worksheet.
        
        Args:
            ws: The worksheet
            table_info: Dict with 'start_row', 'end_row', 'header_row', 'source_row'
            stats: Stats dict to update
            is_10k: True if this is from a 10K (annual) report
        """
        import re
        
        header_row = table_info.get('header_row', 13)
        end_row = table_info.get('end_row', ws.max_row)
        
        # Extract header rows (typically 1-4 rows starting from header_row)
        data_header_rows = []
        for offset in range(4):  # Check up to 4 header rows
            row_idx = header_row + offset
            if row_idx > end_row:
                break
            
            row_values = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = str(cell.value) if cell.value else ''
                
                # For 10K reports: Convert year-only headers to YTD-YYYY
                if is_10k and val.strip() and re.match(r'^20\d{2}$', val.strip()):
                    val = f"YTD-{val.strip()}"
                    # Also update the cell directly
                    cell.value = val
                    stats['cells_formatted'] = stats.get('cells_formatted', 0) + 1
                
                row_values.append(val)
            
            # Check if this row has numeric data (means we've reached data rows)
            has_numeric = any(
                self._is_numeric_value(v) for v in row_values[1:5] if v
            )
            if has_numeric:
                break
            
            # Check if row has any content
            if any(row_values[1:]):
                data_header_rows.append(row_values)
        
        if not data_header_rows:
            return
        
        # Normalize headers
        source_filename = ''
        if hasattr(ws, 'parent') and hasattr(ws.parent, 'path') and ws.parent.path:
            source_filename = str(ws.parent.path)
        
        normalized = normalize_multi_row_headers(data_header_rows, source_filename)
        
        # Extract normalized headers
        l1_per_column = normalized.get('l1_headers', [])
        normalized_headers = normalized.get('normalized_headers', [])
        
        # For 10K reports: Update normalized_headers with YTD- prefix for year-only values
        if is_10k:
            for i, val in enumerate(normalized_headers):
                if val and re.match(r'^20\d{2}$', val.strip()):
                    normalized_headers[i] = f"YTD-{val.strip()}"
        
        num_cols = ws.max_column
        
        def safe_set_cell_value(ws, row, col, value):
            """Safely set cell value, skipping merged cells."""
            from openpyxl.cell.cell import MergedCell
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = value
                return True
            return False
        
        # === REPLACE actual data header rows with normalized Y/Q codes ===
        # Use dynamic row numbers based on header_row
        if len(data_header_rows) >= 1:
            for col_idx in range(2, num_cols + 1):
                orig_val = data_header_rows[0][col_idx - 1] if col_idx - 1 < len(data_header_rows[0]) else ''
                norm_val = normalized_headers[col_idx - 1] if col_idx - 1 < len(normalized_headers) else ''
                
                if orig_val:
                    orig_lower = orig_val.lower().strip()
                    is_date_pattern = any(kw in orig_lower for kw in [
                        'months ended', 'at ', 'as of', 'for the', 
                        'december', 'march', 'june', 'september'
                    ])
                    if is_date_pattern and norm_val:
                        if safe_set_cell_value(ws, header_row, col_idx, norm_val):
                            stats['cells_formatted'] = stats.get('cells_formatted', 0) + 1
        
        # Second header row (if exists)
        if len(data_header_rows) > 1:
            for col_idx in range(2, num_cols + 1):
                orig_val = data_header_rows[1][col_idx - 1] if col_idx - 1 < len(data_header_rows[1]) else ''
                norm_val = normalized_headers[col_idx - 1] if col_idx - 1 < len(normalized_headers) else ''
                
                if norm_val and (norm_val.startswith('Q') or norm_val.startswith('YTD') or '-' in norm_val):
                    if safe_set_cell_value(ws, header_row + 1, col_idx, norm_val):
                        stats['cells_formatted'] = stats.get('cells_formatted', 0) + 1
        
        # === FLATTEN HEADERS for this table ===
        self._flatten_table_headers_dynamic(ws, header_row, data_header_rows, normalized_headers, stats)
        
        # Process data cells for this table
        data_start = header_row + len(data_header_rows) + 1  # After headers + empty separator
        for row_idx in range(data_start, end_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row_idx, col_idx)
                if cell.value is None:
                    continue
                
                original_value = cell.value
                new_value, cell_format = self._process_cell_value(original_value, row_idx, col_idx)
                
                if new_value != original_value:
                    cell.value = new_value
                    stats['cells_formatted'] = stats.get('cells_formatted', 0) + 1
                
                if cell_format:
                    cell.number_format = cell_format
    
    def _is_numeric_value(self, value_str: str) -> bool:
        """Check if a string value represents a number (data, not header)."""
        import re
        
        if not value_str:
            return False
        
        # Clean the value
        clean = value_str.strip()
        
        # Skip common header patterns
        if any(kw in clean.lower() for kw in ['in millions', 'in billions', '$', 'months ended', 'at ']):
            return False
        
        # 4-digit years (2020-2099) are HEADERS, not numeric data
        if re.match(r'^20\d{2}$', clean):
            return False
        
        # Try to parse as number
        try:
            clean_num = clean.replace('%', '').replace(',', '').replace('$', '').replace('-', '').strip()
            if clean_num:
                float(clean_num)
                return True
        except:
            pass
        
        return False
    
    def _process_cell_value(self, value, row_idx: int, col_idx: int):
        """
        Process a single cell value and determine Excel format.
        
        Returns:
            Tuple of (processed_value, excel_format_or_None)
        """
        if value is None:
            return value, None
        
        str_value = str(value).strip()
        
        # Skip empty values
        if not str_value:
            return value, None
        
        # Skip special text values
        if str_value.lower() in ['n/a', '-', '—', 'nm', 'n.a.', '']:
            return value, None
        
        # Already a number? Check if it needs formatting
        if isinstance(value, (int, float)):
            # Determine format based on magnitude (crude heuristic)
            if abs(value) < 1 and value != 0:
                # Likely a percentage (0.155 = 15.5%)
                return value, PERCENT_FORMAT
            else:
                return value, CURRENCY_FORMAT
        
        # Fix OCR broken words and normalize text (for row labels in column A only)
        if isinstance(value, str) and col_idx == 1:
            str_value = ExcelUtils.fix_ocr_broken_words(str_value)
            str_value = ExcelUtils.clean_footnote_references(str_value)
            # Apply text normalization (spaces around dashes, etc.)
            str_value = normalize_text(str_value)
            return str_value, None
        
        # Try percentage conversion first (contains %)
        if isinstance(value, str) and '%' in str_value:
            try:
                pct_str = str_value.replace('%', '').replace(',', '').strip()
                pct_val = float(pct_str)
                return pct_val / 100.0, PERCENT_FORMAT  # 15.5% -> 0.155
            except (ValueError, TypeError):
                pass
        
        # Try currency/negative conversion
        if isinstance(value, str):
            cleaned = ExcelUtils.clean_currency_value(str_value)
            if isinstance(cleaned, (int, float)) and cleaned != value:
                return cleaned, NEGATIVE_CURRENCY_FORMAT if cleaned < 0 else CURRENCY_FORMAT
        
        return str_value if isinstance(value, str) else value, None
    
    def _find_first_data_row_after_source(self, ws) -> int:
        """
        Find the first row with data after the Source(s): marker.
        
        Scans for patterns like "Table Title:", "Source:", "Source(s):" and returns
        the first row after that which contains data.
        
        Returns:
            Row number where data starts (default: 13 if pattern not found)
        """
        source_row = None
        
        # Scan for Source(s): or Source: marker
        for row in range(1, min(ws.max_row + 1, 30)):
            cell_val = ws.cell(row, 1).value
            if cell_val:
                cell_str = str(cell_val).strip().lower()
                if cell_str.startswith('source(s):') or cell_str.startswith('source:'):
                    source_row = row
                    break
        
        if source_row is None:
            # No Source: found, return default
            return 13
        
        # Find first row with data after source row
        for row in range(source_row + 1, min(ws.max_row + 1, source_row + 10)):
            # Check if row has any content
            for col in range(1, min(ws.max_column + 1, 10)):
                cell_val = ws.cell(row, col).value
                if cell_val and str(cell_val).strip():
                    return row
        
        return source_row + 1  # Default to row after source
    
    def _is_key_value_table(self, ws) -> bool:
        """
        Detect if a worksheet contains a key-value table (not a data table with column headers).
        
        Key-value tables have:
        - Row labels in column A (e.g., "Announcement date", "Amount per share")
        - Single data values in column B (not spanning headers)
        - Very few columns with data (typically 2-3)
        - Row labels that look like field names, not data categories
        """
        KEY_VALUE_LABELS = [
            'announcement date', 'amount per share', 'date paid', 'date to be paid',
            'shareholders of record', 'record date', 'ex-dividend date',
            'payment date', 'declaration date'
        ]
        
        # Dynamically find where data starts
        data_start = self._find_first_data_row_after_source(ws)
        
        # Check data rows for key-value patterns
        kv_pattern_count = 0
        total_rows_checked = 0
        
        for row in range(data_start, min(ws.max_row + 1, data_start + 7)):
            col_a = ws.cell(row, 1).value
            col_b = ws.cell(row, 2).value
            col_c = ws.cell(row, 3).value
            
            if not col_a:
                continue
            
            total_rows_checked += 1
            col_a_lower = str(col_a).lower().strip()
            
            # Check if column A looks like a key-value label
            is_kv_label = any(label in col_a_lower for label in KEY_VALUE_LABELS)
            
            # Only count explicit key-value label matches, not generic single-column checks
            if is_kv_label:
                kv_pattern_count += 1
        
        # Require at least 2 explicit key-value labels to treat as key-value table
        return total_rows_checked > 0 and kv_pattern_count >= 2
    
    def _process_data_cells_only(self, ws, stats: Dict):
        """
        Process data cells for numeric formatting only, without header normalization.
        Used for key-value tables that should not have headers modified.
        """
        from openpyxl.styles import numbers
        
        # Dynamically find where data starts
        data_start = self._find_first_data_row_after_source(ws)
        
        # Just process percentage and numeric formatting
        for row in range(data_start, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                if cell.value is None:
                    continue
                
                value = cell.value
                new_value, number_format = self._process_cell_value(value, row, col)
                
                if new_value != value:
                    cell.value = new_value
                    stats['cells_formatted'] = stats.get('cells_formatted', 0) + 1
                
                if number_format:
                    cell.number_format = number_format
    
    def _flatten_table_headers_dynamic(self, ws, header_row: int, data_header_rows: list, normalized_headers: list, stats: dict):
        """
        Flatten multi-level headers for a specific table starting at header_row.
        
        This is the dynamic version that works with any starting row, enabling
        processing of multiple tables per sheet.
        
        Args:
            ws: Worksheet
            header_row: The row where this table's headers start
            data_header_rows: List of header row values
            normalized_headers: List of normalized header values
            stats: Stats dictionary
        """
        from openpyxl.cell.cell import MergedCell
        import re
        
        num_cols = ws.max_column
        num_header_rows = len(data_header_rows)
        
        if num_header_rows == 0:
            return
        
        logger.debug(f"Flattening {num_header_rows} header rows starting at row {header_row}")
        
        # Unmerge any merged cells in the header rows for this table
        merged_ranges_to_unmerge = []
        for merge_range in ws.merged_cells.ranges:
            if header_row <= merge_range.min_row <= header_row + num_header_rows:
                merged_ranges_to_unmerge.append(merge_range)
        
        for merge_range in merged_ranges_to_unmerge:
            try:
                ws.unmerge_cells(str(merge_range))
            except Exception as e:
                logger.debug(f"Could not unmerge {merge_range}: {e}")
        
        if num_header_rows == 1:
            # 1-level: Write normalized headers to the header row
            # DO NOT add empty separator - data starts immediately after
            for col_idx in range(1, num_cols + 1):
                if col_idx < len(normalized_headers) + 1 and normalized_headers[col_idx - 1]:
                    cell = ws.cell(row=header_row, column=col_idx)
                    if not isinstance(cell, MergedCell):
                        cell.value = normalized_headers[col_idx - 1]
            
            # For 1-level headers, the next row is DATA, not a header to remove
            # So we DON'T set it to empty or add a separator
        
        elif num_header_rows == 2:
            # 2-level: Combine into single row + empty separator
            for col_idx in range(1, num_cols + 1):
                if col_idx < len(normalized_headers) + 1 and normalized_headers[col_idx - 1]:
                    cell = ws.cell(row=header_row, column=col_idx)
                    if not isinstance(cell, MergedCell):
                        cell.value = normalized_headers[col_idx - 1]
            
            # Row header_row+1 becomes empty separator
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=header_row + 1, column=col)
                if not isinstance(cell, MergedCell):
                    cell.value = None
            
            stats['headers_flattened'] = stats.get('headers_flattened', 0) + 1
            stats['empty_rows_added'] = stats.get('empty_rows_added', 0) + 1
        
        else:
            # 3+ levels: Build combined headers
            if num_header_rows == 3:
                combined = self._build_combined_headers_3level(data_header_rows, normalized_headers, num_cols)
            else:
                combined = self._build_combined_headers_4level(data_header_rows, normalized_headers, num_cols)
            
            # Write combined headers to header_row
            for col_idx, value in enumerate(combined, start=1):
                cell = ws.cell(row=header_row, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = value
            
            # Row header_row+1 becomes empty separator
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=header_row + 1, column=col)
                if not isinstance(cell, MergedCell):
                    cell.value = None
            
            # Delete remaining header rows (shift data up)
            # We want: header_row (combined) + header_row+1 (empty) + data
            # So delete rows header_row+2 onwards up to the original data start
            rows_to_delete = num_header_rows - 2
            for _ in range(rows_to_delete):
                ws.delete_rows(header_row + 2)
            
            stats['headers_flattened'] = stats.get('headers_flattened', 0) + 1
            stats['empty_rows_added'] = stats.get('empty_rows_added', 0) + 1
        
        logger.debug(f"Flattened headers at row {header_row}")

    def _flatten_table_headers(self, ws, data_header_rows: list, normalized_headers: list, stats: dict):
        """
        Flatten multi-level headers into a single header row with empty separator.
        
        Handles:
        - 1-level: Add empty row separator after header
        - 2-level: L1 (period) + L2 (year) → single row + empty separator
        - 3-level: L1 (period) + L2 (year) + L3 (category) → combined row + empty separator
        - 4-level: L1 (period) + L2 (year) + L3 (year) + L4 (category) → combined row + empty separator
        
        Format Rules:
            At/As of dates      → Q1-2025, Q2-2025, Q3-2025, Q4-2025
            Three Months Ended  → Q1-QTD-2025, Q2-QTD-2025
            Six Months Ended    → Q2-YTD-2025
            Nine Months Ended   → Q3-YTD-2025
            Year Ended          → YTD-2025
        
        This ONLY modifies header rows, NOT data values.
        """
        from openpyxl.cell.cell import MergedCell
        import re
        
        # Count actual header rows (between Source: and first data row)
        header_count = self._count_header_rows(ws)
        
        # Find where data starts (first row with numeric values)
        data_start_row = self._find_data_start_row(ws)
        if not data_start_row:
            return  # Can't determine data start
        
        num_cols = ws.max_column
        header_start = 13  # Headers typically start at row 13
        
        # Calculate rows to flatten
        rows_to_flatten = data_start_row - header_start if data_start_row > header_start else 0
        
        logger.debug(f"Header analysis: {rows_to_flatten} header rows (row 13 to {data_start_row - 1}), data starts at row {data_start_row}")
        
        # Unmerge any merged cells in the header rows
        merged_ranges_to_unmerge = []
        for merge_range in ws.merged_cells.ranges:
            if merge_range.min_row >= header_start and merge_range.min_row < data_start_row:
                merged_ranges_to_unmerge.append(str(merge_range))
        
        for merge_range in merged_ranges_to_unmerge:
            try:
                ws.unmerge_cells(merge_range)
            except Exception as e:
                logger.debug(f"Could not unmerge {merge_range}: {e}")
        
        rows_deleted = 0
        
        if rows_to_flatten <= 1:
            # === 1-LEVEL: Already flat, just add empty row separator ===
            # Insert empty row after header (row 14)
            ws.insert_rows(header_start + 1)
            # Clear the new row
            for col in range(1, num_cols + 1):
                ws.cell(row=header_start + 1, column=col).value = None
            logger.debug("1-level: Added empty row separator after header")
            stats['empty_rows_added'] = stats.get('empty_rows_added', 0) + 1
            return
        
        elif rows_to_flatten == 2:
            # === 2-LEVEL: Period + Year → single row + empty separator ===
            flattened = self._build_flattened_headers(ws, header_start, data_start_row, num_cols, normalized_headers)
            
            # Write flattened headers to row 13
            for col_idx, value in enumerate(flattened, start=1):
                try:
                    cell = ws.cell(row=header_start, column=col_idx)
                    if not isinstance(cell, MergedCell):
                        cell.value = value
                except Exception as e:
                    logger.debug(f"Could not write to row 13, col {col_idx}: {e}")
            
            # Clear row 14 (make it the empty separator) instead of deleting
            for col in range(1, num_cols + 1):
                try:
                    ws.cell(row=header_start + 1, column=col).value = None
                except:
                    pass
            
            logger.debug("2-level: Flattened headers, row 14 is now empty separator")
            
        elif rows_to_flatten == 3:
            # === 3-LEVEL: Period + Year + Category → combined row + empty separator ===
            # Pattern: Row 13 (date), Row 14 (date dup or category labels), Row 15 (categories)
            
            # Collect all header row data
            header_rows_data = []
            for row in range(header_start, data_start_row):
                row_data = [ws.cell(row, c).value for c in range(1, num_cols + 1)]
                header_rows_data.append(row_data)
            
            # Build combined headers: date_code + category
            combined = self._build_combined_headers_3level(
                header_rows_data, normalized_headers, num_cols
            )
            
            # Write combined headers to row 13
            for col_idx, value in enumerate(combined, start=1):
                try:
                    ws.cell(row=header_start, column=col_idx).value = value
                except Exception as e:
                    logger.debug(f"Could not write to row 13, col {col_idx}: {e}")
            
            # Row 14 becomes empty separator
            for col in range(1, num_cols + 1):
                try:
                    ws.cell(row=header_start + 1, column=col).value = None
                except:
                    pass
            
            # Delete remaining header rows (row 15 onwards)
            rows_to_delete = rows_to_flatten - 2
            for _ in range(rows_to_delete):
                ws.delete_rows(header_start + 2)
            rows_deleted = rows_to_delete
            
            logger.debug(f"3-level: Combined headers, deleted {rows_deleted} rows")
            
        else:
            # === 4+ LEVEL: Period + Year + Year + Category → combined row + empty separator ===
            # Collect all header row data
            header_rows_data = []
            for row in range(header_start, data_start_row):
                row_data = [ws.cell(row, c).value for c in range(1, num_cols + 1)]
                header_rows_data.append(row_data)
            
            # Build combined headers for 4+ levels
            combined = self._build_combined_headers_4level(
                header_rows_data, normalized_headers, num_cols
            )
            
            # Write combined headers to row 13
            for col_idx, value in enumerate(combined, start=1):
                try:
                    ws.cell(row=header_start, column=col_idx).value = value
                except Exception as e:
                    logger.debug(f"Could not write to row 13, col {col_idx}: {e}")
            
            # Row 14 becomes empty separator
            for col in range(1, num_cols + 1):
                try:
                    ws.cell(row=header_start + 1, column=col).value = None
                except:
                    pass
            
            # Delete remaining header rows
            rows_to_delete = rows_to_flatten - 2
            for _ in range(rows_to_delete):
                ws.delete_rows(header_start + 2)
            rows_deleted = rows_to_delete
            
            logger.debug(f"4+-level: Combined headers, deleted {rows_deleted} rows")
        
        stats['headers_flattened'] = stats.get('headers_flattened', 0) + 1
        stats['empty_rows_added'] = stats.get('empty_rows_added', 0) + 1
        logger.debug(f"Flattened {rows_to_flatten} header rows, deleted {rows_deleted} rows, added empty separator")
    
    def _build_combined_headers_3level(self, header_rows_data: list, normalized_headers: list, num_cols: int) -> list:
        """
        Build combined headers for 3-level patterns.
        
        Pattern: Period/Date (L1) + Year/Date (L2) + Category (L3)
        Output: "Q3-2025 Level 2 | Q3-2025 Level 3 | Q3-2025 Total"
        
        Key: Propagate the date code across ALL category columns.
        """
        import re
        combined = []
        
        # Find the primary date code from normalized_headers
        primary_date_code = ''
        for norm in normalized_headers:
            if norm and (norm.startswith('Q') or 'YTD' in str(norm) or 'QTD' in str(norm)):
                # Get just the date part (e.g., "Q3-2025" from "Q3-2025 Level 2")
                parts = str(norm).split()
                primary_date_code = parts[0] if parts else norm
                break
        
        # Also try to find date code from header rows
        if not primary_date_code:
            for row_data in header_rows_data:
                for val in row_data:
                    if val and str(val).startswith('Q') and '-' in str(val):
                        primary_date_code = str(val).split()[0]
                        break
                if primary_date_code:
                    break
        
        # Find the category row (usually the last row with unit indicator in col 1)
        category_row_idx = -1
        for i, row_data in enumerate(header_rows_data):
            if row_data[0] and any(kw in str(row_data[0]).lower() for kw in ['$ in', 'in millions', 'in billions', 'fee rate']):
                category_row_idx = i
        
        # If no category row found, use the last row
        if category_row_idx == -1:
            category_row_idx = len(header_rows_data) - 1
        
        category_row = header_rows_data[category_row_idx] if category_row_idx >= 0 else []
        
        for col_idx in range(num_cols):
            if col_idx == 0:
                # Column 1: Keep unit description from category row
                unit_desc = ''
                for row_data in header_rows_data:
                    val = row_data[0] if len(row_data) > 0 else ''
                    if val and any(kw in str(val).lower() for kw in ['$ in', 'in millions', 'in billions', 'fee rate']):
                        unit_desc = str(val)
                        break
                combined.append(unit_desc)
            else:
                # Data columns: Combine date code + category
                # Use the column-specific normalized header if available, otherwise use primary date code
                norm_code = normalized_headers[col_idx] if col_idx < len(normalized_headers) and normalized_headers[col_idx] else ''
                
                # If this column has no norm_code, use the primary date code
                if not norm_code:
                    norm_code = primary_date_code
                
                # Get the category from the category row
                category = str(category_row[col_idx] if col_idx < len(category_row) and category_row[col_idx] else '').strip()
                
                # Skip category if it looks like a date/year or is a duplicate of norm_code
                if category:
                    is_date_pattern = (
                        re.match(r'^20\d{2}$', category) or
                        category.startswith('Q') or
                        'months ended' in category.lower() or
                        category.lower().startswith('at ')
                    )
                    if is_date_pattern or category == norm_code:
                        category = ''
                
                # Build the combined header
                if norm_code and category:
                    # Check if category is already in norm_code to prevent duplicates like "Q3-2025 AAA AAA"
                    if category.lower() in norm_code.lower():
                        combined.append(norm_code)
                    else:
                        combined.append(f"{norm_code} {category}")
                elif norm_code:
                    combined.append(norm_code)
                elif category:
                    # Even if no date code, prepend primary if available
                    if primary_date_code:
                        combined.append(f"{primary_date_code} {category}")
                    else:
                        combined.append(category)
                else:
                    combined.append('')
        
        # Final cleanup: dedupe repeated words and strip footnotes
        combined = self._cleanup_headers(combined)
        
        return combined
    
    def _build_combined_headers_4level(self, header_rows_data: list, normalized_headers: list, num_cols: int) -> list:
        """
        Build combined headers for 4+ level patterns.
        
        Pattern: Period (L1) + Year (L2) + Year (L3 dup) + Category (L4)
        Output: "Q3-QTD-2025 Amortized Cost | Q3-QTD-2025 % of Loans | Q3-QTD-2024 Amortized Cost | Q3-QTD-2024 % of Loans"
        
        Key: Track year column spans to propagate correct date code to each category column.
        """
        import re
        combined = []
        
        # Find year row and build column-to-year mapping
        year_row_idx = -1
        year_row = []
        for i, row_data in enumerate(header_rows_data):
            years_found = sum(1 for v in row_data[1:] if v and re.match(r'^20\d{2}$', str(v).strip()))
            if years_found >= 1:
                year_row_idx = i
                year_row = row_data
                break
        
        # Build column-to-year mapping (propagate year across empty columns)
        col_to_year = {}
        current_year = None
        for col_idx in range(1, num_cols):
            if col_idx < len(year_row):
                val = year_row[col_idx]
                if val and re.match(r'^20\d{2}$', str(val).strip()):
                    current_year = str(val).strip()
            if current_year:
                col_to_year[col_idx] = current_year
        
        # Determine period type from first row
        period_text = ' '.join(str(v or '') for v in header_rows_data[0]).lower() if header_rows_data else ''
        
        def build_date_code(year):
            """Build date code based on period type and year."""
            if not year:
                return ''
            if 'three months' in period_text or '3 months' in period_text:
                if 'september' in period_text or 'sept' in period_text:
                    return f"Q3-QTD-{year}"
                elif 'june' in period_text:
                    return f"Q2-QTD-{year}"
                elif 'march' in period_text:
                    return f"Q1-QTD-{year}"
                elif 'december' in period_text:
                    return f"Q4-QTD-{year}"
            elif 'six months' in period_text or '6 months' in period_text:
                if 'june' in period_text:
                    return f"Q2-YTD-{year}"
                elif 'september' in period_text:
                    return f"Q3-YTD-{year}"
            elif 'nine months' in period_text or '9 months' in period_text:
                if 'september' in period_text:
                    return f"Q3-YTD-{year}"
            elif 'at ' in period_text or 'as of' in period_text:
                if 'september' in period_text or 'sept' in period_text:
                    return f"Q3-{year}"
                elif 'june' in period_text:
                    return f"Q2-{year}"
                elif 'march' in period_text:
                    return f"Q1-{year}"
                elif 'december' in period_text:
                    return f"Q4-{year}"
            return f"Q4-{year}"  # Default
        
        # Find category row (usually has categories like "Amortized Cost", "% of Loans")
        category_row_idx = -1
        for i, row_data in enumerate(header_rows_data):
            non_empty = [str(v).strip() for v in row_data[1:] if v and str(v).strip()]
            if non_empty:
                has_categories = any(
                    kw in ' '.join(non_empty).lower() 
                    for kw in ['amortized', 'cost', 'value', 'level', 'total', '%', 'fair']
                )
                if has_categories:
                    category_row_idx = i
        
        if category_row_idx == -1:
            category_row_idx = len(header_rows_data) - 1
        
        category_row = header_rows_data[category_row_idx] if 0 <= category_row_idx < len(header_rows_data) else []
        
        for col_idx in range(num_cols):
            if col_idx == 0:
                # Column 1: Keep unit description
                unit_desc = ''
                for row_data in header_rows_data:
                    val = row_data[0] if len(row_data) > 0 else ''
                    if val and any(kw in str(val).lower() for kw in ['$ in', 'in millions', 'in billions', 'fee rate']):
                        unit_desc = str(val)
                        break
                combined.append(unit_desc)
            else:
                # Data columns: Get date code from column-to-year mapping
                year = col_to_year.get(col_idx, '')
                norm_code = ''
                
                # First try normalized_headers
                if col_idx < len(normalized_headers) and normalized_headers[col_idx]:
                    norm_code = normalized_headers[col_idx]
                # If no norm_code, build from year
                elif year:
                    norm_code = build_date_code(year)
                
                # Get category
                category = str(category_row[col_idx] if col_idx < len(category_row) and category_row[col_idx] else '').strip()
                
                # Skip category if it looks like a date/year
                if category:
                    is_date_pattern = (
                        re.match(r'^20\d{2}$', category) or
                        category.startswith('Q') or
                        'months ended' in category.lower()
                    )
                    if is_date_pattern:
                        category = ''
                
                # Build combined header - ALWAYS prepend date code if we have it
                if norm_code and category:
                    # Check if category is already in norm_code to prevent duplicates
                    if category.lower() in norm_code.lower():
                        combined.append(norm_code)
                    else:
                        combined.append(f"{norm_code} {category}")
                elif norm_code:
                    combined.append(norm_code)
                elif category:
                    # Still try to get a date code for this column
                    if year:
                        date_code = build_date_code(year)
                        combined.append(f"{date_code} {category}")
                    else:
                        combined.append(category)
                else:
                    combined.append('')
        
        # Final cleanup: dedupe repeated words and strip footnotes
        combined = self._cleanup_headers(combined)
        
        return combined
    
    def _cleanup_headers(self, headers: list) -> list:
        """Clean up combined headers by removing duplicates and footnotes.
        
        - Removes consecutive duplicate words (e.g., 'AAA AAA' -> 'AAA')
        - Strips trailing footnote markers (e.g., ' 1' from 'Netting 1')
        - Preserves meaningful numbers like 'Level 1', 'Tier 1'
        """
        import re
        
        cleaned = []
        for h in headers:
            if not h:
                cleaned.append('')
                continue
            
            h = str(h)
            
            # Dedupe consecutive repeated words (case-insensitive)
            words = h.split()
            deduped = []
            prev_word = None
            for word in words:
                if prev_word is None or word.lower() != prev_word.lower():
                    deduped.append(word)
                    prev_word = word
            h = ' '.join(deduped)
            
            # Strip trailing footnote markers (but preserve Level 1, Tier 1, etc.)
            preserve_patterns = [
                r'\bLevel\s+\d+$',
                r'\bTier\s+\d+$',
                r'\bType\s+\d+$',
            ]
            should_strip = True
            for pattern in preserve_patterns:
                if re.search(pattern, h, re.IGNORECASE):
                    should_strip = False
                    break
            
            if should_strip:
                h = re.sub(r'\s+\d+$', '', h)
            
            cleaned.append(h)
        
        return cleaned
    
    def _count_header_rows(self, ws) -> int:
        """Count header rows between Source: row and first data row."""
        source_row = None
        for row in range(1, 15):
            val = ws.cell(row, 1).value
            if val and 'source' in str(val).lower():
                source_row = row
                break
        
        if not source_row:
            return 0
        
        count = 0
        for row in range(source_row + 1, min(source_row + 8, ws.max_row + 1)):
            # Check if row has numeric data (indicates data, not header)
            has_numbers = False
            for col in range(2, min(ws.max_column + 1, 10)):
                val = ws.cell(row, col).value
                if val is not None:
                    try:
                        float(str(val).replace('%', '').replace(',', '').replace('$', '').replace('-', '0'))
                        has_numbers = True
                        break
                    except:
                        pass
            
            if has_numbers:
                break
            
            # Row has content but no numbers - it's a header row
            if any(ws.cell(row, c).value for c in range(1, min(ws.max_column + 1, 10))):
                count += 1
        
        return count
    
    def _find_data_start_row(self, ws) -> int:
        """
        Find the first row that contains actual data (has numeric values in data columns).
        
        Skips:
        - Empty rows
        - Header rows (period types, years)
        - Section label rows (text only in column A)
        """
        # Dynamically find where headers/data start after Source: marker
        header_start = self._find_first_data_row_after_source(ws)
        
        for row in range(header_start, min(ws.max_row + 1, header_start + 20)):
            # Count numeric values in data columns (columns 2+)
            numeric_count = 0
            for col in range(2, min(ws.max_column + 1, 10)):
                val = ws.cell(row, col).value
                if val is not None and str(val).strip():
                    val_str = str(val).strip()
                    # Skip if it looks like a header pattern
                    if any(kw in val_str.lower() for kw in ['months ended', 'at ', 'as of', 'year ended']):
                        continue
                    # Skip if it's just a year
                    if len(val_str) == 4 and val_str.isdigit():
                        continue
                    # Try to parse as number
                    try:
                        clean_val = val_str.replace('%', '').replace(',', '').replace('$', '').replace('-', '').strip()
                        if clean_val:
                            float(clean_val)
                            numeric_count += 1
                    except:
                        pass
            
            # Need at least 1 numeric value to consider it a data row
            if numeric_count >= 1:
                return row
        
        return None
    
    def _build_flattened_headers(self, ws, header_start: int, data_start: int, 
                                  num_cols: int, normalized_headers: list) -> list:
        """
        Build flattened headers by combining values from all header rows.
        
        Logic:
        - For column 1: Find unit description (e.g., "$ in millions") from header rows
        - For data columns: Use normalized headers if available, otherwise combine row values
        """
        flattened = []
        
        for col_idx in range(1, num_cols + 1):
            if col_idx == 1:
                # Column 1: find unit description or row label header
                # Look for patterns like "$ in millions", "$ in billions", "Fee rate in bps"
                col1_value = ''
                for row in range(header_start, data_start):
                    val = ws.cell(row, 1).value
                    if val:
                        val_str = str(val).strip()
                        # Check if it looks like a unit description
                        val_lower = val_str.lower()
                        if any(kw in val_lower for kw in ['$ in', 'in millions', 'in billions', 'in bps', 'fee rate', 'rate in']):
                            col1_value = val_str
                            break
                
                # If no unit found, use the first non-empty column 1 value from header rows
                if not col1_value:
                    for row in range(header_start, data_start):
                        val = ws.cell(row, 1).value
                        if val and str(val).strip():
                            col1_value = str(val).strip()
                            break
                
                flattened.append(col1_value)
            else:
                # Data columns: prefer normalized header, else combine
                idx = col_idx - 1
                if idx < len(normalized_headers) and normalized_headers[idx]:
                    flattened.append(normalized_headers[idx])
                else:
                    # Combine all header row values for this column
                    parts = []
                    for row in range(header_start, data_start):
                        val = ws.cell(row, col_idx).value
                        if val and str(val).strip():
                            val_str = str(val).strip()
                            # Skip duplicates
                            if val_str not in parts:
                                parts.append(val_str)
                    flattened.append(' '.join(parts) if parts else '')
        
        return flattened
    
    def _is_spanning_header_row(self, row_values: list) -> bool:
        """
        Detect if a row is a spanning header (L2) vs column headers (L3).
        
        Spanning headers have:
        - Period phrases like "Three Months Ended", "At September 30,"
        - Many empty cells (spans across columns)
        - No specific column data values
        
        Returns True if this looks like a spanning header that should be removed.
        """
        if not row_values:
            return False
        
        non_empty_count = sum(1 for v in row_values if v and str(v).strip())
        total_cols = len(row_values)
        
        # If less than 30% of cells are filled, it's likely a spanning header
        fill_ratio = non_empty_count / total_cols if total_cols > 0 else 0
        
        # Check for period phrases
        has_period_phrase = False
        period_patterns = [
            'three months ended', 'six months ended', 'nine months ended',
            'year ended', 'at ', 'as of ', 'for the period'
        ]
        
        for val in row_values:
            if val:
                val_lower = str(val).lower()
                for pattern in period_patterns:
                    if pattern in val_lower:
                        has_period_phrase = True
                        break
        
        # Decision: It's a spanning header if:
        # 1. Has period phrases AND less than 50% fill ratio, OR
        # 2. Less than 30% fill ratio (mostly empty/spanning)
        if has_period_phrase and fill_ratio < 0.5:
            return True
        if fill_ratio < 0.3:
            return True
        
        return False
    
    def _are_duplicate_header_rows(self, row1: list, row2: list) -> bool:
        """
        Check if two header rows are duplicates or very similar.
        
        This handles cases where extraction created duplicate headers in rows 13 and 14.
        If they're the same, we should remove one row.
        
        Returns True if rows are duplicates/similar enough to warrant flattening.
        """
        if not row1 or not row2:
            return False
        
        # Compare non-empty values
        r1_vals = [str(v).strip().lower() for v in row1 if v and str(v).strip()]
        r2_vals = [str(v).strip().lower() for v in row2 if v and str(v).strip()]
        
        if not r1_vals or not r2_vals:
            return False
        
        # Check if rows are identical
        if r1_vals == r2_vals:
            return True
        
        # Check if first row's content is subset of second row
        # (e.g., row 13 has just headers, row 14 has headers + first data)
        if len(r1_vals) <= len(r2_vals):
            overlap_count = sum(1 for v in r1_vals if v in r2_vals)
            if overlap_count >= len(r1_vals) * 0.7:
                return True
        
        return False


# Backward-compatible function for main.py CLI
def run_process(
    source_dir: Optional[str] = None,
    dest_dir: Optional[str] = None
) -> PipelineResult:
    """
    Run processing on extracted xlsx files.
    
    Legacy wrapper maintaining backward compatibility with main.py CLI.
    
    Args:
        source_dir: Override source directory (default: data/processed)
        dest_dir: Override destination directory (default: data/processed)
        
    Returns:
        PipelineResult with processing outcome
    """
    step = ProcessStep(source_dir=source_dir, dest_dir=dest_dir)
    ctx = PipelineContext()
    
    # Validate - if fails, return with SKIPPED status
    if not step.validate(ctx):
        logger.warning("Validation failed - no files to process")
        return PipelineResult(
            step=PipelineStep.PROCESS if hasattr(PipelineStep, 'PROCESS') else PipelineStep.EXTRACT,
            success=False,
            data={},
            message="No files to process - validation failed",
            error="Validation failed: source directory empty or missing",
            metadata={}
        )
    
    result = step.execute(ctx)
    
    return PipelineResult(
        step=PipelineStep.PROCESS if hasattr(PipelineStep, 'PROCESS') else PipelineStep.EXTRACT,
        success=result.success,
        data=result.data,
        message=result.message,
        error=result.error,
        metadata=result.metadata
    )
