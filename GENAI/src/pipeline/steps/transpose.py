"""
Transpose Step - Convert consolidated tables to time-series format.

Step 5 in pipeline: Extract → Process → Process-Advanced → Consolidate → Transpose

Implements operations per docs/pipeline_operations.md:
- Multi-sheet processing
- Column header detection (Y/Q)
- Column → Row pivot
- Chronological ordering
- Index/TOC generation
"""

from pathlib import Path
from typing import Dict, Any, Optional

from src.pipeline.base import StepInterface, StepResult, StepStatus, PipelineContext
from src.pipeline import PipelineResult, PipelineStep
from src.utils import get_logger

# External library imports (moved to module level for efficiency)
import pandas as pd
from openpyxl import load_workbook

logger = get_logger(__name__)


class TransposeStep(StepInterface):
    """
    Transpose step - convert consolidated tables to time-series format.
    
    Implements StepInterface following pipeline pattern.
    
    Reads: data/consolidate/consolidated_tables.xlsx
    Writes: data/transpose/consolidated_tables_transposed.xlsx
    """
    
    name = "transpose"
    
    def __init__(self, source_file: Optional[str] = None, output_file: Optional[str] = None):
        """
        Initialize step with optional file overrides.
        
        Args:
            source_file: Override source file path
            output_file: Override output file path
        """
        self.source_file = source_file
        self.output_file = output_file
    
    def validate(self, context: PipelineContext) -> bool:
        """Validate that source file exists."""
        from src.core import get_paths
        
        if self.source_file:
            source_path = Path(self.source_file)
        else:
            paths = get_paths()
            source_path = Path(paths.data_dir) / "consolidate" / "consolidated_tables.xlsx"
        
        if not source_path.exists():
            logger.warning(f"Source file does not exist: {source_path}")
            return False
        
        return True
    
    def get_step_info(self) -> Dict[str, Any]:
        """Get step metadata."""
        return {
            "name": self.name,
            "description": "Convert consolidated tables to time-series format",
            "reads": ["data/consolidate/consolidated_tables.xlsx"],
            "writes": ["data/transpose/consolidated_tables_transposed.xlsx"],
            "operations": [
                "Multi-sheet processing",
                "Y/Q column detection",
                "Column → Row pivot",
                "Chronological ordering"
            ]
        }
    
    def execute(self, context: PipelineContext) -> StepResult:
        """Execute transpose step."""
        from src.core import get_paths
        from src.infrastructure.extraction.consolidation.consolidated_exporter_transpose import (
            create_transposed_dataframe,
            reconstruct_metadata_from_df
        )
        from openpyxl import load_workbook as openpyxl_load
        from copy import copy
        
        paths = get_paths()
        
        if self.source_file:
            source_path = Path(self.source_file)
        else:
            source_path = Path(paths.data_dir) / "consolidate" / "consolidated_tables.xlsx"
        
        if self.output_file:
            output_path = Path(self.output_file)
        else:
            output_path = Path(paths.data_dir) / "transpose" / "consolidated_tables_transposed.xlsx"
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        stats = {
            'sheets_processed': 0,
            'sheets_copied': 0,
            'sheets_skipped': 0,
            'errors': []
        }
        
        try:
            # Load source workbook
            source_wb = load_workbook(source_path)
            sheet_names = source_wb.sheetnames
            
            # Separate navigation sheets from data sheets
            nav_sheets = [s for s in sheet_names if s in ['TOC', 'Index', 'TOC_Sheet']]
            data_sheets = [s for s in sheet_names if s not in ['TOC', 'Index', 'TOC_Sheet']]
            
            # Reorder: TOC first, then Index, then data sheets
            ordered_nav = []
            if 'TOC' in nav_sheets:
                ordered_nav.append('TOC')
            if 'Index' in nav_sheets:
                ordered_nav.append('Index')
            for s in nav_sheets:
                if s not in ordered_nav:
                    ordered_nav.append(s)
            
            # Process sheets in correct order: Navigation first, then data
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Step 1: Copy navigation sheets FIRST using openpyxl (preserves hyperlinks)
                for sheet_name in ordered_nav:
                    try:
                        # Copy sheet with hyperlinks preserved
                        source_ws = source_wb[sheet_name]
                        # Create new sheet in output workbook
                        dest_ws = writer.book.create_sheet(sheet_name)
                        
                        # Copy each cell including value, style, and hyperlink
                        for row in source_ws.iter_rows():
                            for cell in row:
                                new_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                                if cell.has_style:
                                    new_cell.font = cell.font.copy()
                                    new_cell.fill = cell.fill.copy()
                                    new_cell.alignment = cell.alignment.copy()
                                if cell.hyperlink:
                                    new_cell.hyperlink = cell.hyperlink
                        
                        # Copy column widths
                        for col_letter, col_dim in source_ws.column_dimensions.items():
                            dest_ws.column_dimensions[col_letter].width = col_dim.width
                        
                        stats['sheets_copied'] += 1
                        logger.debug(f"Copied navigation sheet: {sheet_name}")
                    except Exception as e:
                        logger.warning(f"Error copying {sheet_name}: {e}")
                        stats['errors'].append(f"{sheet_name}: {str(e)}")
                
                # Step 2: Process and transpose data sheets
                for sheet_name in data_sheets:
                    try:
                        self._process_sheet(source_path, sheet_name, writer, stats)
                        stats['sheets_processed'] += 1
                    except Exception as e:
                        logger.warning(f"Error processing sheet {sheet_name}: {e}")
                        stats['errors'].append(f"{sheet_name}: {str(e)}")
            
            # Close source workbook
            source_wb.close()
            
            # Apply currency/percentage formatting to transposed output
            try:
                self._apply_number_formatting(output_path)
            except Exception as e:
                logger.warning(f"Could not apply number formatting: {e}")
            
            logger.info(f"Transposed {stats['sheets_processed']} sheets to {output_path}")
            
            return StepResult(
                step_name=self.name,
                status=StepStatus.SUCCESS,
                data={
                    'output_path': str(output_path),
                    **stats
                },
                message=f"Transposed {stats['sheets_processed']} sheets",
                metadata={
                    'source': str(source_path),
                    'output': str(output_path)
                }
            )
            
        except Exception as e:
            logger.error(f"Transpose failed: {e}", exc_info=True)
            return StepResult(
                step_name=self.name,
                status=StepStatus.FAILED,
                error=str(e)
            )
    
    def _process_sheet(self, source_path: Path, sheet_name: str, writer, stats: Dict):
        """Process and transpose a single sheet while preserving metadata.
        
        The metadata rows (Category, Line Items, Column Headers, etc.) are preserved
        at the top of the sheet. Only the table data (after 'Row Label') is transposed.
        """
        from src.infrastructure.extraction.consolidation.consolidated_exporter_transpose import (
            create_transposed_dataframe,
            reconstruct_metadata_from_df,
        )
        
        # Read the sheet with no header (all rows as data)
        df = pd.read_excel(source_path, sheet_name=sheet_name, header=None)
        
        if df.empty:
            logger.warning(f"Sheet {sheet_name} is empty, skipping")
            return
        
        # Find the header row (the one containing 'Row Label')
        header_row_idx = None
        for idx in range(min(20, len(df))):
            first_cell = str(df.iloc[idx, 0]).strip() if pd.notna(df.iloc[idx, 0]) else ''
            if first_cell == 'Row Label':
                header_row_idx = idx
                break
        
        if header_row_idx is None:
            logger.warning(f"Sheet {sheet_name} has no 'Row Label' header, skipping")
            return
        
        # === PRESERVE METADATA ROWS ===
        # Metadata rows are rows 0 to header_row_idx-1 (before 'Row Label')
        metadata_df = df.iloc[:header_row_idx].copy() if header_row_idx > 0 else pd.DataFrame()
        
        # Set proper column headers from header row
        headers = df.iloc[header_row_idx].tolist()
        headers = [str(h) if pd.notna(h) else f'Unnamed_{i}' for i, h in enumerate(headers)]
        
        # Extract data rows (after header row)
        data_df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
        data_df.columns = headers
        
        # Ensure 'Row Label' column exists
        if 'Row Label' not in data_df.columns:
            logger.warning(f"Sheet {sheet_name} missing 'Row Label' column after processing, skipping")
            return
        
        # Reconstruct metadata for transpose
        row_labels, label_to_section, normalized_row_labels = reconstruct_metadata_from_df(data_df)
        
        # Transpose ONLY the data portion
        transposed_df = create_transposed_dataframe(
            data_df, row_labels, label_to_section, normalized_row_labels
        )
        
        # Flatten MultiIndex columns before writing (pandas limitation)
        if isinstance(transposed_df.columns, pd.MultiIndex):
            # NOTE: Do NOT sort columns - this breaks the Dates-first ordering
            # The MultiIndex PerformanceWarning is acceptable to preserve column order
            
            # Convert MultiIndex to flat column names: "Category - Line Item"
            flat_cols = []
            for col in transposed_df.columns:
                if isinstance(col, tuple):
                    # Filter out empty parts and join with ' - '
                    parts = [str(p).strip() for p in col if p and str(p).strip()]
                    flat_cols.append(' - '.join(parts) if parts else 'Unnamed')
                else:
                    flat_cols.append(str(col))
            transposed_df.columns = flat_cols
        
        # === WRITE OUTPUT WITH METADATA PRESERVED ===
        # Note: metadata_df contains "← Back to Index" text from source but hyperlink is lost
        # We need to re-add the hyperlink after writing
        
        if not metadata_df.empty:
            # Write metadata rows first (without header) - includes existing back link text
            metadata_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=0)
            
            # Write transposed data after metadata (with header)
            transposed_startrow = len(metadata_df)
            transposed_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=transposed_startrow)
            
            # Re-add hyperlink to back-to-index cell (pandas loses hyperlinks when reading)
            ws = writer.sheets[sheet_name]
            from openpyxl.styles import Font
            # Find the cell with back-to-index text and add hyperlink
            for r in range(1, min(5, ws.max_row + 1)):
                cell = ws.cell(row=r, column=1)
                if cell.value and '← Back to Index' in str(cell.value):
                    cell.hyperlink = "#Index!A1"
                    cell.font = Font(color="0563C1", underline="single")
                    break
        else:
            # No metadata - write transposed data and add back link
            transposed_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
            
            # Add back-to-index hyperlink
            ws = writer.sheets[sheet_name]
            from openpyxl.styles import Font
            cell = ws.cell(row=1, column=1)
            cell.value = "← Back to Index"
            cell.hyperlink = "#Index!A1"
            cell.font = Font(color="0563C1", underline="single")
    
    def _create_index_sheet(self, writer, total_sheets: int):
        """Create a simple index sheet for the transposed output."""
        
        index_data = {
            'Info': ['Transposed Tables Report'],
            'Value': [f'{total_sheets} sheets processed']
        }
        index_df = pd.DataFrame(index_data)
        index_df.to_excel(writer, sheet_name='Index', index=False)
    
    def _apply_number_formatting(self, output_path: Path) -> None:
        """
        Apply currency and percentage formatting to transposed output.
        
        Uses similar logic to ExcelFormatter.apply_currency_format() but adapted
        for transposed table structure where columns are metric names.
        """
        from openpyxl import load_workbook
        
        # US currency accounting format
        CURRENCY_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        PERCENTAGE_FORMAT = '0.00%'
        
        # Indicators for determining format type
        CURRENCY_INDICATORS = ['$', 'dollar', 'revenue', 'income', 'expense', 'cost', 'assets', 'liabilities']
        PERCENTAGE_INDICATORS = ['%', 'percent', 'ratio', 'margin', 'return', 'rate', 'yield', 'roe', 'roa', 'rotce']
        
        def try_convert_to_number(val):
            """Try to convert value to number, return None if not possible."""
            if isinstance(val, (int, float)):
                return val
            if isinstance(val, str):
                try:
                    # Try float first (handles both int and float strings)
                    return float(val.replace(',', ''))
                except (ValueError, AttributeError):
                    pass
            return None
        
        wb = load_workbook(output_path)
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ['Index', 'TOC', 'TOC_Sheet']:
                continue
            
            ws = wb[sheet_name]
            
            # Find the header row (first row with actual column headers after metadata)
            header_row = None
            for r in range(1, min(15, ws.max_row + 1)):
                cell_val = str(ws.cell(row=r, column=1).value or '').lower()
                # Look for date-related column header names
                if 'dates' in cell_val or 'period' in cell_val or 'date' in cell_val or 'year' in cell_val:
                    header_row = r
                    break
            
            if header_row is None:
                # Fallback: assume first non-empty row with multiple values is header
                for r in range(1, min(15, ws.max_row + 1)):
                    non_empty = sum(1 for c in range(1, min(10, ws.max_column + 1)) if ws.cell(row=r, column=c).value)
                    if non_empty > 2:
                        header_row = r
                        break
            
            if header_row is None:
                continue
            
            data_start_row = header_row + 1
            
            # Get column headers to determine format per column
            for col in range(2, ws.max_column + 1):
                col_header = str(ws.cell(row=header_row, column=col).value or '').lower()
                
                # Skip Source column
                if 'source' in col_header:
                    continue
                
                # Determine if this column should be percentage or currency
                is_pct_column = any(ind in col_header for ind in PERCENTAGE_INDICATORS)
                is_currency_column = any(ind in col_header for ind in CURRENCY_INDICATORS)
                
                if not is_pct_column and not is_currency_column:
                    # Infer from values
                    col_values = []
                    for r in range(data_start_row, min(data_start_row + 10, ws.max_row + 1)):
                        cell = ws.cell(row=r, column=col)
                        num_val = try_convert_to_number(cell.value)
                        if num_val is not None and num_val != 0:
                            col_values.append(num_val)
                    
                    if col_values:
                        all_in_pct_range = all(-1 <= v <= 1 for v in col_values)
                        has_decimal = any(0 < abs(v) < 1 for v in col_values)
                        is_pct_column = all_in_pct_range and has_decimal
                
                # Apply format to all data cells in this column
                format_to_apply = PERCENTAGE_FORMAT if is_pct_column else CURRENCY_FORMAT
                for r in range(data_start_row, ws.max_row + 1):
                    cell = ws.cell(row=r, column=col)
                    num_val = try_convert_to_number(cell.value)
                    if num_val is not None:
                        # Convert string to number and apply format
                        cell.value = num_val
                        cell.number_format = format_to_apply
        
        wb.save(output_path)
        logger.debug(f"Applied number formatting to transposed output")


# Backward-compatible function for main.py CLI
def run_transpose(
    source_file: Optional[str] = None,
    output_file: Optional[str] = None
) -> PipelineResult:
    """
    Run transpose on consolidated tables.
    
    Legacy wrapper maintaining backward compatibility with main.py CLI.
    
    Args:
        source_file: Override source file (default: data/consolidate/consolidated_tables.xlsx)
        output_file: Override output file (default: data/transpose/consolidated_tables_transposed.xlsx)
        
    Returns:
        PipelineResult with processing outcome
    """
    step = TransposeStep(source_file=source_file, output_file=output_file)
    ctx = PipelineContext()
    
    # Validate
    if not step.validate(ctx):
        logger.warning("Validation failed - source file not found")
        return PipelineResult(
            step=PipelineStep.TRANSPOSE,
            success=False,
            data={},
            message="Source file not found - run consolidate first",
            error="Validation failed: consolidated_tables.xlsx not found",
            metadata={}
        )
    
    result = step.execute(ctx)
    
    return PipelineResult(
        step=PipelineStep.TRANSPOSE,
        success=result.success,
        data=result.data,
        message=result.message,
        error=result.error,
        metadata=result.metadata
    )
