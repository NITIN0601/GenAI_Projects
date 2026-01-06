"""
Header Flattener - Reusable multi-level header flattening utilities.

Extracted from process.py for modularity and reusability.

Used by: ProcessStep, any module needing header normalization
"""

from typing import List, Dict, Any, Optional
from openpyxl.cell.cell import MergedCell

from src.utils import get_logger

logger = get_logger(__name__)


class HeaderFlattener:
    """
    Flatten multi-level Excel headers into single normalized rows.
    
    Handles:
    - 1-level headers (single row)
    - 2-level headers (period + year)
    - 3-level headers (period + month/year)
    - 4+ level headers (complex spanning)
    
    Design: Stateless class methods for horizontal scaling.
    """
    
    @classmethod
    def flatten_headers(
        cls,
        ws,
        header_row: int,
        data_header_rows: List[List[str]],
        normalized_headers: List[str],
        stats: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        Flatten multi-level headers for a table starting at header_row.
        
        Args:
            ws: openpyxl worksheet
            header_row: Row where table headers start
            data_header_rows: List of header row values
            normalized_headers: List of normalized header values
            stats: Optional stats dictionary to update
        """
        if stats is None:
            stats = {}
        
        num_cols = ws.max_column
        num_header_rows = len(data_header_rows)
        
        if num_header_rows == 0:
            return
        
        logger.debug(f"Flattening {num_header_rows} header rows starting at row {header_row}")
        
        # Unmerge any merged cells in the header rows
        cls._unmerge_header_cells(ws, header_row, num_header_rows)
        
        if num_header_rows == 1:
            cls._flatten_single_level(ws, header_row, normalized_headers, num_cols)
        elif num_header_rows == 2:
            cls._flatten_two_level(ws, header_row, normalized_headers, num_cols, stats)
        else:
            cls._flatten_multi_level(ws, header_row, data_header_rows, normalized_headers, num_cols, num_header_rows, stats)
        
        logger.debug(f"Flattened headers at row {header_row}")
    
    @classmethod
    def _unmerge_header_cells(cls, ws, header_row: int, num_header_rows: int) -> None:
        """Unmerge any merged cells in header rows."""
        merged_ranges_to_unmerge = []
        for merge_range in ws.merged_cells.ranges:
            if header_row <= merge_range.min_row <= header_row + num_header_rows:
                merged_ranges_to_unmerge.append(merge_range)
        
        for merge_range in merged_ranges_to_unmerge:
            try:
                ws.unmerge_cells(str(merge_range))
            except Exception as e:
                logger.debug(f"Could not unmerge {merge_range}: {e}")
    
    @classmethod
    def _flatten_single_level(cls, ws, header_row: int, normalized_headers: List[str], num_cols: int) -> None:
        """Handle single-level headers."""
        for col_idx in range(1, num_cols + 1):
            if col_idx < len(normalized_headers) + 1 and normalized_headers[col_idx - 1]:
                cell = ws.cell(row=header_row, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = normalized_headers[col_idx - 1]
    
    @classmethod
    def _flatten_two_level(cls, ws, header_row: int, normalized_headers: List[str], num_cols: int, stats: Dict) -> None:
        """Handle two-level headers (most common)."""
        # Write combined headers to first row
        for col_idx in range(1, num_cols + 1):
            if col_idx < len(normalized_headers) + 1 and normalized_headers[col_idx - 1]:
                cell = ws.cell(row=header_row, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.value = normalized_headers[col_idx - 1]
        
        # Second row becomes empty separator
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=header_row + 1, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = None
        
        stats['headers_flattened'] = stats.get('headers_flattened', 0) + 1
        stats['empty_rows_added'] = stats.get('empty_rows_added', 0) + 1
    
    @classmethod
    def _flatten_multi_level(
        cls, ws, header_row: int, data_header_rows: List[List[str]], 
        normalized_headers: List[str], num_cols: int, num_header_rows: int, stats: Dict
    ) -> None:
        """Handle 3+ level headers."""
        # Import here to avoid circular imports
        from src.utils.header_processor import HeaderProcessor
        
        if num_header_rows == 3:
            combined = HeaderProcessor.build_combined_headers_3level(data_header_rows, normalized_headers, num_cols)
        else:
            combined = HeaderProcessor.build_combined_headers_4level(data_header_rows, normalized_headers, num_cols)
        
        # Write combined headers
        for col_idx, value in enumerate(combined, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            if not isinstance(cell, MergedCell):
                cell.value = value
        
        # Add empty separator
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=header_row + 1, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = None
        
        # Delete remaining header rows
        rows_to_delete = num_header_rows - 2
        for _ in range(rows_to_delete):
            ws.delete_rows(header_row + 2)
        
        stats['headers_flattened'] = stats.get('headers_flattened', 0) + 1
        stats['empty_rows_added'] = stats.get('empty_rows_added', 0) + 1


__all__ = ['HeaderFlattener']
