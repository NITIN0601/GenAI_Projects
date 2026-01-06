"""
Header Processor Utility.

Handles multi-row header flattening and normalization for Excel tables.
Extracted from ProcessStep to improve modularity and reduce file size.

Operations:
- Count header rows between Source marker and data start
- Detect multi-level headers (2, 3, 4+ levels)
- Flatten headers into single row with Y/Q date codes
- Build combined headers from period + year + category patterns
"""

from typing import Dict, Any, List, Optional
import re
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell

from src.utils import get_logger
from src.utils.excel_utils import ExcelUtils

logger = get_logger(__name__)


class HeaderProcessor:
    """
    Utility class for processing and flattening multi-row table headers.
    
    Handles common patterns in financial tables:
    - 2-level: Year + Category (e.g., "2025" + "Revenue")
    - 3-level: Period + Year + Category (e.g., "Three Months Ended" + "2025" + "Revenue")
    - 4-level: Period + Year duplicates + Category
    """
    
    # Default constants (can be overridden)
    DEFAULT_HEADER_START_ROW = 13
    
    @classmethod
    def count_header_rows(cls, ws: Worksheet, source_row: int = 12) -> int:
        """
        Count header rows between Source marker and first data row.
        
        Args:
            ws: Worksheet to analyze
            source_row: Row containing "Source:" marker (default 12)
            
        Returns:
            Number of header rows (0 if unable to determine)
        """
        count = 0
        for row in range(source_row + 1, min(source_row + 8, ws.max_row + 1)):
            # Check if row has numeric data (indicates data, not header)
            has_numbers = False
            for col in range(2, min(ws.max_column + 1, 10)):
                val = ws.cell(row, col).value
                if val is not None:
                    try:
                        float(str(val).replace('%', '').replace(',', '').replace('$', '').replace('-', '0'))
                        has_numbers = True
                        break
                    except (ValueError, TypeError):
                        pass
            
            if has_numbers:
                break
            
            # Row has content but no numbers - it's a header row
            if any(ws.cell(row, c).value for c in range(1, min(ws.max_column + 1, 10))):
                count += 1
        
        return count
    
    @classmethod
    def find_data_start_row(cls, ws: Worksheet, start_row: int = 14, max_scan: int = 30) -> Optional[int]:
        """
        Find the first row containing numeric data values.
        
        Args:
            ws: Worksheet to analyze
            start_row: Row to start scanning from
            max_scan: Maximum rows to scan
            
        Returns:
            Row number of first data row, or None if not found
        """
        for row in range(start_row, min(start_row + max_scan, ws.max_row + 1)):
            numeric_count = 0
            for col in range(2, min(ws.max_column + 1, 15)):
                val = ws.cell(row, col).value
                if val is not None:
                    val_str = str(val).strip()
                    if not val_str:
                        continue
                    # Skip 4-digit years (they're headers)
                    if len(val_str) == 4 and val_str.isdigit():
                        continue
                    # Try to parse as number
                    try:
                        clean_val = val_str.replace('%', '').replace(',', '').replace('$', '').replace('-', '').strip()
                        if clean_val:
                            float(clean_val)
                            numeric_count += 1
                    except (ValueError, TypeError):
                        pass
            
            # Need at least 1 numeric value to consider it a data row
            if numeric_count >= 1:
                return row
        
        return None
    
    @classmethod
    def build_combined_headers_3level(
        cls,
        header_rows_data: List[List[Any]],
        normalized_headers: List[str],
        num_cols: int
    ) -> List[str]:
        """
        Build combined headers from 3-level structure: Period + Year + Category.
        
        Pattern: Row 0 (date), Row 1 (date dup or category labels), Row 2 (categories)
        
        Args:
            header_rows_data: List of header row values
            normalized_headers: Pre-normalized header values (Y/Q codes)
            num_cols: Number of columns
            
        Returns:
            List of combined header strings
        """
        combined = []
        
        # Find the primary date code from normalized_headers
        primary_date_code = None
        for h in normalized_headers:
            if h and (h.startswith('Q') or h.startswith('YTD') or '-' in str(h)):
                primary_date_code = h
                break
        
        # Build column-to-date-code mapping from normalized headers
        date_codes_by_col = {}
        if normalized_headers:
            for col_idx, norm_val in enumerate(normalized_headers):
                if norm_val and (str(norm_val).startswith('Q') or str(norm_val).startswith('YTD') or '-' in str(norm_val)):
                    date_codes_by_col[col_idx] = norm_val
        
        for col_idx in range(num_cols):
            # Get date code for this column
            date_code = date_codes_by_col.get(col_idx, primary_date_code)
            
            # Get category from bottom header row
            category = ''
            if len(header_rows_data) >= 1:
                last_row = header_rows_data[-1]
                if col_idx < len(last_row) and last_row[col_idx]:
                    category = str(last_row[col_idx]).strip()
            
            # Build combined header
            if col_idx == 0:
                # First column is label column
                combined.append(category or '')
            elif date_code and category:
                combined.append(f"{date_code} {category}")
            elif date_code:
                combined.append(date_code)
            elif category:
                combined.append(category)
            else:
                combined.append('')
        
        return combined
    
    @classmethod
    def build_combined_headers_4level(
        cls,
        header_rows_data: List[List[Any]],
        normalized_headers: List[str],
        num_cols: int
    ) -> List[str]:
        """
        Build combined headers from 4+ level structure: Period + Year + Year + Category.
        
        Pattern: Row 0 (period), Row 1 (year), Row 2 (year copy), Row 3 (categories)
        
        Args:
            header_rows_data: List of header row values
            normalized_headers: Pre-normalized header values (Y/Q codes)
            num_cols: Number of columns
            
        Returns:
            List of combined header strings
        """
        combined = []
        
        # Find year row and build column-to-year mapping
        year_row_idx = None
        col_to_year = {}
        
        for row_idx, row_data in enumerate(header_rows_data):
            has_years = False
            for col_idx, val in enumerate(row_data):
                if val and re.match(r'^20\d{2}$', str(val).strip()):
                    has_years = True
                    col_to_year[col_idx] = str(val).strip()
            if has_years:
                year_row_idx = row_idx
                break
        
        # Propagate years to adjacent columns (for spanning headers)
        if col_to_year:
            last_year = None
            for col_idx in range(num_cols):
                if col_idx in col_to_year:
                    last_year = col_to_year[col_idx]
                elif last_year:
                    col_to_year[col_idx] = last_year
        
        # Build column-to-date-code mapping from normalized headers
        date_codes_by_col = {}
        if normalized_headers:
            for col_idx, norm_val in enumerate(normalized_headers):
                if norm_val and (str(norm_val).startswith('Q') or str(norm_val).startswith('YTD') or '-' in str(norm_val)):
                    date_codes_by_col[col_idx] = norm_val
        
        for col_idx in range(num_cols):
            date_code = date_codes_by_col.get(col_idx, '')
            
            # Get category from bottom header row
            category = ''
            if header_rows_data:
                last_row = header_rows_data[-1]
                if col_idx < len(last_row) and last_row[col_idx]:
                    category = str(last_row[col_idx]).strip()
            
            # Build combined header
            if col_idx == 0:
                combined.append(category or '')
            elif date_code and category:
                combined.append(f"{date_code} {category}")
            elif date_code:
                combined.append(date_code)
            elif category:
                combined.append(category)
            else:
                combined.append('')
        
        return combined
    
    @classmethod
    def clean_header_values(cls, headers: List[Any]) -> List[str]:
        """
        Clean a list of header values.
        
        - Strips trailing footnote markers (e.g., ' 1' from 'Netting 1')
        - Preserves meaningful numbers like 'Level 1', 'Tier 1'
        
        Args:
            headers: List of raw header values
            
        Returns:
            List of cleaned header strings
        """
        cleaned = []
        for h in headers:
            if h is None:
                cleaned.append('')
                continue
            
            h_str = str(h).strip()
            
            # Clean .0 from float years
            h_str = re.sub(r'(\d{4})\.0$', r'\1', h_str)
            
            # Remove trailing footnote numbers (but not 'Level 1', 'Tier 1')
            # Pattern: space + single digits at end that are clearly footnotes
            if not re.search(r'(level|tier|part|phase|type|class)\s+\d+$', h_str.lower()):
                h_str = re.sub(r'\s+\d{1,2}$', '', h_str)
            
            cleaned.append(h_str)
        
        return cleaned
    
    @classmethod
    def is_spanning_header_row(cls, row_values: List[Any], threshold: float = 0.7) -> bool:
        """
        Detect if a row is a spanning header (mostly empty with few values).
        
        Args:
            row_values: List of cell values in the row
            threshold: Fraction of empty cells to consider "spanning"
            
        Returns:
            True if row appears to be a spanning header
        """
        if not row_values:
            return False
        
        empty_count = sum(1 for v in row_values if not v or str(v).strip() == '')
        total = len(row_values)
        
        return (empty_count / total) >= threshold


__all__ = ['HeaderProcessor']
