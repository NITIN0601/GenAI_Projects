"""
Excel Table Exporter with Index Sheet.

Exports extracted tables to multi-sheet Excel with:
- Index sheet with hyperlinks to each table sheet
- One sheet per unique table title
- Multiple tables with same title stacked with blank rows
- Back-links from each sheet to Index

Usage:
    from src.infrastructure.extraction.formatters.excel_exporter import get_excel_exporter
    
    exporter = get_excel_exporter()
    
    # Per-PDF export
    exporter.export_pdf_tables(tables, "10q0625.pdf")
    # -> data/extracted_raw/10q0625_tables.xlsx
    
    # For consolidated multi-file merge, use:
    from src.infrastructure.extraction.consolidation.consolidated_exporter import get_consolidated_exporter

"""

__version__ = "2.1.0"
__author__ = "dundaymo"

import pandas as pd
from typing import List, Dict, Any, Optional
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import re

from src.utils import get_logger
from src.core import get_paths
from src.infrastructure.extraction.formatters.header_detector import HeaderDetector
from src.infrastructure.extraction.exporters.base_exporter import BaseExcelExporter
# Import from new focused module
from src.infrastructure.extraction.exporters.toc_builder import TOCBuilder
from src.utils.excel_utils import ExcelUtils
from src.utils.metadata_labels import MetadataLabels
from src.utils.metadata_builder import MetadataBuilder
from src.utils.financial_domain import extract_quarter_from_header, extract_year_from_header, convert_year_to_q4_header

from openpyxl import load_workbook

logger = get_logger(__name__)


# =============================================================================
# DYNAMIC HEADER DETECTION HELPERS
# =============================================================================

# Unit patterns that indicate column header rows (case-insensitive)
UNIT_PATTERNS = ['$ in millions', '$ in billions', '$ in thousands', 
                 'in millions', 'in billions', 'in thousands',
                 '$in millions', '$in billions']  # Handle no-space variants

MAX_HEADER_LEVELS = 4  # Track up to 4 header levels above unit row


def is_unit_row(row_values: list) -> bool:
    """
    Detect if row is a unit row (column header starting with $ in millions, etc.)
    
    Args:
        row_values: List of cell values in the row
        
    Returns:
        True if first column matches a unit pattern
    """
    if not row_values:
        return False
    first_col = str(row_values[0]).strip().lower() if row_values[0] else ''
    return any(first_col.startswith(p.lower()) for p in UNIT_PATTERNS)


def is_header_row(row_values: list) -> bool:
    """
    Detect if row is a header row (not a data row).
    
    A header row has:
    - First column empty OR text-only, AND
    - Other columns have text (dates, periods) not pure numeric data
    
    Args:
        row_values: List of cell values in the row
        
    Returns:
        True if row appears to be a header
    """
    if not row_values or len(row_values) < 2:
        return False
    
    # Check first column
    first_col = str(row_values[0]).strip() if row_values[0] else ''
    
    # First col should be empty or non-numeric for spanning headers
    if first_col:
        # If first col has pure numbers (not dates), it's likely data
        first_clean = first_col.replace(',', '').replace('.', '').replace('-', '').replace('$', '').replace('%', '')
        if first_clean.isdigit() and len(first_clean) > 2 and not (1900 <= int(first_clean) <= 2100):
            return False  # Numeric data, not header
    
    # Check if other columns have numeric data (which would make it a data row)
    numeric_count = 0
    text_count = 0
    for val in row_values[1:]:
        if val is None or str(val).strip() == '':
            continue
        val_str = str(val).strip()
        val_clean = val_str.replace(',', '').replace('.', '').replace('-', '').replace('$', '').replace('%', '').replace(' ', '')
        
        # Check if it's a year (1990-2100)
        if val_clean.isdigit() and len(val_clean) == 4 and 1990 <= int(val_clean) <= 2100:
            text_count += 1  # Years are headers
        elif val_clean.lstrip('-').isdigit() and len(val_clean) > 2:
            numeric_count += 1  # Numeric data
        elif val_str:
            text_count += 1  # Text header
    
    # If mostly numeric, it's a data row
    if numeric_count > text_count and numeric_count > 1:
        return False
    
    # If mostly text or empty, it's likely a header
    return text_count > 0 or first_col == ''


class ExcelTableExporter(BaseExcelExporter):
    """
    Export extracted tables to multi-sheet Excel with Index.
    
    Features:
    - Index sheet with Source, PageNo, Table_ID, Location_ID, Table Title, Link
    - One sheet per unique table title (using Table_ID as sheet name)
    - Detailed header structure: Row Headers, Column Headers, Product/Entity
    - Same-title tables stacked with blank separator
    - Bidirectional hyperlinks (Index <-> Sheets)
    
    Inherits shared Excel utilities from BaseExcelExporter.
    Follows singleton pattern consistent with other managers.
    """
    
    def __init__(self):
        """Initialize exporter with paths from PathManager."""
        super().__init__()
        self.paths = get_paths()
        # Step 1 output: extracted_raw (per docs/pipeline_operations.md)
        self.extracted_raw_dir = self.paths.data_dir / "extracted_raw"
        
        # Ensure directory exists
        self.extracted_raw_dir.mkdir(parents=True, exist_ok=True)
    
    def export_pdf_tables(
        self,
        tables: List[Dict[str, Any]],
        source_pdf: str,
        output_dir: Optional[str] = None
    ) -> str:
        """
        Export tables from a single PDF to Excel.
        
        Args:
            tables: List of extracted table dictionaries with:
                - content: Table content (markdown or text)
                - metadata: Dict with page_no, table_title, etc.
            source_pdf: Source PDF filename
            output_dir: Override output directory
            
        Returns:
            Path to created Excel file
        """
        output_dir = Path(output_dir) if output_dir else self.extracted_raw_dir
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate output filename
        pdf_name = Path(source_pdf).stem
        output_path = output_dir / f"{pdf_name}_tables.xlsx"
        
        return self._create_excel_workbook(tables, output_path, source_pdf)
    
    # NOTE: export_combined_tables was removed as it was unused.
    # Use ConsolidatedExporter for cross-file consolidation instead.
    
    def _create_excel_workbook(
        self,
        tables: List[Dict[str, Any]],
        output_path: Path,
        source_label: str
    ) -> str:
        """
        Create Excel workbook with Index and table sheets.
        
        Args:
            tables: List of table dictionaries
            output_path: Output file path
            source_label: Label for source (PDF name or "Combined")
            
        Returns:
            Path to created file
        """
        if not tables:
            logger.warning("No tables to export")
            return ""
        
        try:
            # Group tables by title (returns dict with Table_ID as key)
            tables_by_title = self._group_tables_by_title(tables)
            
            # Create workbook
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Create TOC sheet first (hierarchical section view)
                self._create_toc_sheet(writer, tables)
                
                # Create Index sheet
                self._create_index_sheet(writer, tables, tables_by_title)
                
                # Create table sheets - use Table_ID as sheet name
                for table_id, title_tables in tables_by_title.items():
                    self._create_table_sheet(writer, table_id, title_tables)
            
            # Add hyperlinks (requires reopening workbook)
            self._add_hyperlinks(output_path, tables_by_title)
            
            # Merge repeated header cells to match PDF layout
            self._merge_repeated_header_cells(output_path, tables_by_title)
            
            logger.info(f"Exported {len(tables)} tables to {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Failed to export tables: {e}", exc_info=True)
            return ""
    
    def _group_tables_by_title(
        self,
        tables: List[Dict[str, Any]]
    ) -> Dict[str, List[Dict[str, Any]]]:
        """
        Group tables by their logical table (Section + Title).
        
        IMPORTANT: Tables with the same title but different sections are DIFFERENT tables.
        E.g., "Income Statement Info" under "Institutional Securities" vs "Wealth Management"
        should NOT be merged.
        
        Returns dict with Table_ID as key to avoid sheet name collisions.
        """
        groups = defaultdict(list)
        logical_table_ids = {}
        next_table_id = 1
        
        for table in tables:
            metadata = table.get('metadata', {})
            title = metadata.get('table_title', 'Untitled')
            section = metadata.get('section_name', '')  # Get section name
            
            # Clean title to identify logical table (remove row ranges)
            cleaned_title = self._clean_title_for_grouping(title)
            normalized_title = self._normalize_title_for_grouping(cleaned_title)
            
            # Create grouping key: Section + Title
            # This ensures same-titled tables in different sections stay separate
            # Fix OCR broken words and normalize whitespace
            section_fixed = ExcelUtils.fix_ocr_broken_words(section) if section else ''
            section_normalized = re.sub(r'\s+', ' ', section_fixed).strip().lower() if section_fixed else ''
            if section_normalized:
                grouping_key = f"{section_normalized}::{normalized_title}"
            else:
                grouping_key = f"default::{normalized_title}"
            
            # Assign Table_ID based on unique Section + Title combination
            if grouping_key not in logical_table_ids:
                logical_table_ids[grouping_key] = str(next_table_id)
                next_table_id += 1
            
            table_id = logical_table_ids[grouping_key]
            
            # Store the table_id and section in metadata for later use
            if 'metadata' not in table:
                table['metadata'] = {}
            table['metadata']['_assigned_table_id'] = table_id
            table['metadata']['_cleaned_title'] = cleaned_title
            table['metadata']['_section'] = section
            table['metadata']['_grouping_key'] = grouping_key
            
            # Group by Table_ID instead of sanitized title
            groups[table_id].append(table)
        
        return dict(groups)
    
    def _clean_title_for_grouping(self, title: str) -> str:
        """
        Clean title for grouping - remove row ranges and chunk indicators.
        
        This ensures chunks of the same table are grouped together:
        - "Income Statement (Rows 1-10)" -> "Income Statement"
        - "Balance Sheet (Rows 8-15)" -> "Balance Sheet"
        """
        if not title:
            return "Untitled"
        
        # Remove row ranges like "(Rows 1-10)" or "(Rows 8-17)"
        cleaned = re.sub(r'\s*\(Rows?\s*\d+[-–]\d+\)\s*$', '', title, flags=re.IGNORECASE)
        
        # Remove trailing numbers that might be footnotes
        cleaned = re.sub(r'\s+\d+\s*$', '', cleaned)
        
        # Remove trailing parenthesized numbers
        cleaned = re.sub(r'\s*\(\d+\)\s*$', '', cleaned)
        
        return cleaned.strip() or "Untitled"
    
    def _normalize_title_for_grouping(self, title: str) -> str:
        """
        Normalize title for case-insensitive grouping.
        Delegates to ExcelUtils after cleaning row ranges.
        """
        if not title:
            return "untitled"
        
        # First clean row ranges
        cleaned = self._clean_title_for_grouping(title)
        
        # Use ExcelUtils for the actual normalization
        return ExcelUtils.normalize_title_for_grouping(cleaned) or "untitled"
    
    # NOTE: _sanitize_sheet_name inherited from BaseExcelExporter
    
    def _create_toc_sheet(
        self,
        writer: pd.ExcelWriter,
        tables: List[Dict[str, Any]]
    ) -> None:
        """
        Create TOC sheet with hierarchical section columns.
        
        Delegates to TOCBuilder for centralized logic.
        """
        TOCBuilder.create_toc_sheet(
            writer,
            tables,
            clean_title_func=self._clean_title_for_grouping,
            normalize_title_func=self._normalize_title_for_grouping,
            get_column_letter_func=self._get_column_letter
        )
    
    def _create_index_sheet(
        self,
        writer: pd.ExcelWriter,
        tables: List[Dict[str, Any]],
        tables_by_title: Dict[str, List[Dict[str, Any]]]
    ) -> None:
        """
        Create Index sheet with comprehensive table metadata.
        
        Delegates to TOCBuilder for centralized logic.
        """
        TOCBuilder.create_index_sheet(
            writer,
            tables,
            tables_by_title,
            clean_title_func=self._clean_title_for_grouping,
            normalize_title_func=self._normalize_title_for_grouping,
            get_column_letter_func=self._get_column_letter
        )
    
    # NOTE: _get_column_letter inherited from BaseExcelExporter
    
    def _create_table_sheet(
        self,
        writer: pd.ExcelWriter,
        sheet_name: str,
        tables: List[Dict[str, Any]]
    ) -> None:
        """
        Create sheet for a specific table (identified by Table_ID).
        
        Sheet Structure:
        - Row 1: "← Back to Index" link
        - Row 2: Row Header (Level 1): ...
        - Row 3: Row Header (Level 2 - Sub): ... (if applicable)
        - Row 4: Product/Entity: ...
        - Row 5: Column Headers (Level 1): ...
        - Row 6: Column Headers (Level 2): ... (if applicable)
        - Row 7: <blank>
        - Row 8: Table Title: ...
        - Row 9: <blank>
        - Row 10+: Table data
        
        For multi-page tables with consecutive pages:
        - Consolidates Source line: Source(s): pdf_pg24, pdf_pg25, pdf_pg26
        - Combines all table data into one continuous block
        
        Args:
            writer: Excel writer
            sheet_name: Sheet name (Table_ID, e.g., "1", "2")
            tables: List of table chunks with same logical table
        """
        all_rows = []
        
        # Row 1: Add "Back to Index" link placeholder
        all_rows.append(['← Back to Index'])
        
        # Collect all page numbers to detect multi-page pattern
        all_pages = []
        for table in tables:
            metadata = table.get('metadata', {})
            page_no = metadata.get('page_no')
            if page_no is not None and page_no != 'N/A':
                try:
                    all_pages.append(int(page_no))
                except (ValueError, TypeError):
                    pass
        
        unique_pages = sorted(set(all_pages))
        is_consecutive_multipage = (
            len(unique_pages) > 1 and 
            unique_pages == list(range(unique_pages[0], unique_pages[-1] + 1))
        )
        
        if is_consecutive_multipage:
            # Consolidate multi-page tables into single metadata block
            all_rows.extend(self._generate_consolidated_table_rows(tables, unique_pages))
        else:
            # Original behavior: separate metadata block per chunk
            for i, table in enumerate(tables):
                # Add 2 blank lines before each table (except first)
                if i > 0:
                    all_rows.append([])
                    all_rows.append([])
                
                # Generate metadata + data rows for this table
                table_rows = self._generate_single_table_rows(table)
                all_rows.extend(table_rows)
        
        # Create DataFrame from all rows
        df = pd.DataFrame(all_rows)
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
    
    def _generate_consolidated_table_rows(
        self, 
        tables: List[Dict[str, Any]], 
        pages: List[int]
    ) -> List[List[Any]]:
        """
        Generate consolidated rows for multi-page tables.
        
        Creates ONE metadata block with consolidated Source line,
        then combines all table data.
        """
        if not tables:
            return []
        
        # Use first table for metadata template
        first_table = tables[0]
        metadata = first_table.get('metadata', {})
        display_title = metadata.get('_cleaned_title', metadata.get('table_title', 'Untitled'))
        source_doc = metadata.get('source_doc', 'Unknown')
        
        # Generate rows using first table's metadata
        rows = self._generate_single_table_rows_without_data(first_table)
        
        # Replace Source line with consolidated version
        source_parts = [f"{source_doc}_pg{p}" for p in pages]
        source_str = ', '.join(source_parts)
        if metadata.get('year'):
            source_str += f", {metadata.get('year')}"
        if metadata.get('quarter'):
            source_str += f" {metadata.get('quarter')}"
        
        # Find and replace Source line
        for i, row in enumerate(rows):
            if row and len(row) > 0 and 'Source' in str(row[0]):
                rows[i] = [f"{MetadataLabels.SOURCES} {source_str}"]
                break
        
        # Combine data from all tables with proper separation
        first_chunk = True
        for table in tables:
            table_df = self._parse_table_content(table.get('content', ''))
            if not table_df.empty:
                # Add blank line before each table chunk (except first)
                if not first_chunk:
                    rows.append([])  # Blank line separator
                first_chunk = False
                
                # Add column headers for this chunk (keep raw, Process step will normalize)
                def clean_header(c):
                    if pd.isna(c):
                        return ''
                    if isinstance(c, float) and c == int(c):
                        return str(int(c))
                    return str(c)
                rows.append([clean_header(c) for c in table_df.columns])
                
                # Add data rows
                for _, row in table_df.iterrows():
                    cleaned_row = [ExcelUtils.clean_cell_value(v) if pd.notna(v) else '' for v in row]
                    rows.append(cleaned_row)
        
        return rows
    
    def _generate_single_table_rows_without_data(self, table: Dict[str, Any]) -> List[List[Any]]:
        """Generate metadata rows only (no data) for consolidated tables."""
        # Reuse existing logic but stop before data rows
        full_rows = self._generate_single_table_rows(table)
        
        # Find where data starts (after Source line + blank)
        source_idx = None
        for i, row in enumerate(full_rows):
            if row and len(row) > 0 and 'Source' in str(row[0]):
                source_idx = i
                break
        
        if source_idx is not None:
            # Return up to 2 rows after Source (Source + blank + col headers)
            return full_rows[:source_idx + 3]
        return full_rows[:12]  # Fallback: first 12 rows are metadata

    
    def _detect_column_header_levels(self, content: str) -> Dict[str, Any]:
        """
        Detect multi-level column headers from markdown table content.
        
        Delegates to HeaderDetector class.
        
        Returns:
            {
                'has_multi_level': bool,
                'level_1': List[str],  # Main/spanning headers
                'level_2': List[str]   # Sub-headers (typically with dates/years)
            }
        """
        return HeaderDetector.detect_column_header_levels(content)
    
    def _dedupe_preserve_order(self, items: List[str]) -> List[str]:
        """Deduplicate a list while preserving order. Delegates to HeaderDetector."""
        return HeaderDetector.dedupe_preserve_order(items)
    
    def _parse_table_content(self, content: str) -> pd.DataFrame:
        """Parse markdown/text table content to DataFrame."""
        from src.utils.table_utils import parse_markdown_table
        return parse_markdown_table(content, handle_colon_separator=False)
    
    def _generate_single_table_rows(self, table: Dict[str, Any]) -> List[List[Any]]:
        """
        Generate all rows (metadata + data) for a single table.
        
        This creates a complete block for one table:
        - Row Header (Level 1): ...
        - Row Header (Level 2): ...
        - Product/Entity: ...
        - Column Header (Level 1): ...
        - Column Header (Level 2): ...
        - Year(s): ...
        - [blank]
        - Table Title: ...
        - Source: ...
        - [Column headers row]
        - [Data rows...]
        
        Args:
            table: Single table dict with 'content' and 'metadata'
            
        Returns:
            List of rows to add to the sheet
        """
        rows = []
        metadata = table.get('metadata', {})
        display_title = metadata.get('_cleaned_title', metadata.get('table_title', 'Untitled'))
        
        # Parse table content
        table_df = self._parse_table_content(table.get('content', ''))
        
        # === COLLECT ROW HEADERS ===
        all_row_headers = []      # L1 - section headers
        all_row_sub_headers = []  # L2 - data row labels
        
        if not table_df.empty and len(table_df.columns) > 0:
            for idx, val in enumerate(table_df.iloc[:, 0]):
                if pd.notna(val) and str(val).strip():
                    row_label = str(val).strip()
                    
                    # Check if this row has numeric data in subsequent columns
                    has_numeric_data = False
                    all_same_text = True
                    
                    if len(table_df.columns) > 1:
                        row_data = table_df.iloc[idx, 1:]
                        for cell_val in row_data:
                            if pd.notna(cell_val) and str(cell_val).strip():
                                cell_str = str(cell_val).strip().lower()
                                if cell_str in ['nan', '', '-', '—']:
                                    continue
                                
                                if cell_str != row_label.lower():
                                    all_same_text = False
                                
                                cell_clean = str(cell_val).strip()
                                if (cell_clean.startswith('$') or 
                                    cell_clean.startswith('(') or
                                    (cell_clean and cell_clean[0].isdigit()) or
                                    any(c.isdigit() for c in cell_clean)):
                                    has_numeric_data = True
                                    break
                    
                    is_section_header = (not has_numeric_data) or all_same_text
                    
                    if is_section_header and len(table_df.columns) > 1:
                        all_row_headers.append(row_label)
                    else:
                        all_row_sub_headers.append(row_label)
        
        # Filter headers (keep original values without cleaning - cleaning happens at display time)
        filtered_row_headers = []
        for h in all_row_headers:
            if h and h.lower() not in ['nan', 'none', '']:
                filtered_row_headers.append(h)
        
        filtered_row_sub_headers = []
        for h in all_row_sub_headers:
            if h and h.lower() not in ['nan', 'none', '']:
                filtered_row_sub_headers.append(h)
        
        # Build L2 excluding L1 items
        seen_l1 = set(filtered_row_headers)
        l2_row_headers = [h for h in filtered_row_sub_headers if h not in seen_l1]
        
        # Clean footnotes ONLY for display (preserves original values above for matching)
        display_row_headers = [ExcelUtils.clean_footnote_references(h) for h in filtered_row_headers if h]
        display_l2_row_headers = [ExcelUtils.clean_footnote_references(h) for h in l2_row_headers if h]
        
        # Category (Parent) - section headers (formerly Row Header Level 1)
        row_headers_str = ', '.join(display_row_headers) if display_row_headers else ''
        rows.append([f"{MetadataLabels.CATEGORY_PARENT} {row_headers_str}"])
        
        # Line Items - data row labels (formerly Row Header Level 2)
        row_sub_headers_str = ', '.join(display_l2_row_headers) if display_l2_row_headers else ''
        rows.append([f"{MetadataLabels.LINE_ITEMS} {row_sub_headers_str}"])
        
        # Use shared utility for unit detection
        from src.utils.financial_domain import is_unit_indicator
        
        unique_entities = []
        seen_entities = set()
        for h in display_l2_row_headers:
            if h and h not in seen_entities and not is_unit_indicator(h):
                unique_entities.append(h)
                seen_entities.add(h)
        
        entities_str = ', '.join(unique_entities) if unique_entities else ''
        rows.append([f"{MetadataLabels.PRODUCT_ENTITY} {entities_str}"])
        
        # === COLUMN HEADERS ===
        detected_years = set()
        headers_info = self._detect_column_header_levels(table.get('content', ''))
        source_doc = metadata.get('source_doc', '')
        
        # Helper to filter out data values that aren't valid headers
        def is_valid_header(val: str) -> bool:
            """Check if value is a valid header (not a data value)."""
            if not val or not val.strip():
                return False
            val_clean = val.strip().replace(',', '').replace('.', '')
            # Year patterns are valid (2000-2039)
            if re.match(r'^20[0-3]\d$', val_clean):
                return True
            # Purely numeric (including negative) = data, not header
            if val_clean.lstrip('-').isdigit():
                # But allow 4-digit years
                if len(val_clean.lstrip('-')) != 4:
                    return False
            # Currency values = data
            if val.strip().startswith('$') or val.strip().startswith('('):
                return False
            return True
        
        # Get all header levels and filter out data values
        level_0 = self._dedupe_preserve_order([h for h in headers_info.get('level_0', []) if is_valid_header(h)])
        level_1_raw = self._dedupe_preserve_order([h for h in headers_info.get('level_1', []) if is_valid_header(h)])
        level_2_raw = self._dedupe_preserve_order([h for h in headers_info.get('level_2', []) if is_valid_header(h)])
        
        # Split compound headers in level_1 (e.g., "Average Monthly Balance Three Months Ended March 31,")
        # into L1 (main header) and L2 (period type) components
        # NOTE: Period normalization is deferred to Process step
        level_1 = []
        extracted_l0_from_split = []
        for header in level_1_raw:
            split_result = MetadataBuilder.split_compound_header(str(header))
            if split_result['l1']:
                # Found compound header - L1 is main header, add to level_0
                if split_result['l1'] not in extracted_l0_from_split:
                    extracted_l0_from_split.append(split_result['l1'])
                # L2 is period type, keep in level_1 (raw, Process step will normalize)
                if split_result['l2'] and split_result['l2'] not in level_1:
                    level_1.append(split_result['l2'])
            else:
                # No split needed - keep raw header (Process step will normalize)
                if header and header not in level_1:
                    level_1.append(header)
        
        # Merge extracted L0 headers with existing level_0
        # (extracted from splits come after existing level_0)
        if extracted_l0_from_split:
            for h in extracted_l0_from_split:
                if h not in level_0:
                    level_0.append(h)
        
        # For 10-K reports with year-only headers (no period type), convert to YTD format
        # Only convert if level_1 is empty (no "Three Months Ended" etc.)
        
        # === RECOMBINE SPLIT DATE PARTS IN level_2_raw ===
        # Sometimes dates like "December 31, 2024" are split into ["December 31", "2024"]
        # Recombine adjacent month+day with year items
        recombined_l2 = []
        i = 0
        while i < len(level_2_raw):
            item = str(level_2_raw[i]).strip()
            item_lower = item.lower()
            
            # Check if this item has a month+day pattern (e.g., "December 31")
            has_month_day = re.search(r'(january|february|march|april|may|june|july|august|september|october|november|december|jan|feb|mar|apr|jun|jul|aug|sept|sep|oct|nov|dec)\s+\d{1,2}', item_lower) and not re.search(r'20\d{2}', item)
            
            # Check if next item is a year
            if has_month_day and i + 1 < len(level_2_raw):
                next_item = str(level_2_raw[i + 1]).strip()
                if re.match(r'^20\d{2}$', next_item):
                    # Recombine: "December 31" + "2024" → "December 31, 2024"
                    combined_date = f"{item}, {next_item}"
                    recombined_l2.append(combined_date)
                    i += 2  # Skip next item since we combined it
                    continue
            
            recombined_l2.append(item)
            i += 1
        
        # Use recombined list for further processing
        level_2_raw = recombined_l2
        
        # === COMBINE PERIOD HEADERS WITH RAW DATES (normalization deferred to Process step) ===
        # This combines period text with dates but keeps raw - Process step will normalize
        # If level_1 has period text without years/months (e.g., "Three Months Ended")
        # and level_2_raw has full dates (e.g., "December 31, 2024"), combine them
        combined_l1 = []
        for l1_header in level_1:
            l1_str = str(l1_header).strip()
            # Check if L1 has a period pattern but no year
            has_period = any(p in l1_str.lower() for p in ['months ended', 'at ', 'as of '])
            has_year = bool(re.search(r'20\d{2}', l1_str))
            
            if has_period and not has_year and level_2_raw:
                # Combine with each value from level_2_RAW (keep raw, Process step will normalize)
                for l2_header in level_2_raw:
                    l2_str = str(l2_header).strip()
                    # Case 1: level_2 is year-only (e.g., "2024")
                    if re.match(r'^20\d{2}$', l2_str):
                        combined = f"{l1_str.rstrip(',')} {l2_str}"
                        if combined and combined not in combined_l1:
                            combined_l1.append(combined)
                    # Case 2: level_2 is full date (e.g., "December 31, 2024", "Sept 30, 2025")
                    elif re.search(r'(january|february|march|april|may|june|july|august|september|october|november|december|jan|feb|mar|apr|jun|jul|aug|sept|sep|oct|nov|dec)', l2_str.lower()) and re.search(r'20\d{2}', l2_str):
                        # Combine period type with full date: "Three Months Ended" + "December 31, 2024"
                        combined = f"{l1_str.rstrip(',')} {l2_str}"
                        if combined and combined not in combined_l1:
                            combined_l1.append(combined)
            else:
                # Keep as-is (Process step will normalize if needed)
                if l1_str and l1_str not in combined_l1:
                    combined_l1.append(l1_str)
        
        # Replace level_1 with combined version if we got results
        if combined_l1:
            level_1 = combined_l1
        
        # === HANDLE 3-LEVEL HEADERS: L2 has both period type AND dates ===
        # When level_2_raw contains both period types ("Three Months Ended") and dates ("December 31, 2024")
        # we need to combine them within level_2_raw itself
        # NOTE: Period normalization is deferred to Process step
        period_types_in_l2 = []
        dates_in_l2 = []
        for h in level_2_raw:
            h_str = str(h).strip()
            h_lower = h_str.lower()
            # Check if it's a period type (no year/month date components)
            is_period_type = any(p in h_lower for p in ['months ended', 'at ', 'as of ']) and not re.search(r'(january|february|march|april|may|june|july|august|september|october|november|december|jan|feb|mar|apr|jun|jul|aug|sept|sep|oct|nov|dec)\s+\d', h_lower)
            # Check if it's a full date (month + day + year)
            is_full_date = re.search(r'(january|february|march|april|may|june|july|august|september|october|november|december|jan|feb|mar|apr|jun|jul|aug|sept|sep|oct|nov|dec)\s+\d{1,2}.*20\d{2}', h_lower) is not None
            
            if is_period_type:
                period_types_in_l2.append(h_str)
            elif is_full_date:
                dates_in_l2.append(h_str)
        
        # If we have both period types and dates in L2, combine them (keep raw)
        combined_l2_periods = []
        if period_types_in_l2 and dates_in_l2:
            for period_type in period_types_in_l2:
                for date in dates_in_l2:
                    combined = f"{period_type.rstrip(',')} {date}"
                    if combined and combined not in combined_l2_periods:
                        combined_l2_periods.append(combined)
            
            # Update level_1 with these combined periods if we got results
            if combined_l2_periods:
                for cp in combined_l2_periods:
                    if cp not in level_1:
                        level_1.append(cp)
        
        # Build level_2 (keep raw, Process step will normalize)
        level_2 = []
        for h in level_2_raw:
            h_str = str(h).strip()
            # Skip period types that were already combined above
            if h_str in period_types_in_l2 and combined_l2_periods:
                continue
            # Skip dates that were already combined above  
            if h_str in dates_in_l2 and combined_l2_periods:
                continue
                
            if not level_1:  # No period type headers, just year-only
                converted = convert_year_to_q4_header(h_str, source_doc)
                level_2.append(converted)
            else:
                # Keep raw (Process step will normalize)
                level_2.append(h_str)
        
        # === COMBINE CATEGORY (L0) WITH NORMALIZED PERIODS (L1) ===
        # If level_0 has category labels (e.g., "Average Daily Balance") and level_1 has 
        # normalized periods (e.g., "Q4-QTD-2024"), combine them to produce
        # "Average Daily Balance Q4-QTD-2024"
        if level_0 and level_1:
            combined_headers = []
            for l0_header in level_0:
                l0_str = str(l0_header).strip()
                # Check if L0 is a category label (not a date pattern)
                is_category = not any(p in l0_str.lower() for p in 
                    ['months ended', 'at ', 'as of ', 'q1-', 'q2-', 'q3-', 'q4-', 'ytd-', 'qdt-'])
                
                if is_category and l0_str:
                    # Combine category with each normalized period
                    for l1_header in level_1:
                        l1_str = str(l1_header).strip()
                        # Check if L1 is a normalized period (Qn-YYYY or Qn-QTD-YYYY format)
                        if re.match(r'^Q[1-4]-(QTD|YTD)?-?20\d{2}$', l1_str):
                            combined = f"{l0_str} {l1_str}"
                            if combined not in combined_headers:
                                combined_headers.append(combined)
                        else:
                            # L1 is not normalized, keep as-is
                            if l1_str not in combined_headers:
                                combined_headers.append(l1_str)
            
            # Replace level_1 with combined headers if we got results
            if combined_headers:
                level_1 = combined_headers
                level_0 = []  # Clear L0 since it's now combined into L1
        
        # Extract years from all header levels
        for header in level_0 + level_1 + level_2:
            year_match = re.search(r'(20\d{2})', str(header))
            if year_match:
                detected_years.add(year_match.group(1))
        
        # Column Header L1 - Main Header (top spanning header, Level 0)
        level_0_str = ', '.join(level_0) if level_0 else ''
        rows.append([f"{MetadataLabels.COLUMN_HEADER_L1} {level_0_str}"])
        
        # Column Header L2 - Period Type date periods or main headers (Level 1)
        level_1_str = ', '.join(level_1) if level_1 else ''
        rows.append([f"{MetadataLabels.COLUMN_HEADER_L2} {level_1_str}"])
        
        # Column Header L3 - years or sub-headers (Level 2)
        level_2_str = ', '.join(level_2) if level_2 else ''
        rows.append([f"{MetadataLabels.COLUMN_HEADER_L3} {level_2_str}"])
        
        # Year/Quarter - format as PERIOD_TYPE,YEAR (e.g., QTD3,2024)
        detected_periods = []
        seen_periods = set()
        
        # First, try to get period type from L1 + L0 headers
        for header in level_1 + level_0:
            quarter_type = extract_quarter_from_header(str(header))
            year = extract_year_from_header(str(header))
            
            if quarter_type and year:
                period_key = f"{quarter_type},{year}"
                if period_key not in seen_periods:
                    detected_periods.append(period_key)
                    seen_periods.add(period_key)
            elif quarter_type:
                for y in sorted(detected_years, reverse=True):
                    period_key = f"{quarter_type},{y}"
                    if period_key not in seen_periods:
                        detected_periods.append(period_key)
                        seen_periods.add(period_key)
        
        # If no periods detected from L1/L0, try L2 headers for period type
        # This handles cases like "Average Monthly Balance Three Months Ended March 31"
        # where the L2 header contains the period type but years are in L3
        if not detected_periods and level_1_str:
            l2_quarter_type = extract_quarter_from_header(level_1_str)
            if l2_quarter_type:
                for y in sorted(detected_years, reverse=True):
                    period_key = f"{l2_quarter_type},{y}"
                    if period_key not in seen_periods:
                        detected_periods.append(period_key)
                        seen_periods.add(period_key)
        
        # Fallback: use metadata quarter/year if no headers detected
        if not detected_periods:
            quarter = metadata.get('quarter', '')
            year_from_meta = metadata.get('year', '')
            if quarter and year_from_meta:
                detected_periods.append(f"{quarter},{year_from_meta}")
            elif detected_years:
                # Last resort: use YTD prefix for standalone years from 10-K
                for y in sorted(detected_years, reverse=True):
                    detected_periods.append(f"YTD,{y}")
        
        periods_str = ', '.join(detected_periods) if detected_periods else ''
        rows.append([f"{MetadataLabels.YEAR_QUARTER} {periods_str}"])
        
        # Blank row
        rows.append([])
        
        # Table Title
        rows.append([f"{MetadataLabels.TABLE_TITLE} {display_title}"])
        
        # Source(s) with page number in format: source_pg#
        source_doc = metadata.get('source_doc', 'Unknown')
        page_no = metadata.get('page_no', 'N/A')
        # Format: 10q0625_pg5
        source_info = f"{MetadataLabels.SOURCES} {source_doc}_pg{page_no}"
        if metadata.get('year'):
            source_info += f", {metadata.get('year')}"
        if metadata.get('quarter'):
            source_info += f" {metadata.get('quarter')}"
        rows.append([source_info])
        
        # Empty row after Source for visual separation
        rows.append([])
        
        # === TABLE DATA WITH DYNAMIC HEADER TRACKING ===
        if not table_df.empty:
            # Column headers (first row of table - keep raw, Process step will normalize)
            def clean_header(c):
                if pd.isna(c):
                    return ''
                if isinstance(c, float) and c == int(c):
                    return str(int(c))
                return str(c)
            rows.append([clean_header(c) for c in table_df.columns])
            
            # Track header rows for mid-table replication
            header_stack = []  # Stack of up to MAX_HEADER_LEVELS header rows
            first_unit_row_seen = False  # Track if we've seen the first unit row
            
            # Data rows with L1 cleaning and dynamic header tracking
            for idx, row in table_df.iterrows():
                row_values = row.tolist()
                cleaned_row = []
                is_l1_row = False
                
                # Check if this is a unit row ($ in millions, etc.)
                current_is_unit_row = is_unit_row(row_values)
                
                # If this is a unit row and we've already seen one, insert cached headers
                if current_is_unit_row:
                    if first_unit_row_seen and header_stack:
                        # Insert blank line before repeated headers
                        rows.append([])
                        # Insert cached header rows (up to MAX_HEADER_LEVELS)
                        for header_row in header_stack:
                            rows.append(header_row)
                    first_unit_row_seen = True
                    # Clear header stack after unit row (we'll rebuild for next occurrence)
                    header_stack = []
                
                # Check if this is a header row (to cache for later)
                elif is_header_row(row_values) and first_unit_row_seen:
                    # This is a header row after data - cache it for next unit row
                    header_stack.append(row_values)
                    if len(header_stack) > MAX_HEADER_LEVELS:
                        header_stack.pop(0)  # Keep only last MAX_HEADER_LEVELS
                
                # Check if L1 row (section header within data)
                if len(row) > 1:
                    first_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
                    if first_val:
                        has_numeric = False
                        all_same = True
                        for v in row.iloc[1:]:
                            if pd.notna(v):
                                v_str = str(v).strip()
                                if v_str and v_str.lower() not in ['nan', '', '-', '—']:
                                    if v_str.lower() != first_val.lower():
                                        all_same = False
                                    if (v_str.startswith('$') or v_str.startswith('(') or 
                                        (v_str and v_str[0].isdigit()) or any(c.isdigit() for c in v_str)):
                                        has_numeric = True
                        is_l1_row = all_same or not has_numeric
                
                # Clean and add the row
                for i, v in enumerate(row):
                    if pd.isna(v):
                        cleaned_row.append('')
                    elif is_l1_row and i > 0:
                        first_val = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
                        v_str = str(v).strip()
                        if v_str.lower() == first_val.lower():
                            cleaned_row.append('')
                        else:
                            cleaned_row.append(ExcelUtils.clean_cell_value(v))
                    elif i == 0:
                        # First column (row label) - clean footnotes
                        cleaned_row.append(ExcelUtils.clean_footnote_references(str(v)))
                    else:
                        # Data cells - clean year floats
                        cleaned_row.append(ExcelUtils.clean_cell_value(v))
                rows.append(cleaned_row)
        
        return rows
    
    def _add_hyperlinks(
        self,
        output_path: Path,
        tables_by_title: Dict[str, List[Dict[str, Any]]]
    ) -> None:
        """
        Add hyperlinks to workbook (Index <-> Sheets).
        
        Sheet names are now Table_IDs (e.g., "1", "2", "3"), so linking is simpler.
        """
        try:
            from openpyxl.styles import Font
            
            wb = load_workbook(output_path)
            
            # Add hyperlinks in Index sheet
            index_sheet = wb['Index']
            
            # Find the Link column index dynamically
            header_row = [index_sheet.cell(row=1, column=c).value for c in range(1, index_sheet.max_column + 1)]
            link_col = None
            for i, h in enumerate(header_row, start=1):
                if isinstance(h, str) and h.strip().lower() == 'link':
                    link_col = i
                    break
            
            if link_col is None:
                logger.warning("Link column not found in Index sheet")
                link_col = -1
            
            for row in range(2, index_sheet.max_row + 1):  # Skip header
                if link_col == -1:
                    break
                cell = index_sheet.cell(row=row, column=link_col)
                raw_value = cell.value
                
                # Link value is Table_ID (e.g., "1", "2", "3")
                table_id = None
                if raw_value is not None:
                    try:
                        raw_str = str(raw_value).strip()
                    except Exception:
                        raw_str = ''
                    
                    if raw_str.startswith('→'):
                        table_id = raw_str.lstrip('→').strip()
                    else:
                        table_id = raw_str
                
                if table_id:
                    # Sheet name is simply the Table_ID
                    sheet_name = table_id
                    
                    if sheet_name in wb.sheetnames:
                        cell.hyperlink = f"#'{sheet_name}'!A1"
                        cell.value = f"→ {sheet_name}"
                        cell.font = Font(color="0000FF", underline="single")
                    else:
                        logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            
            # Add hyperlinks in TOC sheet (Sheet column)
            if 'TOC' in wb.sheetnames:
                toc_sheet = wb['TOC']
                
                # Find the Sheet column index dynamically
                toc_header = [toc_sheet.cell(row=1, column=c).value for c in range(1, toc_sheet.max_column + 1)]
                sheet_col = None
                for i, h in enumerate(toc_header, start=1):
                    if isinstance(h, str) and h.strip().lower() == 'sheet':
                        sheet_col = i
                        break
                
                if sheet_col:
                    for row in range(2, toc_sheet.max_row + 1):  # Skip header
                        cell = toc_sheet.cell(row=row, column=sheet_col)
                        raw_value = cell.value
                        
                        table_id = None
                        if raw_value is not None:
                            try:
                                raw_str = str(raw_value).strip()
                            except Exception:
                                raw_str = ''
                            table_id = raw_str
                        
                        if table_id:
                            sheet_name = table_id
                            if sheet_name in wb.sheetnames:
                                cell.hyperlink = f"#'{sheet_name}'!A1"
                                cell.value = f"→ {sheet_name}"
                                cell.font = Font(color="0000FF", underline="single")
            
            # Add back-links in each table sheet
            for sheet_name in tables_by_title.keys():
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    cell = sheet.cell(row=1, column=1)
                    cell.hyperlink = "#'Index'!A1"
                    cell.value = "← Back to Index"
                    cell.font = Font(color="0000FF", underline="single")
            
            wb.save(output_path)
            
        except Exception as e:
            logger.error(f"Failed to add hyperlinks: {e}")
    
    def _merge_repeated_header_cells(
        self,
        output_path: Path,
        tables_by_title: Dict[str, List[Dict[str, Any]]]
    ) -> None:
        """
        Merge adjacent cells with same values in header rows.
        
        This makes Excel output match PDF layout where headers span multiple columns.
        For example: "Three Months Ended" | "Three Months Ended" → merged cell
        """
        try:
            from openpyxl.styles import Alignment
            
            wb = load_workbook(output_path)
            
            for sheet_name in tables_by_title.keys():
                if sheet_name not in wb.sheetnames:
                    continue
                    
                ws = wb[sheet_name]
                
                # Find ALL rows that start with "Source:" and merge their header rows
                # (Multiple tables with same title can be stacked on one sheet)
                for row_num in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row_num, column=1).value
                    if cell_value and MetadataLabels.is_sources(str(cell_value)):
                        # Skip the blank row after Source: - headers start 2 rows after
                        # Sheet structure: Source (row N) -> blank (N+1) -> headers (N+2, N+3, N+4...)
                        # Tables can have up to 3 levels of column headers that need merging
                        for header_offset in range(2, 5):  # Check rows +2, +3, +4
                            header_row = row_num + header_offset
                            if header_row <= ws.max_row:
                                # Only merge if this row looks like a header (has repeated values)
                                self._merge_row_cells(ws, header_row)
                        # Don't break - continue to find more table blocks
            
            wb.save(output_path)
            
        except Exception as e:
            logger.warning(f"Could not merge header cells: {e}")
    
    def _merge_row_cells(self, ws, row_num: int) -> None:
        """Merge adjacent cells with same non-empty values in a row."""
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter
        
        max_col = ws.max_column
        if max_col < 2:
            return
        
        col = 2  # Start from column B (skip column A which is row labels)
        while col <= max_col:
            cell_value = ws.cell(row=row_num, column=col).value
            if not cell_value:
                col += 1
                continue
            
            # Find how many adjacent cells have the same value
            end_col = col
            while end_col < max_col:
                next_value = ws.cell(row=row_num, column=end_col + 1).value
                if next_value and str(next_value).strip() == str(cell_value).strip():
                    end_col += 1
                else:
                    break
            
            # If we found repeated cells, merge them
            if end_col > col:
                start_letter = get_column_letter(col)
                end_letter = get_column_letter(end_col)
                merge_range = f"{start_letter}{row_num}:{end_letter}{row_num}"
                try:
                    ws.merge_cells(merge_range)
                    # Center the merged cell
                    ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center')
                except Exception:
                    pass  # Cell might already be merged
            
            col = end_col + 1

    def merge_processed_files(
        self,
        output_filename: str = "consolidated_tables.xlsx",
        transpose: bool = False
    ) -> dict:
        """
        Merge all processed xlsx files into a single consolidated report.
        
        NOTE: This method delegates to ConsolidatedExcelExporter for cleaner separation.
        For direct access to consolidation features, use:
            from src.infrastructure.extraction.consolidation.consolidated_exporter import get_consolidated_exporter
        
        Args:
            output_filename: Output filename for consolidated report
            transpose: If True, dates become rows (time-series format)
            
        Returns:
            Dict with path, tables_merged, sources_merged, sheet_names
        """
        from src.infrastructure.extraction.consolidation.consolidated_exporter import get_consolidated_exporter
        
        consolidated_exporter = get_consolidated_exporter()
        return consolidated_exporter.merge_processed_files(
            output_filename=output_filename,
            transpose=transpose
        )


# =============================================================================
# SINGLETON PATTERN
# =============================================================================

_excel_exporter: Optional[ExcelTableExporter] = None


def get_excel_exporter() -> ExcelTableExporter:
    """
    Get or create global Excel exporter instance.
    
    Returns:
        ExcelTableExporter singleton instance
    """
    global _excel_exporter
    
    if _excel_exporter is None:
        _excel_exporter = ExcelTableExporter()
    
    return _excel_exporter


def reset_excel_exporter() -> None:
    """Reset the exporter singleton (for testing)."""
    global _excel_exporter
    _excel_exporter = None
