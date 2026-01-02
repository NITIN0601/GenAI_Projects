"""
Block Detection - Find and analyze table blocks in Excel worksheets.

Standalone module for detecting table blocks within Excel sheets.
Used by: table_merger.py
"""

from typing import List, Dict, Any, Set
from openpyxl.worksheet.worksheet import Worksheet

from src.utils import get_logger
from src.utils.metadata_labels import MetadataLabels
from src.utils.financial_domain import (
    TABLE_HEADER_PATTERNS,
    METADATA_BOUNDARY_MARKERS,
    is_new_table_header_row,
)
from src.utils.constants import (
    TABLE_MERGER_MAX_COL_SCAN as MAX_COL_SCAN,
    TABLE_MERGER_MIN_YEAR_COUNT_FOR_HEADER as MIN_YEAR_COUNT_FOR_HEADER,
    YEAR_STRING_LENGTH,
)

logger = get_logger(__name__)


class BlockDetector:
    """
    Detect and analyze table blocks in Excel worksheets.
    
    Handles:
    - Finding table blocks (Source row to next metadata section)
    - Splitting blocks on mid-table headers
    - Identifying header vs data rows
    - Extracting row labels
    """
    
    @classmethod
    def find_table_blocks(cls, ws, extract_labels_func=None) -> List[Dict[str, Any]]:
        """
        Find all table blocks in a worksheet.
        
        A table block:
        - Starts after a "Source:" row
        - Ends before the next metadata section (Row Header) or end of content
        
        Args:
            ws: openpyxl Worksheet
            extract_labels_func: Function to extract row labels (for dependency injection)
        
        Returns:
            List of table block dicts with start/end rows and row labels
        """
        blocks = []
        
        # First pass: find all Source: rows (they mark the end of metadata, start of table data)
        source_rows = []
        metadata_rows = []  # Track "Row Header" rows that start new metadata sections
        
        for row_num in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_num, column=1).value
            if cell_value is None:
                continue
            
            cell_str = str(cell_value).strip()
            
            if MetadataLabels.is_sources(cell_str):
                source_rows.append(row_num)
            # Detect metadata boundary markers using centralized patterns
            elif any(cell_str.lower().startswith(marker) for marker in METADATA_BOUNDARY_MARKERS):
                metadata_rows.append(row_num)
        
        if not source_rows:
            return blocks
        
        # For each Source: row, find the table data range
        for i, source_row in enumerate(source_rows):
            # Table data starts after Source: row (skip completely empty rows)
            data_start = source_row + 1
            while data_start <= ws.max_row:
                # Check all columns for data (not just column 1)
                row_has_data = False
                for col in range(1, ws.max_column + 1):
                    cell_val = ws.cell(row=data_start, column=col).value
                    if cell_val is not None and str(cell_val).strip():
                        row_has_data = True
                        break
                if row_has_data:
                    break
                data_start += 1
            
            if data_start > ws.max_row:
                continue
            
            # Table data ends before next metadata section or next Source/end of sheet
            data_end = ws.max_row
            
            # Find the next metadata section that comes after this source row
            for meta_row in metadata_rows:
                if meta_row > source_row:
                    data_end = meta_row - 1
                    break
            
            # Skip empty rows at the end
            while data_end > data_start:
                has_data = False
                for col in range(1, min(MAX_COL_SCAN, ws.max_column + 1)):
                    if ws.cell(row=data_end, column=col).value is not None:
                        has_data = True
                        break
                if has_data:
                    break
                data_end -= 1
            
            if data_end <= data_start:
                continue
            
            # Find the metadata_start_row for this table block
            prev_source_row = source_rows[i - 1] if i > 0 else 0
            
            # Get all metadata rows that belong to this table
            table_metadata_rows = [m for m in metadata_rows if prev_source_row < m < source_row]
            
            if table_metadata_rows:
                metadata_start = min(table_metadata_rows)
            else:
                metadata_start = source_row
            
            block = {
                'metadata_start_row': metadata_start,
                'source_row': source_row,
                'start_row': data_start,
                'end_row': data_end,
                'row_labels': [],
                'header_row': data_start,
                'data_start_row': data_start
            }
            
            # Detect header rows and data rows
            cls.identify_header_and_data_rows(ws, block)
            
            # Extract row labels if function provided
            if extract_labels_func:
                block['row_labels'] = extract_labels_func(ws, block)
                
                if block['row_labels']:
                    # Check for mid-table column headers that indicate a split is needed
                    split_blocks = cls.split_block_on_new_headers(ws, block, extract_labels_func)
                    blocks.extend(split_blocks)
            else:
                blocks.append(block)
        
        return blocks
    
    @classmethod
    def split_block_on_new_headers(cls, ws: Worksheet, block: Dict, extract_labels_func=None) -> List[Dict]:
        """
        Check if a table block contains new column headers mid-table.
        
        If a row has empty first column but contains date/period patterns
        in other columns, it indicates a new sub-table and the block should be split.
        
        Args:
            ws: Worksheet
            block: Table block dict
            extract_labels_func: Function to extract row labels
            
        Returns:
            List of blocks (original if no split needed, or split blocks)
        """
        data_start = block.get('data_start_row', block['start_row'])
        data_end = block['end_row']
        
        # Scan for new header rows (after at least one data row)
        split_points = []
        seen_data_row = False
        
        for row_num in range(data_start, data_end + 1):
            first_col = ws.cell(row=row_num, column=1).value
            
            # Get all values in the row
            row_values = []
            for col in range(1, min(ws.max_column + 1, 15)):
                row_values.append(ws.cell(row=row_num, column=col).value)
            
            # Check if first column has data (this is a data row)
            first_val = str(first_col).strip() if first_col else ''
            if first_val and first_val.lower() not in ['', 'nan', 'none']:
                seen_data_row = True
                continue
            
            # If we've seen data rows and this row is a new header, mark split point
            if seen_data_row and is_new_table_header_row(row_values, first_col):
                split_points.append(row_num)
        
        # If no splits needed, return original block
        if not split_points:
            return [block]
        
        # Create split blocks
        # IMPORTANT: Only the FIRST split block should inherit the original metadata.
        # Subsequent split blocks should have metadata_start_row = start_row to prevent
        # clearing overlap (which would destroy the first block's data).
        result_blocks = []
        current_start = block['start_row']
        is_first_split = True
        
        for split_row in split_points:
            # Create block ending before the split point
            if split_row > current_start:
                new_block = block.copy()
                new_block['start_row'] = current_start
                new_block['end_row'] = split_row - 1
                new_block['data_start_row'] = current_start
                
                # Only first split block keeps original metadata
                # Subsequent blocks have no metadata to copy (they share with first)
                if not is_first_split:
                    new_block['metadata_start_row'] = current_start
                    new_block['source_row'] = current_start
                    new_block['_is_sub_block'] = True  # Mark as sub-block
                
                if extract_labels_func:
                    new_block['row_labels'] = extract_labels_func(ws, new_block)
                if new_block.get('row_labels'):
                    result_blocks.append(new_block)
                    is_first_split = False
            
            # Next block starts at the split point
            current_start = split_row
        
        # Create final block from last split to end
        if current_start <= data_end:
            new_block = block.copy()
            new_block['start_row'] = current_start
            new_block['end_row'] = data_end
            new_block['data_start_row'] = current_start
            
            # Not first block, so set metadata to own range
            if not is_first_split:
                new_block['metadata_start_row'] = current_start
                new_block['source_row'] = current_start
                new_block['_is_sub_block'] = True
            
            if extract_labels_func:
                new_block['row_labels'] = extract_labels_func(ws, new_block)
            if new_block.get('row_labels'):
                result_blocks.append(new_block)
        
        if result_blocks:
            logger.debug(f"Split table block into {len(result_blocks)} sub-tables at rows {split_points}")
        
        return result_blocks if result_blocks else [block]
    
    @classmethod
    def identify_header_and_data_rows(cls, ws: Worksheet, block: Dict) -> None:
        """
        Identify which rows are headers vs data in a table block.
        Updates block['data_start_row'] to point to first data row.
        
        Header rows typically:
        - Have empty first column but values in other columns (column headers)
        - First column contains currency units like "$ in millions"
        - First column is a year (4 digits)
        - Non-first columns contain year values
        
        Data rows have:
        - Non-empty first column with descriptive text (row labels)
        """
        header_patterns = TABLE_HEADER_PATTERNS
        
        data_start = block['start_row']
        header_rows_found = []
        
        # Only scan first few rows for headers
        max_header_rows = min(4, block['end_row'] - block['start_row'] + 1)
        
        for row_num in range(block['start_row'], block['start_row'] + max_header_rows):
            if row_num > block['end_row']:
                break
                
            row_values = []
            first_col_value = None
            
            for col in range(1, min(MAX_COL_SCAN, ws.max_column + 1)):
                cell_val = ws.cell(row=row_num, column=col).value
                if cell_val:
                    row_values.append(str(cell_val).strip())
                    if col == 1:
                        first_col_value = str(cell_val).strip().lower()
            
            is_header_row = False
            
            # Empty first column but has values in other columns = header row
            if not first_col_value and len(row_values) > 0:
                is_header_row = True
            
            # First column has explicit header pattern
            elif first_col_value:
                if any(pattern in first_col_value for pattern in header_patterns):
                    is_header_row = True
                # First column is a year (4 digits)
                elif first_col_value.isdigit() and len(first_col_value) == YEAR_STRING_LENGTH:
                    is_header_row = True
                else:
                    # This is the start of data rows
                    data_start = row_num
                    break
            
            # Check if non-first columns have multiple year values
            if not is_header_row:
                year_count = sum(1 for v in row_values if v.isdigit() and len(v) == YEAR_STRING_LENGTH)
                if year_count >= MIN_YEAR_COUNT_FOR_HEADER:
                    is_header_row = True
            
            if is_header_row:
                header_rows_found.append(row_num)
                data_start = row_num + 1
        
        block['header_rows'] = header_rows_found
        block['data_start_row'] = data_start
    
    @classmethod
    def extract_row_labels(cls, ws: Worksheet, block: Dict) -> List[str]:
        """
        Extract row labels (Column A values) from a table block's data rows.
        
        Returns:
            List of row labels (normalized for comparison)
        """
        labels = []
        data_start = block.get('data_start_row', block['start_row'])
        
        for row_num in range(data_start, block['end_row'] + 1):
            cell_value = ws.cell(row=row_num, column=1).value
            if cell_value is not None:
                label = str(cell_value).strip().lower()
                if label and label not in ['nan', 'none', '']:
                    labels.append(label)
        
        return labels
