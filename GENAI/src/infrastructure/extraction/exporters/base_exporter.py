"""
Base Excel Exporter with shared functionality.

Both ExcelTableExporter and ConsolidatedExcelExporter inherit from this class
to avoid code duplication while maintaining separate specialized logic.

Import Order (no circular imports):
    constants.py → excel_utils.py → base_exporter.py → excel_exporter.py
                                                      → consolidated_exporter.py
"""

from abc import ABC
from pathlib import Path
from typing import Dict, Any, List, Optional, Union

import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.utils import get_logger
from src.utils.excel_utils import ExcelUtils
from src.utils.constants import (
    EXCEL_MAX_SHEET_NAME_LENGTH,
    EXCEL_MAX_DISPLAY_ITEMS,
)

logger = get_logger(__name__)


class BaseExcelExporter(ABC):
    """
    Base class with shared Excel export functionality.
    
    Provides common methods used by both ExcelTableExporter and ConsolidatedExcelExporter:
    - Utility methods (delegates to ExcelUtils)
    - Cell formatting (currency, headers, links)
    - Index sheet helpers
    - Back-to-index link creation
    
    This class should NOT be instantiated directly - use concrete subclasses.
    """
    
    # =========================================================================
    # STYLING CONSTANTS
    # =========================================================================
    
    # Header styling
    HEADER_FONT = Font(bold=True)
    HEADER_FILL = PatternFill("solid", fgColor="4472C4")  # Blue
    HEADER_FONT_WHITE = Font(bold=True, color="FFFFFF")
    
    # Link styling
    LINK_FONT = Font(color="0563C1", underline="single")
    
    # Number formats
    CURRENCY_FORMAT = '#,##0.00'
    PERCENT_FORMAT = '0.0%'
    INTEGER_FORMAT = '#,##0'
    
    # Border styling
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self):
        """Initialize base exporter."""
        pass

    # =========================================================================
    # UTILITY METHODS (Delegate to ExcelUtils for centralized logic)
    # =========================================================================
    
    def _get_column_letter(self, idx: int) -> str:
        """
        Convert column index to Excel column letter.
        
        Args:
            idx: 0-based column index (0=A, 25=Z, 26=AA)
            
        Returns:
            Excel column letter(s)
        """
        return ExcelUtils.get_column_letter(idx)
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """
        Sanitize string for Excel sheet name.
        
        Removes invalid characters and truncates to max length.
        
        Args:
            name: Raw sheet name
            
        Returns:
            Sanitized sheet name (max 31 chars)
        """
        return ExcelUtils.sanitize_sheet_name(name, EXCEL_MAX_SHEET_NAME_LENGTH)
    
    def _normalize_title_for_grouping(self, title: str) -> str:
        """
        Normalize title for case-insensitive grouping.
        
        Args:
            title: Raw table title
            
        Returns:
            Normalized lowercase title
        """
        return ExcelUtils.normalize_title_for_grouping(title)
    
    def _normalize_row_label(self, label: str) -> str:
        """
        Normalize row label for matching during consolidation.
        
        Removes footnotes, superscripts, and extra whitespace.
        
        Args:
            label: Raw row label
            
        Returns:
            Normalized lowercase label
        """
        return ExcelUtils.normalize_row_label(label)
    
    def _clean_currency_value(self, val) -> Union[float, str]:
        """
        Convert currency string to float, keep special values as string.
        
        Args:
            val: Cell value (string, float, or NaN)
            
        Returns:
            Float for numeric values, string for special values
        """
        return ExcelUtils.clean_currency_value(val)
    
    def _detect_report_type(self, source: str) -> str:
        """
        Detect report type from source filename.
        
        Args:
            source: Source filename (e.g., "10k1224.pdf")
            
        Returns:
            Report type: "10-K", "10-Q", "8-K", or "Unknown"
        """
        return ExcelUtils.detect_report_type(source)
    
    def _clean_year_string(self, val) -> str:
        """
        Remove .0 suffix from year values.
        
        Args:
            val: Value that might be a year
            
        Returns:
            Cleaned string representation
        """
        return ExcelUtils.clean_year_string(val)
    
    def _ensure_string_header(self, val) -> str:
        """
        Ensure header value is a clean string.
        
        Args:
            val: Header value (could be int, float, string, or None)
            
        Returns:
            Clean string suitable for Excel header
        """
        return ExcelUtils.ensure_string_header(val)

    # =========================================================================
    # SHEET CREATION HELPERS
    # =========================================================================
    
    def _add_back_link(
        self, 
        ws: Worksheet, 
        row: int = 1, 
        col: int = 1, 
        text: str = "← Back to Index"
    ) -> None:
        """
        Add a hyperlink back to Index sheet.
        
        Args:
            ws: Worksheet to add link to
            row: Row number for the link
            col: Column number for the link
            text: Display text for the link
        """
        cell = ws.cell(row=row, column=col)
        cell.value = text
        cell.hyperlink = "#Index!A1"
        cell.font = self.LINK_FONT
    
    def _apply_header_style(
        self, 
        ws: Worksheet, 
        row_num: int, 
        col_start: int, 
        col_end: int
    ) -> None:
        """
        Apply header styling to a row range.
        
        Args:
            ws: Worksheet
            row_num: Row to style
            col_start: First column (1-indexed)
            col_end: Last column (1-indexed)
        """
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = self.HEADER_FONT_WHITE
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
    
    def _apply_metadata_label_style(
        self,
        ws: Worksheet,
        row_num: int,
        col: int = 1
    ) -> None:
        """
        Apply styling to metadata label cell (Column A).
        
        Args:
            ws: Worksheet
            row_num: Row number
            col: Column number (default 1)
        """
        cell = ws.cell(row=row_num, column=col)
        cell.font = self.HEADER_FONT
    
    def _is_percentage_row(self, row_values: List) -> bool:
        """
        Determine if a row contains percentage values (should format as %).
        
        A row is considered percentage if all numeric values are between 0 and 1
        (exclusive of boundary to avoid edge cases).
        
        Args:
            row_values: List of cell values in the row
            
        Returns:
            True if row should be formatted as percentage
        """
        numeric_values = []
        for val in row_values:
            if isinstance(val, (int, float)) and not pd.isna(val):
                numeric_values.append(abs(val))
        
        if not numeric_values:
            return False
        
        # If all values are between 0 and 1 (but not all zeros), it's likely percentage
        # Also check that no value is > 1 (which would indicate currency)
        all_small = all(0 <= v <= 1 for v in numeric_values)
        has_non_zero = any(v > 0 for v in numeric_values)
        
        return all_small and has_non_zero
    
    def _apply_number_format(
        self, 
        ws: Worksheet, 
        start_row: int, 
        end_row: int, 
        col_start: int, 
        col_end: int
    ) -> None:
        """
        Apply currency or percentage formatting to data cells based on values.
        
        Uses row-based detection: if all numeric values in a row are 0-1,
        format as percentage; otherwise format as currency.
        
        Args:
            ws: Worksheet
            start_row: First data row
            end_row: Last data row
            col_start: First data column
            col_end: Last data column
        """
        for row_num in range(start_row, end_row + 1):
            # Collect row values
            row_values = []
            for col in range(col_start, col_end + 1):
                cell = ws.cell(row=row_num, column=col)
                if cell.value is not None:
                    row_values.append(cell.value)
            
            # Determine format for this row
            use_percent = self._is_percentage_row(row_values)
            number_format = self.PERCENT_FORMAT if use_percent else self.CURRENCY_FORMAT
            
            # Apply format to numeric cells
            for col in range(col_start, col_end + 1):
                cell = ws.cell(row=row_num, column=col)
                if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                    cell.number_format = number_format
    
    def _format_sources_list(
        self, 
        sources: List[str], 
        max_display: int = None
    ) -> str:
        """
        Format a list of sources with truncation for display.
        
        Args:
            sources: List of source strings
            max_display: Maximum items to show (defaults to EXCEL_MAX_DISPLAY_ITEMS)
            
        Returns:
            Formatted string like "source1, source2... (+3)"
        """
        if max_display is None:
            max_display = EXCEL_MAX_DISPLAY_ITEMS
        
        if not sources:
            return ""
        
        result = ', '.join(sources[:max_display])
        if len(sources) > max_display:
            result += f'... (+{len(sources) - max_display})'
        return result
    
    # =========================================================================
    # HYPERLINK HELPERS
    # =========================================================================
    
    def _create_hyperlink(self, sheet_name: str, cell: str = "A1") -> str:
        """
        Create internal hyperlink string to a sheet.
        
        Args:
            sheet_name: Target sheet name
            cell: Target cell reference (default "A1")
            
        Returns:
            Excel hyperlink string like "#'Sheet Name'!A1"
        """
        # Excel requires sheet names with spaces or special chars to be quoted
        if ' ' in sheet_name or any(c in sheet_name for c in "[]'"):
            # Escape single quotes by doubling them
            escaped_name = sheet_name.replace("'", "''")
            return f"#'{escaped_name}'!{cell}"
        return f"#{sheet_name}!{cell}"
    
    def _add_hyperlink_cell(
        self,
        ws: Worksheet,
        row: int,
        col: int,
        display_text: str,
        target_sheet: str,
        target_cell: str = "A1"
    ) -> None:
        """
        Add a cell with hyperlink to another sheet.
        
        Args:
            ws: Worksheet
            row: Row number
            col: Column number
            display_text: Text to display in the cell
            target_sheet: Sheet name to link to
            target_cell: Cell in target sheet (default "A1")
        """
        cell = ws.cell(row=row, column=col)
        cell.value = display_text
        cell.hyperlink = self._create_hyperlink(target_sheet, target_cell)
        cell.font = self.LINK_FONT

    # =========================================================================
    # DATA VALIDATION HELPERS
    # =========================================================================
    
    def _has_data_values(self, df: pd.DataFrame, start_row: int = 0) -> bool:
        """
        Check if DataFrame has any actual data values (not just headers).
        
        Args:
            df: DataFrame to check
            start_row: Row to start checking from
            
        Returns:
            True if DataFrame has data values
        """
        if df.empty:
            return False
        
        data_df = df.iloc[start_row:] if start_row > 0 else df
        
        # Check if any cell has a non-null, non-empty value
        for col in data_df.columns:
            for val in data_df[col]:
                if val is not None and str(val).strip():
                    # Skip if it's just a header-like value
                    val_str = str(val).strip().lower()
                    if val_str not in ['nan', 'none', '']:
                        return True
        
        return False
