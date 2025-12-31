"""
Index Manager - Manage Index sheet updates for table splits.

Standalone module for managing Excel Index sheet entries.
Used by: table_merger.py
"""

from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.utils import get_logger
from src.utils.metadata_labels import MetadataLabels

logger = get_logger(__name__)


class IndexManager:
    """
    Manage Index sheet entries for split sheets.
    
    Handles:
    - Adding entries for newly split sheets
    - Updating Index when sheets are split
    - Creating hyperlinks to split sheets
    """
    
    @classmethod
    def update_index_for_split_sheets(cls, wb: Workbook, new_sheets: List[str]) -> None:
        """
        Add entries to Index for newly split sheets.
        
        Args:
            wb: openpyxl Workbook
            new_sheets: List of new sheet names
        """
        if 'Index' not in wb.sheetnames or not new_sheets:
            return
        
        index_ws = wb['Index']
        
        # Find the last row with data in Index
        next_row = 1
        for row_num in range(1, index_ws.max_row + 1):
            if index_ws.cell(row=row_num, column=1).value is not None:
                next_row = row_num + 1
        
        for new_sheet_name in new_sheets:
            new_ws = wb[new_sheet_name]
            
            # Extract table title from the new sheet
            title = ''
            for row_num in range(1, 15):
                cell_val = new_ws.cell(row=row_num, column=1).value
                if cell_val and MetadataLabels.TABLE_TITLE in str(cell_val):
                    title = str(cell_val).replace(MetadataLabels.TABLE_TITLE, '').strip()
                    break
            
            if not title:
                title = f"Split Table ({new_sheet_name})"
            
            # Add to Index (column A = sheet name/number, column B = title)
            index_ws.cell(row=next_row, column=1).value = new_sheet_name
            index_ws.cell(row=next_row, column=2).value = title
            
            logger.debug(f"Added '{new_sheet_name}' to Index: {title}")
            next_row += 1
    
    @classmethod
    def update_index_for_splits(cls, wb: Workbook, original_sheet_name: str, split_sheets: List[str]) -> None:
        """
        Update Index when a sheet is split: update original row and insert new rows.
        
        Index structure: Source(A), PageNo(B), Table_ID(C), Location_ID(D), Section(E), Table Title(F), Link(G)
        
        When sheet "9" splits into ["9_1", "9_2"]:
        - Update the original row for "9" -> "9_1" with (Part 1)
        - Insert a new row after it for "9_2" with (Part 2)
        
        Args:
            wb: openpyxl Workbook
            original_sheet_name: Original sheet name (e.g., '9')
            split_sheets: List of split sheet names (e.g., ['9_1', '9_2'])
        """
        if 'Index' not in wb.sheetnames or not split_sheets:
            return
        
        index_ws = wb['Index']
        
        # Find the first row that matches the original sheet name in Table_ID column (C)
        original_row = None
        for row_num in range(2, index_ws.max_row + 1):  # Skip header row 1
            table_id = index_ws.cell(row=row_num, column=3).value  # Column C = Table_ID
            if table_id is not None and str(table_id).strip() == original_sheet_name:
                original_row = row_num
                break
        
        if not original_row:
            logger.debug(f"No Index entry found for sheet '{original_sheet_name}'")
            return
        
        # Get original row data for copying to new rows
        original_data = {
            'source': index_ws.cell(row=original_row, column=1).value,  # Source
            'page_no': index_ws.cell(row=original_row, column=2).value,  # PageNo
            'location_id': index_ws.cell(row=original_row, column=4).value,  # Location_ID
            'section': index_ws.cell(row=original_row, column=5).value,  # Section
            'title': index_ws.cell(row=original_row, column=6).value,  # Table Title
        }
        
        # Step 1: Update the original row with first split sheet (9_1)
        first_split = split_sheets[0]
        title_base = original_data['title'] or 'Split Table'
        
        index_ws.cell(row=original_row, column=3).value = first_split  # Table_ID
        index_ws.cell(row=original_row, column=6).value = f"{title_base} (Part 1)"  # Title
        link_cell = index_ws.cell(row=original_row, column=7)
        link_cell.value = f"→ {first_split}"
        # Create hyperlink to the split sheet
        from openpyxl.worksheet.hyperlink import Hyperlink
        link_cell.hyperlink = Hyperlink(ref=link_cell.coordinate, target=f"#'{first_split}'!A1")
        logger.debug(f"Updated Index row {original_row}: {first_split} (Part 1)")
        
        # Step 2: Insert new rows for additional split sheets (9_2, 9_3, etc.)
        if len(split_sheets) > 1:
            insert_position = original_row + 1
            
            for part_num, split_sheet in enumerate(split_sheets[1:], start=2):
                # Insert a new row at the position
                index_ws.insert_rows(insert_position)
                
                # Copy data from original row
                index_ws.cell(row=insert_position, column=1).value = original_data['source']
                index_ws.cell(row=insert_position, column=2).value = original_data['page_no']
                index_ws.cell(row=insert_position, column=3).value = split_sheet  # Table_ID
                index_ws.cell(row=insert_position, column=4).value = f"{original_sheet_name}_{part_num}"  # Location_ID
                index_ws.cell(row=insert_position, column=5).value = original_data['section']
                index_ws.cell(row=insert_position, column=6).value = f"{title_base} (Part {part_num})"
                link_cell = index_ws.cell(row=insert_position, column=7)
                link_cell.value = f"→ {split_sheet}"
                # Create hyperlink to the split sheet
                from openpyxl.worksheet.hyperlink import Hyperlink
                from openpyxl.styles import Font
                link_cell.hyperlink = Hyperlink(ref=link_cell.coordinate, target=f"#'{split_sheet}'!A1")
                # Apply blue hyperlink styling
                link_cell.font = Font(color='000000FF', underline='single')
                
                logger.debug(f"Inserted Index row {insert_position}: {split_sheet} (Part {part_num})")
                insert_position += 1
        
        # Step 3: Delete any remaining rows that still have old Table_ID
        rows_to_delete = []
        for row_num in range(2, index_ws.max_row + 1):
            table_id = index_ws.cell(row=row_num, column=3).value
            if table_id is not None and str(table_id).strip() == original_sheet_name:
                rows_to_delete.append(row_num)
        
        # Delete rows in reverse order to maintain correct indices
        for row_num in reversed(rows_to_delete):
            index_ws.delete_rows(row_num, 1)
            logger.debug(f"Deleted duplicate Index row {row_num}: Table_ID was '{original_sheet_name}'")
