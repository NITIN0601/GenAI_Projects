"""
Excel Table Exporter with Index Sheet.

Exports extracted tables to multi-sheet Excel with:
- Index sheet with hyperlinks to each table sheet
- One sheet per unique table title
- Multiple tables with same title stacked with blank rows
- Back-links from each sheet to Index

Usage:
    from src.infrastructure.extraction.formatters.excel_exporter import get_excel_exporter
    
    exporter = get_excel_exporter()
    
    # Per-PDF export
    exporter.export_pdf_tables(tables, "10q0625.pdf")
    # -> data/processed/10q0625_tables.xlsx
    
    # Combined export (after all PDFs processed)
    exporter.export_combined_tables(all_tables)
    # -> data/extracted/all_tables_combined.xlsx
    
    # For consolidated multi-file merge, use:
    from src.infrastructure.extraction.formatters.consolidated_exporter import get_consolidated_exporter
"""

__version__ = "2.1.0"
__author__ = "dundaymo"

import pandas as pd
from typing import List, Dict, Any, Optional
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import re

from src.utils import get_logger
from src.core import get_paths
from src.infrastructure.extraction.formatters.header_detector import HeaderDetector
from src.infrastructure.extraction.formatters.excel_utils import ExcelUtils

logger = get_logger(__name__)


class ExcelTableExporter:
    """
    Export extracted tables to multi-sheet Excel with Index.
    
    Features:
    - Index sheet with Source, PageNo, Table_ID, Location_ID, Table Title, Link
    - One sheet per unique table title (using Table_ID as sheet name)
    - Detailed header structure: Row Headers, Column Headers, Product/Entity
    - Same-title tables stacked with blank separator
    - Bidirectional hyperlinks (Index <-> Sheets)
    
    Follows singleton pattern consistent with other managers.
    """
    
    def __init__(self):
        """Initialize exporter with paths from PathManager."""
        self.paths = get_paths()
        self.processed_dir = self.paths.data_dir / "processed"
        self.extracted_dir = self.paths.data_dir / "extracted"
        
        # Ensure directories exist
        self.processed_dir.mkdir(parents=True, exist_ok=True)
        self.extracted_dir.mkdir(parents=True, exist_ok=True)
    
    def export_pdf_tables(
        self,
        tables: List[Dict[str, Any]],
        source_pdf: str,
        output_dir: Optional[str] = None
    ) -> str:
        """
        Export tables from a single PDF to Excel.
        
        Args:
            tables: List of extracted table dictionaries with:
                - content: Table content (markdown or text)
                - metadata: Dict with page_no, table_title, etc.
            source_pdf: Source PDF filename
            output_dir: Override output directory
            
        Returns:
            Path to created Excel file
        """
        output_dir = Path(output_dir) if output_dir else self.processed_dir
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate output filename
        pdf_name = Path(source_pdf).stem
        output_path = output_dir / f"{pdf_name}_tables.xlsx"
        
        return self._create_excel_workbook(tables, output_path, source_pdf)
    
    def export_combined_tables(
        self,
        all_tables: List[Dict[str, Any]],
        output_dir: Optional[str] = None
    ) -> str:
        """
        Export tables from all PDFs to single combined Excel.
        
        Groups tables by title across all sources.
        
        Args:
            all_tables: List of all extracted tables from all PDFs
            output_dir: Override output directory
            
        Returns:
            Path to created Excel file
        """
        output_dir = Path(output_dir) if output_dir else self.extracted_dir
        output_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d")
        output_path = output_dir / f"all_tables_combined_{timestamp}.xlsx"
        
        return self._create_excel_workbook(all_tables, output_path, "Combined")
    
    def _create_excel_workbook(
        self,
        tables: List[Dict[str, Any]],
        output_path: Path,
        source_label: str
    ) -> str:
        """
        Create Excel workbook with Index and table sheets.
        
        Args:
            tables: List of table dictionaries
            output_path: Output file path
            source_label: Label for source (PDF name or "Combined")
            
        Returns:
            Path to created file
        """
        if not tables:
            logger.warning("No tables to export")
            return ""
        
        try:
            # Group tables by title (returns dict with Table_ID as key)
            tables_by_title = self._group_tables_by_title(tables)
            
            # Create workbook
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Create Index sheet first
                self._create_index_sheet(writer, tables, tables_by_title)
                
                # Create table sheets - use Table_ID as sheet name
                for table_id, title_tables in tables_by_title.items():
                    self._create_table_sheet(writer, table_id, title_tables)
            
            # Add hyperlinks (requires reopening workbook)
            self._add_hyperlinks(output_path, tables_by_title)
            
            logger.info(f"Exported {len(tables)} tables to {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Failed to export tables: {e}", exc_info=True)
            return ""
    
    def _group_tables_by_title(
        self,
        tables: List[Dict[str, Any]]
    ) -> Dict[str, List[Dict[str, Any]]]:
        """
        Group tables by their logical table (Section + Title).
        
        IMPORTANT: Tables with the same title but different sections are DIFFERENT tables.
        E.g., "Income Statement Info" under "Institutional Securities" vs "Wealth Management"
        should NOT be merged.
        
        Returns dict with Table_ID as key to avoid sheet name collisions.
        """
        groups = defaultdict(list)
        logical_table_ids = {}
        next_table_id = 1
        
        for table in tables:
            metadata = table.get('metadata', {})
            title = metadata.get('table_title', 'Untitled')
            section = metadata.get('section_name', '')  # Get section name
            
            # Clean title to identify logical table (remove row ranges)
            cleaned_title = self._clean_title_for_grouping(title)
            normalized_title = self._normalize_title_for_grouping(cleaned_title)
            
            # Create grouping key: Section + Title
            # This ensures same-titled tables in different sections stay separate
            if section and section.strip():
                grouping_key = f"{section.strip()}::{normalized_title}"
            else:
                grouping_key = f"Default::{normalized_title}"
            
            # Assign Table_ID based on unique Section + Title combination
            if grouping_key not in logical_table_ids:
                logical_table_ids[grouping_key] = str(next_table_id)
                next_table_id += 1
            
            table_id = logical_table_ids[grouping_key]
            
            # Store the table_id and section in metadata for later use
            if 'metadata' not in table:
                table['metadata'] = {}
            table['metadata']['_assigned_table_id'] = table_id
            table['metadata']['_cleaned_title'] = cleaned_title
            table['metadata']['_section'] = section
            table['metadata']['_grouping_key'] = grouping_key
            
            # Group by Table_ID instead of sanitized title
            groups[table_id].append(table)
        
        return dict(groups)
    
    def _clean_title_for_grouping(self, title: str) -> str:
        """
        Clean title for grouping - remove row ranges and chunk indicators.
        
        This ensures chunks of the same table are grouped together:
        - "Income Statement (Rows 1-10)" -> "Income Statement"
        - "Balance Sheet (Rows 8-15)" -> "Balance Sheet"
        """
        if not title:
            return "Untitled"
        
        # Remove row ranges like "(Rows 1-10)" or "(Rows 8-17)"
        cleaned = re.sub(r'\s*\(Rows?\s*\d+[-–]\d+\)\s*$', '', title, flags=re.IGNORECASE)
        
        # Remove trailing numbers that might be footnotes
        cleaned = re.sub(r'\s+\d+\s*$', '', cleaned)
        
        # Remove trailing parenthesized numbers
        cleaned = re.sub(r'\s*\(\d+\)\s*$', '', cleaned)
        
        return cleaned.strip() or "Untitled"
    
    def _normalize_title_for_grouping(self, title: str) -> str:
        """
        Normalize title for case-insensitive grouping.
        Delegates to ExcelUtils after cleaning row ranges.
        """
        if not title:
            return "untitled"
        
        # First clean row ranges
        cleaned = self._clean_title_for_grouping(title)
        
        # Use ExcelUtils for the actual normalization
        return ExcelUtils.normalize_title_for_grouping(cleaned) or "untitled"
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize string for Excel sheet name. Delegates to ExcelUtils."""
        return ExcelUtils.sanitize_sheet_name(name)
    
    def _create_index_sheet(
        self,
        writer: pd.ExcelWriter,
        tables: List[Dict[str, Any]],
        tables_by_title: Dict[str, List[Dict[str, Any]]]
    ) -> None:
        """
        Create Index sheet with comprehensive table metadata.
        
        Columns: Source, PageNo, Table_ID, Location_ID, Section, Table Title, Link
        
        Location_ID = <pageNo>_<Table_IndexID_perpage>
        Section is included to show why tables with same title might be separate.
        """
        index_data = []
        
        # Track tables per page to generate Location_ID
        page_table_counts = {}
        
        # Track unique logical tables to assign Table_ID (Section + Title)
        logical_table_ids = {}
        next_table_id = 1
        
        for table in tables:
            metadata = table.get('metadata', {})
            title = metadata.get('table_title', 'Untitled')
            section = metadata.get('section_name', '')  # Get section name
            
            # Clean title to identify logical table (remove row ranges)
            cleaned_title = self._clean_title_for_grouping(title)
            normalized_title = self._normalize_title_for_grouping(cleaned_title)
            
            # Create grouping key: Section + Title (consistent with _group_tables_by_title)
            if section and section.strip():
                grouping_key = f"{section.strip()}::{normalized_title}"
            else:
                grouping_key = f"Default::{normalized_title}"
            
            # Assign Table_ID based on unique Section + Title combination
            if grouping_key not in logical_table_ids:
                logical_table_ids[grouping_key] = str(next_table_id)
                next_table_id += 1
            
            table_id = logical_table_ids[grouping_key]
            
            # Get page number
            page_no = metadata.get('page_no', 'N/A')
            
            # Generate Location_ID as <pageNo>_<TableIndexID_perpage>
            if page_no != 'N/A':
                if page_no not in page_table_counts:
                    page_table_counts[page_no] = 0
                page_table_counts[page_no] += 1
                location_id = f"{page_no}_{page_table_counts[page_no]}"
            else:
                location_id = ''
            
            # Required columns per spec (now includes Section)
            index_data.append({
                'Source': metadata.get('source_doc', 'Unknown'),
                'PageNo': page_no,
                'Table_ID': table_id,
                'Location_ID': location_id,
                'Section': section,  # NEW: Section column
                'Table Title': title,
                'Link': table_id  # Use Table_ID as link target (sheet name)
            })
        
        df = pd.DataFrame(index_data)
        
        # Ensure columns in correct order (now includes Section)
        REQUIRED_COLUMNS = ['Source', 'PageNo', 'Table_ID', 'Location_ID', 'Section', 'Table Title', 'Link']
        cols_to_keep = [c for c in REQUIRED_COLUMNS if c in df.columns]
        if cols_to_keep:
            df = df[cols_to_keep]
        
        df.to_excel(writer, sheet_name='Index', index=False)
        
        # Auto-adjust column widths with smart sizing
        worksheet = writer.sheets['Index']
        for idx, col in enumerate(df.columns):
            col_letter = self._get_column_letter(idx)
            
            max_content_len = df[col].astype(str).map(len).max() if len(df) > 0 else 0
            header_len = len(col)
            max_len = max(max_content_len, header_len) + 2
            
            # Apply reasonable limits based on column type
            if col == 'Table Title':
                max_len = min(max_len, 60)
            elif col == 'Section':
                max_len = min(max_len, 25)  # Section column
            elif col in ['Source', 'Table_ID', 'Location_ID']:
                max_len = min(max_len, 25)
            elif col == 'PageNo':
                max_len = min(max_len, 10)
            elif col == 'Link':
                max_len = min(max_len, 15)
            else:
                max_len = min(max_len, 20)
            
            worksheet.column_dimensions[col_letter].width = max_len
    
    def _get_column_letter(self, idx: int) -> str:
        """Convert column index to Excel column letter. Delegates to ExcelUtils."""
        return ExcelUtils.get_column_letter(idx)
    
    def _create_table_sheet(
        self,
        writer: pd.ExcelWriter,
        sheet_name: str,
        tables: List[Dict[str, Any]]
    ) -> None:
        """
        Create sheet for a specific table (identified by Table_ID).
        
        Sheet Structure:
        - Row 1: "← Back to Index" link
        - Row 2: Row Header (Level 1): ...
        - Row 3: Row Header (Level 2 - Sub): ... (if applicable)
        - Row 4: Product/Entity: ...
        - Row 5: Column Headers (Level 1): ...
        - Row 6: Column Headers (Level 2): ... (if applicable)
        - Row 7: <blank>
        - Row 8: Table Title: ...
        - Row 9: <blank>
        - Row 10+: Table data
        
        Args:
            writer: Excel writer
            sheet_name: Sheet name (Table_ID, e.g., "1", "2")
            tables: List of table chunks with same logical table
        """
        all_rows = []
        
        # Row 1: Add "Back to Index" link placeholder
        all_rows.append(['← Back to Index'])
        
        # Collect headers and metadata from ALL chunks
        if tables:
            first_table = tables[0]
            metadata = first_table.get('metadata', {})
            
            # Get the cleaned title (without row ranges) for display
            display_title = metadata.get('_cleaned_title', metadata.get('table_title', 'Untitled'))
            
            # === COLLECT ROW HEADERS FROM ALL CHUNKS ===
            all_row_headers = []
            all_row_sub_headers = []
            has_row_hierarchy = False
            
            for table in tables:
                table_metadata = table.get('metadata', {})
                table_df = self._parse_table_content(table.get('content', ''))
                
                # Try to get row headers from metadata first
                row_headers_meta = table_metadata.get('row_headers', '')
                if row_headers_meta:
                    if ',' in row_headers_meta:
                        headers = [h.strip() for h in row_headers_meta.split(',')]
                    else:
                        headers = [row_headers_meta]
                    all_row_headers.extend(headers)
                elif not table_df.empty and len(table_df.columns) > 0:
                    # Extract from first column of data
                    for idx, val in enumerate(table_df.iloc[:, 0]):
                        if pd.notna(val) and str(val).strip():
                            row_label = str(val).strip()
                            
                            # Check if this row has any data in subsequent columns
                            has_data = False
                            if len(table_df.columns) > 1:
                                row_data = table_df.iloc[idx, 1:]
                                for cell_val in row_data:
                                    if pd.notna(cell_val) and str(cell_val).strip() and str(cell_val).strip().lower() not in ['nan', '', '-', '—']:
                                        has_data = True
                                        break
                            
                            if not has_data and len(table_df.columns) > 1:
                                # This is a section header (no data in row) → Level 1
                                all_row_headers.append(row_label)
                                has_row_hierarchy = True
                            else:
                                # This is a line item with data → Level 2
                                all_row_sub_headers.append(row_label)
            
            # Filter out invalid values (keep duplicates for Level 1/2)
            filtered_row_headers = [h for h in all_row_headers if h and h.lower() not in ['nan', 'none', '']]
            filtered_row_sub_headers = [h for h in all_row_sub_headers if h and h.lower() not in ['nan', 'none', '']]
            
            # Row 2: Row Header (Level 1) - section headers without data (keeps duplicates)
            row_headers_str = ', '.join(filtered_row_headers) if filtered_row_headers else ''
            all_rows.append([f"Row Header (Level 1): {row_headers_str}"])
            
            # Row 3: Row Header (Level 2) - line items with data (keeps duplicates)
            if filtered_row_sub_headers:
                row_sub_headers_str = ', '.join(filtered_row_sub_headers)
                all_rows.append([f"Row Header (Level 2): {row_sub_headers_str}"])
            else:
                all_rows.append([f"Row Header (Level 2):"])
            
            # Row 4: Product/Entity (unique row headers for this table)
            # Exclude unit indicators like "$ in million", "in millions", etc.
            UNIT_PATTERNS = [
                '$ in million', '$ in billion', '$ in thousand',
                'in millions', 'in billions', 'in thousands',
                'dollars in millions', 'dollars in billions',
                '(in millions)', '(in billions)', '(in thousands)',
                'amounts in millions', 'amounts in billions',
            ]
            
            def is_unit_indicator(text: str) -> bool:
                """Check if text is a unit indicator."""
                text_lower = text.lower().strip()
                for pattern in UNIT_PATTERNS:
                    if pattern in text_lower:
                        return True
                # Also exclude if it starts with $ and contains 'in'
                if text_lower.startswith('$') and ' in ' in text_lower:
                    return True
                return False
            
            all_unique_entities = []
            seen_entities = set()
            # Only Level 2 (line items with data) are actual products/entities
            # Level 1 section headers (like "Net revenues:") are just categories
            # Product/Entity is UNIQUE (deduplicated)
            for entity in filtered_row_sub_headers:
                if entity and entity not in seen_entities and not is_unit_indicator(entity):
                    seen_entities.add(entity)
                    all_unique_entities.append(entity)
            
            entities_str = ', '.join(all_unique_entities) if all_unique_entities else ''
            all_rows.append([f"Product/Entity: {entities_str}"])
            
            # === COLLECT COLUMN HEADERS FROM ALL CHUNKS ===
            # Parse actual table content to detect multi-level headers
            all_main_headers = []
            all_sub_headers = []
            detected_years = set()
            
            for table in tables:
                content = table.get('content', '')
                
                # Parse the markdown table to detect header levels
                headers_info = self._detect_column_header_levels(content)
                
                if headers_info['has_multi_level']:
                    all_main_headers.extend(headers_info['level_1'])
                    all_sub_headers.extend(headers_info['level_2'])
                else:
                    # Single level - put in level 1
                    all_main_headers.extend(headers_info['level_1'])
                
                # Extract years from sub-headers (e.g., "2024", "March 31, 2023")
                for header in headers_info.get('level_2', []) + headers_info.get('level_1', []):
                    year_match = re.search(r'(20\d{2})', str(header))
                    if year_match:
                        detected_years.add(year_match.group(1))
            
            # Deduplicate headers while preserving order
            unique_main = self._dedupe_preserve_order(all_main_headers)
            unique_sub = self._dedupe_preserve_order(all_sub_headers)
            
            # Row 5: Column Header (Level 1) - main/spanning headers
            if unique_main:
                main_headers_str = ', '.join(unique_main)
                all_rows.append([f"Column Header (Level 1): {main_headers_str}"])
            else:
                all_rows.append([f"Column Header (Level 1):"])
            
            # Row 6: Column Header (Level 2) - sub-headers with dates/years
            if unique_sub:
                sub_headers_str = ', '.join(unique_sub)
                all_rows.append([f"Column Header (Level 2): {sub_headers_str}"])
            else:
                all_rows.append([f"Column Header (Level 2):"])
            
            # NEW: Row 7: Year(s) detected from column headers
            if detected_years:
                years_str = ', '.join(sorted(detected_years, reverse=True))
                all_rows.append([f"Year(s): {years_str}"])
            
            # Row 7: Blank row
            all_rows.append([])
            
            # Row 8: Table Title
            all_rows.append([f"Table Title: {display_title}"])
            
            # Row 9: Blank row
            all_rows.append([])
        
        # Row 10+: Table data (stacked if multiple chunks)
        for i, table in enumerate(tables):
            metadata = table.get('metadata', {})
            
            # Add source header for each table occurrence
            source_info = f"Source: {metadata.get('source_doc', 'Unknown')}, Page {metadata.get('page_no', 'N/A')}"
            if metadata.get('year'):
                source_info += f", {metadata.get('year')}"
            if metadata.get('quarter'):
                source_info += f" {metadata.get('quarter')}"
            
            all_rows.append([source_info])
            
            # Parse table content
            table_df = self._parse_table_content(table.get('content', ''))
            
            if not table_df.empty:
                # Add header row
                all_rows.append(list(table_df.columns))
                
                # Add data rows
                for _, row in table_df.iterrows():
                    all_rows.append(list(row))
            
            # Add blank rows between tables (except after last)
            if i < len(tables) - 1:
                all_rows.append([])
                all_rows.append([])
        
        # Create DataFrame from all rows
        df = pd.DataFrame(all_rows)
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
    
    def _detect_column_header_levels(self, content: str) -> Dict[str, Any]:
        """
        Detect multi-level column headers from markdown table content.
        
        Delegates to HeaderDetector class.
        
        Returns:
            {
                'has_multi_level': bool,
                'level_1': List[str],  # Main/spanning headers
                'level_2': List[str]   # Sub-headers (typically with dates/years)
            }
        """
        return HeaderDetector.detect_column_header_levels(content)
    
    def _dedupe_preserve_order(self, items: List[str]) -> List[str]:
        """Deduplicate a list while preserving order. Delegates to HeaderDetector."""
        return HeaderDetector.dedupe_preserve_order(items)
    
    def _parse_table_content(self, content: str) -> pd.DataFrame:
        """Parse markdown/text table content to DataFrame."""
        from src.utils.table_utils import parse_markdown_table
        return parse_markdown_table(content, handle_colon_separator=False)
    
    def _add_hyperlinks(
        self,
        output_path: Path,
        tables_by_title: Dict[str, List[Dict[str, Any]]]
    ) -> None:
        """
        Add hyperlinks to workbook (Index <-> Sheets).
        
        Sheet names are now Table_IDs (e.g., "1", "2", "3"), so linking is simpler.
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font
            
            wb = load_workbook(output_path)
            
            # Add hyperlinks in Index sheet
            index_sheet = wb['Index']
            
            # Find the Link column index dynamically
            header_row = [index_sheet.cell(row=1, column=c).value for c in range(1, index_sheet.max_column + 1)]
            link_col = None
            for i, h in enumerate(header_row, start=1):
                if isinstance(h, str) and h.strip().lower() == 'link':
                    link_col = i
                    break
            
            if link_col is None:
                logger.warning("Link column not found in Index sheet")
                link_col = -1
            
            for row in range(2, index_sheet.max_row + 1):  # Skip header
                if link_col == -1:
                    break
                cell = index_sheet.cell(row=row, column=link_col)
                raw_value = cell.value
                
                # Link value is Table_ID (e.g., "1", "2", "3")
                table_id = None
                if raw_value is not None:
                    try:
                        raw_str = str(raw_value).strip()
                    except Exception:
                        raw_str = ''
                    
                    if raw_str.startswith('→'):
                        table_id = raw_str.lstrip('→').strip()
                    else:
                        table_id = raw_str
                
                if table_id:
                    # Sheet name is simply the Table_ID
                    sheet_name = table_id
                    
                    if sheet_name in wb.sheetnames:
                        cell.hyperlink = f"#'{sheet_name}'!A1"
                        cell.value = f"→ {sheet_name}"
                        cell.font = Font(color="0000FF", underline="single")
                    else:
                        logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            
            # Add back-links in each table sheet
            for sheet_name in tables_by_title.keys():
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    cell = sheet.cell(row=1, column=1)
                    cell.hyperlink = "#'Index'!A1"
                    cell.value = "← Back to Index"
                    cell.font = Font(color="0000FF", underline="single")
            
            wb.save(output_path)
            
        except Exception as e:
            logger.error(f"Failed to add hyperlinks: {e}")

    def merge_processed_files(
        self,
        output_filename: str = "consolidated_tables.xlsx",
        transpose: bool = False
    ) -> dict:
        """
        Merge all processed xlsx files into a single consolidated report.
        
        NOTE: This method delegates to ConsolidatedExcelExporter for cleaner separation.
        For direct access to consolidation features, use:
            from src.infrastructure.extraction.formatters.consolidated_exporter import get_consolidated_exporter
        
        Args:
            output_filename: Output filename for consolidated report
            transpose: If True, dates become rows (time-series format)
            
        Returns:
            Dict with path, tables_merged, sources_merged, sheet_names
        """
        from src.infrastructure.extraction.formatters.consolidated_exporter import get_consolidated_exporter
        
        consolidated_exporter = get_consolidated_exporter()
        return consolidated_exporter.merge_processed_files(
            output_filename=output_filename,
            transpose=transpose
        )


# =============================================================================
# SINGLETON PATTERN
# =============================================================================

_excel_exporter: Optional[ExcelTableExporter] = None


def get_excel_exporter() -> ExcelTableExporter:
    """
    Get or create global Excel exporter instance.
    
    Returns:
        ExcelTableExporter singleton instance
    """
    global _excel_exporter
    
    if _excel_exporter is None:
        _excel_exporter = ExcelTableExporter()
    
    return _excel_exporter


def reset_excel_exporter() -> None:
    """Reset the exporter singleton (for testing)."""
    global _excel_exporter
    _excel_exporter = None
