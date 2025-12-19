"""
Consolidated Excel Exporter.

Merges tables from multiple processed Excel files into a single consolidated report.
Tables are merged horizontally (same rows, columns for each period).

Features:
- Multi-file merge with row alignment
- Section + Title + Structure-based grouping
- Quarter format conversion (Q1, 2025)
- Date-sorted columns (newest first)
- Optional transpose (dates as rows)
"""

import glob
import re
import pandas as pd
from typing import List, Dict, Any, Optional
from pathlib import Path
from collections import defaultdict

from src.utils import get_logger
from src.core import get_paths
from src.infrastructure.extraction.formatters.date_utils import DateUtils
from src.infrastructure.extraction.formatters.excel_utils import ExcelUtils

logger = get_logger(__name__)


class ConsolidatedExcelExporter:
    """
    Export consolidated tables from multiple quarterly/annual reports.
    
    Merges tables by Section + Title with horizontal column alignment.
    """
    
    def __init__(self):
        """Initialize exporter with paths from PathManager."""
        paths = get_paths()
        self.processed_dir = Path(paths.PROCEEDINGS_DIR) if hasattr(paths, 'PROCEEDINGS_DIR') else Path("data/processed")
        self.extracted_dir = Path(paths.EXTRACTED_DIR) if hasattr(paths, 'EXTRACTED_DIR') else Path("data/extracted")
        
        # Ensure directories exist
        self.processed_dir.mkdir(parents=True, exist_ok=True)
        self.extracted_dir.mkdir(parents=True, exist_ok=True)
    
    def merge_processed_files(
        self,
        output_filename: str = "consolidated_tables.xlsx",
        transpose: bool = False
    ) -> dict:
        """
        Merge all processed xlsx files into a single consolidated report.
        
        Reads all *_tables.xlsx files from data/processed/, merges tables
        by title (row-aligned), and adds metadata columns.
        
        Args:
            output_filename: Output filename for consolidated report
            transpose: If True, dates become rows (time-series format)
            
        Returns:
            Dict with path, tables_merged, sources_merged, sheet_names
        """
        from openpyxl import load_workbook
        
        # Find all processed xlsx files
        pattern = str(self.processed_dir / "*_tables.xlsx")
        xlsx_files = glob.glob(pattern)
        
        # Filter out temp files (start with ~$)
        xlsx_files = [f for f in xlsx_files if not Path(f).name.startswith('~$')]
        
        if not xlsx_files:
            logger.warning(f"No processed xlsx files found in {self.processed_dir}")
            return {}
        
        logger.info(f"Found {len(xlsx_files)} processed files to merge")
        
        # Collect all tables from all files with metadata
        all_tables_by_full_title = defaultdict(list)
        title_to_sheet_name = {}
        
        for xlsx_path in xlsx_files:
            try:
                self._process_xlsx_file(
                    xlsx_path, 
                    all_tables_by_full_title, 
                    title_to_sheet_name
                )
            except Exception as e:
                logger.error(f"Error reading {xlsx_path}: {e}")
        
        if not all_tables_by_full_title:
            logger.warning("No tables collected for merging")
            return {}
        
        # Assign numeric sheet names (1, 2, 3...) like individual xlsx files
        sheet_number = 1
        for normalized_key in all_tables_by_full_title.keys():
            mapping = title_to_sheet_name.get(normalized_key, {})
            # Use simple numeric ID as sheet name
            mapping['unique_sheet_name'] = str(sheet_number)
            title_to_sheet_name[normalized_key] = mapping
            sheet_number += 1
        
        # Create consolidated output
        output_path = self.extracted_dir / output_filename
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Create Master Index
                self._create_consolidated_index(writer, all_tables_by_full_title, title_to_sheet_name)
                
                # Create merged sheets
                for normalized_key, tables in all_tables_by_full_title.items():
                    mapping = title_to_sheet_name.get(normalized_key, {})
                    sheet_name = mapping.get('unique_sheet_name')
                    if sheet_name:
                        self._create_merged_table_sheet(writer, sheet_name, tables, transpose=transpose)
            
            # Add hyperlinks
            self._add_hyperlinks(output_path, all_tables_by_full_title, title_to_sheet_name)
            
            logger.info(f"Created consolidated report at {output_path}")
            logger.info(f"  - Tables merged: {len(all_tables_by_full_title)}")
            logger.info(f"  - Sources: {len(xlsx_files)}")
            
            return {
                'path': str(output_path),
                'tables_merged': len(all_tables_by_full_title),
                'sources_merged': len(xlsx_files),
                'sheet_names': [m.get('unique_sheet_name', '') for m in title_to_sheet_name.values()]
            }
            
        except Exception as e:
            logger.error(f"Failed to create consolidated report: {e}", exc_info=True)
            return {}
    
    def _process_xlsx_file(
        self,
        xlsx_path: str,
        all_tables_by_full_title: Dict[str, List],
        title_to_sheet_name: Dict[str, dict]
    ) -> None:
        """Process a single xlsx file and add tables to collection."""
        # Read Index sheet to get metadata
        index_df = pd.read_excel(xlsx_path, sheet_name='Index')
        
        # Get all sheet names except Index
        xl = pd.ExcelFile(xlsx_path)
        table_sheets = [s for s in xl.sheet_names if s != 'Index']
        
        # Create mapping from Link to full Table Title
        link_to_full_title = {}
        for _, idx_row in index_df.iterrows():
            full_title = str(idx_row.get('Table Title', ''))
            link = str(idx_row.get('Link', ''))
            if link.startswith('→ '):
                link = link[2:]
            if link and full_title:
                link_to_full_title[link] = full_title
        
        for sheet_name in table_sheets:
            full_title = link_to_full_title.get(sheet_name, sheet_name)
            
            # Find metadata for this sheet
            matching_rows = index_df[
                index_df['Table Title'].astype(str) == full_title
            ]
            
            if len(matching_rows) > 0:
                row = matching_rows.iloc[0]
                section = str(row.get('Section', '')) if 'Section' in row.index else ''
                metadata = {
                    'source': row.get('Source', Path(xlsx_path).stem),
                    'year': row.get('Year', ''),
                    'quarter': row.get('Quarter', ''),
                    'report_type': self._detect_report_type(str(row.get('Source', ''))),
                    'section': section,
                }
            else:
                section = ''
                metadata = {
                    'source': Path(xlsx_path).stem,
                    'year': '',
                    'quarter': '',
                    'report_type': '',
                    'section': '',
                }
            
            # Read table content (skip header rows)
            try:
                table_df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=None)
                
                # Dynamically find where data starts by looking for "Source:" row
                # Individual xlsx structure:
                #   Row 0: ← Back to Index
                #   Row 1-6: Metadata (Row Header, Column Header, etc.)
                #   Row 7: Year(s) or blank
                #   Row 8: Table Title
                #   Row 9: Blank
                #   Row 10+: "Source: ..." followed by actual table data
                data_start_row = 0
                for idx, row in table_df.iterrows():
                    first_cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
                    if first_cell.startswith('Source:'):
                        data_start_row = idx
                        break
                
                # If Source: not found, fallback to row 10 or after metadata
                if data_start_row == 0:
                    # Look for first row that looks like table data
                    for idx, row in table_df.iterrows():
                        first_cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
                        # Skip metadata patterns
                        if any(first_cell.startswith(p) for p in 
                               ['← Back', 'Row Header', 'Column Header', 'Product', 
                                'Table Title', 'Year(s)', 'Source:']):
                            continue
                        if first_cell.strip() == '':
                            continue
                        # This looks like data - start here
                        data_start_row = idx
                        break
                
                # Slice from data start
                if data_start_row > 0 and len(table_df) > data_start_row:
                    table_df = table_df.iloc[data_start_row:].reset_index(drop=True)
                
                # Compute row signature for strict matching
                raw_rows = table_df.iloc[:, 0].astype(str).tolist() if not table_df.empty else []
                norm_rows = []
                for r in raw_rows:
                    r_clean = r.strip()
                    if not r_clean: 
                        continue
                    r_lower = r_clean.lower()
                    if r_lower == 'row label': 
                        continue
                    if r_lower.startswith('source:'): 
                        continue
                    if r_lower.startswith('page '): 
                        continue
                    if r_lower.startswith('$ in'): 
                        continue
                    norm_rows.append(self._normalize_row_label(r_clean))
                
                row_sig = "|".join(norm_rows)
                
                # Create grouping key: Section + Title + Structure
                normalized_title = self._normalize_title_for_grouping(full_title)
                if section and section.strip():
                    normalized_key = f"{section.strip()}::{normalized_title}::{row_sig}"
                else:
                    normalized_key = f"Default::{normalized_title}::{row_sig}"
                
                all_tables_by_full_title[normalized_key].append({
                    'data': table_df,
                    'metadata': metadata,
                    'sheet_name': sheet_name,
                    'original_title': full_title,
                    'section': section,
                    'row_sig': row_sig
                })
                
                if normalized_key not in title_to_sheet_name:
                    clean_display = full_title.replace('→ ', '').strip() if full_title.startswith('→') else full_title
                    if section and section.strip():
                        display_with_section = f"{section.strip()} - {clean_display}"
                    else:
                        display_with_section = clean_display
                    
                    title_to_sheet_name[normalized_key] = {
                        'display_title': display_with_section,
                        'section': section.strip() if section else '',
                        'original_title': full_title,
                    }
                    
            except Exception as e:
                logger.debug(f"Skipping sheet {sheet_name}: {e}")
    
    def _detect_report_type(self, source: str) -> str:
        """Detect report type from source filename. Delegates to ExcelUtils."""
        return ExcelUtils.detect_report_type(source)
    
    def _normalize_row_label(self, label: str) -> str:
        """Normalize row label for matching. Delegates to ExcelUtils."""
        return ExcelUtils.normalize_row_label(label)
    
    def _normalize_title_for_grouping(self, title: str) -> str:
        """Normalize title for grouping. Delegates to ExcelUtils."""
        return ExcelUtils.normalize_title_for_grouping(title)
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize for Excel sheet name. Delegates to ExcelUtils."""
        return ExcelUtils.sanitize_sheet_name(name)
    
    def _get_column_letter(self, idx: int) -> str:
        """Convert column index to letter. Delegates to ExcelUtils."""
        return ExcelUtils.get_column_letter(idx)
    
    def _create_consolidated_index(
        self,
        writer: pd.ExcelWriter,
        tables_by_full_title: Dict[str, List[Dict[str, Any]]],
        title_to_sheet_name: Dict[str, dict]
    ) -> None:
        """Create master index for consolidated report."""
        index_data = []
        
        for normalized_key, tables in tables_by_full_title.items():
            sources = set()
            years = set()
            quarters = set()
            report_types = set()
            sections = set()
            source_refs = []  # List of "Q1,2025_page" references
            
            for t in tables:
                meta = t['metadata']
                section = t.get('section', '') or meta.get('section', '')
                if section and str(section).strip() and str(section).lower() != 'nan':
                    sections.add(str(section).strip())
                if meta.get('source'):
                    sources.add(str(meta['source']))
                if meta.get('year') and str(meta.get('year')) not in ['nan', 'NaN', '']:
                    years.add(str(int(meta['year'])) if isinstance(meta.get('year'), float) else str(meta.get('year')))
                if meta.get('quarter') and str(meta.get('quarter')) not in ['nan', 'NaN', '']:
                    quarters.add(str(meta['quarter']))
                if meta.get('report_type'):
                    report_types.add(str(meta['report_type']))
                
                # Build source reference: Q1,2025_page or FY2024_page
                year_val = str(int(meta['year'])) if isinstance(meta.get('year'), float) else str(meta.get('year', ''))
                quarter_val = str(meta.get('quarter', ''))
                page_val = str(meta.get('page', ''))
                
                if quarter_val and year_val:
                    ref = f"{quarter_val},{year_val}"
                elif year_val:
                    ref = f"FY{year_val}"
                else:
                    ref = str(meta.get('source', 'Unknown'))[:10]
                
                if page_val and page_val != 'nan':
                    ref += f"_p{page_val}"
                
                if ref and ref not in source_refs:
                    source_refs.append(ref)
            
            # Filter out NaN values
            quarters = {q for q in quarters if q and q.lower() != 'nan'}
            years = {y for y in years if y and y.lower() != 'nan'}
            sources = {s for s in sources if s and s.lower() != 'nan'}
            report_types = {r for r in report_types if r and r.lower() != 'nan'}
            sections = {s for s in sections if s and s.lower() != 'nan'}
            
            mapping = title_to_sheet_name.get(normalized_key, {})
            display_title = mapping.get('display_title', normalized_key)
            original_title = mapping.get('original_title', display_title)
            sheet_name = mapping.get('unique_sheet_name', self._sanitize_sheet_name(normalized_key))
            section_str = mapping.get('section', '') or ', '.join(sorted(sections))
            
            # Format source references (limit to 5 to avoid very long strings)
            source_refs_str = ', '.join(source_refs[:5])
            if len(source_refs) > 5:
                source_refs_str += f'... (+{len(source_refs) - 5})'
            
            index_data.append({
                '#': 0,  # Will be set after sorting
                'Section': section_str if section_str else '-',
                'Table Title': original_title if original_title else display_title,
                'TableCount': len(tables),
                'Sources': source_refs_str,
                'Report Types': ', '.join(sorted(report_types)),
                'Years': ', '.join(sorted(years, reverse=True)),
                'Quarters': ', '.join(sorted(quarters)),
                'Link': '',  # Will be set after sorting
                '_normalized_key': normalized_key  # Hidden key for sheet mapping
            })
        
        df = pd.DataFrame(index_data)
        # Sort by Section, then by Table Title
        df = df.sort_values(['Section', 'Table Title'])
        # Re-number after sorting and set Link to match #
        df['#'] = range(1, len(df) + 1)
        df['Link'] = df['#'].astype(str)
        
        # Update title_to_sheet_name with new numeric sheet names based on sorted order
        for idx, row in df.iterrows():
            normalized_key = row['_normalized_key']
            sheet_num = str(row['#'])
            if normalized_key in title_to_sheet_name:
                title_to_sheet_name[normalized_key]['unique_sheet_name'] = sheet_num
        
        # Drop the hidden column before saving
        df = df.drop(columns=['_normalized_key'])
        df.to_excel(writer, sheet_name='Index', index=False)
        
        # Auto-adjust column widths
        worksheet = writer.sheets['Index']
        for idx, col in enumerate(df.columns):
            col_letter = self._get_column_letter(idx)
            max_content_len = df[col].astype(str).map(len).max() if len(df) > 0 else 0
            header_len = len(col)
            max_len = max(max_content_len, header_len) + 2
            
            if col == '#':
                max_len = 5
            elif col == 'Section':
                max_len = min(max_len, 35)
            elif col == 'Table Title':
                max_len = min(max_len, 50)
            elif col == 'TableCount':
                max_len = 10
            elif col == 'Sources':
                max_len = min(max_len, 45)  # Wider for Q1,2025_p48 format
            elif col == 'Report Types':
                max_len = min(max_len, 15)
            elif col in ['Years', 'Quarters']:
                max_len = min(max_len, 20)
            elif col == 'Link':
                max_len = min(max_len, 32)
            else:
                max_len = min(max_len, 20)
            
            worksheet.column_dimensions[col_letter].width = max_len
    
    def _create_merged_table_sheet(
        self,
        writer: pd.ExcelWriter,
        title: str,
        tables: List[Dict[str, Any]],
        transpose: bool = False
    ) -> None:
        """Create merged sheet with tables from multiple sources."""
        if not tables:
            return
        
        all_columns = {}
        row_labels = []
        normalized_row_labels = {}
        
        for table_info in tables:
            df = table_info['data']
            meta = table_info['metadata']
            
            if df.empty or len(df.columns) < 2:
                continue
            
            # First column is row labels
            for val in df.iloc[:, 0].tolist():
                if pd.notna(val):
                    norm_val = self._normalize_row_label(val)
                    if norm_val and norm_val not in normalized_row_labels:
                        normalized_row_labels[norm_val] = val
                        row_labels.append(val)
            
            # Extract column data
            for col_idx in range(1, len(df.columns)):
                header_parts = []
                for row_idx in range(min(5, len(df))):
                    val = df.iloc[row_idx, col_idx]
                    if pd.notna(val):
                        val_str = str(val).strip()
                        if val_str and val_str != 'nan' and val_str not in ['$', '$,', '%', '%,', '-', '—']:
                            is_year = bool(re.match(r'^20[2-3]\d$', val_str))
                            is_currency = val_str.startswith('$') and any(c.isdigit() for c in val_str)
                            is_percentage = '%' in val_str and any(c.isdigit() for c in val_str)
                            num_digits = sum(c.isdigit() for c in val_str)
                            is_large_number = num_digits > 4 and val_str.replace(',', '').replace('.', '').replace('-', '').replace('(', '').replace(')', '').isdigit()
                            
                            is_data = (is_currency or is_percentage or is_large_number) and not is_year
                            
                            if not is_data:
                                header_parts.append(val_str)
                            else:
                                break
                
                if header_parts:
                    seen = set()
                    unique_parts = []
                    for p in header_parts:
                        if p.lower() not in seen:
                            seen.add(p.lower())
                            unique_parts.append(p)
                    header = ' '.join(unique_parts)
                    header = DateUtils.extract_date_from_header(header)
                else:
                    header = f"Col_{col_idx}"
                
                sort_key = DateUtils.parse_date_from_header(header)
                norm_header = header.lower().strip()
                
                source_row_labels = df.iloc[:, 0].tolist()
                col_values = df.iloc[:, col_idx].tolist()
                
                row_data_map = {}
                for row_label, value in zip(source_row_labels, col_values):
                    if pd.notna(row_label):
                        norm_row = self._normalize_row_label(row_label)
                        if norm_row:
                            row_data_map[norm_row] = value
                
                non_empty_data = [v for v in row_data_map.values() if pd.notna(v) and str(v).strip() and str(v).strip() != 'nan']
                non_empty_data = [x for x in non_empty_data if str(x).strip() != header]
                
                if not non_empty_data:
                    continue
                
                # Check for duplicates
                is_duplicate = False
                col_data_str = str(sorted([(k, str(v).strip()) for k, v in row_data_map.items()]))
                
                for existing_key, existing_val in all_columns.items():
                    existing_data_str = str(sorted([(k, str(v).strip()) for k, v in existing_val.get('row_data_map', {}).items()]))
                    if col_data_str == existing_data_str:
                        is_duplicate = True
                        break
                
                if not is_duplicate:
                    final_key = norm_header
                    counter = 1
                    while final_key in all_columns:
                        final_key = f"{norm_header}_{counter}"
                        counter += 1
                    
                    all_columns[final_key] = {
                        'header': header,
                        'row_data_map': row_data_map,
                        'sort_key': sort_key,
                        'source': meta.get('source', '')
                    }
        
        if not all_columns:
            pd.DataFrame([{'← Back to Index': ''}]).to_excel(writer, sheet_name=title, index=False)
            return
        
        # Sort columns by date (DESCENDING - newest first)
        sorted_cols = sorted(
            all_columns.items(),
            key=lambda x: x[1]['sort_key'],
            reverse=True  # Newest first
        )
        
        # Build result dataframe
        result_data = {'Row Label': row_labels}
        used_headers = {"row label"}
        
        for _, col_info in sorted_cols:
            original_header = col_info['header']
            row_data_map = col_info.get('row_data_map', {})
            
            # Convert to Qn, YYYY format
            header = DateUtils.convert_to_quarter_format(original_header)
            
            counter = 1
            while header.lower() in used_headers:
                header = f"{original_header} ({counter})"
                counter += 1
            used_headers.add(header.lower())
            
            col_data = []
            for original_label in row_labels:
                norm_label = self._normalize_row_label(original_label)
                value = row_data_map.get(norm_label, '')
                col_data.append(value)
            
            result_data[header] = col_data
        
        result_df = pd.DataFrame(result_data)
        
        # Transpose if requested
        if transpose and len(result_df.columns) > 1:
            result_df = result_df.set_index('Row Label')
            result_df = result_df.T
            result_df.index.name = 'Period'
            result_df = result_df.reset_index()
            
            if 'Period' not in result_df.columns and len(result_df.columns) > 0:
                result_df = result_df.rename(columns={result_df.columns[0]: 'Period'})
        
        # Add back-link as first row
        first_col = result_df.columns[0] if len(result_df.columns) > 0 else 'Row Label'
        back_link_row = pd.DataFrame([{first_col: '← Back to Index'}])
        result_df = pd.concat([back_link_row, result_df], ignore_index=True)
        
        result_df.to_excel(writer, sheet_name=title, index=False)
    
    def _add_hyperlinks(
        self,
        output_path: Path,
        all_tables_by_full_title: Dict[str, List],
        title_to_sheet_name: Dict[str, dict]
    ) -> None:
        """Add hyperlinks to consolidated workbook."""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font
            
            wb = load_workbook(output_path)
            
            # Add hyperlinks in Index sheet
            index_sheet = wb['Index']
            
            # Find Link column dynamically (same as individual xlsx logic)
            header_row = [index_sheet.cell(row=1, column=c).value for c in range(1, index_sheet.max_column + 1)]
            link_col = None
            for i, h in enumerate(header_row, start=1):
                if isinstance(h, str) and h.strip().lower() == 'link':
                    link_col = i
                    break
            
            if link_col is None:
                logger.warning("Link column not found in consolidated Index sheet")
            else:
                for row in range(2, index_sheet.max_row + 1):  # Skip header
                    cell = index_sheet.cell(row=row, column=link_col)
                    raw_value = cell.value
                    
                    # Extract sheet name (handle → prefix if present)
                    sheet_name = None
                    if raw_value is not None:
                        raw_str = str(raw_value).strip()
                        if raw_str.startswith('→'):
                            sheet_name = raw_str.lstrip('→').strip()
                        else:
                            sheet_name = raw_str
                    
                    if sheet_name and sheet_name in wb.sheetnames:
                        cell.hyperlink = f"#'{sheet_name}'!A1"
                        cell.value = f"→ {sheet_name}"
                        cell.font = Font(color="0000FF", underline="single")
                    elif sheet_name:
                        logger.debug(f"Sheet '{sheet_name}' not found in consolidated workbook")
            
            # Add back-links in each data sheet
            for sheet_name in wb.sheetnames:
                if sheet_name == 'Index':
                    continue
                ws = wb[sheet_name]
                cell = ws.cell(row=1, column=1)
                # Set back-link regardless of current value
                cell.value = "← Back to Index"
                cell.hyperlink = "#'Index'!A1"
                cell.font = Font(color="0000FF", underline="single")
            
            wb.save(output_path)
            logger.info(f"Added hyperlinks to consolidated workbook: {len(wb.sheetnames) - 1} sheets")
            
        except Exception as e:
            logger.warning(f"Could not add hyperlinks: {e}")


# =============================================================================
# SINGLETON PATTERN
# =============================================================================

_consolidated_exporter: Optional[ConsolidatedExcelExporter] = None


def get_consolidated_exporter() -> ConsolidatedExcelExporter:
    """
    Get or create global ConsolidatedExcelExporter instance.
    
    Returns:
        ConsolidatedExcelExporter singleton instance
    """
    global _consolidated_exporter
    
    if _consolidated_exporter is None:
        _consolidated_exporter = ConsolidatedExcelExporter()
    
    return _consolidated_exporter


def reset_consolidated_exporter() -> None:
    """Reset the exporter singleton (for testing)."""
    global _consolidated_exporter
    _consolidated_exporter = None
