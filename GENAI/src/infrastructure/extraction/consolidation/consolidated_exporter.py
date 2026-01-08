"""
Consolidated Excel Exporter.

Merges tables from multiple processed Excel files into a single consolidated report.
Tables are merged horizontally (same rows, columns for each period).

Features:
- Multi-file merge with row alignment
- Section + Title + Structure-based grouping
- Quarter format conversion (Q1, 2025)
- Date-sorted columns (newest first)
- Optional transpose (dates as rows)
"""

import glob
import re
import pandas as pd
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
from collections import defaultdict

from openpyxl import load_workbook

from src.utils import get_logger
from src.core import get_paths
from src.utils.date_utils import DateUtils
from src.utils.excel_utils import ExcelUtils
from src.utils.metadata_labels import MetadataLabels, TableMetadata
from src.utils.metadata_builder import MetadataBuilder
from src.utils.financial_domain import (
    DATE_HEADER_PATTERNS,
    extract_quarter_from_header,
    extract_year_from_header,
    is_year_value,
)
from src.infrastructure.extraction.consolidation.consolidated_exporter_transpose import *
from src.infrastructure.extraction.exporters.base_exporter import BaseExcelExporter
# Import from new focused modules
from src.infrastructure.extraction.consolidation.table_detection import TableDetector
from src.infrastructure.extraction.consolidation.table_grouping import TableGrouper
from src.infrastructure.extraction.consolidation.excel_formatting import ExcelFormatter
from src.infrastructure.extraction.consolidation.period_type_detector import PeriodTypeDetector
from src.utils.constants import (
    CONSOLIDATION_YEAR_MIN,
    CONSOLIDATION_YEAR_MAX,
)
from config.settings import settings

logger = get_logger(__name__)

# =============================================================================
# CONSTANTS (from settings for .env configurability, with fallback to constants.py)
# =============================================================================
VALID_YEAR_RANGE = (
    getattr(settings, 'EXTRACTION_YEAR_MIN', CONSOLIDATION_YEAR_MIN), 
    getattr(settings, 'EXTRACTION_YEAR_MAX', CONSOLIDATION_YEAR_MAX)
)


class ConsolidatedExcelExporter(BaseExcelExporter):
    """
    Export consolidated tables from multiple quarterly/annual reports.
    
    Merges tables by Section + Title with horizontal column alignment.
    
    Inherits shared Excel utilities from BaseExcelExporter.
    """
    
    def __init__(self):
        """Initialize exporter with paths from PathManager."""
        super().__init__()
        paths = get_paths()
        # Step 4: reads from processed_advanced, outputs to consolidate/
        self.processed_dir = Path(paths.data_dir) / "processed_advanced" if hasattr(paths, 'data_dir') else Path("data/processed_advanced")
        self.consolidate_dir = Path(paths.data_dir) / "consolidate" if hasattr(paths, 'data_dir') else Path("data/consolidate")
        
        # Ensure directories exist
        self.processed_dir.mkdir(parents=True, exist_ok=True)
        self.consolidate_dir.mkdir(parents=True, exist_ok=True)

    
    def _parse_year_quarter_from_source(self, source: str) -> Tuple[str, str]:
        """
        Parse year and quarter from source filename pattern.
        
        Patterns:
        - 10k1224.pdf → 2024, Q4 (10-K filed Dec 2024)
        - 10q0325.pdf → 2025, Q1 (10-Q filed Mar 2025) 
        - 10q0624.pdf → 2024, Q2 (10-Q filed Jun 2024)
        
        Args:
            source: Source filename (e.g., '10k1224.pdf', '10q0325_tables')
            
        Returns:
            Tuple of (year, quarter) where quarter is 'Q4' for 10-K or 'Q1'-'Q4' for 10-Q
        """
        # Pattern: 10k MMYY or 10q MMYY
        match = re.search(r'(10[kq])(\d{2})(\d{2})', source.lower())
        if match:
            report_type = match.group(1)  # 10k or 10q
            month = int(match.group(2))
            year = 2000 + int(match.group(3))
            
            if report_type == '10k':
                return str(year), 'Q4'
            else:
                # Map month to quarter
                quarter_map = {3: 'Q1', 6: 'Q2', 9: 'Q3', 12: 'Q4', 1: 'Q4', 2: 'Q4'}
                # Handle edge cases: Jan/Feb filings are typically Q4 of previous year
                if month in [1, 2]:
                    year -= 1
                quarter = quarter_map.get(month, f'Q{(month-1)//3 + 1}')
                return str(year), quarter
        
        return '', ''
    
    def merge_processed_files(
        self,
        output_filename: str = "consolidated_tables.xlsx",
        transpose: bool = False
    ) -> dict:
        """
        Merge all processed xlsx files into a single consolidated report.
        
        Reads all *_tables.xlsx files from data/processed/, merges tables
        by title (row-aligned), and adds metadata columns.
        
        Always generates BOTH:
        - Regular consolidated output (consolidated_tables.xlsx)
        - Transposed output (consolidated_tables_transposed.xlsx)
        
        Args:
            output_filename: Output filename for consolidated report
            transpose: If True, only generate transposed file (legacy, ignored)
            
        Returns:
            Dict with path, tables_merged, sources_merged, sheet_names
        """
        
        # Find all processed xlsx files
        pattern = str(self.processed_dir / "*_tables.xlsx")
        xlsx_files = glob.glob(pattern)
        
        # Filter out temp files (start with ~$)
        xlsx_files = [f for f in xlsx_files if not Path(f).name.startswith('~$')]
        
        if not xlsx_files:
            logger.warning(f"No processed xlsx files found in {self.processed_dir}")
            return {}
        
        logger.info(f"Found {len(xlsx_files)} processed files to merge")
        
        # Collect all tables from all files with metadata
        all_tables_by_full_title = defaultdict(list)
        title_to_sheet_name = {}
        
        for xlsx_path in xlsx_files:
            try:
                self._process_xlsx_file(
                    xlsx_path, 
                    all_tables_by_full_title, 
                    title_to_sheet_name
                )
            except Exception as e:
                logger.error(f"Error reading {xlsx_path}: {e}")
        
        if not all_tables_by_full_title:
            logger.warning("No tables collected for merging")
            return {}
        
        # Generate BOTH outputs: regular and transposed
        results = {}
        
        for is_transposed in [False, True]:
            # Determine output filename
            if is_transposed:
                base_name = Path(output_filename).stem
                ext = Path(output_filename).suffix or ".xlsx"
                current_output_filename = f"{base_name}_transposed{ext}"
                logger.info("Creating transposed consolidated report...")
            else:
                current_output_filename = output_filename
                logger.info("Creating regular consolidated report...")
            
            # PASS 1: Evaluate which tables have data (without creating sheets)
            non_empty_tables = {}
            for normalized_key, tables in all_tables_by_full_title.items():
                if self._evaluate_table_has_data(tables, transpose=is_transposed):
                    non_empty_tables[normalized_key] = tables
            
            logger.info(f"Found {len(non_empty_tables)} non-empty tables out of {len(all_tables_by_full_title)}")
            
            # Sort tables by page number for TOC-based ordering
            def get_sort_key(k):
                mapping = title_to_sheet_name.get(k, {})
                min_page = mapping.get('min_page', 9999)
                return (
                    min_page,
                    mapping.get('section', ''),
                    mapping.get('original_title', k)
                )
            
            sorted_keys = sorted(non_empty_tables.keys(), key=get_sort_key)
            
            # Assign sheet numbers ONLY to non-empty tables (no gaps)
            sheet_number = 1
            for normalized_key in sorted_keys:
                mapping = title_to_sheet_name.get(normalized_key, {})
                mapping['unique_sheet_name'] = str(sheet_number)
                title_to_sheet_name[normalized_key] = mapping
                sheet_number += 1
            
            # Create output file - transposed goes to transpose/ folder
            if is_transposed:
                transpose_dir = self.consolidate_dir.parent / "transpose"
                transpose_dir.mkdir(parents=True, exist_ok=True)
                output_path = transpose_dir / current_output_filename
            else:
                output_path = self.consolidate_dir / current_output_filename

            
            try:
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    # PASS 2: Create sheets with period type splitting
                    created_tables = {}
                    split_sheet_mapping = {}  # Track split sheets: sheet_name -> period_type
                    
                    for normalized_key in sorted_keys:
                        tables = non_empty_tables[normalized_key]
                        mapping = title_to_sheet_name.get(normalized_key, {})
                        sheet_name = mapping.get('unique_sheet_name')
                        if sheet_name:
                            # Use split method - may create multiple sheets (_1, _2, _3)
                            created_sheets = self._create_merged_table_sheets_with_split(
                                writer, 
                                sheet_name, 
                                tables, 
                                transpose=is_transposed,
                                title_to_sheet_name=title_to_sheet_name,
                                normalized_key=normalized_key
                            )
                            
                            if created_sheets:
                                # Track all created sheets
                                split_sheet_mapping.update(created_sheets)
                                
                                # If split occurred, update the mapping for Index
                                if len(created_sheets) > 1:
                                    # Multiple sheets created - mark original as "split parent"
                                    mapping['was_split'] = True
                                    mapping['split_sheets'] = list(created_sheets.keys())
                                    for split_sheet_name in created_sheets:
                                        created_tables[f"{normalized_key}::{created_sheets[split_sheet_name]}"] = tables
                                else:
                                    # Single sheet created - update mapping to use the actual sheet name
                                    # This fixes Index links when period-type splitting creates suffix like _3
                                    actual_sheet_name = list(created_sheets.keys())[0]
                                    mapping['unique_sheet_name'] = actual_sheet_name
                                    created_tables[normalized_key] = tables
                            else:
                                logger.debug(f"Sheet {sheet_name} was not created (no data after processing)")
                    
                    # Create TOC_Sheet with hierarchical section columns
                    self._create_toc_sheet(writer, created_tables, title_to_sheet_name)
                    
                    # Create Index with ONLY successfully created sheets
                    self._create_consolidated_index(writer, created_tables, title_to_sheet_name)
                
                # Add hyperlinks
                self._add_hyperlinks(output_path, created_tables, title_to_sheet_name)
                
                # Apply currency formatting to numeric cells
                self._apply_currency_format(output_path, created_tables, title_to_sheet_name)
                
                # Validate Index links match sheet names
                validation_result = self._validate_index_sheet_links(output_path)
                if not validation_result['valid']:
                    logger.warning(f"Validation failed - missing sheets: {validation_result['missing_sheets']}")
                else:
                    logger.info("Validation passed: all Index links have matching sheets")
                
                logger.info(f"Created {'transposed ' if is_transposed else ''}consolidated report at {output_path}")
                logger.info(f"  - Tables merged: {len(created_tables)}")
                logger.info(f"  - Sources: {len(xlsx_files)}")
                
                key = 'transposed' if is_transposed else 'regular'
                results[key] = {
                    'path': str(output_path),
                    'tables_merged': len(created_tables),
                    'sources_merged': len(xlsx_files),
                    'sheet_names': [title_to_sheet_name[k].get('unique_sheet_name', '') for k in created_tables.keys()],
                    'validation': validation_result
                }
                
            except Exception as e:
                logger.error(f"Failed to create {'transposed ' if is_transposed else ''}consolidated report: {e}", exc_info=True)
        
        # Create Consolidated_Index.xlsx with all Index sheets from source files
        index_result = self.consolidate_index_sheets(xlsx_files)
        
        # Return combined results (backward compat: use regular as main result)
        if 'regular' in results:
            result = results['regular'].copy()
            result['transposed_path'] = results.get('transposed', {}).get('path')
            result['index_path'] = index_result.get('path') if index_result else None
            return result
        return results.get('transposed', {})

    
    def consolidate_index_sheets(
        self,
        xlsx_files: List[str],
        output_filename: str = "Consolidated_Index.xlsx"
    ) -> dict:
        """
        Consolidate all Index sheets from source XLSX files into a single file.
        
        Each source file's Index sheet is copied as a separate sheet named:
        - Index_1 (from first source)
        - Index_2 (from second source)
        - etc.
        
        Note: Internal hyperlinks (links to sheets within the source file) will NOT work
        in the consolidated file since those target sheets don't exist in this file.
        External hyperlinks (URLs) will continue to work.
        
        Args:
            xlsx_files: List of source XLSX file paths
            output_filename: Output filename for consolidated index
            
        Returns:
            Dict with path, sources_count, and sheet_names
        """
        if not xlsx_files:
            logger.warning("No xlsx files provided for index consolidation")
            return {}
        
        output_path = self.consolidate_dir / output_filename
        
        try:
            # Sort files for consistent ordering (by filename)
            sorted_files = sorted(xlsx_files, key=lambda x: Path(x).name)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                sheet_names = []
                source_mapping = []
                
                for idx, xlsx_path in enumerate(sorted_files, start=1):
                    try:
                        # Read the Index sheet from source file
                        index_df = pd.read_excel(xlsx_path, sheet_name='Index')
                        
                        # Get source filename for reference (remove _tables suffix)
                        source_name = Path(xlsx_path).stem
                        
                        # Use source name as sheet name (e.g., "10k1224" or "10q0325")
                        # Remove common suffixes and sanitize for Excel
                        clean_name = source_name.replace('_tables', '')
                        # Excel sheet names max 31 chars, no special chars
                        sheet_name = clean_name[:31].replace('/', '_').replace('\\', '_')
                        
                        # Write Index directly to the sheet (sheet name already identifies source)
                        index_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        sheet_names.append(sheet_name)
                        source_mapping.append({
                            'sheet': sheet_name,
                            'source': source_name,
                            'rows': len(index_df)
                        })
                        
                        logger.debug(f"Added {sheet_name} from {source_name} ({len(index_df)} rows)")
                        
                    except Exception as e:
                        logger.warning(f"Could not read Index from {xlsx_path}: {e}")
                        continue
                
                # Create a summary sheet listing all sources
                if source_mapping:
                    summary_data = []
                    for item in source_mapping:
                        summary_data.append({
                            'Sheet': item['sheet'],
                            'Source File': item['source'],
                            'Table Count': item['rows']
                        })
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            logger.info(f"Created Consolidated_Index.xlsx at {output_path}")
            logger.info(f"  - Sources consolidated: {len(sheet_names)}")
            
            return {
                'path': str(output_path),
                'sources_count': len(sheet_names),
                'sheet_names': sheet_names,
                'source_mapping': source_mapping
            }
            
        except Exception as e:
            logger.error(f"Failed to create Consolidated_Index.xlsx: {e}", exc_info=True)
            return {}


    def _process_xlsx_file(
        self,
        xlsx_path: str,
        all_tables_by_full_title: Dict[str, List],
        title_to_sheet_name: Dict[str, dict]
    ) -> None:
        """Process a single xlsx file and add tables to collection."""
        # Read Index sheet to get metadata
        index_df = pd.read_excel(xlsx_path, sheet_name='Index')
        
        # DEDUPLICATE Index entries: Source files may have duplicate rows for same sheet
        # This happens when tables are split but Index is updated multiple times
        # Keep first occurrence to avoid inflating TableCount
        if 'Link' in index_df.columns:
            index_df['_link_clean'] = index_df['Link'].astype(str).str.replace('→ ', '', regex=False).str.strip()
            original_count = len(index_df)
            index_df = index_df.drop_duplicates(subset=['_link_clean'], keep='first')
            dedup_count = original_count - len(index_df)
            if dedup_count > 0:
                logger.debug(f"Deduplicated {dedup_count} Index entries from {Path(xlsx_path).name}")
            index_df = index_df.drop(columns=['_link_clean'])
        
        # Get all sheet names except Index
        xl = pd.ExcelFile(xlsx_path)
        table_sheets = [s for s in xl.sheet_names if s != 'Index']
        
        # Create mapping from Link to full Table Title
        link_to_full_title = {}
        for _, idx_row in index_df.iterrows():
            full_title = str(idx_row.get('Table Title', ''))
            link = str(idx_row.get('Link', ''))
            if link.startswith('→ '):
                link = link[2:]
            if link and full_title:
                link_to_full_title[link] = full_title
        
        for sheet_name in table_sheets:
            full_title = link_to_full_title.get(sheet_name, sheet_name)
            
            # Find metadata for this SPECIFIC sheet by matching Link/Table_ID
            # CRITICAL: Don't match by Table Title alone because multiple tables 
            # can have the same title (e.g., "Income Statement Information" for IS/WM/IM)
            # The Link column contains "→ {sheet_name}" format
            matching_rows = index_df[
                index_df['Link'].astype(str).str.replace('→ ', '', regex=False).str.strip() == sheet_name
            ]
            
            # Fallback: try Table_ID column if Link matching fails
            if len(matching_rows) == 0 and 'Table_ID' in index_df.columns:
                matching_rows = index_df[
                    index_df['Table_ID'].astype(str) == sheet_name
                ]
            
            if len(matching_rows) > 0:
                row = matching_rows.iloc[0]
                section = str(row.get('Section', '')) if 'Section' in row.index else ''
                source_name = str(row.get('Source', Path(xlsx_path).stem))
                
                # Get year/quarter from Index columns, or parse from filename
                year_val = row.get('Year', '')
                quarter_val = row.get('Quarter', '')
                
                # If year/quarter not in Index, parse from source filename
                if not year_val or str(year_val) in ['', 'nan', 'NaN']:
                    parsed_year, parsed_quarter = self._parse_year_quarter_from_source(source_name)
                    year_val = parsed_year
                    quarter_val = parsed_quarter
                
                metadata = {
                    'source': source_name,
                    'year': year_val,
                    'quarter': quarter_val,
                    'report_type': self._detect_report_type(source_name),
                    'section': section,
                    'page': row.get('PageNo', ''),
                }
            else:
                section = ''
                source_name = Path(xlsx_path).stem
                # Parse year/quarter from filename
                parsed_year, parsed_quarter = self._parse_year_quarter_from_source(source_name)
                
                metadata = {
                    'source': source_name,
                    'year': parsed_year,
                    'quarter': parsed_quarter,
                    'report_type': self._detect_report_type(source_name),
                    'section': '',
                    'page': '',
                }
            
            # Read table content (skip header rows)
            try:
                table_df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=None)
                
                # Extract Main Header (Level 0) from source metadata if present
                main_header = ''
                for idx in range(min(10, len(table_df))):  # Metadata is in first 10 rows
                    first_cell = str(table_df.iloc[idx, 0]) if pd.notna(table_df.iloc[idx, 0]) else ''
                    # Check for both old and new label names using MetadataLabels
                    if first_cell.startswith(MetadataLabels.COLUMN_HEADER_L1) or first_cell.startswith('Main Header:') or first_cell.startswith('Column Header (Level 0):'):
                        # Extract value after the colon
                        main_header = first_cell.split(':', 1)[1].strip() if ':' in first_cell else ''
                        break
                
                # Store main_header in metadata for consolidated output
                metadata['main_header'] = main_header
                
                # CRITICAL: Extract Category and Line Items fingerprints from ORIGINAL table_df
                # BEFORE slicing (which removes metadata rows at line 436)
                # This ensures IS/WM/IM Income Statements have different fingerprints
                original_category_fp = ""
                original_line_items_fp = ""
                
                for idx in range(min(10, len(table_df))):
                    first_cell = str(table_df.iloc[idx, 0]) if pd.notna(table_df.iloc[idx, 0]) else ''
                    if first_cell.startswith(MetadataLabels.CATEGORY_PARENT) or first_cell.startswith('Category'):
                        cat_value = first_cell.split(':', 1)[1].strip() if ':' in first_cell else ''
                        if cat_value:
                            cat_parts = [p.strip() for p in cat_value.split(',') if p.strip()]
                            if cat_parts:
                                original_category_fp = re.sub(r'[^a-z0-9]', '', cat_parts[0].lower())[:20]
                    elif first_cell.startswith(MetadataLabels.LINE_ITEMS) or first_cell.startswith('Line Items:'):
                        li_value = first_cell.split(':', 1)[1].strip() if ':' in first_cell else ''
                        if li_value:
                            li_parts = [p.strip().lower() for p in li_value.split(',') if p.strip()]
                            li_parts = [p for p in li_parts if p not in ['$ in millions', 'in millions', 'except per share data']]
                            if li_parts:
                                fp_items = li_parts[:5]
                                original_line_items_fp = '|'.join(re.sub(r'[^a-z0-9]', '', item)[:20] for item in fp_items)
                
                # Store fingerprints in metadata for use by subtables
                metadata['category_fingerprint'] = original_category_fp
                metadata['line_items_fingerprint'] = original_line_items_fp
                
                # Dynamically find where data starts by looking for "Source:" row
                # Individual xlsx structure:
                #   Row 0: ← Back to Index
                #   Row 1-6: Metadata (Row Header, Column Header, etc.)
                #   Row 7: Year(s) or blank
                #   Row 8: Table Title
                #   Row 9: Blank
                #   Row 10+: "Source: ..." followed by actual table data
                data_start_row = 0
                for idx, row in table_df.iterrows():
                    first_cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
                    if first_cell.startswith(MetadataLabels.SOURCES):
                        data_start_row = idx
                        break
                
                # If Source: not found, fallback to row 10 or after metadata
                if data_start_row == 0:
                    # Look for first row that looks like table data
                    for idx, row in table_df.iterrows():
                        first_cell = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
                        # Skip metadata patterns
                        if any(first_cell.startswith(p) for p in 
                               [MetadataLabels.BACK_LINK, 'Row Header', 'Column Header', 'Product', 
                                MetadataLabels.TABLE_TITLE, MetadataLabels.COLUMN_HEADER_L3, MetadataLabels.SOURCES]):
                            continue
                        if first_cell.strip() == '':
                            continue
                        # This looks like data - start here
                        data_start_row = idx
                        break
                
                # Slice from data start
                if data_start_row > 0 and len(table_df) > data_start_row:
                    table_df = table_df.iloc[data_start_row:].reset_index(drop=True)
                
                # Split embedded sub-tables (e.g., "Investments" + "Income (loss)" in same table)
                subtables = self._split_into_subtables(table_df, 0)
                
                # Process each sub-table as a separate table entry
                for subtable_idx, (subtable_df, _) in enumerate(subtables):
                    if subtable_df.empty or len(subtable_df) < 2:
                        continue
                    
                    # Compute row signature for this sub-table
                    raw_rows = subtable_df.iloc[:, 0].astype(str).tolist() if not subtable_df.empty else []
                    norm_rows = []
                    for r in raw_rows:
                        r_clean = r.strip()
                        if not r_clean: 
                            continue
                        r_lower = r_clean.lower()
                        if r_lower == 'row label': 
                            continue
                        if r_lower.startswith(MetadataLabels.SOURCES.lower()): 
                            continue
                        if r_lower.startswith('page '): 
                            continue
                        if r_lower.startswith('$ in'): 
                            continue
                        # Skip metadata row patterns (these vary between files)
                        if r_lower.startswith('row header'):
                            continue
                        if r_lower.startswith('column header'):
                            continue
                        if r_lower.startswith('product/entity'):
                            continue
                        if r_lower.startswith('table title'):
                            continue
                        if r_lower.startswith('year(s)'):
                            continue
                        if r_lower.startswith('year/quarter'):
                            continue
                        if r_lower.startswith('category'):
                            continue
                        if r_lower.startswith('line items'):
                            continue
                        if r_lower.startswith('← back'):
                            continue
                        # Skip embedded header patterns
                        if self._is_embedded_header_row(r_clean, []):
                            continue
                        norm_rows.append(self._normalize_row_label(r_clean))
                    
                    row_sig = "|".join(norm_rows)
                    
                    # Store normalized row labels for 80% overlap comparison
                    # This is used by TableGrouper.calculate_row_label_overlap()
                    metadata['row_labels'] = norm_rows.copy()
                    
                    # Skip if no valid rows found
                    if not row_sig:
                        continue
                    
                    # Extract date suffix from title for column header context
                    # e.g., "Borrowings at March 31, 2025" → date_code = "Q1-2025"
                    _, title_date_suffix, title_date_code = ExcelUtils.extract_title_date_suffix(full_title)
                    if title_date_code:
                        metadata['title_date_code'] = title_date_code  # e.g., "Q1-2025"
                        metadata['title_date_suffix'] = title_date_suffix  # e.g., "at March 31, 2025"
                    
                    # Create grouping key: Section + Title + [SubtableIdx] + Structure
                    # Note: normalized_title now has date suffix stripped for better merging
                    normalized_title = self._normalize_title_for_grouping(full_title)
                    subtable_suffix = f"::sub{subtable_idx}" if len(subtables) > 1 else ""
                    
                    # Use full row signature to ensure exact structure matching
                    # Tables with any row differences will remain separate
                    row_sig_full = "|".join(norm_rows)
                    
                    # Normalize section field for grouping
                    # Fixes OCR broken words like 'Manageme nt' -> 'Management'
                    normalized_section = ""
                    if section and section.strip():
                        # Fix OCR broken words and normalize whitespace
                        fixed_section = ExcelUtils.fix_ocr_broken_words(section.strip())
                        normalized_section = re.sub(r'\s+', ' ', fixed_section).lower()
                    
                    # Extract BOTH Category AND Line Items for structure-based merging
                    # Tables with same structure should merge (even from different sources)
                    # Tables with different structure should NOT merge (even from same source/title)
                    
                    # Use pre-extracted fingerprints from original table_df (before slicing)
                    # This ensures IS/WM/IM Income Statements maintain their distinct fingerprints
                    category_fingerprint = metadata.get('category_fingerprint', '')
                    line_items_fingerprint = metadata.get('line_items_fingerprint', '')
                    
                    # Fallback: Try to extract from subtable_df if not in metadata
                    # (This handles edge cases where metadata extraction failed)
                    if not category_fingerprint and not line_items_fingerprint:
                        for idx in range(min(10, len(subtable_df))):
                            first_cell = str(subtable_df.iloc[idx, 0]) if pd.notna(subtable_df.iloc[idx, 0]) else ''
                            if first_cell.startswith(MetadataLabels.CATEGORY_PARENT) or first_cell.startswith('Category'):
                                cat_value = first_cell.split(':', 1)[1].strip() if ':' in first_cell else ''
                                if cat_value:
                                    cat_parts = [p.strip() for p in cat_value.split(',') if p.strip()]
                                    if cat_parts:
                                        category_fingerprint = re.sub(r'[^a-z0-9]', '', cat_parts[0].lower())[:20]
                            elif first_cell.startswith(MetadataLabels.LINE_ITEMS) or first_cell.startswith('Line Items:'):
                                li_value = first_cell.split(':', 1)[1].strip() if ':' in first_cell else ''
                                if li_value:
                                    li_parts = [p.strip().lower() for p in li_value.split(',') if p.strip()]
                                    li_parts = [p for p in li_parts if p not in ['$ in millions', 'in millions', 'except per share data']]
                                    if li_parts:
                                        fp_items = li_parts[:5]
                                        line_items_fingerprint = '|'.join(re.sub(r'[^a-z0-9]', '', item)[:20] for item in fp_items)
                    
                    # Combined structure fingerprint: Category + Line Items
                    structure_fingerprint = f"{category_fingerprint}_{line_items_fingerprint}"
                    
                    # Extract header structure pattern for merge grouping
                    # Tables only merge if they have the same L1/L2/L3 structure
                    # This prevents merging L3_ONLY (point-in-time) with L2_L3 (period-based)
                    header_pattern = self._get_header_structure_pattern(subtable_df)
                    logger.debug(f"Table '{normalized_title}' has header pattern: {header_pattern}, structure: {structure_fingerprint}")
                    
                    # UPDATED: Grouping key uses 80% fuzzy matching on Section+Title
                    # Structure fingerprint and header_pattern must match exactly
                    # Section+Title can fuzzy match at 80% threshold
                    section_title_combo = f"{normalized_section}|{normalized_title}" if normalized_section else normalized_title
                    
                    # Exact key for structure+pattern (these must match exactly)
                    structure_key = f"{structure_fingerprint}::{header_pattern}"
                    
                    # Find best matching group using fuzzy Section+Title matching
                    # Use 80% threshold for title AND row label overlap (per merge condition spec)
                    # Structure fingerprint ensures only same-structure tables merge
                    best_match_key = self._find_fuzzy_matching_group(
                        all_tables_by_full_title,
                        section_title_combo,
                        structure_key,
                        threshold=0.80,
                        row_labels=norm_rows  # Pass row labels for 80% overlap check
                    )
                    
                    if best_match_key:
                        normalized_key = best_match_key
                    else:
                        # No fuzzy match found, create new key
                        normalized_key = f"{section_title_combo}::{structure_key}"
                    
                    # DEBUG: Log Reconciliations keys
                    if 'reconciliation' in normalized_title.lower():
                        logger.info(f"RECON KEY: sheet={sheet_name}, combo={section_title_combo[:40]}, struct={structure_key[:30]}, KEY={normalized_key[:80]}")
                    
                    # Debug logging removed - was: INC_STMT logger.warning

                    
                    all_tables_by_full_title[normalized_key].append({
                        'data': subtable_df,
                        'metadata': metadata,
                        'sheet_name': sheet_name,
                        'original_title': full_title if len(subtables) == 1 else f"{full_title} (Part {subtable_idx + 1})",
                        'section': section,
                        'row_sig': row_sig,
                        'subtable_idx': subtable_idx
                    })
                    
                    if normalized_key not in title_to_sheet_name:
                        clean_display = full_title.replace('→ ', '').strip() if full_title.startswith('→') else full_title
                        if len(subtables) > 1:
                            clean_display = f"{clean_display} (Part {subtable_idx + 1})"
                        if section and section.strip():
                            display_with_section = f"{section.strip()} - {clean_display}"
                        else:
                            display_with_section = clean_display
                        
                        # Parse page number for sorting (prefer 10-Q over 10-K for TOC order)
                        page_val = metadata.get('page', '')
                        report_type = metadata.get('report_type', '')
                        is_10q = '10-Q' in str(report_type).upper() or '10q' in str(metadata.get('source', '')).lower()
                        
                        try:
                            page_num = int(page_val) if page_val and str(page_val) != 'nan' else 9999
                        except (ValueError, TypeError):
                            page_num = 9999
                        
                        title_to_sheet_name[normalized_key] = {
                            'display_title': display_with_section,
                            'section': section.strip() if section else '',
                            'original_title': full_title,
                            'min_page': page_num,  # For TOC-based sorting
                            'has_10q_page': is_10q,  # Prefer 10-Q page numbers
                        }
                    else:
                        # Update min_page: prefer 10-Q pages over 10-K pages
                        page_val = metadata.get('page', '')
                        report_type = metadata.get('report_type', '')
                        is_10q = '10-Q' in str(report_type).upper() or '10q' in str(metadata.get('source', '')).lower()
                        
                        try:
                            page_num = int(page_val) if page_val and str(page_val) != 'nan' else 9999
                        except (ValueError, TypeError):
                            page_num = 9999
                        
                        existing_entry = title_to_sheet_name[normalized_key]
                        existing_has_10q = existing_entry.get('has_10q_page', False)
                        existing_min = existing_entry.get('min_page', 9999)
                        
                        # Priority: 10-Q page > 10-K page (regardless of page number)
                        if is_10q and not existing_has_10q:
                            # Replace with 10-Q page
                            existing_entry['min_page'] = page_num
                            existing_entry['has_10q_page'] = True
                        elif is_10q and existing_has_10q:
                            # Both are 10-Q, use min page
                            if page_num < existing_min:
                                existing_entry['min_page'] = page_num
                        elif not is_10q and not existing_has_10q:
                            # Both are 10-K, use min page
                            if page_num < existing_min:
                                existing_entry['min_page'] = page_num
                        # If existing is 10-Q and new is 10-K, keep existing
                    
            except Exception as e:
                logger.debug(f"Skipping sheet {sheet_name}: {e}")
    
    def _detect_report_type(self, source: str) -> str:
        """Detect report type from source filename. Delegates to ExcelUtils."""
        return ExcelUtils.detect_report_type(source)
    
    # NOTE: _normalize_row_label inherited from BaseExcelExporter
    # NOTE: _normalize_title_for_grouping - keeping as it may have custom logic
    
    def _normalize_title_for_grouping(self, title: str) -> str:
        """Normalize title for grouping. Delegates to ExcelUtils."""
        return ExcelUtils.normalize_title_for_grouping(title)
    
    # NOTE: _sanitize_sheet_name inherited from BaseExcelExporter
    # NOTE: _get_column_letter inherited from BaseExcelExporter
    # NOTE: _clean_currency_value inherited from BaseExcelExporter
    
    def _normalize_header_for_deduplication(self, header: str) -> str:
        """
        Normalize header for deduplication comparison.
        
        Delegates to TableGrouper for centralized logic.
        """
        return TableGrouper.normalize_header_for_deduplication(header)
    
    def _evaluate_table_has_data(self, tables: List[Dict[str, Any]], transpose: bool = False) -> bool:
        """
        Evaluate if merged tables would have data (without creating a sheet).
        
        This mirrors the logic in _create_merged_table_sheet but only checks 
        if columns with data exist.
        
        Args:
            tables: List of table dictionaries to merge
            transpose: If True, check for transpose format
            
        Returns:
            True if merged table would have data, False otherwise
        """
        if not tables:
            return False
        
        all_columns = {}
        
        for table_info in tables:
            df = table_info['data']
            meta = table_info['metadata']
            
            if df.empty or len(df.columns) < 2:
                continue
            
            # Check for data columns (simplified version of _create_merged_table_sheet logic)
            l1_row, l2_row, data_start = self._find_header_rows(df)
            
            for col_idx in range(1, len(df.columns)):
                # Get column header - try multiple sources
                header = ''
                
                # Try L2 row first (usually years/dates)
                if l2_row is not None and l2_row < len(df):
                    val = df.iloc[l2_row, col_idx]
                    if pd.notna(val) and str(val).strip() and str(val).strip() != 'nan':
                        header = ExcelUtils.clean_year_string(str(val))
                
                # Fallback: Try L1 row (period descriptions like "Three Months Ended")
                if not header and l1_row is not None and l1_row < len(df):
                    val = df.iloc[l1_row, col_idx]
                    if pd.notna(val) and str(val).strip() and str(val).strip() != 'nan':
                        header = str(val).strip()
                
                # Last fallback: Use DataFrame column name if not generic
                if not header:
                    col_name = str(df.columns[col_idx])
                    if col_name and not col_name.startswith('Unnamed') and col_name != str(col_idx):
                        header = col_name
                
                # Final fallback: Col_N
                if not header:
                    header = f"Col_{col_idx}"
                
                # Check if column has actual data values
                col_values = df.iloc[:, col_idx].tolist()
                non_empty = [v for v in col_values if pd.notna(v) and str(v).strip() and str(v).strip() != 'nan']
                # Filter out header-like values
                non_empty = [x for x in non_empty if str(x).strip() != header]
                non_empty = [x for x in non_empty if not (
                    str(x).startswith('$') and 'in' in str(x).lower() or
                    str(x).isdigit() and len(str(x)) == 4 and 2000 <= int(str(x)) <= 2099
                )]
                
                if non_empty:
                    all_columns[header] = True
        
        return len(all_columns) > 0
    
    def _validate_index_sheet_links(self, output_path: Path) -> dict:
        """
        Verify all Index links have matching sheets.
        
        Args:
            output_path: Path to the consolidated xlsx file
            
        Returns:
            Dict with missing_sheets, orphan_sheets, and valid flag
        """
        try:
            
            wb = load_workbook(output_path)
            sheet_names = set(wb.sheetnames) - {'Index'}
            
            index_df = pd.read_excel(output_path, sheet_name='Index')
            
            # Extract link values (removing → prefix)
            link_values = set()
            for link in index_df['Link'].dropna().astype(str).tolist():
                clean_link = link.replace('→ ', '').strip()
                if clean_link:
                    link_values.add(clean_link)
            
            missing_sheets = link_values - sheet_names
            orphan_sheets = sheet_names - link_values
            
            return {
                'missing_sheets': missing_sheets,
                'orphan_sheets': orphan_sheets,
                'valid': len(missing_sheets) == 0,
                'total_links': len(link_values),
                'total_sheets': len(sheet_names)
            }
            
        except Exception as e:
            logger.error(f"Validation error: {e}")
            return {
                'missing_sheets': set(),
                'orphan_sheets': set(),
                'valid': False,
                'error': str(e)
            }

    
    def _is_embedded_header_row(self, row_label: str, row_values: list) -> bool:
        """
        Detect if a row is an embedded sub-table header rather than a data row.
        
        Delegates to TableDetector for centralized logic.
        """
        return TableDetector.is_embedded_header_row(row_label, row_values)
    
    def _find_fuzzy_matching_group(
        self,
        all_tables_by_full_title: Dict[str, List],
        section_title_combo: str,
        structure_key: str,
        threshold: float = 0.80,
        row_labels: Optional[List[str]] = None
    ) -> Optional[str]:
        """
        Find an existing group key that fuzzy-matches using 80% conditions.
        
        Delegates to TableGrouper for centralized logic.
        Checks both title similarity (80%) and row label overlap (80%).
        """
        return TableGrouper.find_fuzzy_matching_group(
            all_tables_by_full_title, section_title_combo, structure_key, threshold, row_labels
        )
    
    def _split_into_subtables(self, df: pd.DataFrame, data_start_idx: int = 0) -> List[Tuple[pd.DataFrame, int]]:
        """
        Split a DataFrame at embedded header rows into separate sub-tables.
        
        Delegates to TableDetector for centralized logic.
        """
        return TableDetector.split_into_subtables(df, data_start_idx)
    
    def _get_header_structure_pattern(self, df: pd.DataFrame) -> str:
        """
        Detect the header structure pattern of a table for merge grouping.
        
        Delegates to TableDetector for centralized logic.
        """
        return TableDetector.get_header_structure_pattern(df)
    
    def _find_header_rows(self, df: pd.DataFrame) -> Tuple[int, int, int]:
        """
        Dynamically find L1 header, L2 header, and data start row indices.
        
        Delegates to TableDetector for centralized logic.
        """
        return TableDetector.find_header_rows(df, VALID_YEAR_RANGE)
    
    def _create_consolidated_index(
        self,
        writer: pd.ExcelWriter,
        tables_by_full_title: Dict[str, List[Dict[str, Any]]],
        title_to_sheet_name: Dict[str, dict]
    ) -> None:
        """Create master index for consolidated report."""
        index_data = []
        
        for normalized_key, tables in tables_by_full_title.items():
            sources = set()
            years = set()
            quarters = set()
            report_types = set()
            sections = set()
            source_refs = []  # List of "Q1,2025_page" references
            
            for t in tables:
                meta = t['metadata']
                section = t.get('section', '') or meta.get('section', '')
                if section and str(section).strip() and str(section).lower() != 'nan':
                    sections.add(str(section).strip())
                if meta.get('source'):
                    sources.add(str(meta['source']))
                if meta.get('year') and str(meta.get('year')) not in ['nan', 'NaN', '']:
                    years.add(str(int(meta['year'])) if isinstance(meta.get('year'), float) else str(meta.get('year')))
                if meta.get('quarter') and str(meta.get('quarter')) not in ['nan', 'NaN', '']:
                    quarters.add(str(meta['quarter']))
                if meta.get('report_type'):
                    report_types.add(str(meta['report_type']))
                
                # Build source reference: Q1,2025_page or Q4,2024_page
                year_val = str(int(meta['year'])) if isinstance(meta.get('year'), float) else str(meta.get('year', ''))
                quarter_val = str(meta.get('quarter', ''))
                page_val = str(meta.get('page', ''))
                
                if quarter_val and year_val:
                    ref = f"{quarter_val},{year_val}"
                elif year_val:
                    ref = f"Q4,{year_val}"
                else:
                    ref = str(meta.get('source', 'Unknown'))[:10]
                
                if page_val and page_val != 'nan':
                    ref += f"_p{page_val}"
                
                if ref and ref not in source_refs:
                    source_refs.append(ref)
            
            # Filter out NaN values
            quarters = {q for q in quarters if q and q.lower() != 'nan'}
            years = {y for y in years if y and y.lower() != 'nan'}
            sources = {s for s in sources if s and s.lower() != 'nan'}
            report_types = {r for r in report_types if r and r.lower() != 'nan'}
            sections = {s for s in sections if s and s.lower() != 'nan'}
            
            mapping = title_to_sheet_name.get(normalized_key, {})
            display_title = mapping.get('display_title', normalized_key)
            original_title = mapping.get('original_title', display_title)
            sheet_name = mapping.get('unique_sheet_name', self._sanitize_sheet_name(normalized_key))
            section_str = mapping.get('section', '') or ', '.join(sorted(sections))
            
            # Format source references (limit to 5 to avoid very long strings)
            source_refs_str = ', '.join(source_refs[:5])
            if len(source_refs) > 5:
                source_refs_str += f'... (+{len(source_refs) - 5})'
            
            # Fix OCR broken words and normalize whitespace in Section and Title
            section_clean = ExcelUtils.fix_ocr_broken_words(section_str) if section_str else ''
            section_clean = re.sub(r'\s+', ' ', section_clean).strip() if section_clean else '-'
            
            title_text = original_title if original_title else display_title
            title_clean = ExcelUtils.fix_ocr_broken_words(title_text) if title_text else ''
            title_clean = re.sub(r'\s+', ' ', title_clean).strip() if title_clean else ''
            
            # FIX: Count UNIQUE source files, not total table entries
            unique_source_files = set()
            for table in tables:
                source = table.get('source', '') or table.get('source_file', '')
                if source:
                    unique_source_files.add(source)
            table_count = len(unique_source_files) if unique_source_files else len(tables)
            
            index_data.append({
                '#': 0,  # Will be set after sorting
                'Section': section_clean,
                'Table Title': title_clean,
                'TableCount': table_count,  # Now counts unique source files
                'Sources': source_refs_str,
                'Report Types': ', '.join(sorted(report_types)),
                'Years': ', '.join(sorted(years, reverse=True)),
                'Quarters': ', '.join(sorted(quarters)),
                'Link': '',  # Will be set after sorting
                '_normalized_key': normalized_key  # Hidden key for sheet mapping
            })
        
        # Validation check: Log debug for tables with TableCount > source_count
        # This helps identify potential merge issues for manual inspection
        # Calculate source_count from the tables we just processed
        all_sources = set()
        for entry in index_data:
            sources_str = entry.get('Sources', '')
            for src in sources_str.split(','):
                src = src.strip()
                if src and not src.startswith('(+'):
                    all_sources.add(src.split('_')[0])  # Get Q1,2025 part
        source_count = len(all_sources) if all_sources else 4  # Default to 4 if can't determine
        
        # After table splitting in process_advanced, TC > source_count is expected
        # Tables with multiple period types get split (e.g., 66_1, 66_2, 66_3, 66_4)
        # Changed from warning to debug since this is expected behavior
        if source_count > 0:
            for entry in index_data:
                if entry['TableCount'] > source_count * 4:  # Only log if > 4x expected (truly unusual)
                    logger.debug(
                        f"TableCount validation: '{entry['Table Title'][:50]}' has "
                        f"TC={entry['TableCount']} (source_count={source_count}). "
                        f"This can occur due to subtable splits."
                    )
        
        df = pd.DataFrame(index_data)
        # Keep the order from input (already sorted by page number in merge_processed_files)
        # Set # to row number for display (1-based)
        df['#'] = range(1, len(df) + 1)
        
        # Set Link to use the PRE-ASSIGNED sheet name from title_to_sheet_name
        # (NOT re-numbering based on sorted position)
        df['Link'] = df['_normalized_key'].map(
            lambda k: title_to_sheet_name.get(k, {}).get('unique_sheet_name', '')
        )
        
        # Drop the hidden column before saving
        df = df.drop(columns=['_normalized_key'])
        df.to_excel(writer, sheet_name='Index', index=False)

        
        # Auto-adjust column widths (using settings for configurability)
        worksheet = writer.sheets['Index']
        for idx, col in enumerate(df.columns):
            col_letter = self._get_column_letter(idx)
            max_content_len = df[col].astype(str).map(len).max() if len(df) > 0 else 0
            header_len = len(col)
            calculated_len = max(max_content_len, header_len) + 2
            
            # Use configured width from settings, or default
            configured_width = settings.INDEX_COLUMN_WIDTHS.get(
                col, settings.INDEX_COLUMN_WIDTHS.get('_default', 20)
            )
            max_len = min(calculated_len, configured_width) if col != '#' else configured_width
            
            worksheet.column_dimensions[col_letter].width = max_len
    
    def _create_toc_sheet(
        self,
        writer: pd.ExcelWriter,
        tables_by_full_title: Dict[str, List[Dict[str, Any]]],
        title_to_sheet_name: Dict[str, dict]
    ) -> None:
        """
        Create TOC sheet with hierarchical section columns.
        
        Columns: Table of Contents | Main Section | Section1 | Section2 | Section3 | Section4 | Table Title | Sheet | Sources
        
        Uses SectionHierarchyTracker for pattern-based classification.
        """
        from src.infrastructure.extraction.helpers.docling_helper import SectionHierarchyTracker
        
        # Initialize tracker with config-based patterns
        tracker = SectionHierarchyTracker()
        
        toc_data = []
        
        # First pass: collect all section names for pre-scan
        all_sections = []
        for normalized_key, tables in tables_by_full_title.items():
            mapping = title_to_sheet_name.get(normalized_key, {})
            section_str = mapping.get('section', '')
            
            if not section_str:
                for t in tables:
                    section = t.get('section', '') or t.get('metadata', {}).get('section', '')
                    if section and str(section).strip() and str(section).lower() != 'nan':
                        section_str = str(section).strip()
                        break
            
            if section_str:
                all_sections.append((1, section_str))  # Page 1 as placeholder
        
        # Pre-scan to identify repeating headers
        tracker.pre_scan_headers(all_sections)
        
        # Second pass: process each table group
        for normalized_key, tables in tables_by_full_title.items():
            mapping = title_to_sheet_name.get(normalized_key, {})
            display_title = mapping.get('display_title', normalized_key)
            original_title = mapping.get('original_title', display_title)
            sheet_name = mapping.get('unique_sheet_name', str(len(toc_data) + 1))
            section_str = mapping.get('section', '')
            
            # Extract section from tables if not in mapping
            if not section_str:
                for t in tables:
                    section = t.get('section', '') or t.get('metadata', {}).get('section', '')
                    if section and str(section).strip() and str(section).lower() != 'nan':
                        section_str = str(section).strip()
                        break
            
            # Use tracker to classify section
            if section_str:
                hierarchy = tracker.process_header(section_str.strip(), 1)
            else:
                hierarchy = tracker.get_current_hierarchy()
            
            # Get clean title
            title_text = original_title if original_title else display_title
            title_clean = ExcelUtils.fix_ocr_broken_words(str(title_text)) if title_text else ''
            title_clean = re.sub(r'\s+', ' ', title_clean).strip()
            title_clean = re.sub(r'^Table Title:\s*', '', title_clean, flags=re.IGNORECASE)
            
            # Get sources
            source_refs = []
            for t in tables:
                meta = t.get('metadata', {})
                year_val = str(int(meta['year'])) if isinstance(meta.get('year'), float) else str(meta.get('year', ''))
                quarter_val = str(meta.get('quarter', ''))
                page_val = str(meta.get('page', ''))
                
                if quarter_val and year_val:
                    ref = f"{quarter_val},{year_val}"
                elif year_val:
                    ref = f"Q4,{year_val}"
                else:
                    ref = str(meta.get('source', ''))[:10]
                
                if page_val and page_val != 'nan':
                    ref += f"_p{page_val}"
                
                if ref and ref not in source_refs:
                    source_refs.append(ref)
            
            sources_str = ', '.join(source_refs[:3])
            if len(source_refs) > 3:
                sources_str += '...'
            
            toc_data.append({
                'Table of Contents': hierarchy.get('toc', '') or '-',
                'Main Section': hierarchy.get('main', '') or '-',
                'Section1': hierarchy.get('section1', '') or '-',
                'Section2': hierarchy.get('section2', '') or '-',
                'Section3': hierarchy.get('section3', '') or '-',
                'Section4': hierarchy.get('section4', '') or '-',
                'Table Title': title_clean,
                'Sheet': sheet_name,
                'Sources': sources_str
            })
        
        # Create DataFrame and write to sheet
        df = pd.DataFrame(toc_data)
        
        # Sort by hierarchy
        df = df.sort_values(
            by=['Table of Contents', 'Main Section', 'Section1', 'Section2', 'Section3', 'Section4', 'Table Title'],
            key=lambda x: x.astype(str).str.lower().fillna('')
        )
        
        df.to_excel(writer, sheet_name='TOC', index=False)
        
        # Auto-adjust column widths
        worksheet = writer.sheets['TOC']
        for idx, col in enumerate(df.columns):
            col_letter = self._get_column_letter(idx)
            max_content_len = df[col].astype(str).map(len).max() if len(df) > 0 else 0
            header_len = len(col)
            calculated_len = max(max_content_len, header_len) + 2
            max_len = min(calculated_len, 50)
            worksheet.column_dimensions[col_letter].width = max_len
        
        logger.info(f"Created TOC sheet with {len(df)} entries")
    
    def _create_merged_table_sheets_with_split(
        self,
        writer: pd.ExcelWriter,
        base_sheet_name: str,
        tables: List[Dict[str, Any]],
        transpose: bool = False,
        title_to_sheet_name: Dict[str, dict] = None,
        normalized_key: Optional[str] = None
    ) -> Dict[str, str]:
        """
        Create merged sheets with period type splitting.
        
        If tables have columns with multiple period types (e.g., Qn-YYYY and Qn-QTD-YYYY),
        split into separate sheets:
        - {sheet_name}_1 for point_in_time (Qn-YYYY)
        - {sheet_name}_2 for period_based (Qn-QTD-YYYY, Qn-YTD-YYYY)
        - {sheet_name}_3 for annual (YYYY)
        
        Args:
            writer: Excel writer
            base_sheet_name: Base sheet name (e.g., "5")
            tables: List of table dicts to merge
            transpose: If True, transpose the output
            title_to_sheet_name: Mapping dict to update with split sheet info
            normalized_key: The normalized key for this table group
            
        Returns:
            Dict mapping created sheet names to their period types
        """
        if not tables:
            return {}
        
        # First, collect all column headers by running a pre-analysis pass
        all_column_headers = set()
        
        for table_info in tables:
            df = table_info['data']
            if df.empty:
                continue
            
            # Find header rows
            l1_row, l2_row, data_start = self._find_header_rows(df)
            
            # Collect column headers from L2 row (where period codes are)
            header_row_idx = l2_row if l2_row is not None else (l1_row if l1_row is not None else 0)
            if header_row_idx < len(df):
                for col_idx in range(1, len(df.columns)):
                    val = df.iloc[header_row_idx, col_idx]
                    if pd.notna(val) and str(val).strip():
                        all_column_headers.add(str(val).strip())
        
        # Classify headers by period type (filter out corrupt headers)
        period_types_found = set()
        valid_headers = set()
        for header in all_column_headers:
            # Skip corrupt numeric headers (e.g., "1212447" which is data, not a header)
            if not PeriodTypeDetector.is_valid_period_header(header):
                logger.debug(f"Sheet {base_sheet_name}: Filtered corrupt header: {header}")
                continue
            
            valid_headers.add(header)
            period_type = PeriodTypeDetector.classify_header(header)
            if period_type != 'other':
                period_types_found.add(period_type)
        
        # If NO period types found at all, this is a non-period table (e.g., SEC registration)
        # Don't merge - just use the most recent source table
        if len(period_types_found) == 0:
            logger.debug(f"Sheet {base_sheet_name}: No period codes found, using single source (no merge)")
            # Use only the first (most recent) table
            single_table = [tables[0]] if tables else tables
            created = self._create_merged_table_sheet(writer, base_sheet_name, single_table, transpose)
            if created:
                return {base_sheet_name: 'no_period'}
            return {}
        
        # If only one period type, create single sheet without suffix
        if len(period_types_found) == 1:
            created = self._create_merged_table_sheet(writer, base_sheet_name, tables, transpose)
            if created:
                return {base_sheet_name: list(period_types_found)[0]}
            return {}
        
        # Multiple period types - need to split
        logger.info(f"Sheet {base_sheet_name}: Splitting by period types: {period_types_found}")
        
        created_sheets = {}
        
        # For each period type, filter the column data and create separate sheet
        for period_type in sorted(period_types_found):
            suffix = PeriodTypeDetector.get_suffix(period_type)
            split_sheet_name = f"{base_sheet_name}{suffix}"
            
            # Ensure sheet name <= 31 chars
            if len(split_sheet_name) > 31:
                split_sheet_name = f"{base_sheet_name[:28]}{suffix}"
            
            # Create the sheet with period type filter
            # We pass the period_type to filter columns during sheet creation
            created = self._create_merged_table_sheet(
                writer, 
                split_sheet_name, 
                tables, 
                transpose,
                period_type_filter=period_type
            )
            
            if created:
                created_sheets[split_sheet_name] = period_type
                
                # Update title_to_sheet_name mapping for Index
                if title_to_sheet_name and normalized_key:
                    label = PeriodTypeDetector.get_label(period_type)
                    original_mapping = title_to_sheet_name.get(normalized_key, {})
                    
                    # Create new entry for split sheet
                    split_key = f"{normalized_key}::{period_type}"
                    title_to_sheet_name[split_key] = {
                        'unique_sheet_name': split_sheet_name,
                        'display_title': f"{original_mapping.get('display_title', '')} {label}",
                        'section': original_mapping.get('section', ''),
                        'original_title': original_mapping.get('original_title', ''),
                        'period_type': period_type,
                        'is_split_sheet': True,
                        'parent_key': normalized_key
                    }
        
        return created_sheets
    
    def _create_merged_table_sheet(
        self,
        writer: pd.ExcelWriter,
        title: str,
        tables: List[Dict[str, Any]],
        transpose: bool = False,
        period_type_filter: Optional[str] = None
    ) -> bool:
        """Create merged sheet with tables from multiple sources.
        
        Args:
            writer: Excel writer
            title: Sheet name
            tables: List of table dicts to merge
            transpose: If True, transpose the output
            period_type_filter: If provided, only include columns matching this period type
                                ('point_in_time', 'period_based', 'annual')
        
        Returns:
            True if sheet has data, False if empty (sheet not created)
        """
        if not tables:
            return False
        
        all_columns = {}
        row_labels = []
        normalized_row_labels = {}
        # Track section headers for disambiguation of duplicate row labels
        # (e.g., "Institutional Securities" appears under multiple sections)
        label_to_section = {}  # Maps normalized label to its section
        
        for table_info in tables:
            df = table_info['data']
            meta = table_info['metadata']
            
            if df.empty or len(df.columns) < 2:
                continue
            
            # First, find the "Row Label" header row to know where data starts
            # This ensures we only collect actual data row labels, not headers
            row_label_row_idx = None
            for idx, val in enumerate(df.iloc[:, 0].tolist()):
                if pd.notna(val) and str(val).strip() == 'Row Label':
                    row_label_row_idx = idx
                    break
            
            # If "Row Label" not found, find data start dynamically
            if row_label_row_idx is None:
                # Look for Source(s) row, then skip to first data row
                source_row_idx = None
                unit_row_idx = None
                first_data_row_idx = None
                
                for idx, val in enumerate(df.iloc[:, 0].tolist()):
                    val_str = str(val).strip() if pd.notna(val) else ''
                    val_lower = val_str.lower()
                    
                    # Find Source(s) row
                    if val_str.startswith(MetadataLabels.SOURCES):
                        source_row_idx = idx
                    # Find unit header row ($ in millions, etc.)
                    elif val_lower.startswith('$ in'):
                        unit_row_idx = idx
                    # First non-metadata row after Source(s) is the data start
                    elif source_row_idx is not None and val_str:
                        # Skip blank rows and metadata patterns
                        if not val_str.startswith(('←', 'Category', 'Line Items', 'Product', 'Column Header', 'Year', 'Table Title')):
                            if not val_lower.startswith('$ in'):
                                first_data_row_idx = idx
                                break
                
                # Determine data_start_for_labels
                if first_data_row_idx is not None:
                    # We found the first data row - use its index directly
                    data_start_for_labels = first_data_row_idx
                elif unit_row_idx is not None:
                    # Data starts after unit row
                    data_start_for_labels = unit_row_idx + 1
                else:
                    # Fallback to _find_header_rows
                    _, _, data_start_fallback = self._find_header_rows(df)
                    data_start_for_labels = data_start_fallback
            else:
                # Data rows start AFTER the "Row Label" row
                data_start_for_labels = row_label_row_idx + 1
            
            # First pass: identify section headers (rows with label but no data)
            current_section = ''
            
            # Collect row labels starting from data_start_for_labels
            for row_idx in range(data_start_for_labels, len(df)):
                val = df.iloc[row_idx, 0]
                if pd.notna(val):
                    val_str = str(val).strip()
                    # Skip metadata patterns that might appear in data
                    if val_str.startswith((MetadataLabels.SOURCES, 'Page ')):
                        continue
                    if val_str.startswith('$ in') or val_str.startswith('$,'):
                        continue
                    if val_str.startswith('← Back') or val_str == 'Row Label':
                        continue
                    if val_str.startswith(('Category', 'Line Items', 'Product', 'Period', 'Year', 'Table Title', 'Column Header')):
                        continue
                    
                    # Skip embedded header rows (sub-table headers like "Three Months Ended")
                    row_values = df.iloc[row_idx].tolist() if row_idx < len(df) else []
                    if self._is_embedded_header_row(val_str, row_values):
                        continue
                    
                    # Check if this is a section header (has label but no data values)
                    has_data = any(pd.notna(row_values[i]) and str(row_values[i]).strip() 
                                   and str(row_values[i]).strip() != 'nan'
                                   for i in range(1, len(row_values)))
                    
                    if not has_data:
                        # This is a section header - update current section and add to row_labels
                        current_section = val_str
                        norm_val = self._normalize_row_label(val)
                        if norm_val and norm_val not in normalized_row_labels:
                            normalized_row_labels[norm_val] = val_str
                            # Store the key (norm_val) in row_labels, not the display label
                            row_labels.append(norm_val)
                            label_to_section[norm_val] = ''  # Section headers have no parent section
                    else:
                        # This is a data row - use NORMALIZED section prefix for unique matching
                        # Normalizing the section ensures consistent matching across source files
                        # (handles OCR variations like "Institutional Securities" vs "Institutional securities")
                        norm_val = self._normalize_row_label(val)
                        norm_section = self._normalize_row_label(current_section) if current_section else ''
                        # Create a normalized section-prefixed key for disambiguation
                        section_key = f"{norm_section}::{norm_val}" if norm_section else norm_val
                        
                        if section_key not in normalized_row_labels:
                            normalized_row_labels[section_key] = val_str
                            # Store the section_key in row_labels for unique identification
                            row_labels.append(section_key)
                            label_to_section[section_key] = current_section
            
            # Extract column data with ORIGINAL headers from source
            # Use the same data_start we found for row labels to ensure consistency
            l1_row, l2_row, _ = self._find_header_rows(df)
            # Use data_start_for_labels to sync with row label collection
            data_start = data_start_for_labels
            
            # Get Level 0 header (Main Header) from table metadata
            # This comes from the source file's "Main Header:" row
            level0_header_from_meta = meta.get('main_header', '')
            
            # Track spanning L1 headers for cell merging later
            current_l1_header = ''
            l1_header_spans = {}  # {header: [start_col, end_col]}
            
            # Track current section for this specific table's data extraction
            current_section_for_data = ''
            
            for col_idx in range(1, len(df.columns)):
                # Get Level 1 header (spanning header row)
                level1_header = ''
                if l1_row is not None and l1_row < len(df):
                    val = df.iloc[l1_row, col_idx]
                    if pd.notna(val) and str(val).strip() and str(val).strip() != 'nan':
                        val_str = ExcelUtils.clean_year_string(val)
                        # L1 headers are descriptive text, not currency values
                        if not (val_str.startswith('$') and any(c.isdigit() for c in val_str)):
                            level1_header = val_str
                            current_l1_header = val_str
                            # Start tracking this spanning header
                            if val_str not in l1_header_spans:
                                l1_header_spans[val_str] = [col_idx, col_idx]
                    elif current_l1_header:
                        # Empty cell - extend the current spanning header
                        level1_header = current_l1_header
                        l1_header_spans[current_l1_header][1] = col_idx
                
                # Get Level 2 header (sub-header row with dates/years)
                level2_header = ''
                if l2_row is not None and l2_row < len(df):
                    val = df.iloc[l2_row, col_idx]
                    if pd.notna(val) and str(val).strip() and str(val).strip() != 'nan':
                        val_str = ExcelUtils.clean_year_string(val)
                        level2_header = val_str
                
                # Build the full header for display
                # Priority: If both exist, combine them. If only L2, use L2. 
                header_parts = []
                if level1_header and level1_header not in ['%', '%Change', 'nan', '$', '$ in millions']:
                    header_parts.append(level1_header)
                if level2_header and level2_header not in ['nan', '', '$ in millions']:
                    header_parts.append(level2_header)
                
                if header_parts:
                    header = ' '.join(header_parts)
                else:
                    # Fallback 1: Use DataFrame column name if not generic
                    col_name = str(df.columns[col_idx])
                    if col_name and not col_name.startswith('Unnamed') and col_name != str(col_idx):
                        header = col_name
                    else:
                        # Fallback 2: Use column index as header (will be deduplicated later)
                        header = f"Column {col_idx}"
                
                # Final cleanup
                header = ExcelUtils.clean_year_string(header)
                
                # Skip corrupt headers (large numeric values that are data, not headers)
                if not PeriodTypeDetector.is_valid_period_header(header):
                    logger.debug(f"Skipping corrupt header in column {col_idx}: {header}")
                    continue
                
                sort_key = DateUtils.parse_date_from_header(header)
                
                # Normalize header for deduplication comparison
                # This handles case, month abbreviations, punctuation, and decimal years
                norm_header = self._normalize_header_for_deduplication(header)
                
                # Get row labels and values starting from data_start (not from row 0!)
                # This prevents header row values from being extracted as data
                source_row_labels = df.iloc[data_start:, 0].tolist()
                col_values = df.iloc[data_start:, col_idx].tolist()
                
                row_data_map = {}
                current_section_for_data = ''  # Track current section during data extraction
                
                for rel_idx, (row_label, value) in enumerate(zip(source_row_labels, col_values)):
                    row_idx = data_start + rel_idx  # Actual row index in DataFrame
                    if pd.notna(row_label):
                        row_label_str = str(row_label).strip()
                        
                        # Get full row values to detect embedded headers
                        row_values = df.iloc[row_idx].tolist() if row_idx < len(df) else []
                        
                        # Skip embedded header rows (sub-table headers like "Three Months Ended")
                        if self._is_embedded_header_row(row_label, row_values):
                            continue
                        
                        # Check if this row has data values (non-empty cells after the label)
                        has_data = any(pd.notna(row_values[i]) and str(row_values[i]).strip() 
                                       and str(row_values[i]).strip() != 'nan'
                                       for i in range(1, len(row_values)))
                        
                        norm_row = self._normalize_row_label(row_label)
                        
                        if not has_data:
                            # This is a section header - update current section
                            current_section_for_data = row_label_str
                            if norm_row:
                                row_data_map[norm_row] = value  # Section headers stay as simple keys
                        else:
                            # This is a data row - use NORMALIZED section prefix for matching
                            # Normalizing ensures consistent matching across source files
                            if norm_row:
                                norm_section = self._normalize_row_label(current_section_for_data) if current_section_for_data else ''
                                section_key = f"{norm_section}::{norm_row}" if norm_section else norm_row
                                row_data_map[section_key] = value
                
                non_empty_data = [v for v in row_data_map.values() if pd.notna(v) and str(v).strip() and str(v).strip() != 'nan']
                non_empty_data = [x for x in non_empty_data if str(x).strip() != header]
                # Exclude header row values (units like "$ in billions", "$ in millions", years)
                # These are not actual data
                non_empty_data = [x for x in non_empty_data if not (
                    str(x).startswith('$') and 'in' in str(x).lower() or  # $ in millions/billions
                    str(x).isdigit() and len(str(x)) == 4 and 2000 <= int(str(x)) <= 2099  # Years like 2025
                )]
                
                if not non_empty_data:
                    continue
                
                # Period type filter - skip columns not matching the filter
                if period_type_filter:
                    col_period_type = PeriodTypeDetector.classify_header(header)
                    if col_period_type != period_type_filter and col_period_type != 'other':
                        # Column has a different period type - skip it
                        continue
                
                # Check for duplicates - skip if SAME header AND same data already exists
                # This prevents multiple "At December 31, 2024" columns with identical data
                is_duplicate = False
                col_data_str = str(sorted([(k, str(v).strip()) for k, v in row_data_map.items()]))
                
                # Check if this header already exists - if so, merge data instead of duplicating
                matching_key = None
                for existing_key, existing_val in all_columns.items():
                    existing_data_str = str(sorted([(k, str(v).strip()) for k, v in existing_val.get('row_data_map', {}).items()]))
                    existing_header = existing_val.get('header', '').lower().strip()
                    
                    # Check if headers match (normalized comparison)
                    if existing_header == norm_header:
                        # If data is also identical, skip entirely (true duplicate)
                        if col_data_str == existing_data_str:
                            is_duplicate = True
                            logger.debug(f"Removing exact duplicate column: '{header}' (same as '{existing_val.get('header')}')")
                            break
                        else:
                            # Headers match but data differs - merge data into existing column
                            matching_key = existing_key
                            break
                
                # If we found a matching header with different data, merge
                if matching_key and not is_duplicate:
                    existing_data = all_columns[matching_key].get('row_data_map', {})
                    # Merge: fill empty values from new data
                    for row_key, new_val in row_data_map.items():
                        existing_val = existing_data.get(row_key, '')
                        # Use new value if existing is empty
                        if not existing_val or str(existing_val).strip() in ['', 'nan']:
                            if new_val and str(new_val).strip() not in ['', 'nan']:
                                existing_data[row_key] = new_val
                    all_columns[matching_key]['row_data_map'] = existing_data
                    is_duplicate = True  # Mark as handled (don't add new column)
                    logger.debug(f"Merged data for duplicate header: '{header}'")
                
                if not is_duplicate:
                    all_columns[norm_header] = {
                        'header': header,
                        'level0_header': level0_header_from_meta,  # Main Header (e.g., 'Average Monthly Balance')
                        'level1_header': level1_header,  # Period Type (e.g., 'Three Months Ended')
                        'level2_header': level2_header,  # Years/Dates (e.g., '2024')
                        'row_data_map': row_data_map,
                        'sort_key': sort_key,
                        'source': meta.get('source', ''),
                        'page': meta.get('page', ''),  # Track page number for source per column
                    }
        
        if not all_columns:
            # Don't create empty sheets - just return False
            return False
        
        # Sort columns by date (ASCENDING - chronological order: oldest first)
        # e.g., Dec 2024, Mar 2025, Jun 2025, Sep 2025
        sorted_cols = sorted(
            all_columns.items(),
            key=lambda x: x[1]['sort_key'],
            reverse=False  # Oldest first (chronological)
        )
        
        # Build result dataframe
        # row_labels now contains section_keys, convert back to display labels
        display_row_labels = [
            ExcelUtils.clean_footnote_references(normalized_row_labels.get(key, key)) 
            for key in row_labels
        ]
        result_data = {'Row Label': display_row_labels}
        used_headers = {"row label"}
        
        # Track Level 0, Level 1 and Level 2 headers separately for multi-level output
        # Naming: L0 = "Column Header L1", L1 = "Column Header L2", L2 = "Column Header L3"
        level0_headers = ['']  # First column is Row Label (no L0 header)
        level1_headers = ['']  # First column is Row Label (no L1 header)
        level2_headers = ['Row Label']  # Row Label is the L2/L3 header for first column
        sources_per_column = ['']  # First column is Row Label (no source)
        
        for _, col_info in sorted_cols:
            original_header = col_info['header']
            row_data_map = col_info.get('row_data_map', {})
            level0 = col_info.get('level0_header', '')
            level1 = col_info.get('level1_header', '')
            level2 = col_info.get('level2_header', '')
            source = col_info.get('source', '')
            
            # Keep original header format (no conversion to Qn, YYYY)
            header = original_header
            
            # Track if this header already exists - we'll merge data instead of creating duplicates
            is_duplicate_header = header.lower() in used_headers
            
            # If header exists, we'll merge; if new, add it
            if not is_duplicate_header:
                used_headers.add(header.lower())
            
            # Store Level 0 header (Main Header like "Average Monthly Balance")
            l0_val = ExcelUtils.ensure_string_header(level0) if level0 else ''
            
            # Store Level 1 and Level 2 headers
            # If L1 is the same as L2 (both are just years), set L1 to empty to avoid duplication
            l1_val = ExcelUtils.ensure_string_header(level1) if level1 else ''
            l2_val = ExcelUtils.ensure_string_header(level2) if level2 else header
            
            # 10-K Special Case: If L2 is just a year and source is 10-K, prefix with "YTD, "
            # e.g., "2024" -> "YTD, 2024"
            is_10k = '10k' in source.lower() if source else False
            if l2_val and l2_val.isdigit() and len(l2_val) == 4 and is_10k:
                if not l1_val:  # Only add YTD if L1 is empty (no period type specified)
                    l2_val = f"YTD, {l2_val}"
            
            # If L1 is just a year (same as L2), clear it to avoid duplicate headers
            if l1_val and l2_val and l1_val == l2_val:
                l1_val = ''
            # If L1 is a year and L2 is also a year, L1 should be empty
            if l1_val and l1_val.replace(',', '').strip().isdigit():
                # Check if it looks like just a year (4 digits)
                clean_l1 = l1_val.replace(',', '').strip()
                if len(clean_l1) == 4 and clean_l1.isdigit():
                    l1_val = ''  # Years belong in L2 only
            
            # COMBINE L1 and L2 if both present (e.g., "Three Months Ended June 30," + "2023")
            # This creates complete headers like "Three Months Ended June 30, 2023"
            # NOTE: Period normalization is deferred to Process step
            if l1_val and l2_val and l2_val.isdigit() and len(l2_val) == 4:
                # L1 is period description (Three Months Ended...), L2 is year
                combined = f"{l1_val.rstrip(',')} {l2_val}"
                # Keep raw combined value (Process step will normalize)
                l2_val = combined
                l1_val = ''  # Clear L1 since we've combined
            
            # 10-Q Case: If L2 is just a year and no L1 context, derive quarter from source
            # e.g., 10q0624 -> June 2024 -> Q2-2024, 10q0925 -> Sept 2025 -> Q3-2025
            if l2_val and l2_val.isdigit() and len(l2_val) == 4 and not l1_val:
                is_10q = '10q' in source.lower() if source else False
                if is_10q:
                    # Extract month from source filename (e.g., 10q0624 -> 06 -> June -> Q2)
                    source_lower = source.lower()
                    month_map = {
                        '03': 'Q1', '0325': 'Q1',  # March -> Q1
                        '06': 'Q2', '0624': 'Q2', '0625': 'Q2',  # June -> Q2
                        '09': 'Q3', '0924': 'Q3', '0925': 'Q3',  # Sept -> Q3
                        '12': 'Q4', '1224': 'Q4', '1225': 'Q4',  # Dec -> Q4
                    }
                    derived_quarter = None
                    for pattern, quarter in month_map.items():
                        if pattern in source_lower:
                            derived_quarter = quarter
                            break
                    if derived_quarter:
                        l2_val = f"{derived_quarter}-{l2_val}"
            
            # Only add headers for NEW columns (not duplicates we're merging)
            if not is_duplicate_header:
                level0_headers.append(l0_val)
                level1_headers.append(l1_val)
                level2_headers.append(l2_val)
                # Build source per column: "10q0624_p45" format
                source_str = source.replace('.pdf', '').replace('.xlsx', '') if source else ''
                page_str = str(col_info.get('page', '')).strip()
                if source_str and page_str and page_str not in ['', 'nan']:
                    sources_per_column.append(f"{source_str}_p{page_str}")
                elif source_str:
                    sources_per_column.append(source_str)
                else:
                    sources_per_column.append('')
            
            col_data = []
            # row_labels now contains section_keys (e.g., "Average common equity::institutional securities")
            # We can look them up directly in row_data_map
            for section_key in row_labels:
                # Get the value using the section_key directly
                value = row_data_map.get(section_key, '')
                
                # Clean currency values: '$1,234' -> 1234.0, keep 'N/A' and '-' as strings
                cleaned_value = self._clean_currency_value(value)
                # Also clean year floats (2024.0 -> '2024')
                if isinstance(cleaned_value, str):
                    cleaned_value = ExcelUtils.clean_cell_value(cleaned_value)
                # Ensure numeric values are float (not int) for table data
                elif isinstance(cleaned_value, int):
                    cleaned_value = float(cleaned_value)
                col_data.append(cleaned_value)
            
            # MERGE logic for duplicate headers: update empty cells with new values
            if is_duplicate_header and header in result_data:
                existing_col = result_data[header]
                merged_col = []
                for i, (existing_val, new_val) in enumerate(zip(existing_col, col_data)):
                    # Keep existing value if present, otherwise use new value
                    if existing_val == '' or (isinstance(existing_val, str) and existing_val.lower() == 'nan'):
                        merged_col.append(new_val)
                    else:
                        merged_col.append(existing_val)
                result_data[header] = merged_col
            else:
                result_data[header] = col_data
        
        # Store header info for later cell merging
        # Deduplicate adjacent identical headers for spanning (keep first, empty rest)
        # e.g., ["At Dec 31, 2024", "At Dec 31, 2024", "At Dec 31, 2024"] -> ["At Dec 31, 2024", "", ""]
        
        # Deduplicate Level 0 headers
        deduped_level0 = []
        prev_l0 = None
        for l0 in level0_headers:
            if l0 and l0 == prev_l0:
                deduped_level0.append('')  # Duplicate - empty it for spanning
            else:
                deduped_level0.append(l0)
                prev_l0 = l0 if l0 else prev_l0
        
        # Deduplicate Level 1 headers
        deduped_level1 = []
        prev_l1 = None
        for l1 in level1_headers:
            if l1 and l1 == prev_l1:
                deduped_level1.append('')  # Duplicate - empty it for spanning
            else:
                deduped_level1.append(l1)
                prev_l1 = l1 if l1 else prev_l1
        
        self._last_level0_headers = deduped_level0
        self._last_level1_headers = deduped_level1
        self._last_level2_headers = level2_headers
        
        result_df = pd.DataFrame(result_data)
        
        # Remove duplicate columns AND empty columns (keep first occurrence)
        # 1. Remove columns where entire content is identical to another column
        # 2. Remove columns that are entirely empty (no data values)
        cols_to_keep = ['Row Label']  # Always keep first column
        seen_data = {}
        for col in result_df.columns:
            if col == 'Row Label':
                continue
            col_data = tuple(result_df[col].fillna('').astype(str).tolist())
            
            # Check if column is entirely empty
            is_empty = all(v == '' or v.lower() == 'nan' for v in col_data)
            if is_empty:
                continue  # Skip empty columns
            
            if col_data not in seen_data:
                seen_data[col_data] = col
                cols_to_keep.append(col)
            # else: skip duplicate column
        
        if len(cols_to_keep) < len(result_df.columns):
            result_df = result_df[cols_to_keep]
        
        # Ensure column names are strings (prevent Excel from converting '2024' to 2024.0)
        # Use centralized ExcelUtils.ensure_string_header
        result_df.columns = [ExcelUtils.ensure_string_header(c) for c in result_df.columns]
        
        # Transpose if requested - creates multi-level column headers
        if transpose and len(result_df.columns) > 1:
            # Debug logging removed
            # First, identify category (section) rows and their associated line items
            # Category rows have text in first column but rest is empty
            category_to_items = {}  # Maps category -> list of line items under it
            current_category = "General"  # Default category if none found
            
            for row_label in row_labels:
                if row_label:
                    label_str = str(row_label).strip()
                    
                    # Check if this row is a category header (appears in section rows)
                    # Category rows are typically empty in data columns
                    is_category = False
                    for table_info in tables:
                        df_check = table_info['data']
                        if len(df_check) > 3:
                            for idx in range(3, len(df_check)):
                                row = df_check.iloc[idx]
                                first_cell = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
                                if first_cell == label_str:
                                    rest_empty = all(pd.isna(v) or str(v).strip() in ['', 'nan'] 
                                                    for v in row.iloc[1:])
                                    if rest_empty:
                                        is_category = True
                                        break
                            break
                    
                    if is_category:
                        current_category = label_str
                        if current_category not in category_to_items:
                            category_to_items[current_category] = []
                    else:
                        # This is a line item under the current category
                        if current_category not in category_to_items:
                            category_to_items[current_category] = []
                        category_to_items[current_category].append(label_str)
            
            # Transpose the DataFrame
            result_df = result_df.set_index('Row Label')
            result_df = result_df.T
            result_df.index.name = 'Dates'
            result_df = result_df.reset_index()
            
            # Create multi-level column headers
            # Build (Category, LineItem) tuples for each column
            # Use title() for proper capitalization
            new_columns = []
            for col in result_df.columns:
                col_str = str(col)
                if col_str == 'Dates':
                    new_columns.append(('', 'Dates'))  # First column
                else:
                    # Find which category this line item belongs to
                    found_category = 'General'
                    for category, items in category_to_items.items():
                        if col_str in items or ExcelUtils.clean_footnote_references(col_str) in items:
                            found_category = category
                            break
                    # Apply title() for proper capitalization
                    category_display = found_category.title() if found_category else ''
                    line_item_display = ExcelUtils.clean_footnote_references(col_str).title()
                    new_columns.append((category_display, line_item_display))
            
            # Create MultiIndex columns
            result_df.columns = pd.MultiIndex.from_tuples(new_columns, names=['Category', 'Line Item'])
            
            # Convert Dates column values to QnYYYY format for traceability
            # e.g., 'Three Months Ended March 31, 2024' -> '3QTD2024'
            dates_col = ('', 'Dates')
            # Check if column exists without triggering lexsort warning
            col_list = list(result_df.columns)
            if dates_col in col_list:
                col_idx = col_list.index(dates_col)
                for row_idx in range(len(result_df)):
                    val = result_df.iloc[row_idx, col_idx]
                    if val is not None and str(val).strip():
                        result_df.iloc[row_idx, col_idx] = MetadataBuilder.convert_to_qn_format(str(val))
        
        # Build metadata rows (matching individual file structure)
        # Collect metadata from all sources
        all_sources = sorted(set(t['metadata'].get('source', '') for t in tables if t['metadata'].get('source')))
        all_years = sorted(set(str(t['metadata'].get('year', '')) for t in tables if t['metadata'].get('year')), reverse=True)
        all_quarters = sorted(set(str(t['metadata'].get('quarter', '')) for t in tables if t['metadata'].get('quarter')))
        # Collect Main Headers from source tables (Level 0 spanning headers like "Average Monthly Balance")
        all_main_headers = sorted(set(
            str(t['metadata'].get('main_header', '')).strip() 
            for t in tables 
            if t['metadata'].get('main_header') and str(t['metadata'].get('main_header')).strip()
        ))
        original_title = tables[0].get('original_title', title) if tables else title
        section = tables[0].get('section', '') if tables else ''
        
        # Row Header Level 1: Section rows where first cell has text but rest is empty
        # These are rows like "Revenues" where only the first cell has content
        # Collect RAW values first, clean for display later
        row_headers_l1_raw = []
        for table_info in tables:
            df = table_info['data']
            if len(df) > 3:  # Skip if too few rows
                for idx in range(3, len(df)):  # Start after header rows
                    row = df.iloc[idx]
                    first_cell = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
                    # Check if rest of row is empty
                    rest_empty = all(pd.isna(v) or str(v).strip() == '' or str(v).strip() == 'nan' 
                                    for v in row.iloc[1:])
                    if first_cell and rest_empty and first_cell not in ['Row Label', 'nan']:
                        if first_cell not in row_headers_l1_raw:
                            row_headers_l1_raw.append(first_cell)
            break  # Only check first table since they should have same structure
        
        # Row Header Level 2: All data row labels (non-section rows) - keep raw for matching
        row_headers_l2_raw = []
        for label in row_labels:
            if label and not str(label).startswith('$'):
                label_str = str(label).strip()
                # Skip if it's a section header (in L1)
                if label_str not in row_headers_l1_raw and label_str not in row_headers_l2_raw:
                    row_headers_l2_raw.append(label_str)
        
        # NOW apply footnote cleaning ONLY for display purposes
        row_headers_l1 = [ExcelUtils.clean_footnote_references(h) for h in row_headers_l1_raw if h]
        row_headers_l2 = [ExcelUtils.clean_footnote_references(h) for h in row_headers_l2_raw if h]
        
        # Product/Entity: Unique entities from L2 row headers (deduplicated, first 5), cleaned for display
        products = row_headers_l2[:5]
        
        # Column Headers - from the stored level1_headers and level2_headers
        col_headers_l1_display = []  # Spanning headers like "Three Months Ended..."
        col_headers_l2_display = []  # Sub-headers like "December 31, 2024", "2024"
        
        if hasattr(self, '_last_level1_headers') and self._last_level1_headers:
            for h in self._last_level1_headers[1:]:  # Skip first (Row Label column)
                if h and str(h).strip() and str(h).strip() not in ['Row Label', '$ in millions']:
                    h_clean = str(h).strip()
                    if h_clean not in col_headers_l1_display:
                        col_headers_l1_display.append(h_clean)
        
        if hasattr(self, '_last_level2_headers') and self._last_level2_headers:
            for h in self._last_level2_headers[1:]:  # Skip first (Row Label column)
                if h and str(h).strip() and str(h).strip() not in ['Row Label', '$ in millions']:
                    h_clean = str(h).strip()
                    # Clean "At March 31, 2025" -> "March 31, 2025"
                    h_clean = re.sub(r'^At\s+', '', h_clean)
                    if h_clean not in col_headers_l2_display:
                        col_headers_l2_display.append(h_clean)
        
        # === BUILD METADATA ROWS (Column A only) ===
        # Rows 1-12: Simple key-value metadata in column A
        # Row 13+: Actual table with headers (L0, L1, L2) then data
        
        def _convert_header_value(val):
            """Convert header value to clean string (period normalization deferred to Process step)."""
            if val is None or val == '':
                return ''
            clean_val = ExcelUtils.ensure_string_header(val)
            
            # NOTE: Period normalization (e.g., "Three Months Ended" -> "Q2-QTD-2024")
            # is now done in Process step, not here
            
            # Handle multi-period headers separated by ",, " - just clean them up
            # (e.g., "Three Months Ended June 30,, Six Months Ended June 30,")
            if ',, ' in clean_val or (clean_val.count('Months Ended') > 1):
                # Split on ",, " or "Ended [Date]," pattern
                import re
                parts = re.split(r',\s*,\s*', clean_val)
                if len(parts) == 1:
                    # Try splitting on repeated period patterns
                    parts = re.split(r'(?<=\d{4}),\s*(?=[A-Z])', clean_val)
                
                cleaned_parts = []
                for part in parts:
                    part = part.strip().rstrip(',')
                    if not part:
                        continue
                    cleaned_parts.append(part)
                
                # Remove duplicates while preserving order
                seen = set()
                unique_parts = []
                for p in cleaned_parts:
                    if p not in seen:
                        seen.add(p)
                        unique_parts.append(p)
                return ', '.join(unique_parts) if unique_parts else clean_val
            
            return clean_val
        
        # Prepare column header lists with converted values
        l0_headers = [_convert_header_value(h) for h in self._last_level0_headers]
        l1_headers = [_convert_header_value(h) for h in self._last_level1_headers]
        l2_headers = [_convert_header_value(h) for h in self._last_level2_headers]
        
        # Ensure header lists match column count (pad or trim)
        col_count = len(result_df.columns)
        
        def _ensure_length(headers, target_len):
            """Ensure headers list has exactly target_len elements."""
            if len(headers) < target_len:
                return headers + [''] * (target_len - len(headers))
            elif len(headers) > target_len:
                return headers[:target_len]
            return headers
        
        l0_headers = _ensure_length(l0_headers, col_count)
        l1_headers = _ensure_length(l1_headers, col_count)
        l2_headers = _ensure_length(l2_headers, col_count)
        
        # Build table title
        display_title = f"{section} - {original_title}" if section else original_title
        display_title = ExcelUtils.fix_ocr_broken_words(display_title)
        display_title = re.sub(r'\s+', ' ', display_title).strip()
        
        # Create TableMetadata container for MetadataBuilder
        # Ensure sources_per_column matches column count
        sources_per_col_padded = _ensure_length(sources_per_column, col_count)
        
        table_metadata = TableMetadata(
            category_parent=row_headers_l1[:5] if row_headers_l1 else [],
            line_items=row_headers_l2[:10] if row_headers_l2 else [],
            product_entity=products[:5] if products else [],
            column_header_l1=l0_headers,
            column_header_l2=l1_headers,
            column_header_l3=l2_headers,
            sources_per_column=sources_per_col_padded,  # Per-column source tracking
            table_title=display_title,
            sources=all_sources,
            section=''  # Already included in display_title
        )
        
        # Build metadata using MetadataBuilder (12 rows with summaries in column A)
        columns = list(result_df.columns)
        first_col = columns[0] if columns else 'Row Label'
        metadata_df = MetadataBuilder.build_metadata_dataframe(table_metadata, columns, first_col)
        
        # Build table header rows (L0, L1, L2 - spread across columns)
        # These will be part of the actual table, not metadata
        has_l0 = any(h for h in l0_headers[1:] if h)  # Skip first col
        has_l1 = any(h for h in l1_headers[1:] if h)  # Skip first col
        
        header_frames = []
        if has_l0:
            level0_df = pd.DataFrame([l0_headers], columns=columns, dtype=str)
            header_frames.append(level0_df)
        if has_l1:
            level1_df = pd.DataFrame([l1_headers], columns=columns, dtype=str)
            header_frames.append(level1_df)
        # L2 (years/dates) is always present as table header
        level2_df = pd.DataFrame([l2_headers], columns=columns, dtype=str)
        header_frames.append(level2_df)
        
        # Source per Column row: Shows source file and page for each column
        # Only add if there are actual sources per column
        has_sources_per_col = any(s for s in sources_per_col_padded[1:] if s)
        if has_sources_per_col:
            # First cell is label, rest are per-column sources
            sources_row_data = sources_per_col_padded.copy()
            sources_row_data[0] = 'Source:'
            sources_df = pd.DataFrame([sources_row_data], columns=columns, dtype=str)
            header_frames.append(sources_df)
        
        
        # === CONCATENATE ALL FRAMES ===
        # For TRANSPOSED output: Keep metadata, but transpose the data portion
        # For REGULAR output: metadata (12 rows) + table headers (L0/L1/L2) + data
        if transpose:
            # Transposed output: Flatten MultiIndex columns
            # Keep category prefix ONLY if it's a real category (not 'General')
            if isinstance(result_df.columns, pd.MultiIndex):
                flat_cols = []
                for col in result_df.columns:
                    if isinstance(col, tuple):
                        cat = col[0] if len(col) > 0 else ''
                        line_item = col[1] if len(col) > 1 else str(col[0]) if len(col) > 0 else str(col)
                        # Only add category prefix if it's a real category (not General)
                        if cat and cat.strip() and cat.strip().lower() != 'general':
                            flat_cols.append(f"{cat} - {line_item}")
                        else:
                            flat_cols.append(str(line_item))
                    else:
                        flat_cols.append(str(col))
                result_df.columns = flat_cols
            
            # Build transposed output directly:
            # Use same number of columns as transposed data
            n_cols = len(result_df.columns)
            
            # Create metadata rows with padding
            metadata_rows = []
            metadata_items = [
                f"Category (Parent): {', '.join(row_headers_l1[:3]) if row_headers_l1 else ''}",
                f"Line Items: {', '.join(row_headers_l2[:5]) if row_headers_l2 else ''}",
                f"Product/Entity: {', '.join(products[:3]) if products else ''}",
                f"Column Header L1: {', '.join(col_headers_l1_display[:3]) if col_headers_l1_display else ''}",
                f"Column Header L2: {', '.join(col_headers_l2_display[:3]) if col_headers_l2_display else ''}",  
                f"Column Header L3: {', '.join(list(result_df.columns)[:5])}",
                f"Year/Quarter: {', '.join(all_quarters)}",
                "",
                f"Table Title: {display_title} (Transposed)",
                f"Source(s): {', '.join(all_sources)}",
                "",
                "",
            ]
            
            for item in metadata_items:
                row = [item] + [''] * (n_cols - 1)
                metadata_rows.append(row)
            
            # Create metadata DataFrame
            col_names = list(result_df.columns)
            metadata_df_trans = pd.DataFrame(metadata_rows, columns=col_names)
            
            # Add header row (column names as data)
            header_row = pd.DataFrame([col_names], columns=col_names)
            
            # Stack: metadata + header + transposed data
            final_df = pd.concat([metadata_df_trans, header_row, result_df], ignore_index=True)
            
            # Write directly
            final_df.to_excel(writer, sheet_name=title, index=False, header=False)
            return True
        
        # REGULAR output: metadata + header rows + data
        frames_to_concat = [metadata_df] + header_frames + [result_df]
        
        # Ensure all frames have compatible column structures before concat
        # This fixes "cannot join with no overlapping index names" errors
        normalized_frames = []
        for i, frame in enumerate(frames_to_concat):
            # Reset index to avoid index join issues
            frame = frame.reset_index(drop=True)
            # Flatten MultiIndex columns if present
            if isinstance(frame.columns, pd.MultiIndex):
                frame.columns = [' '.join(str(c) for c in col).strip() if isinstance(col, tuple) else str(col) for col in frame.columns]
            # Ensure columns are simple strings
            frame.columns = [str(c) for c in frame.columns]
            normalized_frames.append(frame)
        
        # Now concatenate with matching column names
        try:
            final_df = pd.concat(normalized_frames, ignore_index=True)
        except (ValueError, pd.errors.InvalidIndexError) as e:
            # Fallback: align columns by position if names don't match or have duplicates
            logger.warning(f"Column issue during concat, using positional alignment: {e}")
            # Deduplicate column names first
            aligned_frames = []
            for frame in normalized_frames:
                # Create unique column names by appending index for duplicates
                seen = {}
                new_cols = []
                for col in frame.columns:
                    col_str = str(col)
                    if col_str in seen:
                        seen[col_str] += 1
                        new_cols.append(f"{col_str}_{seen[col_str]}")
                    else:
                        seen[col_str] = 0
                        new_cols.append(col_str)
                frame.columns = new_cols
                aligned_frames.append(frame)
            
            try:
                final_df = pd.concat(aligned_frames, ignore_index=True)
            except Exception as e2:
                # Last resort: use positional indexing
                logger.warning(f"Fallback to positional: {e2}")
                for i, frame in enumerate(aligned_frames):
                    frame.columns = list(range(len(frame.columns)))
                final_df = pd.concat(aligned_frames, ignore_index=True)
                final_df.columns = list(range(len(final_df.columns)))
        
        # Track header row count for cell merging later
        header_row_count = 12 + len(header_frames)
        
        # Fix header rows: Convert any float years (e.g., 2025.0) back to strings
        # Extend range to cover Row Label row which may be at row 13-15
        fix_up_to_row = min(20, len(final_df))  # Cover all possible header rows
        for row_idx in range(fix_up_to_row):
            for col_idx in range(len(final_df.columns)):
                val = final_df.iloc[row_idx, col_idx]
                if isinstance(val, float) and not pd.isna(val):
                    if val == int(val):
                        final_df.iloc[row_idx, col_idx] = str(int(val))

        
        # Write without pandas headers (we've included them in the data)
        final_df.to_excel(writer, sheet_name=title, index=False, header=False)
        return True  # Sheet has data
    
    def _add_hyperlinks(
        self,
        output_path: Path,
        all_tables_by_full_title: Dict[str, List],
        title_to_sheet_name: Dict[str, dict]
    ) -> None:
        """
        Add hyperlinks to consolidated workbook.
        
        Delegates to ExcelFormatter for centralized logic.
        """
        ExcelFormatter.add_hyperlinks(output_path, all_tables_by_full_title, title_to_sheet_name)
    
    def _apply_currency_format(
        self,
        output_path: Path,
        all_tables_by_full_title: Dict[str, List],
        title_to_sheet_name: Dict[str, dict]
    ) -> None:
        """
        Apply US currency number format to numeric data cells.
        
        Delegates to ExcelFormatter for centralized logic.
        """
        ExcelFormatter.apply_currency_format(output_path, all_tables_by_full_title, title_to_sheet_name)


# =============================================================================
# SINGLETON PATTERN
# =============================================================================

_consolidated_exporter: Optional[ConsolidatedExcelExporter] = None


def get_consolidated_exporter() -> ConsolidatedExcelExporter:
    """
    Get or create global ConsolidatedExcelExporter instance.
    
    Returns:
        ConsolidatedExcelExporter singleton instance
    """
    global _consolidated_exporter
    
    if _consolidated_exporter is None:
        _consolidated_exporter = ConsolidatedExcelExporter()
    
    return _consolidated_exporter


def reset_consolidated_exporter() -> None:
    """Reset the exporter singleton (for testing)."""
    global _consolidated_exporter
    _consolidated_exporter = None
