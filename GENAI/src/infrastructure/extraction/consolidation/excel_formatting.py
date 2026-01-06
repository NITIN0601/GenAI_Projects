"""
Excel Formatting - Apply formatting to consolidated Excel workbooks.

Standalone module for Excel formatting operations.
Used by: consolidated_exporter.py
"""

import re
from pathlib import Path
from typing import Dict, List, Any

from src.utils import get_logger
from src.utils.excel_utils import ExcelUtils

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE

logger = get_logger(__name__)


class ExcelFormatter:
    """
    Apply formatting to Excel workbooks.
    
    Handles:
    - Hyperlinks (Index ↔ Data sheets)
    - Currency and percentage formatting
    - Header cell merging
    - Year value cleanup
    """
    
    # Currency indicators in row labels - these should NEVER be percentages
    CURRENCY_INDICATORS = [
        'per share', 'eps', 'dividend', 'price', 'book value',
        'tangible book', 'revenue', 'income', 'expense', 'cost',
        'asset', 'liability', 'equity', 'cash', 'debt', 'loan',
        'deposit', 'fee', 'commission', 'compensation', 'salary',
        '$ in', 'in millions', 'in billions', 'in thousands'
    ]
    
    # Percentage indicators - these are definitely percentages
    PERCENTAGE_INDICATORS = [
        'ratio', 'roe', 'rotce', 'roa', 'margin', 'rate', 'yield',
        'efficiency', 'leverage', 'tier 1', 'tier 2', 'cet1', 
        'percentage', 'percent', '%', 'return on'
    ]
    
    @classmethod
    def add_hyperlinks(
        cls,
        output_path: Path,
        all_tables_by_full_title: Dict[str, List],
        title_to_sheet_name: Dict[str, dict]
    ) -> None:
        """Add hyperlinks to consolidated workbook."""
        try:
            wb = load_workbook(output_path)
            
            # Add hyperlinks in Index sheet
            if 'Index' in wb.sheetnames:
                index_sheet = wb['Index']
                
                # Find Link column dynamically
                header_row = [index_sheet.cell(row=1, column=c).value for c in range(1, index_sheet.max_column + 1)]
                link_col = None
                for i, h in enumerate(header_row, start=1):
                    if isinstance(h, str) and h.strip().lower() == 'link':
                        link_col = i
                        break
                
                if link_col is None:
                    logger.warning("Link column not found in consolidated Index sheet")
                else:
                    for row in range(2, index_sheet.max_row + 1):  # Skip header
                        cell = index_sheet.cell(row=row, column=link_col)
                        raw_value = cell.value
                        
                        # Extract sheet name (handle → prefix if present)
                        sheet_name = None
                        if raw_value is not None:
                            raw_str = str(raw_value).strip()
                            if raw_str.startswith('→'):
                                sheet_name = raw_str.lstrip('→').strip()
                            else:
                                sheet_name = raw_str
                        
                        if sheet_name and sheet_name in wb.sheetnames:
                            cell.hyperlink = f"#'{sheet_name}'!A1"
                            cell.value = f"→ {sheet_name}"
                            cell.font = Font(color="0000FF", underline="single")
                        elif sheet_name:
                            logger.debug(f"Sheet '{sheet_name}' not found in consolidated workbook")
            
            # Add back-links in each data sheet
            for sheet_name in wb.sheetnames:
                if sheet_name == 'Index':
                    continue
                ws = wb[sheet_name]
                cell = ws.cell(row=1, column=1)
                # Set back-link regardless of current value
                cell.value = "← Back to Index"
                cell.hyperlink = "#'Index'!A1"
                cell.font = Font(color="0000FF", underline="single")
            
            wb.save(output_path)
            logger.info(f"Added hyperlinks to consolidated workbook: {len(wb.sheetnames) - 1} sheets")
            
        except Exception as e:
            logger.warning(f"Could not add hyperlinks: {e}")
    
    @classmethod
    def apply_currency_format(
        cls,
        output_path: Path,
        all_tables_by_full_title: Dict[str, List],
        title_to_sheet_name: Dict[str, dict]
    ) -> None:
        """Apply US currency number format to numeric data cells."""
        try:
            wb = load_workbook(output_path)
            
            # US currency accounting format: $#,##0.00 for positive, ($#,##0.00) for negative
            CURRENCY_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
            PERCENTAGE_FORMAT = '0.00%'
            
            for sheet_name in wb.sheetnames:
                if sheet_name in ['Index', 'TOC']:
                    continue
                
                ws = wb[sheet_name]
                
                # Dynamically find header and data rows by looking for "Row Label" marker
                row_label_row = None
                for r in range(1, min(20, ws.max_row + 1)):
                    cell_val = ws.cell(row=r, column=1).value
                    if cell_val and str(cell_val).strip() == 'Row Label':
                        row_label_row = r
                        break
                
                if row_label_row is None:
                    row_label_row = 13  # Fallback to expected position
                
                # Dynamically detect header rows
                header_rows = cls._detect_header_rows(ws, row_label_row)
                data_start_row = row_label_row + 1
                
                # Fix header rows - convert year floats to int
                cls._fix_header_years(ws, header_rows)
                
                # Merge cells for spanning headers
                cls._merge_spanning_headers(ws, header_rows)
                
                # Apply currency/percentage format using HYBRID detection
                cls._apply_number_formats(ws, data_start_row, CURRENCY_FORMAT, PERCENTAGE_FORMAT)
            
            wb.save(output_path)
            logger.info(f"Applied currency/percentage formatting to consolidated workbook")
            
        except Exception as e:
            logger.warning(f"Could not apply currency formatting: {e}")
    
    @classmethod
    def _detect_header_rows(cls, ws, row_label_row: int) -> List[int]:
        """Detect header rows by scanning backwards from Row Label row."""
        header_rows = []
        for r in range(row_label_row - 1, max(0, row_label_row - 4), -1):
            if r < 1:
                continue
            cell_val = ws.cell(row=r, column=2).value
            if cell_val and str(cell_val).strip():
                header_rows.insert(0, r)
            else:
                # Check other columns - might be a spanning header
                has_content = False
                for c in range(2, min(10, ws.max_column + 1)):
                    if ws.cell(row=r, column=c).value:
                        has_content = True
                        break
                if has_content:
                    header_rows.insert(0, r)
        return header_rows
    
    @classmethod
    def _fix_header_years(cls, ws, header_rows: List[int]) -> None:
        """Convert year floats to int in header rows."""
        for header_row in header_rows:
            if header_row > 0:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=header_row, column=col)
                    if cell.value is not None:
                        if isinstance(cell.value, float) and cell.value == int(cell.value):
                            cell.value = int(cell.value)
                        elif isinstance(cell.value, str):
                            cleaned = ExcelUtils.clean_year_string(cell.value)
                            if cleaned.isdigit() and len(cleaned) == 4:
                                year = int(cleaned)
                                if 2000 <= year <= 2099:
                                    cell.value = year
    
    @classmethod
    def _merge_spanning_headers(cls, ws, header_rows: List[int]) -> None:
        """Merge cells for spanning headers in header rows."""
        
        for header_row in header_rows:
            if header_row < 1:
                continue
            col = 2  # Start from column B
            while col <= ws.max_column:
                cell_value = ws.cell(row=header_row, column=col).value
                if cell_value and str(cell_value).strip():
                    # Find how many consecutive columns have same/empty value
                    span_end = col
                    for next_col in range(col + 1, ws.max_column + 1):
                        next_value = ws.cell(row=header_row, column=next_col).value
                        if not next_value or str(next_value).strip() == '' or str(next_value).strip() == str(cell_value).strip():
                            span_end = next_col
                            if next_value and str(next_value).strip() == str(cell_value).strip():
                                ws.cell(row=header_row, column=next_col).value = ''
                        else:
                            break
                    
                    # Merge if span is more than 1 column
                    if span_end > col:
                        ws.merge_cells(start_row=header_row, start_column=col, end_row=header_row, end_column=span_end)
                        ws.cell(row=header_row, column=col).alignment = Alignment(horizontal='center')
                    
                    col = span_end + 1
                else:
                    col += 1
    
    @classmethod
    def _apply_number_formats(cls, ws, data_start_row: int, currency_fmt: str, pct_fmt: str) -> None:
        """Apply currency or percentage format to data cells."""
        for row in range(data_start_row, ws.max_row + 1):
            row_label = str(ws.cell(row=row, column=1).value or '').lower()
            
            # Collect all numeric values in this row
            row_values = []
            for col in range(2, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)) and cell.value is not None:
                    row_values.append(cell.value)
            
            # HYBRID DETECTION: Percentage first, then currency
            is_pct_label = any(ind in row_label for ind in cls.PERCENTAGE_INDICATORS)
            is_currency_label = any(ind in row_label for ind in cls.CURRENCY_INDICATORS)
            
            if is_pct_label:
                is_percentage_row = True
            elif is_currency_label:
                is_percentage_row = False
            else:
                # Value-based detection for unlabeled rows
                is_percentage_row = False
                if row_values:
                    non_zero_values = [v for v in row_values if v != 0]
                    if non_zero_values:
                        all_in_pct_range = all(-1 <= v <= 1 for v in non_zero_values)
                        has_decimal = any(0 < abs(v) < 1 for v in non_zero_values)
                        is_percentage_row = all_in_pct_range and has_decimal
            
            # Apply appropriate format
            for col in range(2, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)) and cell.value is not None:
                    if is_percentage_row:
                        cell.number_format = pct_fmt
                    else:
                        cell.number_format = currency_fmt
